from fastapi import FastAPI,APIRouter, Depends, HTTPException, Request
from sqlalchemy.orm import Session
from pydantic import BaseModel
import copy
import time
from app.utils.logging_utils import (
    auth_logger, 
    FittbotHTTPException, 
    SecuritySeverity,
    EventType,
    log_exceptions
)
from app.models.marketingmodels import GymDatabase
from app.models.fittbot_models import GymJoinRequest,GymDetails,FittbotGymMembership,GymBusinessPayment,TrainerProfile,GymFees,ClientFittbotAccess,ClientNextXp,GymOwner,BlockedUsers,ClientGym,OldGymData,Gym_Feedback,Client,Attendance,TemplateDiet,FeeHistory,Expenditure,ClientScheduler,DietTemplate,WorkoutTemplate,ClientActual,GymHourlyAgg,DailyGymHourlyAgg,Gym,GymAnalysis,GymMonthlyData,GymPlans,GymBatches,Trainer,TemplateWorkout,Post,Comment,Like,Feedback,Message,RewardGym,LeaderboardOverall,LeaderboardDaily,RewardQuest,LeaderboardMonthly,RewardPrizeHistory,GymLocation,ClientTarget,Report,PostMedia,AboutToExpire, AccountDetails,GymEnquiry,Brochures,FeesReceipt, GymImportData, EnquiryEstimates, ManualClient
from app.models.fittbot_payments_models import Payment as FittbotPayment, Payout
from app.fittbot_api.v1.payments.models.entitlements import Entitlement
from app.fittbot_api.v1.payments.models.orders import OrderItem
#from models import Brochures,ClientNextXp,GymOwner,BlockedUsers,ClientGym,OldGymData,Gym_Feedback,Client,Attendance,TemplateDiet,FeeHistory,Expenditure,ClientScheduler,DietTemplate,WorkoutTemplate,ClientActual,GymHourlyAgg,Gym,GymAnalysis,GymMonthlyData,GymPlans,GymBatches,Trainer,TemplateWorkout,Post,Comment,Like,Feedback,Message,RewardGym,LeaderboardOverall,LeaderboardDaily,RewardQuest,LeaderboardMonthly,RewardPrizeHistory,GymLocation,ClientTarget,Report,PostMedia,AboutToExpire, AccountDetails,GymEnquiry,FeesReceipt, GymImportData
from app.utils.hashing import verify_password
from app.models.database import get_db
from datetime import datetime, timedelta
from app.utils.redis_config import get_redis
from redis.asyncio import Redis
from sqlalchemy.sql import extract,func
import json
from fastapi import Query

from datetime import datetime, date
from typing import Dict,List,Any,Optional,Tuple
import pytz
from typing import Optional
from sqlalchemy.future import select
from sqlalchemy import desc
import requests
from fastapi import FastAPI, UploadFile, File, Form
from dotenv import load_dotenv
from app.utils.security import (
    verify_password, create_access_token, create_refresh_token,
    refresh_tokens_store, get_password_hash, SECRET_KEY, ALGORITHM
)
from app.utils.request_auth import authenticate_identity
from jose import jwt, JWTError
from sqlalchemy import or_,and_,asc,desc
import random
import string
import os
import traceback
from email.message import EmailMessage
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
from enum import Enum
import uuid
#from utils.aes_encryption import decrypt_uuid

import csv
import re
from io import BytesIO, StringIO
from openpyxl import load_workbook
import boto3
import calendar
from app.utils.aws_retry import invoke_lambda_with_retry



load_dotenv()
app = FastAPI()

router = APIRouter(prefix="/owner", tags=["Gymowner"])

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

MEDIA_CACHE_TTL_SECONDS = 3600
EMPTY_MEDIA_CACHE_TTL_SECONDS = 60


def resolve_post_actor(
    request: Request,
    db: Session,
    *,
    gym_id: int,
    declared_role: Optional[str],
    provided_client_id: Optional[int],
) -> Tuple[int, str]:
    normalized_role = (declared_role or "").lower() if declared_role else None
    actor_id, token_role = authenticate_identity(
        request,
        declared_role=normalized_role,
        provided_user_id=provided_client_id if normalized_role == "client" else None,
    )

    if token_role == "owner":
        gym_exists = (
            db.query(Gym.gym_id)
            .filter(Gym.gym_id == gym_id, Gym.owner_id == actor_id)
            .first()
        )
        if not gym_exists:
            raise HTTPException(status_code=403, detail="Owner not associated with this gym")

    return actor_id, token_role


def enforce_post_mutation_permissions(
    *,
    token_role: str,
    actor_id: int,
    post: Post,
) -> None:
    if token_role == "client":
        if post.client_id != actor_id:
            raise HTTPException(status_code=403, detail="Clients can only modify their own posts")
        return

    if token_role == "owner":
        return

    raise HTTPException(status_code=403, detail="Unsupported role for this operation")
lambda_client = boto3.client(
    "lambda",
    aws_access_key_id=os.getenv("AWS_ACCESS_KEY_ID"),
    aws_secret_access_key=os.getenv("AWS_SECRET_ACCESS_KEY"),
    aws_session_token=None,       
    region_name="ap-south-2"
)

LAMBDA_FUNCTION_NAME = "receipt_mail" 

class GymHomeRequest(BaseModel):
    gym_id: int

@router.post("/gym/home")
async def get_gym_home_data(req: GymHomeRequest, db: Session = Depends(get_db), redis: Redis = Depends(get_redis)):
    start_time = time.time()
    
    try:

        response = {}

        today = datetime.now().date()
        attendance_key = f"gym:{req.gym_id}:attendance:{today.strftime('%Y-%m-%d')}"
        attendance_data = await redis.hgetall(attendance_key)

        if not attendance_data:
            attended_clients = db.query(
                Attendance.client_id, 
                Attendance.in_time, 
                Attendance.out_time, 
                Client.name
            ).join(Client, Attendance.client_id == Client.client_id).filter(
                Attendance.date == today,
                Client.gym_id == req.gym_id
            ).all()

            attendance_details = [
                {
                    "client_id": client.client_id,
                    "name": client.name,
                    "in_time": client.in_time.strftime('%H:%M') if client.in_time else "",
                    "out_time": client.out_time.strftime('%H:%M') if client.out_time else ""
                }
                for client in attended_clients
            ]
            
            current_count = len(attended_clients)
            expected_count = db.query(Client).filter(
                Client.gym_id == req.gym_id,
                Client.status == 'active'
            ).count()

            await redis.hset(
                attendance_key,
                mapping={
                    "current_count": current_count,
                    "expected_count": expected_count,
                    "details": ",".join(
                        f"{detail['client_id']}|{detail['name']}|{detail['in_time']}|{detail['out_time']}" 
                        for detail in attendance_details
                    )
                }
            )
            await redis.expire(attendance_key, 86400)  

        else:
            print("Data fetched from Redis.")
            if not attendance_data.get("details"):
                print("No attendance details found in the data.")
                attendance_details = []
                current_count = int(attendance_data.get("current_count", 0))
                expected_count = int(attendance_data.get("expected_count", 0))
            else:
                try:
                    attendance_details = [
                        {
                            "name": entry.split("|")[1],  
                            "in_time": entry.split("|")[2],
                            "out_time": entry.split("|")[3]
                        }
                        for entry in attendance_data["details"].split(",")
                    ]
                    current_count = int(attendance_data["current_count"])
                    expected_count = int(attendance_data["expected_count"])
                except Exception as e:
                    print(f"Error parsing attendance details: {e}")
                    attendance_details = []
                    current_count = int(attendance_data.get("current_count", 0))
                    expected_count = int(attendance_data.get("expected_count", 0))


        attendance_response = {
            "current_count": current_count,
            "expected_count": expected_count,
            "details": attendance_details  
        }
        response["attendance"] = attendance_response

        members_key = f"gym:{req.gym_id}:members"
        members_data = await redis.hgetall(members_key)

        if not members_data:
            print("No members data in Redis, fetching from database.")
            total_members = db.query(Client).filter(Client.gym_id == req.gym_id).count()
            active_members = db.query(Client).filter(
                Client.gym_id == req.gym_id, Client.status == "active"
            ).count()

            current_month = datetime.now().month
            last_month = (datetime.now() - timedelta(days=30)).month

            last_month_count = db.query(FeeHistory).filter(
                FeeHistory.gym_id == req.gym_id,
                extract("month", FeeHistory.payment_date) == last_month
            ).count()

            current_month_count = db.query(FeeHistory).filter(
                FeeHistory.gym_id == req.gym_id,
                extract("month", FeeHistory.payment_date) == current_month
            ).count()
            access_true_count = db.query(Client).filter(
                Client.gym_id == req.gym_id, Client.access == True
            ).count()

            retention_rate = (current_month_count / last_month_count * 100) if last_month_count > 0 else 0
            unpaid_count = total_members - active_members

            attendance_records = db.query(Attendance).filter(
                Attendance.gym_id == req.gym_id
            ).all()

            total_visits = len(attendance_records)
            unique_days = len(set(record.date for record in attendance_records))
            average_visits = int(total_visits / unique_days if unique_days > 0 else 0)

            average_age = int(db.query(func.avg(Client.age)).filter(
                Client.gym_id == req.gym_id
            ).scalar() or 0)


            await redis.hset(
                members_key,
                mapping={
                    "total_members": int(total_members),  
                    "active_members": int(active_members),  
                    "retention_rate": float(retention_rate),  
                    "unpaid_count": int(unpaid_count),  
                    "average_visits": float(average_visits),  
                    "average_age": int(average_age),  
                    "access_count":int(access_true_count)
                },
            )

            await redis.expire(members_key, 86400)

            members_data = {
                "total_members": total_members,
                "active_members": active_members,
                "retention_rate": retention_rate,
                "unpaid_count": unpaid_count,
                "average_visits": average_visits,
                "average_age": average_age,
                "access_count":int(access_true_count)

            }
        else:
            print("Members data fetched from Redis.")
            members_data = {
                "total_members": int(members_data["total_members"]),
                "active_members": int(members_data["active_members"]),
                "retention_rate": float(members_data["retention_rate"]),
                "unpaid_count": int(members_data["unpaid_count"]),
                "average_visits": float(members_data["average_visits"]),
                "average_age": int(members_data["average_age"]),
                "access_count":int(members_data["access_count"])
            }

        response["members"] = members_data

        invoice_key = f'gym{req.gym_id}:invoice_data'
        invoice_data = await redis.get(invoice_key)

        if not invoice_data:
            sent_invoices = db.query(AboutToExpire).filter(AboutToExpire.gym_id == req.gym_id, AboutToExpire.mail_status == True, AboutToExpire.expired==False).all()
            unsent_invoices = db.query(AboutToExpire).filter(AboutToExpire.gym_id == req.gym_id, AboutToExpire.mail_status == False, AboutToExpire.expired==False).all()

            invoice_data ={}
            
            sent_data=[{
                "expiry_id" : invoice.expiry_id,
                "client_id" : invoice.client_id,
                "gym_id" : invoice.gym_id,
                "client_name" : invoice.client_name,
                "gym_name": invoice.gym_name,
                "gym_logo": invoice.gym_logo,
                "gym_contact": invoice.gym_contact,
                "gym_location": invoice.gym_location,
                "plan_id" : invoice.plan_id,
                "plan_description" : invoice.plan_description,
                "fees" : invoice.fees,
                "discount" : invoice.discount,
                "discounted_fees":invoice.discounted_fees,
                "due_date" : str(invoice.due_date),
                "invoice_number" :invoice.invoice_number,
                "client_contact" : invoice.client_contact,
                "bank_details" : invoice.bank_details,
                "ifsc_code" :invoice.ifsc_code,
                "account_holder_name" : invoice.account_holder_name,
                "paid":invoice.paid,
                "mail_send":invoice.mail_status,
                "expired":invoice.expired,
                "email":invoice.email
            } for invoice in sent_invoices]

            unsent_data=[{
                "expiry_id" : invoice.expiry_id,
                "client_id" : invoice.client_id,
                "gym_id" : invoice.gym_id,
                "client_name" : invoice.client_name,
                "gym_name": invoice.gym_name,
                "gym_logo": invoice.gym_logo,
                "gym_contact": invoice.gym_contact,
                "gym_location": invoice.gym_location,
                "plan_id" : invoice.plan_id,
                "plan_description" : invoice.plan_description,
                "fees" : invoice.fees,
                "discount" : invoice.discount,
                "discounted_fees":invoice.discounted_fees,
                "due_date" : str(invoice.due_date),
                "invoice_number" :invoice.invoice_number,
                "client_contact" : invoice.client_contact,
                "bank_details" : invoice.bank_details,
                "ifsc_code" :invoice.ifsc_code,
                "account_holder_name" : invoice.account_holder_name,
                "paid":invoice.paid,
                "mail_send":invoice.mail_status,
                "expired":invoice.expired,
                "email":invoice.email
            } for invoice in unsent_invoices]

            invoice_data["send"]=sent_data
            invoice_data["unsend"]=unsent_data

            response["invoice_data"]= invoice_data
            await redis.set(invoice_key,json.dumps(invoice_data))
            await redis.expire(invoice_key, 86400)

        else:
            response["invoice_data"] = json.loads(invoice_data)

        unpaid_key = f'gym{req.gym_id}:unpaid_members'
        unpaid_data = await redis.get(unpaid_key)

        if not unpaid_data:
            unpaid_sent_invoices = db.query(AboutToExpire).filter(AboutToExpire.gym_id == req.gym_id, AboutToExpire.mail_status == True, AboutToExpire.expired==True).all()
            unpaid_unsent_invoices = db.query(AboutToExpire).filter(AboutToExpire.gym_id == req.gym_id, AboutToExpire.mail_status == False, AboutToExpire.expired==True).all()

            unpaid_data ={}
            
            unpaid_sent_data=[{
                "expiry_id" : invoice.expiry_id,
                "client_id" : invoice.client_id,
                "gym_id" : invoice.gym_id,
                "client_name" : invoice.client_name,
                "gym_name": invoice.gym_name,
                "gym_logo": invoice.gym_logo,
                "gym_contact": invoice.gym_contact,
                "gym_location": invoice.gym_location,
                "plan_id" : invoice.plan_id,
                "plan_description" : invoice.plan_description,
                "fees" : invoice.fees,
                "discount" : invoice.discount,
                "discounted_fees":invoice.discounted_fees,
                "due_date" : str(invoice.due_date),
                "invoice_number" :invoice.invoice_number,
                "client_contact" : invoice.client_contact,
                "bank_details" : invoice.bank_details,
                "ifsc_code" :invoice.ifsc_code,
                "account_holder_name" : invoice.account_holder_name,
                "paid":invoice.paid,
                "mail_send":invoice.mail_status,
                "expired":invoice.expired,
                "email":invoice.email
            } for invoice in unpaid_sent_invoices]

            unpaid_unsent_data=[{
                "expiry_id" : invoice.expiry_id,
                "client_id" : invoice.client_id,
                "gym_id" : invoice.gym_id,
                "client_name" : invoice.client_name,
                "gym_name": invoice.gym_name,
                "gym_logo": invoice.gym_logo,
                "gym_contact": invoice.gym_contact,
                "gym_location": invoice.gym_location,
                "plan_id" : invoice.plan_id,
                "plan_description" : invoice.plan_description,
                "fees" : invoice.fees,
                "discount" : invoice.discount,
                "discounted_fees":invoice.discounted_fees,
                "due_date" : str(invoice.due_date),
                "invoice_number" :invoice.invoice_number,
                "client_contact" : invoice.client_contact,
                "bank_details" : invoice.bank_details,
                "ifsc_code" :invoice.ifsc_code,
                "account_holder_name" : invoice.account_holder_name,
                "paid":invoice.paid,
                "mail_send":invoice.mail_status,
                "expired":invoice.expired,
                "email":invoice.email
            } for invoice in unpaid_unsent_invoices]

            unpaid_data["send"]=unpaid_sent_data
            unpaid_data["unsend"]=unpaid_unsent_data

            response["unpaid_data"]= unpaid_data
            await redis.set(unpaid_key,json.dumps(unpaid_data))
            await redis.expire(unpaid_key, 86400)

        else:
            response["unpaid_data"] = json.loads(unpaid_data)

        new_clients_key = f"gym:{req.gym_id}:new_clients"
        new_clients_data = await redis.hgetall(new_clients_key)

        if not new_clients_data:
            print("No new clients data in Redis, fetching from database.")
            current_month = datetime.now().month
            current_year = datetime.now().year
            new_clients = db.query(Client).filter(
                Client.gym_id == req.gym_id,
                extract("month", Client.joined_date) == current_month,
                extract("year", Client.joined_date) == current_year
            ).all()

            if not new_clients:
                print("No new clients found for this month.")
                await redis.hset(
                    new_clients_key,
                    mapping={
                        "total_new_entrant": 0,
                        "average_age": 0,
                        "training_type_summary": "{}",
                        "batch_summary": "{}",
                        "details": "[]",  
                    }
                )
                await redis.expire(new_clients_key, 86400)
            else:
                
                total_clients = len(new_clients)
                total_age = sum(client.age for client in new_clients if client.age)
                average_age = int(round(total_age / total_clients, 2) if total_clients else 0)
                training_type_summary = {}
                batch_summary = {}

                client_details = [
                    {
                        "name": client.name,
                        "age": client.age,
                        "joining_date": client.joined_date.strftime("%Y-%m-%d")
                    }
                    for client in new_clients
                ]

                for client in new_clients:
                    training= db.query(GymPlans).filter(GymPlans.id==client.training_id).first()
                    
                    if training:
                        training_type=training.plans
                    else:
                        training_type=None

                    training_type_summary[training_type] = training_type_summary.get(training_type, 0) + 1
                    
                    batch_t=db.query(GymBatches).filter(GymBatches.batch_id==client.batch_id).first()
                    if batch_t:
                        batch=batch_t.batch_name
                    else:
                        batch=None
                    batch_summary[batch] = batch_summary.get(batch, 0) + 1

                await redis.hset(
                    new_clients_key,
                    mapping={
                        "total_new_entrant": total_clients,
                        "average_age": average_age,
                        "training_type_summary": str(training_type_summary),
                        "batch_summary": str(batch_summary),
                        "details": json.dumps(client_details),  
                    }
                )
                await redis.expire(new_clients_key, 86400)

                new_clients_data = {
                    "total_new_entrant": total_clients,
                    "average_age": average_age,
                    "training_type_summary": training_type_summary,
                    "batch_summary": batch_summary,
                    "details": client_details,
                }
        else:
            print("New clients data fetched from Redis.")
            new_clients_data = {
                "total_new_entrant": int(new_clients_data["total_new_entrant"]),
                "average_age": float(new_clients_data["average_age"]),
                "training_type_summary": eval(new_clients_data["training_type_summary"]),
                "batch_summary": eval(new_clients_data["batch_summary"]),
                "details": json.loads(new_clients_data["details"]), 
            }

        response["new_entrant"] = new_clients_data

        final_result = {}
        collection_key = f"gym:{req.gym_id}:collection"
        cached_data = await redis.hgetall(collection_key)
        if cached_data:
            print("Collection data fetched from Redis.")
            income_list = json.loads(cached_data["income_details"])
            total_income = float(cached_data["total_income"])
            expenditure_list = json.loads(cached_data["expenditure_details"])
            total_expenditure = float(cached_data["total_expenditure"])
            profit = float(cached_data["profit"])

        else:
            print("No collection data in Redis, querying the database.")
        
            current_month = datetime.now().month
            current_year = datetime.now().year
            income_details = db.query(
                FeeHistory.client_id,
                FeeHistory.fees_paid,
                FeeHistory.payment_date,
                Client.name,
                Client.training_id
            ).join(Client, FeeHistory.client_id == Client.client_id).filter(
                FeeHistory.gym_id == req.gym_id,
                extract("month", FeeHistory.payment_date) == current_month,
                extract("year", FeeHistory.payment_date) == current_year
            ).all()

            income_list = [
                {
                    "client_name": record.name,
                    "type": record.training_id,
                    "fees_paid": record.fees_paid,
                    "date": record.payment_date.strftime("%Y-%m-%d")
                }
                for record in income_details
            ]
            total_income = sum(record.fees_paid for record in income_details)

            expenditure_details = db.query(
                Expenditure.expenditure_id,
                Expenditure.expenditure_type,
                Expenditure.amount,
                Expenditure.date
            ).filter(
                Expenditure.gym_id == req.gym_id,
                extract("month", Expenditure.date) == current_month,
                extract("year", Expenditure.date) == current_year
            ).all()

            expenditure_list = [
                {   
                    "id":record.expenditure_id,
                    "type": record.expenditure_type,
                    "amount": record.amount,
                    "date": record.date.strftime("%Y-%m-%d")
                }
                for record in expenditure_details
            ]
            total_expenditure = sum(record.amount for record in expenditure_details)

            profit = total_income - total_expenditure

            await redis.hset(
                collection_key,
                mapping={
                    "income_details": json.dumps(income_list),
                    "total_income": total_income,
                    "expenditure_details": json.dumps(expenditure_list),
                    "total_expenditure": total_expenditure,
                    "profit": profit
                }
            )
            await redis.expire(collection_key, 86400)  

        final_result["income_details"] = income_list
        final_result["total_income"] = total_income
        final_result["expenditure_details"] = expenditure_list
        final_result["total_expenditure"] = total_expenditure
        final_result["profit"] = profit
        response["final_result"] = final_result

        analytics_key = f"gym:{req.gym_id}:analytics"
        analytics_data = await redis.hgetall(analytics_key)

        if analytics_data:
            print("Data fetched from Redis (Analytics Key).")
            response["analytics_summary"] = json.loads(analytics_data["details"])
        else:
            print("Fetching data from database...")
            today = datetime.now().date()

            current_clients = db.query(
                Attendance.client_id, Attendance.in_time, Client.name, Client.training_id, Client.goals, Attendance.muscle
            ).join(Client, Attendance.client_id == Client.client_id).filter(
                Attendance.date == today,
                Attendance.out_time.is_(None),
                Client.gym_id == req.gym_id
            ).all()

            goals_summary = {}
            training_type_summary = {}
            muscle_summary = {}

            for client in current_clients:
            
                if client.goals not in goals_summary:
                    goals_summary[client.goals] = {"count": 0, "clients": []}
                goals_summary[client.goals]["count"] += 1
                goals_summary[client.goals]["clients"].append(client.name)

                training= db.query(GymPlans).filter(GymPlans.id==client.training_id).first()
                    
                if training:
                    training_type=training.plans
                    print("training_type is",training_type)
                else:
                    training_type=None
                    print("training_type is",training_type)


                if training_type not in training_type_summary:
                    training_type_summary[training_type] = {"count": 0, "clients": []}
                
                training_type_summary[training_type]["count"] += 1
                training_type_summary[training_type]["clients"].append(client.name)

                for muscle in client.muscle:
                    if muscle not in muscle_summary:
                        muscle_summary[muscle] = {"count": 0, "clients": []}
                    muscle_summary[muscle]["count"] += 1
                    muscle_summary[muscle]["clients"].append(client.name)

        
            analytics_summary = {
                "goals_summary": goals_summary,
                "training_type_summary": training_type_summary,
                "muscle_summary": muscle_summary,
                "total_present": len(current_clients)
            }

            await redis.hset(
                analytics_key,
                mapping={"details": json.dumps(analytics_summary)}
            )
            await redis.expire(analytics_key, 86400)  

            response["analytics_summary"] = analytics_summary
        

        
        return {"data":response,"status":200,"message":"Gym Data listed successfully"}

    except Exception as e:
        auth_logger.error("gym_home_data_failed", 
                         error=repr(e), 
                         gym_id=req.gym_id if req else None,
                         execution_time=time.time() - start_time)
        raise FittbotHTTPException(
            status_code=500,
            detail="Internal server error occurred while fetching gym home data",
            error_code="GYM_HOME_DATA_ERROR",
            log_data={"error": repr(e), "gym_id": req.gym_id if req else None}
        )
    finally:
        auth_logger.debug("gym_home_data_exit", execution_time=time.time() - start_time)
    
    
class UpdateInvoiceRequest(BaseModel):
    expiry_id : int
    gym_id : int
    discount :float


@router.put("/update-invoice")
async def update_invoice(request: UpdateInvoiceRequest, db:Session=Depends(get_db), redis:Redis= Depends(get_redis)):
    try:
        invoice = db.query(AboutToExpire).filter(AboutToExpire.expiry_id == request.expiry_id).first()

        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice Not found")
        
        invoice.discount = request.discount
        invoice.discounted_fees = invoice.fees-( invoice.fees * (request.discount/100))
        db.commit()

        invoice_key = f'gym{request.gym_id}:invoice_data'
        unpaid_key = f'gym{request.gym_id}:unpaid_members'

        if await redis.exists(invoice_key):
            await redis.delete(invoice_key)

        if await redis.exists(unpaid_key):
            await redis.delete(unpaid_key)

        return{
            "status":200,
            "message":"Invoice updated successfully"
        }        

    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occured : {str(e)}")


class GymClientRequest(BaseModel):
    gym_id: int
 
 
class GymClientRequest(BaseModel):
    gym_id: int
 


def serialize_client(client,batch_map, training_map, db, is_old=False, is_punched_out=True, role=None, trainer_id=None, gym_id=None):
    from datetime import date

    try:
        gym_client = db.query(ClientGym).filter(ClientGym.client_id == client.client_id, ClientGym.gym_id==gym_id).first()
        gym_client_id = gym_client.gym_client_id if gym_client else ""
        admission_number = gym_client.admission_number if gym_client else ""
    except:
        gym_client_id = ""
        admission_number=""

    # Determine status based on membership expiry date
    today = date.today()
    membership_status = "inactive"  # Default status
    starts_at = None
    expires_at = None
    plan_id = None  # Store plan_id from FittbotGymMembership
    latest_membership_id = None

    if is_old:
        # For old clients, check OldGymData.starts_at and expires_at
        try:
            if hasattr(client, 'starts_at') and client.starts_at:
                starts_at = str(client.starts_at) if client.starts_at else None
            if hasattr(client, 'expires_at') and client.expires_at:
                expires_at = str(client.expires_at) if client.expires_at else None
                if client.expires_at >= today:
                    membership_status = "active"
                    latest_membership_id=None
                else:
                    latest_membership_id=None
                    membership_status = "inactive"
            else:
                # No expires_at - fallback to client.status
                membership_status = client.status if hasattr(client, 'status') else "inactive"
        except Exception as e:
            print(f"Error checking old client status for client {client.client_id}: {e}")
            membership_status = client.status if hasattr(client, 'status') else "inactive"
    else:

        try:
      
            membership = db.query(FittbotGymMembership).filter(
                FittbotGymMembership.client_id == str(client.client_id),
                FittbotGymMembership.gym_id == str(gym_id),
                FittbotGymMembership.status != "upcoming"
            ).order_by(FittbotGymMembership.id.desc()).first()

            if membership:
                # Check if membership is active based on expiry date
                if membership.expires_at and membership.expires_at >= today:
                    membership_status = "active"
                else:
                    membership_status = "inactive"

                # Set starts_at, expires_at, plan_id for both active and inactive
                starts_at = str(membership.joined_at) if membership.joined_at else None
                expires_at = str(membership.expires_at) if membership.expires_at else None
                plan_id = int(membership.plan_id) if membership.plan_id else None
                latest_membership_id= int(membership.id) if membership.id else None
            else:
                latest_membership_id=None
                membership_status="inactive"

        except Exception as e:
            print(f"Error checking membership status for client {client.client_id}: {e}")
            # Fallback to client.status if membership check fails
            membership_status = client.status if hasattr(client, 'status') else "inactive"

    # Get training name and goal with proper mapping
    training_name = training_map.get(plan_id if plan_id else client.training_id, "unknown")
    goal_value = client.goals if not client.goals == 'maintain' else 'body_recomposition'

    client_data = {
        "client_id": client.client_id,
        "id": client.client_id,  # Alias for client_id
        "gym_id": client.gym_id,
        "name": client.name,
        "gym_client_id": gym_client_id,
        "admission_number": admission_number,
        "location": client.location,
        "email": client.email,
        "contact": client.contact,
        "lifestyle": client.lifestyle,
        "medical_issues": client.medical_issues,
        "batch": batch_map.get(client.batch_id, "unknown"),
        "batch_id": client.batch_id,
        "training_type": training_name,
        "training": training_name,  # Alias for training_type (used by ClientInformation)
        "training_id": plan_id if plan_id else client.training_id,
        "plan_id": plan_id if plan_id else client.training_id,  # Alias for training_id
        "profile": client.profile,
        "age": client.age,
        "goals": goal_value,
        "goal": goal_value,  # Alias for goals (used by ClientInformation)
        "gender": client.gender,
        "height": client.height,
        "weight": client.weight,
        "bmi": client.bmi,
        "joined_date": str(client.joined_date.strftime("%Y-%m-%d")) if client.joined_date else None,
        "status": membership_status,  # Use membership-based status
        "is_old_client": is_old,
        "is_punched_out": is_punched_out,
        "data_sharing": client.data_sharing if not is_old else False,
        "starts_at": starts_at,
        "expires_at": expires_at,
        "latest_membership_id":latest_membership_id
    }
 
    if role == "trainer" and trainer_id and gym_id:
        try:
            # Updated to use TrainerProfile instead of TrainerPermissions
            trainer_profile = db.query(TrainerProfile).filter(
                and_(
                    TrainerProfile.trainer_id == trainer_id,
                    TrainerProfile.gym_id == gym_id
                )
            ).first()
           
            client_data["trainer_permissions"] = {
                "can_view_client_data": trainer_profile.can_view_client_data if trainer_profile else False,
                "personal_trainer": trainer_profile.personal_trainer if trainer_profile else False,
                "profile_id": trainer_profile.profile_id if trainer_profile else None,
            }
        except Exception as e:
            print(f"Error getting trainer permissions: {e}")
            client_data["trainer_permissions"] = {
                "can_view_client_data": False,
                "personal_trainer": False,
                "profile_id": None,
            }
 
    return client_data
 


 
@router.get("/trainer/assigned_clients")
async def get_trainer_assigned_clients(
    trainer_id: int = Query(...),
    gym_id: int = Query(...),
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis),
):
    redis_key = f"trainer:{trainer_id}:gym:{gym_id}:assigned_clients"
    try:
        cached_data = await redis.get(redis_key)
        if cached_data:
            parsed_data = json.loads(cached_data)
            # Sort using category and index to preserve DB order
            # Category 0 = Regular clients, 1 = Old clients, 2 = Imported clients
            def sort_key(c):
                category = c.get("_sort_category", 999)  # Default to 999 for any uncategorized
                index = c.get("_sort_index", 0)
                return (category, index)

            client_data_sorted = sorted(
                parsed_data["client_data"],
                key=sort_key
            )

            # Add map_client_id sequentially in final sorted order
            for map_index, client in enumerate(client_data_sorted, start=1):
                client['map_client_id'] = map_index

            return {
                "status": 200,
                "message": "Trainer assigned clients retrieved successfully from cache.",
                "data": client_data_sorted,
                "gym_data": parsed_data['gym_data'],
                "trainer_permissions": parsed_data.get('trainer_permissions', {})
            }
 
        trainer_profile = db.query(TrainerProfile).filter(
            and_(
                TrainerProfile.trainer_id == trainer_id,
                TrainerProfile.gym_id == gym_id
            )
        ).first()
 
        if not trainer_profile:
            return {
                "status": 404,
                "message": "Trainer profile not found for this gym"
            }
 
        gym = db.query(Gym).filter(Gym.gym_id == gym_id).first()

        if not gym:
            raise HTTPException(status_code=404, detail=f"Gym with id {gym_id} not found")

        account = db.query(AccountDetails).filter(AccountDetails.gym_id == gym_id).first()

        gym_data = {
            "gym_id": gym.gym_id,
            "name": gym.name,
            "location": gym.location,
            "cover_pic": gym.cover_pic,
            "logo": gym.logo,
            "subscription_start_date": str(gym.subscription_start_date) if gym.subscription_start_date else None,
            "subscription_end_date": str(gym.subscription_end_date) if gym.subscription_end_date else None,
            "account_number": account.account_number if account else None,
            "account_holdername": account.account_holdername if account else None,
            "account_ifsccode": account.account_ifsccode if account else None,
            "account_branch": account.account_branch if account else None,
            "account_id": account.account_id if account else None,
            "upi_id": account.upi_id if account else None,
            "gst_number": account.gst_number if account else None
        }
 
        today = date.today()
        incomplete_attendance = db.query(Attendance.client_id).filter(
            Attendance.gym_id == gym_id,
            Attendance.date == today,
            or_(
                and_(Attendance.in_time.isnot(None), Attendance.out_time.is_(None)),
                and_(Attendance.in_time_2.isnot(None), Attendance.out_time_2.is_(None)),
                and_(Attendance.in_time_3.isnot(None), Attendance.out_time_3.is_(None))
            )
        ).distinct().all()
 
        punched_in_client_ids = {row.client_id for row in incomplete_attendance}
 
        batches = db.query(GymBatches.batch_id, GymBatches.batch_name).filter(GymBatches.gym_id == gym_id).all()
        plans = db.query(GymPlans.id, GymPlans.plans).filter(GymPlans.gym_id == gym_id).all()
 
        batch_map = {batch.batch_id: batch.batch_name for batch in batches}
        training_map = {plan.id: plan.plans for plan in plans}
 
        assigned_client_ids_profile = db.query(ClientScheduler.client_id).filter(
            and_(
                ClientScheduler.gym_id == gym_id,
                ClientScheduler.assigned_trainer_profile == trainer_profile.profile_id
            )
        ).all() if hasattr(ClientScheduler, 'assigned_trainer_profile') else []
 
        assigned_client_ids_trainer = db.query(ClientScheduler.client_id).filter(
            and_(
                ClientScheduler.gym_id == gym_id,
                ClientScheduler.assigned_trainer == trainer_id
            )
        ).all()
 
        assigned_client_ids = list(set(
            [row.client_id for row in assigned_client_ids_profile] +
            [row.client_id for row in assigned_client_ids_trainer]
        ))
 
        if not assigned_client_ids:
            return {
                "status": 201,
                "message": "No clients assigned to this trainer"
            }
       
 
        clients = (
            db.query(Client)
            .filter(
                and_(
                    Client.gym_id == gym_id,
                    Client.client_id.in_(assigned_client_ids)
                )
            )
            .order_by(Client.client_id.desc())
            .all()
        )
 
        old_clients = (
            db.query(OldGymData)
            .filter(
                and_(
                    OldGymData.gym_id == gym_id,
                    OldGymData.client_id.in_(assigned_client_ids)
                )
            )
            .all()
        )
 
        if not clients and not old_clients:
            return {
                "status": 201,
                "message": "No clients found assigned to this trainer"
            }
 
        client_data = [
            serialize_client(
                client,
                batch_map,
                training_map,
                db,
                is_old=False,
                is_punched_out=(client.client_id not in punched_in_client_ids),
                role="trainer",
                trainer_id=trainer_id,
                gym_id=gym_id
            )
            for client in clients
        ]
 
        old_client_data = [
            serialize_client(
                client,
                batch_map,
                training_map,
                db,
                is_old=True,
                role="trainer",
                trainer_id=trainer_id,
                gym_id=gym_id
            )
            for client in old_clients
        ]
 
        combined_data = client_data + old_client_data
 
        trainer_permissions = {
            "can_view_client_data": trainer_profile.can_view_client_data,
            "personal_trainer": trainer_profile.personal_trainer,
            "profile_id": trainer_profile.profile_id,
            "trainer_name": trainer_profile.full_name,
            "specialization": trainer_profile.specializations
        }
 
        # Add map_client_id sequentially in final sorted order
        for map_index, client in enumerate(combined_data, start=1):
            client['map_client_id'] = map_index

        cache_data = {
            'client_data': combined_data,
            'gym_data': gym_data,
            'trainer_permissions': trainer_permissions
        }

        await redis.set(redis_key, json.dumps(cache_data), ex=86400)

        return {
            "status": 200,
            "message": "Trainer assigned clients retrieved successfully.",
            "data": combined_data,
            'gym_data': gym_data,
            'trainer_permissions': trainer_permissions
        }
 
    except Exception as e:
        print(f"Unexpected error: {str(e)}")
        raise HTTPException(status_code=500, detail="An unexpected error occurred.")
 
 
@router.get("/gym/client")
async def get_clients(
    gym_id: int = Query(...),
    role: str = Query(None),  
    trainer_id: int = Query(None),  
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis),
):
   
    redis_key = f"gym:{gym_id}:clientdata"
    try:

 
        gym = db.query(Gym).filter(Gym.gym_id == gym_id).first()

        if not gym:
            raise HTTPException(status_code=404, detail=f"Gym with id {gym_id} not found")

        account = db.query(AccountDetails).filter(AccountDetails.gym_id == gym_id).first()

        gym_data = {
            "gym_id": gym.gym_id,
            "name": gym.name,
            "location": gym.location,
            "cover_pic": gym.cover_pic,
            "logo": gym.logo,
            "subscription_start_date": str(gym.subscription_start_date) if gym.subscription_start_date else None,
            "subscription_end_date": str(gym.subscription_end_date) if gym.subscription_end_date else None,
            "account_number": account.account_number if account else None,
            "account_holdername": account.account_holdername if account else None,
            "account_ifsccode": account.account_ifsccode if account else None,
            "account_branch": account.account_branch if account else None,
            "account_id": account.account_id if account else None,
            "upi_id": account.upi_id if account else None,
            "gst_number": account.gst_number if account else None
        }
 
        today = date.today()
        incomplete_attendance = db.query(Attendance.client_id).filter(
            Attendance.gym_id == gym_id,
            Attendance.date == today,
            or_(
                and_(Attendance.in_time.isnot(None), Attendance.out_time.is_(None)),
                and_(Attendance.in_time_2.isnot(None), Attendance.out_time_2.is_(None)),
                and_(Attendance.in_time_3.isnot(None), Attendance.out_time_3.is_(None))
            )
        ).distinct().all()
 
        punched_in_client_ids = {row.client_id for row in incomplete_attendance}
 
        batches = db.query(GymBatches.batch_id, GymBatches.batch_name).filter(GymBatches.gym_id == gym_id).all()
        plans = db.query(GymPlans.id, GymPlans.plans).filter(GymPlans.gym_id == gym_id).all()
 
        batch_map = {batch.batch_id: batch.batch_name for batch in batches}
        training_map = {plan.id: plan.plans for plan in plans}
 
        clients = (
            db.query(Client)
            .filter(Client.gym_id == gym_id)
            .order_by(Client.client_id.desc())
            .all()
        )
 
        old_clients = db.query(OldGymData).filter(OldGymData.gym_id == gym_id).all()
 
        client_data = [
            serialize_client(
                client,
                batch_map,
                training_map,
                db,
                is_old=False,
                is_punched_out=(client.client_id not in punched_in_client_ids),
                role=role,
                trainer_id=trainer_id,
                gym_id=gym_id
            )
            for client in clients
        ]
 
        old_client_data = [
            serialize_client(
                client,
                batch_map,
                training_map,
                db,
                is_old=True,
                role=role,
                trainer_id=trainer_id,
                gym_id=gym_id
            )
            for client in old_clients
        ]
 
        # Fetch imported clients data
        imported_clients = db.query(GymImportData).filter(GymImportData.gym_id == gym_id).all()

        import_client_data = []

        for client in imported_clients:

            import_membership = db.query(FittbotGymMembership).filter(
                FittbotGymMembership.client_id == f"import_{client.import_id}",
                FittbotGymMembership.gym_id == str(gym_id)
            ).order_by(FittbotGymMembership.id.desc()).first()

            client_dict = get_data_as_dict(client)
            client_dict['import_status'] = True

            if import_membership:
                client_dict['starts_at'] = str(import_membership.joined_at) if import_membership.joined_at else client_dict.get('joined_at')
                client_dict['expires_at'] = str(import_membership.expires_at) if import_membership.expires_at else client_dict.get('expires_at')
                client_dict['latest_membership_id'] = import_membership.id

            import_client_data.append(client_dict)

        # Fetch manual CRM clients
        manual_clients = db.query(ManualClient).filter(ManualClient.gym_id == gym_id).order_by(ManualClient.id.desc()).all()

        manual_client_data = []
        for client in manual_clients:
            # Get FittbotGymMembership for manual client (stored with client_id="manual_{id}")
            manual_membership = db.query(FittbotGymMembership).filter(
                FittbotGymMembership.client_id == f"manual_{client.id}",
                FittbotGymMembership.gym_id == str(gym_id)
            ).order_by(FittbotGymMembership.id.desc()).first()

            latest_membership_id = manual_membership.id if manual_membership else None
            starts_at = str(manual_membership.joined_at) if manual_membership and manual_membership.joined_at else str(client.joined_at) if client.joined_at else None
            expires_at = str(manual_membership.expires_at) if manual_membership and manual_membership.expires_at else str(client.expires_at) if client.expires_at else None

            # Determine membership status based on expiry date (consistent with regular clients)
            membership_status = "inactive"
            if manual_membership and manual_membership.expires_at:
                if manual_membership.expires_at >= today:
                    membership_status = "active"
            elif client.expires_at:
                if client.expires_at >= today:
                    membership_status = "active"

            # Get batch and training names for this client
            batch_name = batch_map.get(client.batch_id) if client.batch_id else None
            training_name = training_map.get(client.plan_id) if client.plan_id else None

            manual_client_data.append({
                "client_id": client.id,
                "id": client.id,  # Alias for client_id (used by some frontend components)
                "name": client.name,
                "contact": client.contact,
                "email": client.email,
                "gender": client.gender,
                "date_of_birth": str(client.date_of_birth) if client.date_of_birth else None,
                "age": client.age,
                "height": client.height,
                "weight": client.weight,
                "bmi": client.bmi,
                "goal": client.goal,  # Goal field (weight_gain, weight_loss, etc.)
                "goals": client.goal,  # Alias for compatibility
                "profile": None,  # Manual clients don't have profile pics yet
                "dp": client.dp,  # Manual client profile photo
                "admission_number": client.admission_number,
                "batch_id": client.batch_id,
                "batch_name": batch_name,
                "batch": batch_name,  # Alias for batch_name (used by ClientInformation)
                "training_id": client.plan_id,
                "training_name": training_name,
                "training": training_name,  # Alias for training_name (used by ClientInformation)
                "training_type": training_name,  # Another alias for compatibility
                "plan_id": client.plan_id,  # Alias for training_id
                "joined_date": str(client.joined_at) if client.joined_at else None,
                "starts_at": starts_at,
                "expires_at": expires_at,
                "admission_fee": client.admission_fee,
                "monthly_fee": client.monthly_fee,
                "total_paid": client.total_paid,
                "balance_due": client.balance_due,
                "last_payment_date": str(client.last_payment_date) if client.last_payment_date else None,
                "status": membership_status,  # Use membership-based status (same as regular clients)
                "notes": client.notes,
                "entry_type": "manual",
                "manual_client": True,
                "is_punched_out": True,  # Manual clients don't have punch-in tracking yet
                "is_old_client": False,  # Manual clients are not old clients
                "data_sharing": False,  # Manual clients don't use the app
                "latest_membership_id": latest_membership_id,
                "gym_id": gym_id,  # Include gym_id for API calls
            })

        seen_contacts = {}
        combined_data = []

        # Helper function to normalize contact number
        def normalize_contact(contact):
            if not contact:
                return None
            # Convert to string and strip whitespace
            normalized = str(contact).strip()
            # Remove any non-digit characters
            normalized = ''.join(filter(str.isdigit, normalized))
            return normalized if normalized else None

        # Track unique_id to prevent collisions
        seen_unique_ids = set()
        unique_id_counter = 1

        # Priority 1: Regular clients from Client table
        for idx, client in enumerate(client_data):
            contact = normalize_contact(client.get('contact') or client.get('client_contact'))
            if contact:
                if contact not in seen_contacts:
                    seen_contacts[contact] = True
                    # Generate guaranteed unique_id
                    unique_id = f"client_{client.get('client_id')}_{idx}"
                    while unique_id in seen_unique_ids:
                        unique_id = f"client_{client.get('client_id')}_{unique_id_counter}"
                        unique_id_counter += 1
                    seen_unique_ids.add(unique_id)
                    client['unique_id'] = unique_id
                    client['_sort_category'] = 0  # Regular clients = category 0
                    client['_sort_index'] = idx
                    combined_data.append(client)
            else:
                # If no contact, still add but with unique_id
                unique_id = f"client_{client.get('client_id')}_{idx}"
                while unique_id in seen_unique_ids:
                    unique_id = f"client_{client.get('client_id')}_{unique_id_counter}"
                    unique_id_counter += 1
                seen_unique_ids.add(unique_id)
                client['unique_id'] = unique_id
                client['_sort_category'] = 0  # Regular clients = category 0
                client['_sort_index'] = idx
                combined_data.append(client)

        # Priority 2: Old clients from OldGymData table
        for idx, client in enumerate(old_client_data):
            contact = normalize_contact(client.get('contact') or client.get('client_contact'))
            if contact:
                if contact not in seen_contacts:
                    seen_contacts[contact] = True
                    # Generate guaranteed unique_id
                    unique_id = f"old_{client.get('client_id')}_{idx}"
                    while unique_id in seen_unique_ids:
                        unique_id = f"old_{client.get('client_id')}_{unique_id_counter}"
                        unique_id_counter += 1
                    seen_unique_ids.add(unique_id)
                    client['unique_id'] = unique_id
                    client['_sort_category'] = 1  # Old clients = category 1
                    client['_sort_index'] = idx
                    combined_data.append(client)
            else:
                # If no contact, still add but with unique_id
                unique_id = f"old_{client.get('client_id')}_{idx}"
                while unique_id in seen_unique_ids:
                    unique_id = f"old_{client.get('client_id')}_{unique_id_counter}"
                    unique_id_counter += 1
                seen_unique_ids.add(unique_id)
                client['unique_id'] = unique_id
                client['_sort_category'] = 1  # Old clients = category 1
                client['_sort_index'] = idx
                combined_data.append(client)

        # Priority 3: Imported clients from GymImportData table
        for idx, client in enumerate(import_client_data):
            contact = normalize_contact(client.get('contact') or client.get('client_contact'))
            if contact:
                if contact not in seen_contacts:
                    seen_contacts[contact] = True
                    # Generate guaranteed unique_id
                    unique_id = f"import_{client.get('import_id')}_{idx}"
                    while unique_id in seen_unique_ids:
                        unique_id = f"import_{client.get('import_id')}_{unique_id_counter}"
                        unique_id_counter += 1
                    seen_unique_ids.add(unique_id)
                    client['unique_id'] = unique_id
                    client['_sort_category'] = 2  # Imported clients = category 2
                    client['_sort_index'] = idx
                    combined_data.append(client)
            else:
                # If no contact, still add but with unique_id
                unique_id = f"import_{client.get('import_id')}_{idx}"
                while unique_id in seen_unique_ids:
                    unique_id = f"import_{client.get('import_id')}_{unique_id_counter}"
                    unique_id_counter += 1
                seen_unique_ids.add(unique_id)
                client['unique_id'] = unique_id
                client['_sort_category'] = 2  # Imported clients = category 2
                client['_sort_index'] = idx
                combined_data.append(client)

        # Priority 4: Manual CRM clients from ManualClient table
        for idx, client in enumerate(manual_client_data):
            contact = normalize_contact(client.get('contact'))
            if contact:
                if contact not in seen_contacts:
                    seen_contacts[contact] = True
                    # Generate guaranteed unique_id
                    unique_id = f"manual_{client.get('client_id')}_{idx}"
                    while unique_id in seen_unique_ids:
                        unique_id = f"manual_{client.get('client_id')}_{unique_id_counter}"
                        unique_id_counter += 1
                    seen_unique_ids.add(unique_id)
                    client['unique_id'] = unique_id
                    client['_sort_category'] = 3  # Manual clients = category 3
                    client['_sort_index'] = idx
                    combined_data.append(client)
            else:
                # If no contact, still add but with unique_id
                unique_id = f"manual_{client.get('client_id')}_{idx}"
                while unique_id in seen_unique_ids:
                    unique_id = f"manual_{client.get('client_id')}_{unique_id_counter}"
                    unique_id_counter += 1
                seen_unique_ids.add(unique_id)
                client['unique_id'] = unique_id
                client['_sort_category'] = 3  # Manual clients = category 3
                client['_sort_index'] = idx
                combined_data.append(client)

        # Add enumerated id starting from 1 for all combined data
        # This id is guaranteed to be unique and sequential
        for index, client in enumerate(combined_data, start=1):
            client['id'] = index

        # Debug: Check for duplicate unique_ids and ids
        unique_ids_check = [client.get('unique_id') for client in combined_data]
        ids_check = [client.get('id') for client in combined_data]

        # Check for duplicates
        from collections import Counter
        unique_id_counts = Counter(unique_ids_check)
        id_counts = Counter(ids_check)

        duplicate_unique_ids = {k: v for k, v in unique_id_counts.items() if v > 1}
        duplicate_ids = {k: v for k, v in id_counts.items() if v > 1}

        if duplicate_unique_ids:
            print(f"WARNING: Duplicate unique_ids found: {duplicate_unique_ids}")
        if duplicate_ids:
            print(f"WARNING: Duplicate ids found: {duplicate_ids}")



        trainer_permissions = {}
        if role == "trainer" and trainer_id and gym_id:
            trainer_profile = db.query(TrainerProfile).filter(
                and_(
                    TrainerProfile.trainer_id == trainer_id,
                    TrainerProfile.gym_id == gym_id
                )
            ).first()

            if trainer_profile:
                trainer_permissions = {
                    "can_view_client_data": trainer_profile.can_view_client_data,
                    "personal_trainer": trainer_profile.personal_trainer,
                    "profile_id": trainer_profile.profile_id,
                    "trainer_name": trainer_profile.full_name,
                    "specialization": trainer_profile.specializations

                }
            else:
                trainer_permissions = {
                    "can_view_client_data": False,
                    "personal_trainer": False,
                    "profile_id": None,
                    "trainer_name": None,
                    "specialization": None
                }

        # Sort combined_data alphabetically by name (case-insensitive)
        combined_data.sort(key=lambda x: (x.get('name') or '').lower())

        # Add map_client_id sequentially in final sorted order
        for map_index, client in enumerate(combined_data, start=1):
            client['map_client_id'] = map_index

        clients_data = {
            'client_data': combined_data,
            'gym_data': gym_data,
            'trainer_permissions': trainer_permissions
        }

        await redis.set(redis_key, json.dumps(clients_data), ex=86400)
        

        return {
            "status": 200,
            "message": "Clients retrieved successfully.",
            "data": combined_data,
            'gym_data': gym_data,
            'trainer_permissions': trainer_permissions
        }

    except Exception as e:
        import traceback
        print(f"Unexpected error: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail="An unexpected error occurred.")
 
class EditExpiry(BaseModel):
    latest_membership_id: Optional[int] = None
    manual_client_id: Optional[int] = None  # For manual CRM clients
    client_id: Optional[int] = None  # Regular client id (fallback)
    gym_id: Optional[int] = None
    starts_at: date
    expires_at: date



@router.put("/edit_expiry")
async def edit_expiry(editRequest:EditExpiry, db: Session = Depends(get_db), redis: Redis = Depends(get_redis)):
    try:
        cache_gym_id = None

        # Handle manual clients
        if editRequest.manual_client_id:
            manual_client = db.query(ManualClient).filter(ManualClient.id == editRequest.manual_client_id).first()
            if manual_client:
                manual_client.joined_at = editRequest.starts_at
                manual_client.expires_at = editRequest.expires_at
                cache_gym_id = manual_client.gym_id
                # Also update FittbotGymMembership if exists
                manual_membership = db.query(FittbotGymMembership).filter(
                    FittbotGymMembership.client_id == f"manual_{editRequest.manual_client_id}"
                ).order_by(FittbotGymMembership.id.desc()).first()
                if manual_membership:
                    manual_membership.joined_at = editRequest.starts_at
                    manual_membership.expires_at = editRequest.expires_at
                    cache_gym_id = cache_gym_id or manual_membership.gym_id
                db.commit()
                if cache_gym_id:
                    try:
                        await redis.delete(f"gym:{cache_gym_id}:members")
                    except Exception as cache_err:
                        print(f"[edit_expiry] Failed to clear members cache for gym {cache_gym_id}: {cache_err}")
                return {"status": 200}
            else:
                raise HTTPException(status_code=404, detail="Manual client not found")

        # Handle regular clients with membership
        if editRequest.latest_membership_id:
            edit_expiry = db.query(FittbotGymMembership).filter(FittbotGymMembership.id == editRequest.latest_membership_id).first()
            if edit_expiry:
                edit_expiry.joined_at = editRequest.starts_at
                edit_expiry.expires_at = editRequest.expires_at
                db.commit()
                cache_gym_id = edit_expiry.gym_id
                if cache_gym_id:
                    try:
                        await redis.delete(f"gym:{cache_gym_id}:members")
                    except Exception as cache_err:
                        print(f"[edit_expiry] Failed to clear members cache for gym {cache_gym_id}: {cache_err}")
                return {"status": 200}
            else:
                raise HTTPException(status_code=404, detail="Membership record not found")

        # Fallback: locate latest membership by client_id and gym_id
        if editRequest.client_id and editRequest.gym_id:
            membership = (
                db.query(FittbotGymMembership)
                .filter(
                    FittbotGymMembership.client_id == str(editRequest.client_id),
                    FittbotGymMembership.gym_id == str(editRequest.gym_id),
                )
                .order_by(FittbotGymMembership.id.desc())
                .first()
            )
            if membership:
                membership.joined_at = editRequest.starts_at
                membership.expires_at = editRequest.expires_at
                db.commit()
                cache_gym_id = membership.gym_id or editRequest.gym_id
                if cache_gym_id:
                    try:
                        await redis.delete(f"gym:{cache_gym_id}:members")
                    except Exception as cache_err:
                        print(f"[edit_expiry] Failed to clear members cache for gym {cache_gym_id}: {cache_err}")
                return {"status": 200}
            else:
                raise HTTPException(status_code=404, detail="Active membership not found for client")

        raise HTTPException(status_code=400, detail="Either latest_membership_id or manual_client_id is required")

    except HTTPException:
        db.rollback()
        raise

    except Exception as e:
        db.rollback()
        print(f"Unexpected error: {str(e)}")
        raise FittbotHTTPException(
            status_code=500,
            detail=f"An unexpected error occurred: {str(e)}",
            error_code="SYSTEM_UNEXPECTED_ERROR"
        )




@router.get("/gym/client_data")
async def get_client_data(
    client_id: int = Query(...),
    gym_id: int = Query(...),
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis),
):

    try:
        fees_redis_key = f"{client_id}:{gym_id}:fees"
        diet_redis_key = f"{client_id}:{gym_id}:diettemplate"
        workout_redis_key = f"{client_id}:{gym_id}:workouttemplate"
        client_actual_redis_key = f"{client_id}:{gym_id}:clientactual"

        fee_data = await redis.get(fees_redis_key)
        if fee_data:
            fee_history = json.loads(fee_data)
        else:
            fee_history_records = (
                db.query(FeeHistory)
                .filter(FeeHistory.gym_id == gym_id, FeeHistory.client_id == client_id)
                .all()
            )
            fee_history = [
                {"payment_date": record.payment_date.strftime("%Y-%m-%d"), "fees_paid": record.fees_paid}
                for record in fee_history_records
            ]
            await redis.set(fees_redis_key, json.dumps(fee_history), ex=86400)

        diet_data = await redis.get(diet_redis_key)
        if diet_data:
            diet_variants = json.loads(diet_data)
        else:
            diet_records = db.query(DietTemplate).filter(DietTemplate.client_id == client_id).all()
            diet_variants = {}
            for record in diet_records:
                variant = record.diet_variant
                if variant not in diet_variants:
                    diet_variants[variant] = []
                diet_variants[variant].append({
                    "time_slot": record.time_slot,
                    "meal_type": record.meal_type,
                    "diet_type": record.diet_type,
                    "calories": record.calories,
                    "protein": record.protein,
                    "fat": record.fat,
                    "carbs": record.carbs,
                    "notes": record.notes,
                })
            await redis.set(diet_redis_key, json.dumps(diet_variants), ex=86400)

        workout_data = await redis.get(workout_redis_key)
        if workout_data:
            workout_days = json.loads(workout_data)
        else:
            workout_records = db.query(WorkoutTemplate).filter(WorkoutTemplate.client_id == client_id).all()
            workout_days = {}
            for record in workout_records:
                day = record.day
                if day not in workout_days:
                    workout_days[day] = []
                workout_days[day].append({
                    "workout_name": record.workout_name,
                    "sets": record.sets,
                    "reps": record.reps,
                    "weights": [record.weight_1, record.weight_2, record.weight_3, record.weight_4],
                    "muscle_group": record.muscle_group,
                    "duration": record.duration,
                    "rest_time": record.rest_time,
                    "notes": record.notes,
                })
            await redis.set(workout_redis_key, json.dumps(workout_days), ex=86400)

        client_actual_data = await redis.get(client_actual_redis_key)
        if client_actual_data:
            client_actual = json.loads(client_actual_data)
        else:
            current_month = datetime.now().month
            current_year = datetime.now().year

            client_actual_records = (
                db.query(ClientActual)
                .filter(ClientActual.client_id == client_id)
                .filter(ClientActual.date.between(f"{current_year}-{current_month}-01", f"{current_year}-{current_month}-31"))
                .all()
            )
            client_actual = [
                {
                    "date": record.date.strftime("%Y-%m-%d"),
                    "weight": record.weight,
                    "calories": record.calories,
                    "protein": record.protein,
                    "carbs": record.carbs,
                    "fats": record.fats,
                    "steps": record.steps,
                    "burnt_calories": record.burnt_calories,
                    "water_intake": record.water_intake,
                    "sleep_hours": record.sleep_hours,
                }
                for record in client_actual_records
            ]
            await redis.set(client_actual_redis_key, json.dumps(client_actual), ex=86400)


        response_data = {
            "fee_history": fee_history,
            "diet_variants": diet_variants,
            "workout_days": workout_days,
            "client_actual": client_actual,
        }
        print("response data is",response_data)
        return {"data":response_data,"status":200,"message":"Client data listed successfully"}
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected error occured:{str(e)}")



@router.get("/gym/get_fee_details")
async def get_fee_details( training_id:int, db: Session = Depends(get_db), redis: Redis = Depends(get_redis)):
    try:
        training= db.query(GymPlans).filter(GymPlans.id==training_id).first()

        if not training:
            raise HTTPException(status_code=404, detail="Plan not found")

        return {
            "status": 200,
            "message": "Data fetched successfully.",
            "data": training
        }
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        return {
            "status": 500,
            "message": "An unexpected error occurred.",
            "data": None
        }


@router.get("/gym/fee_history")
async def get_fee_history(
    client_id: str,  # Accept string to support both regular (e.g., "123") and manual clients (e.g., "manual_5")
    gym_id: int,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis)
):

    try:

        if not client_id or not client_id.strip():
            raise FittbotHTTPException(
                status_code=400,
                detail="Invalid client_id",
                error_code="INVALID_CLIENT_ID",
                log_data={"client_id": client_id},
            )

        if not isinstance(gym_id, int) or gym_id <= 0:
            raise FittbotHTTPException(
                status_code=400,
                detail="Invalid gym_id",
                error_code="INVALID_GYM_ID",
                log_data={"gym_id": gym_id},
            )

        memberships = (
            db.query(FittbotGymMembership, GymPlans.plans)
            .outerjoin(GymPlans, GymPlans.id == FittbotGymMembership.plan_id)
            .filter(
                FittbotGymMembership.client_id == client_id,
                FittbotGymMembership.gym_id == str(gym_id),
                FittbotGymMembership.status != "upcoming"
            )
            .order_by(FittbotGymMembership.joined_at.desc())
            .all()
        )

        if not memberships:
            return {
                "status": 200,
                "message": "No fee history found for this client",
                "data": []
            }


        fee_history = []
        for membership, plan_name in memberships:
            fee_history.append({
                "id": membership.id,
                "plan_name": plan_name if plan_name else "Unknown Plan",
                "amount": float(membership.amount) if membership.amount else 0.0,
                "joined_at": str(membership.joined_at) if membership.joined_at else None,
                "expires_at": str(membership.expires_at) if membership.expires_at else None,
                "type": membership.type if membership.type else "normal",  # normal, gym_membership, personal_training, etc.
            })

  
        return {
            "status": 200,
            "message": "Fee history fetched successfully",
            "data": fee_history
        }

    except FittbotHTTPException:
        
        raise

    except Exception as e:

        raise FittbotHTTPException(
            status_code=500,
            detail="Failed to fetch fee history",
            error_code="FEE_HISTORY_FETCH_ERROR",
            log_data={"client_id": client_id, "gym_id": gym_id, "error": repr(e)},
        )


class UpdateFeeEntryRequest(BaseModel):
    membership_id: int
    gym_id: int
    amount: float
    joined_at: str  # YYYY-MM-DD
    expires_at: str  # YYYY-MM-DD


@router.put("/gym/fee_history/update")
async def update_fee_entry(
    request: UpdateFeeEntryRequest,
    db: Session = Depends(get_db),
):
    """Update a FittbotGymMembership entry (fee history record)"""
    try:
        # Find the membership record
        membership = db.query(FittbotGymMembership).filter(
            FittbotGymMembership.id == request.membership_id,
            FittbotGymMembership.gym_id == str(request.gym_id)
        ).first()

        if not membership:
            raise FittbotHTTPException(
                status_code=404,
                detail="Fee entry not found",
                error_code="FEE_ENTRY_NOT_FOUND",
                log_data={"membership_id": request.membership_id},
            )

        # Store old amount for income adjustment
        old_amount = float(membership.amount or 0)
        new_amount = float(request.amount)
        amount_diff = new_amount - old_amount

        # Update the membership
        membership.amount = new_amount
        membership.joined_at = datetime.strptime(request.joined_at, "%Y-%m-%d").date()
        membership.expires_at = datetime.strptime(request.expires_at, "%Y-%m-%d").date()

        # Adjust GymMonthlyData income if amount changed
        if amount_diff != 0:
            month_tag = membership.joined_at.strftime("%Y-%m") if membership.joined_at else datetime.now().strftime("%Y-%m")
            monthly_record = db.query(GymMonthlyData).filter(
                GymMonthlyData.gym_id == request.gym_id,
                GymMonthlyData.month_year.like(f"{month_tag}%")
            ).first()
            if monthly_record:
                monthly_record.income = float(monthly_record.income or 0) + amount_diff

        db.commit()

        return {
            "status": 200,
            "message": "Fee entry updated successfully",
            "data": {
                "id": membership.id,
                "amount": float(membership.amount),
                "joined_at": str(membership.joined_at),
                "expires_at": str(membership.expires_at)
            }
        }

    except FittbotHTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise FittbotHTTPException(
            status_code=500,
            detail="Failed to update fee entry",
            error_code="FEE_ENTRY_UPDATE_ERROR",
            log_data={"membership_id": request.membership_id, "error": repr(e)},
        )


@router.delete("/gym/fee_history/delete")
async def delete_fee_entry(
    membership_id: int,
    gym_id: int,
    db: Session = Depends(get_db),
):
    """Delete a FittbotGymMembership entry (fee history record)"""
    try:
        # Find the membership record
        membership = db.query(FittbotGymMembership).filter(
            FittbotGymMembership.id == membership_id,
            FittbotGymMembership.gym_id == str(gym_id)
        ).first()

        if not membership:
            raise FittbotHTTPException(
                status_code=404,
                detail="Fee entry not found",
                error_code="FEE_ENTRY_NOT_FOUND",
                log_data={"membership_id": membership_id},
            )

        # Adjust GymMonthlyData income
        amount_to_subtract = float(membership.amount or 0)
        if amount_to_subtract > 0:
            month_tag = membership.joined_at.strftime("%Y-%m") if membership.joined_at else datetime.now().strftime("%Y-%m")
            monthly_record = db.query(GymMonthlyData).filter(
                GymMonthlyData.gym_id == gym_id,
                GymMonthlyData.month_year.like(f"{month_tag}%")
            ).first()
            if monthly_record:
                monthly_record.income = max(0, float(monthly_record.income or 0) - amount_to_subtract)

        # Delete the membership record
        db.delete(membership)
        db.commit()

        return {
            "status": 200,
            "message": "Fee entry deleted successfully"
        }

    except FittbotHTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise FittbotHTTPException(
            status_code=500,
            detail="Failed to delete fee entry",
            error_code="FEE_ENTRY_DELETE_ERROR",
            log_data={"membership_id": membership_id, "error": repr(e)},
        )


def update_to_current_month_and_year(given_date):
    if isinstance(given_date, str):
        given_date = datetime.strptime(given_date, "%Y-%m-%d").date()
    
    today = date.today()
    updated_date = given_date.replace(year=today.year, month=today.month)
    return updated_date.strftime("%Y-%m-%d")

def start_of_current_month():
    today = datetime.today()
    start_date = today.replace(day=1)
    return start_date.strftime("%Y-%m-%d")
 






DEFAULT_ANALYSIS: Dict[str, dict] = {
    "gender": {},
    "goal_data": {},
    "expenditure": {},
    "goal_income": {},
    "training_data": {},
    "training_income": {},
    "expenditure_data": {},
}


def _bump_analysis(
    db: Session,
    gym_id: int,
    goal_name: str,
    training_name: str,
    amount: float,
) -> None:

    row = (
        db.query(GymAnalysis)
        .filter(GymAnalysis.gym_id == gym_id)
        .with_for_update()
        .first()
    )

    if row is None:
        data = copy.deepcopy(DEFAULT_ANALYSIS)      
    else:
        raw = row.analysis
        if isinstance(raw, dict):                    
            data = raw
        else:                                        
            try:
                data = json.loads(raw) if raw else copy.deepcopy(DEFAULT_ANALYSIS)
            except json.JSONDecodeError:
                print(f"[WARN] Corrupt analysis JSON for gym {gym_id}; aborting bump.")
                return

    goal_dict  = data.setdefault("goal_data", {})
    train_dict = data.setdefault("training_data", {})

    goal_dict[goal_name]       = goal_dict.get(goal_name, 0)      + amount
    train_dict[training_name]  = train_dict.get(training_name, 0) + amount

    analysis_obj = data                   

    now = datetime.now()

    if row is None:
        db.add(
            GymAnalysis(
                gym_id        = gym_id,
                analysis_type = "revenue",
                analysis_name = "gym_revenue_analysis",
                value         = 0.0,
                analysis      = analysis_obj,
                created_at    = now,
                updated_at    = now,
            )
        )
    else:
        row.analysis   = analysis_obj    
        row.updated_at = now

    db.commit()


class PostFees(BaseModel):
    client_id: int
    gym_id: int
    plan_id:int
    batch_id:int
    type: str
    fees:int
    gst_percentage:int
    gst_type:str
    payment_method:Optional[str]=None
    payment_reference_number:Optional[str]=None
    payment_type: Optional[str]=None
    membership_id:Optional[int]= None
    total_amount:float
    joined_date: Optional[str] = None  # Format: "YYYY-MM-DD"
    expires_at: Optional[str] = None   # Format: "YYYY-MM-DD"
    request_type: Optional[bool]=False


@router.post("/gym/update_fee_status")
async def update_fee_status(
    post_fees: PostFees,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis),
):
    try:
        print(f"[update_fee_status] Received payload: {post_fees.dict()}")

        client = db.query(Client).filter(Client.client_id == post_fees.client_id).first()
        if not client:
            raise HTTPException(404, "Client not found")
        print(f"[update_fee_status] Located client: client_id={client.client_id}, gym_id={client.gym_id}")

        gym_id    = post_fees.gym_id

        if client.gym_id:
            if client.gym_id!=gym_id:
                admission_num=db.query(ClientGym).filter(ClientGym.client_id==client.client_id).first()
                client=db.query(Client).filter(Client.client_id==client.client_id).first()

                # Get expires_at from FittbotGymMembership (latest joined_at)
                latest_membership = (
                    db.query(FittbotGymMembership)
                    .filter(
                        FittbotGymMembership.client_id == str(client.client_id),
                        FittbotGymMembership.gym_id == str(client.gym_id)
                    )
                    .order_by(desc(FittbotGymMembership.joined_at))
                    .first()
                )
                latest_membership_id=latest_membership.id
                expires_at_value = latest_membership.expires_at if latest_membership else None
                starts_at_value = latest_membership.joined_at if (latest_membership and latest_membership.joined_at) else None
                print(f"[update_fee_status] Latest historical membership for transfer: starts_at={starts_at_value}, expires_at={expires_at_value}")

                old_row_data = {
                "gym_client_id":  admission_num.gym_client_id if admission_num.gym_client_id else None,
                "client_id":      client.client_id,
                "gym_id":         client.gym_id,
                "name":           client.name,
                "profile":        client.profile,
                "location":       client.location,
                "email":          client.email,
                "contact":        client.contact,
                "lifestyle":      client.lifestyle,
                "medical_issues": client.medical_issues,
                "batch_id":       client.batch_id,
                "training_id":    client.training_id,
                "age":            client.age,
                "goals":          client.goals,
                "gender":         client.gender,
                "height":         client.height,
                "weight":         client.weight,
                "bmi":            client.bmi,
                "joined_date":    client.joined_date,
                "status":         client.status,
                "dob":            client.dob,
                "admission_number": admission_num.admission_number if admission_num else None,
                "starts_at":      starts_at_value,
                "expires_at":     expires_at_value,
                "latest_membership_id":latest_membership_id
                }

                db.add(OldGymData(**old_row_data))
                print(f"[update_fee_status] Archived prior gym association for client_id={client.client_id}")

        client.gym_id=gym_id
        client.batch_id=post_fees.batch_id
        client.training_id=post_fees.plan_id
        print(
            "[update_fee_status] Updated client gym linkage: "
            f"gym_id={client.gym_id}, batch_id={client.batch_id}, training_id={client.training_id}"
        )

        db.flush()
        print("[update_fee_status] Initial client reassignment flushed.")


        old_data=db.query(OldGymData).filter(OldGymData.client_id==client.client_id, OldGymData.gym_id==gym_id).first()
        if old_data:
                db.delete(old_data)
                db.flush()
                print("[update_fee_status] Removed duplicate old data record after reassignment.")

        client_id = post_fees.client_id
        fees      = post_fees.fees
        plan_id   = post_fees.plan_id
        fee_type  = post_fees.type
        total_amount = post_fees.total_amount
        print(
            "[update_fee_status] Fee processing inputs: "
            f"client_id={client_id}, plan_id={plan_id}, fee_type={fee_type}, "
            f"fees={fees}, total_amount={total_amount}"
        )

        client.status = "active"
        client.training_id = plan_id
        db.flush()

        db.add(
            FeeHistory(
                gym_id=gym_id,
                client_id=client_id,
                fees_paid=total_amount,
                type=fee_type,
                payment_date=date.today(),
            )
        )
        print("[update_fee_status] Fee history entry staged.")

        month_tag = datetime.now().strftime("%Y-%m")
        rec = (
            db.query(GymMonthlyData)
            .filter(
                GymMonthlyData.gym_id == gym_id,
                GymMonthlyData.month_year.like(f"{month_tag}%"),
            )
            .first()
        )
        if rec:
            rec.income += fees
            print(f"[update_fee_status] Updated GymMonthlyData for {month_tag}: +{fees} income.")
        else:
            db.add(
                GymMonthlyData(
                    gym_id=gym_id,
                    month_year=datetime.now().strftime("%Y-%m-%d"),
                    income=fees,
                    expenditure=0,
                    new_entrants=0,
                )
            )
            print(f"[update_fee_status] Created GymMonthlyData record for {month_tag} with income={fees}.")

        training_name = (
            db.query(GymPlans).filter(GymPlans.id == plan_id).one().plans
        )
        goal_name = client.goals or "unknown_goal"
        print(f"[update_fee_status] Calling analysis bump with goal={goal_name}, training_name={training_name}.")
        _bump_analysis(db, gym_id, goal_name, training_name, fees)


        gym        = db.query(Gym).filter(Gym.gym_id == gym_id).first()
        if not gym:
            raise HTTPException(404, "Gym not found")

        gym_owner  = db.query(GymOwner).filter(GymOwner.owner_id == gym.owner_id).first()
        plan       = db.query(GymPlans).filter(GymPlans.id == client.training_id).first()
        account    = db.query(AccountDetails).filter(AccountDetails.gym_id == gym_id).first()

        discount_pct = (fees / plan.amount) * 100
        today_date = date.today()

        import_data = db.query(GymImportData).filter(
            GymImportData.gym_id == gym_id,
            GymImportData.client_contact == client.contact
        ).first()
        import_data_expiry = None
        if import_data and getattr(import_data, "expires_at", None):
            expiry_value = import_data.expires_at
            if isinstance(expiry_value, date):
                import_data_expiry = expiry_value
            else:
                try:
                    import_data_expiry = datetime.strptime(str(expiry_value)[:10], "%Y-%m-%d").date()
                except ValueError:
                    import_data_expiry = None


        import_data_joined_at = None
        if import_data and getattr(import_data, "joined_at", None):
            joined_value = import_data.joined_at
            if isinstance(joined_value, date):
                import_data_joined_at = joined_value
            else:
                try:
                    import_data_joined_at = datetime.strptime(str(joined_value)[:10], "%Y-%m-%d").date()
                except ValueError:
                    import_data_joined_at = None

        def _parse_date(value):
            if isinstance(value, date):
                return value
            if value:
                try:
                    return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
                except ValueError:
                    return None
            return None
        provided_expiry_date = _parse_date(post_fees.expires_at)

        def _add_months_safe(base: date, months: int) -> date:
            months = int(months or 0)
            month_index = base.month - 1 + months
            year = base.year + month_index // 12
            month = (month_index % 12) + 1
            max_day = calendar.monthrange(year, month)[1]
            return date(year, month, min(base.day, max_day))

        # Determine cycle_start_date and calculated_expiry_date based on payment_type
        cycle_start_date: date
        calculated_expiry_date: Optional[date] = None
        actual_joined_at: date = today_date  # Default to today, will be set properly in normal flow
        payment_type = post_fees.payment_type

        if payment_type in ("gym_membership", "personal_training"):
            # For gym_membership and pt: always use today as joined_date
            cycle_start_date = today_date

            # Calculate expiry from today
            plan_duration_months = int(plan.duration) if plan and plan.duration is not None else 0
            calculated_expiry_date = _add_months_safe(cycle_start_date, plan_duration_months)

            # Add bonus if exists
            if plan and plan.bonus and plan.bonus_type:
                bonus_type = str(plan.bonus_type).lower().strip()
                if bonus_type in ("day", "2"):
                    calculated_expiry_date = calculated_expiry_date + timedelta(days=int(plan.bonus))
                elif bonus_type in ("month", "1"):
                    calculated_expiry_date = _add_months_safe(calculated_expiry_date, int(plan.bonus))

        else:
            # For normal/renewal: check if import data has valid expiry
            import_expiry_valid = (
                import_data_expiry is not None
                and import_data_expiry >= today_date
            )

            if import_expiry_valid:
                # Use import expiry directly, don't recalculate
                calculated_expiry_date = import_data_expiry
                # Use import joined_at if available, otherwise use today's date
                if import_data_joined_at is not None:
                    cycle_start_date = import_data_joined_at
                else:
                    cycle_start_date = today_date
            else:
                # NEW 4-CASE LOGIC for FittbotGymMembership joined_at and expires_at

                # Parse provided joined_date from request if available
                provided_joined_date = None
                if post_fees.joined_date:
                    provided_joined_date = _parse_date(post_fees.joined_date)
                    print(f"[DEBUG] provided_joined_date (parsed): {provided_joined_date}")
                else:
                    print(f"[DEBUG] No joined_date provided in request")

                # Determine actual_joined_at and expiry_calculation_base based on 4 cases
                expiry_calculation_base: date

                if client.expiry == "start_of_the_month":
                    print(f"[DEBUG] Mode: start_of_the_month")
                    # Case 1 & 2: start_of_month mode
                    if provided_joined_date:
                        # Case 2: start_of_month + joining_date GIVEN
                        print(f"[DEBUG] CASE 2: start_of_month + joining_date GIVEN")
                        actual_joined_at = provided_joined_date
                        expiry_calculation_base = date(provided_joined_date.year, provided_joined_date.month, 1)
                    else:
                        # Case 1: start_of_month + NO joining_date given
                        print(f"[DEBUG] CASE 1: start_of_month + NO joining_date")
                        actual_joined_at = today_date
                        expiry_calculation_base = date(today_date.year, today_date.month, 1)
                        print(f"[DEBUG] actual_joined_at: {actual_joined_at}")
                        print(f"[DEBUG] expiry_calculation_base (1st of month): {expiry_calculation_base}")
                else:
                    print(f"[DEBUG] Mode: joining_date (or default)")
                    # Case 3 & 4: joining_date mode (or default)
                    if provided_joined_date:
                        # Case 4: joined_date + joining_date GIVEN
                        print(f"[DEBUG] CASE 4: joining_date + joining_date GIVEN")
                        actual_joined_at = provided_joined_date
                        expiry_calculation_base = provided_joined_date
                        print(f"[DEBUG] actual_joined_at: {actual_joined_at}")
                        print(f"[DEBUG] expiry_calculation_base: {expiry_calculation_base}")
                    else:
                        # Case 3: joined_date + NO joining_date given
                        print(f"[DEBUG] CASE 3: joining_date + NO joining_date")
                        actual_joined_at = today_date
                        expiry_calculation_base = today_date
                        print(f"[DEBUG] actual_joined_at: {actual_joined_at}")
                        print(f"[DEBUG] expiry_calculation_base: {expiry_calculation_base}")

                cycle_start_date = actual_joined_at  # For receipt and other uses
                print(f"[DEBUG] cycle_start_date: {cycle_start_date}")

                # Calculate expiry from expiry_calculation_base
                plan_duration_months = int(plan.duration) if plan and plan.duration is not None else 0
                print(f"[DEBUG] plan_duration_months: {plan_duration_months}")
                calculated_expiry_date = _add_months_safe(expiry_calculation_base, plan_duration_months)
                print(f"[DEBUG] calculated_expiry_date (after duration): {calculated_expiry_date}")

                # Add bonus if exists
                if plan and plan.bonus and plan.bonus_type:
                    bonus_type = str(plan.bonus_type).lower().strip()
                    print(f"[DEBUG] bonus: {plan.bonus}, bonus_type: {bonus_type}")
                    if bonus_type in ("day", "2"):
                        calculated_expiry_date = calculated_expiry_date + timedelta(days=int(plan.bonus))
                        print(f"[DEBUG] calculated_expiry_date (after day bonus): {calculated_expiry_date}")
                    elif bonus_type in ("month", "1"):
                        calculated_expiry_date = _add_months_safe(calculated_expiry_date, int(plan.bonus))
                        print(f"[DEBUG] calculated_expiry_date (after month bonus): {calculated_expiry_date}")
                else:
                    print(f"[DEBUG] No bonus to apply")

                # If caller provided an explicit expiry date, honor it
                if provided_expiry_date:
                    calculated_expiry_date = provided_expiry_date
                    print(f"[DEBUG] Override calculated_expiry_date with provided_expiry_date: {calculated_expiry_date}")


        if not calculated_expiry_date:
            calculated_expiry_date = cycle_start_date


        due_date = calculated_expiry_date

        new_receipt = FeesReceipt(
            client_id               = client.client_id,
            gym_id                  = gym_id,
            client_name             = client.name,
            gym_name                = gym.name,
            gym_logo                = gym.logo,
            gym_contact             = gym_owner.contact_number,
            gym_location            = gym.location,
            plan_id                 = client.training_id,
            plan_description        = plan.plans,
            fees                    = plan.amount,
            discount                = discount_pct,
            discounted_fees         = fees,
            due_date                = due_date,
            invoice_number          = None,  # Will be set after flush
            client_contact          = client.contact,
            bank_details            = account.account_number if account else "",
            ifsc_code               = account.account_ifsccode if account else "",
            account_holder_name     = account.account_holdername if account else "",
            invoice_date            = datetime.now().date(),
            payment_method          = post_fees.payment_method,
            gst_number              = account.gst_number if account else "",
            bank_name               = account.bank_name if account else "",
            branch                  = account.account_branch if account else "",
            client_email            = client.email,
            mail_status             = False,
            payment_date            = cycle_start_date,
            payment_reference_number= post_fees.payment_reference_number,
            created_at              = datetime.now(),
            update_at               = datetime.now(),
            gst_percentage          = post_fees.gst_percentage,
            gst_type                = post_fees.gst_type,
            total_amount            = total_amount,
            fees_type="Renewal"
        )
        db.add(new_receipt)
        db.flush()  # Get receipt_id without committing
        print(f"[update_gstatus] Created FeesReceipt placeholder id={new_receipt.receipt_id}.")

        # Generate invoice number based on per-gym count
        gym_receipt_count = db.query(FeesReceipt).filter(FeesReceipt.gym_id == gym_id).count()
        new_receipt.invoice_number = f"{gym.location[:3].upper()}-{gym.gym_id}-{gym_receipt_count}"
        print(f"[update_fee_status] Assigned invoice_number={new_receipt.invoice_number} (count={gym_receipt_count}).")

        db.flush()
        db.refresh(new_receipt)
        print("[update_fee_status] Fee receipt flushed.")


        payload = {
            "invoice_data": {
              
                "receipt_id":        new_receipt.receipt_id,
                "client_id":         new_receipt.client_id,
                "gym_id":            new_receipt.gym_id,
                "client_name":       new_receipt.client_name,
                "gym_name":          new_receipt.gym_name,
                "gym_logo":          new_receipt.gym_logo,
                "gym_contact":       new_receipt.gym_contact,
                "gym_location":      new_receipt.gym_location,
                "plan_id":           new_receipt.plan_id,
                "plan_description":  new_receipt.plan_description,
                "fees":              int(new_receipt.fees),
                "discount":          int(new_receipt.discount),
                "discounted_fees":   int(new_receipt.discounted_fees),
                "due_date":          str(new_receipt.due_date),
                "invoice_number":    new_receipt.invoice_number,
                "client_contact":    new_receipt.client_contact,
                "bank_details":      new_receipt.bank_details,
                "ifsc_code":         new_receipt.ifsc_code,
                "account_holder":    new_receipt.account_holder_name,
                "invoice_date":      str(new_receipt.invoice_date),
                "payment_method":    new_receipt.payment_method,
                "gst_number":        new_receipt.gst_number,
                "client_email":      new_receipt.client_email,
                "payment_date":      str(new_receipt.payment_date),
                "payment_reference": new_receipt.payment_reference_number,
                "gst_type":          new_receipt.gst_type,
                "gst_percentage":    new_receipt.gst_percentage,
                "discounted_price":  new_receipt.fees - new_receipt.discounted_fees,
                "branch":            new_receipt.branch,
                "invoice_type":      "receipt",
            }
        }
        try:
            # Use Lambda retry utility - automatically retries on throttling/errors
            invoke_lambda_with_retry(
                lambda_client,
                FunctionName=LAMBDA_FUNCTION_NAME,
                InvocationType="Event",
                Payload=json.dumps(payload).encode(),
            )
            print("[update_fee_status] Receipt dispatch enqueued to Lambda.")
        except Exception as e:
            print(f"[update_fee_status] Failed to enqueue receipt: {e}")
            raise HTTPException(500, f"Could not enqueue receipt for mailing: {e}")

        duration = plan_duration_months
        fees=db.query(GymFees).filter(GymFees.client_id==client.client_id).first()
        print(f"[update_fee_status] Plan duration for membership update: {duration} months.")

        start_dt = cycle_start_date
        end_dt = calculated_expiry_date
        print(
            "[update_fee_status] Fee schedule aligned: "
            f"start_dt={start_dt}, end_dt={end_dt}"
        )

        if fees:
            fees.start_date=start_dt
            fees.end_date= end_dt
            print("[update_fee_status] Updated existing GymFees range.")
        else:
            print("[update_fee_status] Creating new GymFees range.")
            db.add(
                GymFees(
                    client_id=client.client_id,
                    start_date=start_dt,
                    end_date=end_dt
                )
            )
        db.flush()


        patterns = [
            f"gym{gym_id}:feesReceipt:*",
            f"{client_id}:fees",
            f"gym:{gym_id}:members",
            f"gym:{gym_id}:collection",
            f"gym:{gym_id}:monthly_data",
            f"gym:{gym_id}:analysis",
            f"gym:{gym_id}:hourlyagg",
            f"gym:{gym_id}:clientdata",
            f"gym:{gym_id}:custom_interval",
            f"gym:{gym_id}:specific_month_year",
            f"gym:{gym_id}:current_month*", 
            f"gym:{gym_id}:overall",
        ]

        for pat in patterns:
            async for key in redis.scan_iter(match=pat):
                print(f"[update_fee_status] Clearing Redis key={key} for pattern={pat}")
                await redis.delete(key)

        print(f"[update_fee_status] payment_type={post_fees.payment_type}")

        if post_fees.payment_type in ("gym_membership", "personal_training"):

            membership_status = db.query(FittbotGymMembership).filter(
                FittbotGymMembership.id == post_fees.membership_id
            ).first()

            # Expire any other active memberships for this client+gym (excluding the one being activated)
            if membership_status:
                other_active_memberships = db.query(FittbotGymMembership).filter(
                    FittbotGymMembership.client_id == str(post_fees.client_id),
                    FittbotGymMembership.gym_id == str(post_fees.gym_id),
                    FittbotGymMembership.status == "active",
                    FittbotGymMembership.id != post_fees.membership_id  # Exclude the one being activated
                ).all()

                for other_membership in other_active_memberships:
                    other_membership.status = "expired"
                    print(f"[update_fee_status] Expired other membership id={other_membership.id}")

                if other_active_memberships:
                    db.flush()
                    print(f"[update_fee_status] Expired {len(other_active_memberships)} other active membership(s)")

            if membership_status and membership_status.entitlement_id:
             
                entitlement = db.query(Entitlement).filter(
                    Entitlement.id == membership_status.entitlement_id
                ).first()

                if entitlement and entitlement.order_item_id:
                 
                    order_item = db.query(OrderItem).filter(
                        OrderItem.id == entitlement.order_item_id
                    ).first()

                    if order_item and order_item.order_id:
                       
                        payment_row = (
                            db.query(FittbotPayment)
                            .filter(FittbotPayment.entitlement_id == order_item.order_id)
                            .first()
                        )
                        if payment_row:
                          
                            payout = Payout(
                                payment_id=payment_row.id,
                                gym_id=int(post_fees.gym_id),
                                gym_owner_id=None,
                                amount_gross=payment_row.amount_net,
                                amount_net=payment_row.amount_net,
                                status="ready_for_transfer",
                            )
                            db.add(payout)
                            print(f"[update_fee_status] Created Payout: payment_id={payment_row.id}, gym_id={post_fees.gym_id}, order_id={order_item.order_id}")
                        else:
                            print(f"[update_fee_status] FittbotPayment not found for order_id={order_item.order_id}")
                    else:
                        print(f"[update_fee_status] OrderItem not found for order_item_id={entitlement.order_item_id if entitlement else 'N/A'}")
                else:
                    print(f"[update_fee_status] Entitlement not found for entitlement_id={membership_status.entitlement_id}")

            membership_status.status = "active"
            membership_status.joined_at = actual_joined_at
            membership_status.amount = total_amount
            membership_status.expires_at = calculated_expiry_date
            db.flush()
            print(f"[update_fee_status] Updated membership_id={membership_status.id} expiry to {calculated_expiry_date}.")

        else:
            # Expire any existing active memberships for this client+gym before creating new one
            existing_active_memberships = db.query(FittbotGymMembership).filter(
                FittbotGymMembership.client_id == str(post_fees.client_id),
                FittbotGymMembership.gym_id == str(post_fees.gym_id),
                FittbotGymMembership.status == "active"
            ).all()

            for existing_membership in existing_active_memberships:
                existing_membership.status = "expired"
                print(f"[update_fee_status] Expired existing membership id={existing_membership.id}")

            if existing_active_memberships:
                db.flush()
                print(f"[update_fee_status] Expired {len(existing_active_memberships)} existing active membership(s)")

            print(
                "[update_fee_status] Creating new membership: "
                f"joined_at={actual_joined_at}, expires_at={calculated_expiry_date}, "
                f"duration={plan_duration_months}, "
                f"bonus={plan.bonus if plan else None}, bonus_type={plan.bonus_type if plan else None}"
            )
            print(f"[DEBUG] ===== ABOUT TO CREATE FittbotGymMembership RECORD =====")
            print(f"[DEBUG] gym_id: {post_fees.gym_id}")
            print(f"[DEBUG] client_id: {post_fees.client_id}")
            print(f"[DEBUG] plan_id: {post_fees.plan_id}")
            print(f"[DEBUG] type: normal")
            print(f"[DEBUG] amount: {post_fees.total_amount}")
            print(f"[DEBUG] purchased_at: {datetime.now()}")
            print(f"[DEBUG] status: active")
            print(f"[DEBUG] joined_at VALUE TO BE INSERTED: {actual_joined_at}")
            print(f"[DEBUG] expires_at VALUE TO BE INSERTED: {calculated_expiry_date}")
            print(f"[DEBUG] ===== CREATING RECORD NOW =====")

            new_membership = FittbotGymMembership(
                gym_id=str(post_fees.gym_id),
                client_id=str(post_fees.client_id),
                plan_id=post_fees.plan_id,
                type="normal",
                amount=post_fees.total_amount,
                purchased_at=datetime.now(),
                status="active",
                joined_at=actual_joined_at,
                expires_at=calculated_expiry_date,
            )
            db.add(new_membership)
            db.flush()
            print(f"[DEBUG] AFTER FLUSH - Record ID: {new_membership.id}")
            print(f"[DEBUG] AFTER FLUSH - joined_at from object: {new_membership.joined_at}")
            print(f"[DEBUG] AFTER FLUSH - expires_at from object: {new_membership.expires_at}")
            print(f"[update_fee_status] Created new membership record id={new_membership.id if hasattr(new_membership, 'id') else 'unknown'}, joined_at={actual_joined_at}, expires_at={calculated_expiry_date}.")

        # Invalidate analytics and monthly data caches (for real-time data)
        redis_key_analysis = f"gym:{post_fees.gym_id}:analysis"
        if await redis.exists(redis_key_analysis):
            await redis.delete(redis_key_analysis)

        redis_key_monthly = f"gym:{post_fees.gym_id}:monthly_data"
        if await redis.exists(redis_key_monthly):
            await redis.delete(redis_key_monthly)


    
        join_request = db.query(GymJoinRequest).filter(
                GymJoinRequest.gym_id == post_fees.gym_id,
                GymJoinRequest.client_id == client.client_id,
                GymJoinRequest.status=="pending"
                ).order_by(GymJoinRequest.id.desc()).first()

        if join_request:
                join_request.status = "onboarded"
                join_request.updated_at = datetime.now()

        # Final atomic commit - all changes committed together
        db.commit()
        print("[update_fee_status] All changes committed atomically.")

        return {"message": "Fee status updated and history recorded", "status": 200}

    except Exception as e:
        db.rollback()
        print(f"[update_fee_status] Error: {e}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred.")
     



@router.get("/gym/hourly_agg")
async def get_hourly_aggregation(
    gym_id: int = Query(...),
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis)
):
    try:
        response = {}
        today = date.today()
        redis_key_hourly = f"gym:{gym_id}:daily_hourlyagg:{today.strftime('%Y-%m-%d')}"
        cached_hourly_data = await redis.get(redis_key_hourly)
        if cached_hourly_data:
            print("Daily hourly aggregation data fetched from Redis.")
            hourly_agg_data = json.loads(cached_hourly_data)
        else:
            print("Daily hourly aggregation data not found in Redis, querying DailyGymHourlyAgg table.")

            db_record = db.query(DailyGymHourlyAgg).filter(
                DailyGymHourlyAgg.gym_id == gym_id,
                DailyGymHourlyAgg.agg_date == today
            ).first()
            if not db_record:
                hourly_agg_data = {}
            else:
                hourly_agg_data = {
                    "4-6": db_record.col_4_6 or 0,
                    "6-8": db_record.col_6_8 or 0,
                    "8-10": db_record.col_8_10 or 0,
                    "10-12": db_record.col_10_12 or 0,
                    "12-14": db_record.col_12_14 or 0,
                    "14-16": db_record.col_14_16 or 0,
                    "16-18": db_record.col_16_18 or 0,
                    "18-20": db_record.col_18_20 or 0,
                    "20-22": db_record.col_20_22 or 0,
                    "22-24": db_record.col_22_24 or 0,
                }
            await redis.set(redis_key_hourly, json.dumps(hourly_agg_data), ex=3600)  # 1 hour cache

        response["hourly_agg"] = hourly_agg_data
            

        # REAL-TIME ANALYSIS - Fetch from FittbotGymMembership, Client, and Expenditure tables
        redis_key = f"gym:{gym_id}:analysis"
        cached = await redis.get(redis_key)

        if cached:
            print("Analysis data fetched from Redis.")
            analysis = json.loads(cached)
        else:
            print("Analysis data not found in Redis, calculating real-time from database.")

            def to_list(section: Dict[str, float]) -> List[Dict[str, float]]:
                return [{"name": k, "value": v} for k, v in section.items()]

            # 1. GENDER ANALYSIS - From Client table
            gender_data = {}
            clients = db.query(Client).filter(Client.gym_id == gym_id).all()

            for client in clients:
                gender = client.gender if client.gender else "Unknown"
                gender_data[gender] = gender_data.get(gender, 0) + 1

            # Convert to percentage
            total_clients = sum(gender_data.values())
            if total_clients > 0:
                gender_percentage = {k: round((v / total_clients) * 100, 2) for k, v in gender_data.items()}
            else:
                gender_percentage = {}

            # 2. EXPENDITURE ANALYSIS - From Expenditure table
            expenditure_data = {}
            expenditures = db.query(Expenditure).filter(Expenditure.gym_id == gym_id).all()

            for exp in expenditures:
                exp_type = exp.expenditure_type if exp.expenditure_type else "Other"
                expenditure_data[exp_type] = expenditure_data.get(exp_type, 0) + float(exp.amount or 0)

            # 3. GOAL_INCOME ANALYSIS - From FittbotGymMembership + Client (by goal)
            goal_income_data = {}

            # Get all memberships with their client goals
            memberships = db.query(FittbotGymMembership).filter(
                FittbotGymMembership.gym_id == str(gym_id)
            ).all()

            for membership in memberships:
                # Get client's goal
                client = db.query(Client).filter(Client.client_id == membership.client_id).first()
                if client and client.goals:
                    goal = client.goals
                    amount = float(membership.amount or 0)
                    goal_income_data[goal] = goal_income_data.get(goal, 0) + amount

            # Convert goal_income to percentage
            total_goal_income = sum(goal_income_data.values())
            if total_goal_income > 0:
                goal_income_percentage = {k: round((v / total_goal_income) * 100, 2) for k, v in goal_income_data.items()}
            else:
                goal_income_percentage = {}

            analysis: Dict[str, List[Dict[str, float]]] = {
                "gender": to_list(gender_percentage),
                "expenditure": to_list(expenditure_data),
                "goal_income": to_list(goal_income_percentage),
                "training_income": []  # Not needed as per requirement
            }

            print("Real-time analysis calculated:", analysis)

            # Cache for 10 minutes (600 seconds) - shorter TTL for real-time data
            await redis.set(redis_key, json.dumps(analysis), ex=600)

        response["analysis"] = analysis

 
        # REAL-TIME MONTHLY DATA - Fetch from FittbotGymMembership and Expenditure tables
        redis_key_monthly = f"gym:{gym_id}:monthly_data"
        cached_monthly_data = await redis.get(redis_key_monthly)

        if cached_monthly_data:
            print("Monthly data fetched from Redis.")
            monthly_data = json.loads(cached_monthly_data)
        else:
            print("Monthly data not found in Redis, calculating real-time from database.")

            from sqlalchemy import extract
            from datetime import datetime
            from collections import defaultdict

            # Get all memberships and group by month-year
            memberships = db.query(FittbotGymMembership).filter(
                FittbotGymMembership.gym_id == str(gym_id),
                FittbotGymMembership.joined_at.isnot(None)
            ).all()

            # Get all expenditures and group by month-year
            expenditures = db.query(Expenditure).filter(
                Expenditure.gym_id == gym_id
            ).all()

            # Dictionary to hold monthly aggregated data
            # Using set to track unique client_ids per month
            monthly_dict = defaultdict(lambda: {"income": 0, "expenditure": 0, "new_entrants": set()})

            # Process memberships - income and new_entrants
            for membership in memberships:
                if membership.joined_at:
                    month_key = membership.joined_at.strftime("%Y-%m")
                    monthly_dict[month_key]["income"] += float(membership.amount or 0)
                    monthly_dict[month_key]["new_entrants"].add(membership.client_id)  # Track unique client_ids

            # Process expenditures
            for exp in expenditures:
                if exp.date:
                    month_key = exp.date.strftime("%Y-%m")
                    monthly_dict[month_key]["expenditure"] += float(exp.amount or 0)

            # Group by year and format
            grouped_data = {}
            for month_key, data in monthly_dict.items():
                year = int(month_key.split("-")[0])
                if year not in grouped_data:
                    grouped_data[year] = []
                grouped_data[year].append({
                    "month_year": month_key,
                    "income": data["income"],
                    "expenditure": data["expenditure"],
                    "new_entrants": len(data["new_entrants"])  # Convert set to count of unique clients
                })

            # Sort months within each year
            for year in grouped_data:
                grouped_data[year] = sorted(grouped_data[year], key=lambda x: x["month_year"])

            # Get last 2 years
            if grouped_data:
                sorted_years = sorted(grouped_data.keys(), reverse=True)
                filtered_years = sorted_years[:2] if len(sorted_years) > 1 else sorted_years
                monthly_data = {year: grouped_data[year] for year in filtered_years}
            else:
                monthly_data = {}

            print("Real-time monthly data calculated:", monthly_data)

            # Cache for 10 minutes (600 seconds)
            await redis.set(redis_key_monthly, json.dumps(monthly_data), ex=600)

        response["monthly_data"] = monthly_data
 
        print("Response prepared:", response)
        return {
            "status": 200,
            "message": "Data fetched successfully.",
            "data": response
        }
 
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred.")

@router.get("/gym/plans_and_batches")
async def get_plans_and_batches(
    gym_id: int = Query(...),
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis)
):

    try:   
        response = {}
        redis_key_plans = f"gym:{gym_id}:plans"
        redis_key_batches = f"gym:{gym_id}:batches"
        cached_plans = await redis.get(redis_key_plans)
        if cached_plans:
            plans_data = json.loads(cached_plans)
            response["plans"] = plans_data

        else:
            print("Plans data not found in Redis, querying the database.")
            plans_records = db.query(GymPlans).filter(GymPlans.gym_id == gym_id).all()
            if plans_records: 
                plans_data = [
                    {"id": record.id, "plans": record.plans, "gym_id":record.gym_id, "amount": record.amount,"duration":record.duration, "description":record.description, "personal_training":record.personal_training, "services":record.services} 
                    for index, record in enumerate(plans_records)
                ]
                print("setting redis")
                await redis.set(redis_key_plans, json.dumps(plans_data), ex=86400)
                print("not setting redis")
                response["plans"] = plans_data
            else:
                response["plans"] = []
    
        cached_batches = await redis.get(redis_key_batches)
        if cached_batches:
            print("Batches data fetched from Redis.")
            batches_data = json.loads(cached_batches)
            response["batches"] = batches_data

        else:
            print("Batches data not found in Redis, querying the database.")
            batches_records = db.query(GymBatches).filter(GymBatches.gym_id == gym_id).all()
            if not batches_records:
                response["batches"] = []
            else:
    
                batches_data = [
                    {"id": record.batch_id, "batch_name": record.batch_name,"timing":record.timing, "description":record.description}
                    for index, record in enumerate(batches_records)
                ]
                await redis.set(redis_key_batches, json.dumps(batches_data), ex=86400)
    
                response["batches"] = batches_data

        print("response is",response)

    
        return {"data":response,"status":200,"message":"Plans and batches listed successfully"}
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected error occured: {str(e)}")


class AddPlanRequest(BaseModel):
    gym_id: int
    plan_name: str
    amount: int
    duration:int
    personal_training:bool
    services:Optional[List[str]]=None
    description:Optional[str]=None

class AddBatchRequest(BaseModel):
    gym_id: int
    batch_name: str
    timing: str
    description:Optional[str]=None

@router.post("/gym/add_plan")
async def add_plan(
    request: AddPlanRequest,  
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis)
):
    try:
        gym_id = request.gym_id
        plan_name = request.plan_name
        amount = request.amount
        duration=request.duration
        description = request.description
        personal_training=request.personal_training
        services=request.services

        new_plan = GymPlans(gym_id=gym_id, plans=plan_name, amount=amount,duration=duration, personal_training=personal_training, description=description, services=services)
        db.add(new_plan)
        db.commit()

        db.refresh(new_plan)
        redis_key_plans = f"gym:{gym_id}:plans"
        if await redis.exists(redis_key_plans):
            await redis.delete(redis_key_plans)

        return {"status": 200, "message": "Plan added successfully."}
    except Exception as e:
        db.rollback()  
        print(f"Error adding plan: {e}")
        raise HTTPException(status_code=500, detail=f"Error adding plan: {str(e)}")


@router.post("/gym/add_batch")
async def add_batch(
    request: AddBatchRequest,  
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis)
):
    try:

        gym_id = request.gym_id
        batch_name = request.batch_name
        timing = request.timing
        description=request.description

        new_batch = GymBatches(gym_id=gym_id, batch_name=batch_name, timing=timing, description=description)
        db.add(new_batch)
        db.commit()

        db.refresh(new_batch)
        redis_key_batches = f"gym:{gym_id}:batches"
        if await redis.exists(redis_key_batches):
            await redis.delete(redis_key_batches)

        return {"status": 200, "message": "Batch added successfully."}
    except Exception as e:
        db.rollback()  
        print(f"Error adding batch: {e}")
        raise HTTPException(status_code=500, detail=f"Error adding batch: {str(e)}")


class EditPlanRequest(BaseModel):
    id: int
    plans: Optional[str] = None
    amount: Optional[int] = None
    duration: Optional[int] = None
    personal_training:bool
    services:Optional[List[str]]=None
    description: Optional[str]=None

@router.put("/gym/edit_plan")
async def edit_plan(
    request: EditPlanRequest,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis)
):
    try:

        plan = db.query(GymPlans).filter(GymPlans.id == request.id).first()
        if not plan:
            raise HTTPException(status_code=404, detail="Plan not found")

        if request.plans is not None:
            plan.plans = request.plans
        if request.amount is not None:
            plan.amount = request.amount
        if request.duration is not None:
            plan.duration = request.duration
        if request.description is not None:
            plan.description = request.description

        plan.personal_training = request.personal_training

        if request.services is not None:
            plan.services = request.services

        db.commit()

        gym_id = plan.gym_id
        redis_key_plans = f"gym:{gym_id}:plans"
        existing_plans = await redis.get(redis_key_plans)
        if existing_plans:
            await redis.delete(redis_key_plans)

        return {"status": 200, "message": "Plan updated successfully."}
    except Exception as e:
        db.rollback()
        print(f"Error updating plan: {e}")
        raise HTTPException(status_code=500, detail=f"Error updating plan: {str(e)}")

@router.delete("/gym/delete_plan")
async def delete_plan(
    id: int,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis)
):
    try:
        plan = db.query(GymPlans).filter(GymPlans.id == id).first()
        if not plan:
            raise HTTPException(status_code=404, detail="Plan not found")

        gym_id = plan.gym_id
        redis_key_plans = f"gym:{gym_id}:plans"
        existing_plans = await redis.get(redis_key_plans)
        if existing_plans:
            await redis.delete(redis_key_plans)
        db.delete(plan)
        db.commit()
        return {"status": 200, "message": "Plan deleted successfully."}
    except Exception as e:
        db.rollback()
        print(f"Error deleting plan: {e}")
        raise HTTPException(status_code=500, detail=f"Error deleting plan: {str(e)}")


class EditBatchRequest(BaseModel):
    batch_id: int
    batch_name: Optional[str] = None
    timing: Optional[str] = None
    description : Optional[str]=None

@router.put("/gym/edit_batch")
async def edit_batch(
    request: EditBatchRequest,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis)
):
    try:
        batch = db.query(GymBatches).filter(GymBatches.batch_id == request.batch_id).first()
        if not batch:
            raise HTTPException(status_code=404, detail="Batch not found")

        if request.batch_name is not None:
            batch.batch_name = request.batch_name
        if request.timing is not None:
            batch.timing = request.timing
        if request.description is not None:
            batch.description= request.description

        db.commit()
        redis_key_batches = f"gym:{batch.gym_id}:batches"
        existing_batches = await redis.get(redis_key_batches)
        if existing_batches:
            await redis.delete(redis_key_batches)

        return {"status": 200, "message": "Batch updated successfully."}
    except Exception as e:
        db.rollback()
        print(f"Error updating batch: {e}")
        raise HTTPException(status_code=500, detail=f"Error updating batch: {str(e)}")

@router.delete("/gym/delete_batch")
async def delete_batch(
    batch_id: int,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis)
):
    try:
        batch = db.query(GymBatches).filter(GymBatches.batch_id == batch_id).first()
        if not batch:
            raise HTTPException(status_code=404, detail="Batch not found")

        gym_id = batch.gym_id
        redis_key_batches = f"gym:{gym_id}:batches"
        existing_batches = await redis.get(redis_key_batches)
        if existing_batches:
            await redis.delete(redis_key_batches)
        db.delete(batch)
        db.commit()

        return {"status": 200, "message": "Batch deleted successfully."}
    except Exception as e:
        db.rollback()
        print(f"Error deleting batch: {e}")
        raise HTTPException(status_code=500, detail=f"Error deleting batch: {str(e)}")



class ExpiryType(str, Enum):
    joining_date = "joining_date"
    start_of_the_month = "start_of_the_month"


class AddClientDataRequest(BaseModel):
    gym_id: int
    full_name: str
    date_of_birth: Optional[date] = None
    gender: str
    contact: str
    email: Optional[str] = None
    height: Optional[float] = None
    weight: Optional[float] = None
    bmi: Optional[float] = None
    job_nature: Optional[str] = None
    fitness_goal: Optional[str] = None
    training_type: int
    batch_type: int
    admission_fee:Optional[int]=None
    discounted_fee:int
    expiry: ExpiryType
    payment_method:str
    payment_reference_number:Optional[str]=None
    admission_number:Optional[str]=None
    fee_collection_start_date: Optional[date]=None
    original_fee:Optional[float]=0
    discount_type:Optional[str]=""
    discount_amount:Optional[float]=0
    discount_percentage:Optional[float]=0
    gst_type:Optional[str]=""
    gst_percentage:Optional[float]=0
    total_amount:Optional[float]=0
    membership_id:Optional[int]=None
    entry_type:str


def generate_password(length=8):
    characters = string.ascii_letters + string.digits
    password = ''.join(random.choice(characters) for _ in range(length))
    return password

def send_welcome_email(user_name, company_name, mobile_number, default_password, login_url, recipient_email, support_email):
    sender_email = os.getenv("SMTP_EMAIL")
    sender_password = os.getenv("SMTP_PASSWORD")
    sender_email = os.getenv("SMTP_EMAIL")
    sender_password = os.getenv("SMTP_PASSWORD")
    msg = EmailMessage()
    subject = f"Welcome to {company_name}! - Your Account Details"
    body = f"""
    <html>
    <body>
        <p>Hello {user_name},</p>
 
        <p>Welcome to <b>{company_name}</b>! Your account has been successfully created. Please find your login details below:</p>
 
        <p><b>Registered Mobile Number:</b> {mobile_number}<br>
        <b>Default Password:</b> {default_password}</p>
 
        <p>For your security, we recommend that you change your default password immediately after your first login.</p>
 
        <p>You can access your account here: <a href="{login_url}">{login_url}</a></p>
 
        <p>If you have any questions or need further assistance, please feel free to contact our support team at <a href="mailto:{support_email}">{support_email}</a>.</p>
 
        <p>Thank you for joining us!</p>
 
        <p>Best regards,<br>
        The {company_name} Team</p>
    </body>
    </html>
    """
 
    
    msg = MIMEMultipart()
    msg["From"] = sender_email
    msg["To"] = recipient_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "html"))
 
    try:
        server = smtplib.SMTP(os.getenv("SMTP_SERVER"), os.getenv("SMTP_PORT"))
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, recipient_email, msg.as_string())
        server.quit()
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False

def send_welcome_sms(user_name,phone_number,default_password,login_url):
    api_key = os.getenv("OTHER_API_KEY")
    sender_id = "NFCFIT"
    entity_id = "1701174022473316577"
    template_id = "1707174123946514804"
    user_name="Champion"
    encoded_message = f"Welcome {user_name}, Your registered mobile number is {phone_number} and your temporary password is {default_password}. Please reset the password immediately after login. You can access your account here {login_url}. Regards- NFCFIT"
 
    url = (
        f"http://pwtpl.com/sms/V1/send-sms-api.php?"
        f"apikey={api_key}&senderid={sender_id}&templateid={template_id}"
        f"&entityid={entity_id}&number={phone_number}&message={encoded_message}&format=json"
    )
 
    try:
        response = requests.get(url)
        if response.status_code == 200:
            json_response = response.json()
            print("SMS API Response:", json_response)
            if json_response.get('status') == 'OK':
                print("SMS sent successfully!")
                return True
            else:
                print("Failed to send SMS. Please check the API credentials and parameters.")
                return False
        else:
            print("HTTP error occurred:", response.status_code)
    except Exception as e:
        print("An error occurred:", e)



def _pick_next_reward(
    ladder: List[RewardGym],
    current_xp: int,
):
    for tier in ladder:
        if tier.xp > current_xp:
            return tier
    return None



class AddClientDataRequest(BaseModel):
    gym_id: int
    full_name: str
    date_of_birth: Optional[date] = None
    gender: str
    contact: str
    email: Optional[str] = None
    height: Optional[float] = None
    weight: Optional[float] = None
    bmi: Optional[float] = None
    job_nature: Optional[str] = None
    fitness_goal: Optional[str] = None
    training_type: int
    batch_type: int
    admission_fee:Optional[int]=None
    discounted_fee:int
    expiry: ExpiryType
    payment_method:str
    payment_reference_number:Optional[str]=None
    admission_number:Optional[str]=None
    fee_collection_start_date: Optional[date]=None
    original_fee:Optional[float]=0
    discount_type:Optional[str]=""
    discount_amount:Optional[float]=0
    discount_percentage:Optional[float]=0
    gst_type:Optional[str]=""
    gst_percentage:Optional[float]=0
    total_amount:Optional[float]=0
    membership_id:Optional[int]=None
    entry_type:str
    old_client:Optional[bool]=False
    new_expiry_date: Optional[date]=None
    request_type: Optional[bool]=False


@router.post("/gym/add_client_data")
async def add_client_data(
    request: AddClientDataRequest,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis)
):
    
    try:
        payload_snapshot = request.dict()
        print(f"[add_client_data] Received payload: {payload_snapshot}")
        gym_id = request.gym_id
        contact = request.contact
        training_type = request.training_type
        batch_type = request.batch_type
        expiry = request.expiry
        admission_fee=request.admission_fee if request.admission_fee is not None else 0
        discounted_fee=request.discounted_fee
        admission_number=request.admission_number
        fee_start_date=request.fee_collection_start_date if request.fee_collection_start_date else date.today()
        original_fee = request.original_fee
        discount_type = request.discount_type
        discount_amount = request.discount_amount
        discount_percentage = request.discount_percentage
        gst_type = request.gst_type
        gst_percentage = request.gst_percentage
        total_amount = request.total_amount
        final_admission_fees = (admission_fee + ((admission_fee * gst_percentage)/100))  if gst_type == "exclusive" and admission_fee > 0 else admission_fee
        entry_type=request.entry_type
        
        
        print(
            "[add_client_data] Derived basics: "
            f"gym_id={gym_id}, contact={contact}, training_type={training_type}, "
            f"batch_type={batch_type}, expiry={expiry}"
        )
        print(
            "[add_client_data] Fee summary: "
            f"admission_fee={admission_fee}, discounted_fee={discounted_fee}, "
            f"admission_number={admission_number}, fee_start_date={fee_start_date}"
        )
        print(
            "[add_client_data] Discount/GST summary: "
            f"original_fee={original_fee}, discount_type={discount_type}, "
            f"discount_amount={discount_amount}, discount_percentage={discount_percentage}, "
            f"gst_type={gst_type}, gst_percentage={gst_percentage}, total_amount={total_amount}, "
            f"final_admission_fees={final_admission_fees}, entry_type={entry_type}"
        )

        # Check if there's matching data in GymImportData based on gym_id and contact
        import_data = db.query(GymImportData).filter(
            GymImportData.gym_id == gym_id,
            GymImportData.client_contact == contact
        ).first()
        
        import_data_expiry = None
        if import_data and getattr(import_data, "expires_at", None):
            expiry_value = import_data.expires_at
            if isinstance(expiry_value, date):
                import_data_expiry = expiry_value
            else:
                try:
                    import_data_expiry = datetime.strptime(str(expiry_value)[:10], "%Y-%m-%d").date()
                except ValueError:
                    import_data_expiry = None

        import_data_joined_at = None
        if import_data and getattr(import_data, "joined_at", None):
            joined_value = import_data.joined_at
            if isinstance(joined_value, date):
                import_data_joined_at = joined_value
            else:
                try:
                    import_data_joined_at = datetime.strptime(str(joined_value)[:10], "%Y-%m-%d").date()
                except ValueError:
                    import_data_joined_at = None

        # If import data exists and has admission_number, use it
        if import_data and import_data.admission_number:
            if not admission_number:  # Only use import admission_number if not already provided
                admission_number = import_data.admission_number
                print(f"Using admission_number from GymImportData: {admission_number}")

        cursor = b'0'
        pattern = f"gym:{gym_id}:*"
        print(f"[add_client_data] Clearing Redis cache with pattern: {pattern}")
        while cursor:
            cursor, keys = await redis.scan(cursor=cursor, match=pattern, count=100)
            if keys:
                await redis.delete(*keys)
        print("[add_client_data] Redis cache cleared for gym scope.")

        existing_client = db.query(Client).filter(Client.contact == contact).first()
        print(f"[add_client_data] Existing client lookup result: {bool(existing_client)}")

        if existing_client:
            print(f"[add_client_data] Processing existing client_id={existing_client.client_id}")
            if existing_client.gym_id is not None:
                print(f"[add_client_data] Existing client already linked to gym_id={existing_client.gym_id}")
                admission_num=db.query(ClientGym).filter(ClientGym.client_id==existing_client.client_id).first()

                # Get expires_at from FittbotGymMembership (latest joined_at)
                latest_membership = (
                    db.query(FittbotGymMembership)
                    .filter(
                        FittbotGymMembership.client_id == str(existing_client.client_id),
                        FittbotGymMembership.gym_id == str(existing_client.gym_id)
                    )
                    .order_by(desc(FittbotGymMembership.joined_at))
                    .first()
                )
                expires_at_value = latest_membership.expires_at if latest_membership else None
                starts_at_value = latest_membership.joined_at if (latest_membership and latest_membership.joined_at) else None

                old_row_data = {
                "gym_client_id":  admission_num.gym_client_id if admission_num.gym_client_id else None,
                "gym_id":         existing_client.gym_id,
                "name":           existing_client.name,
                "profile":        existing_client.profile,
                "location":       existing_client.location,
                "email":          existing_client.email,
                "contact":        existing_client.contact,
                "lifestyle":      existing_client.lifestyle,
                "medical_issues": existing_client.medical_issues,
                "batch_id":       existing_client.batch_id,
                "training_id":    existing_client.training_id,
                "age":            existing_client.age,
                "goals":          existing_client.goals,
                "gender":         existing_client.gender,
                "height":         existing_client.height,
                "weight":         existing_client.weight,
                "bmi":            existing_client.bmi,
                "joined_date":    existing_client.joined_date,
                "status":         existing_client.status,
                "dob":            existing_client.dob,
                "admission_number": admission_num.admission_number if admission_num else None,
                "starts_at":      starts_at_value,
                "expires_at":     expires_at_value,
                }


                db.add(OldGymData(**old_row_data))
                # db.commit()  # REMOVED - commit happens at the end               
                        

            gym_record=db.query(Gym).filter(Gym.gym_id == gym_id).first()
            if gym_record and gym_record.name and gym_record.name.strip():
                first_two = gym_record.name.strip().upper()[:2]
            else:
                first_two = ''.join(random.choices(string.ascii_uppercase, k=2))
            print(f"[add_client_data] gym_client_id prefix resolved to: {first_two}")

            client_count = (
                db.query(func.count())
                .select_from(Client)
                .filter(Client.gym_id == gym_id)
                .scalar()
            )

            old_count = (
                db.query(func.count())
                .select_from(OldGymData)
                .filter(OldGymData.gym_id == gym_id)
                .scalar()
            )

            running_number = (client_count or 0) + (old_count or 0) + 1

            existing_old_client=db.query(OldGymData).filter(OldGymData.gym_id==gym_id,OldGymData.contact==contact).first()


            if existing_old_client:
                print(f"[add_client_data] Re-using historical gym_client_id={existing_old_client.gym_client_id}")
                gym_client_id=existing_old_client.gym_client_id
                admission_number=existing_old_client.admission_number
                db.delete(existing_old_client)

            else:
                try:
                    # Find the maximum running number for this gym in a single query
                    prefix_pattern = f"{first_two}-{gym_id}-%"
                    existing_ids = db.query(ClientGym.gym_client_id).filter(
                        ClientGym.gym_id == gym_id,
                        ClientGym.gym_client_id.like(prefix_pattern)
                    ).all()

                    # Extract numbers from existing gym_client_ids and find max
                    max_number = 0
                    for (id_value,) in existing_ids:
                        try:
                            # Split by '-' and get the last part (the number)
                            parts = id_value.split('-')
                            if len(parts) == 3:
                                num = int(parts[2])
                                max_number = max(max_number, num)
                        except (ValueError, IndexError):
                            continue

                    # Use max + 1 as the next running number
                    running_number = max_number + 1
                    gym_client_id = f"{first_two}-{gym_id}-{running_number}"
                    print(f"[add_client_data] Generated unique gym_client_id={gym_client_id} (max was {max_number})")
                except Exception as e:
                    print(f"[add_client_data] Error generating gym_client_id: {str(e)}. Setting to None.")
                    gym_client_id = None


            # Check if ClientGym mapping already exists for this client_id AND gym_id combination
            gym_client = db.query(ClientGym).filter(
                ClientGym.client_id == existing_client.client_id,
                ClientGym.gym_id == gym_id
            ).first()

            if gym_client:
                print(f"[add_client_data] ClientGym mapping already exists for client_id={existing_client.client_id} and gym_id={gym_id}. Leaving it unchanged.")
                # Leave existing mapping unchanged
            else:
                print(f"[add_client_data] Creating new ClientGym mapping for client_id={existing_client.client_id} and gym_id={gym_id}.")
                new_data = ClientGym(
                    client_id=existing_client.client_id,
                    gym_client_id=gym_client_id,
                    gym_id=gym_id,
                    admission_number=admission_number if admission_number else None
                )
                db.add(new_data)
                # db.commit()  # REMOVED - commit happens at the end



            existing_client.gym_id = gym_id
            existing_client.batch_id = batch_type
            existing_client.training_id = training_type
            existing_client.expiry = expiry
            existing_client.status="active"
            existing_client.access = True
            print(
                "[add_client_data] Updated client core attributes: "
                f"gym_id={existing_client.gym_id}, batch_id={existing_client.batch_id}, "
                f"training_id={existing_client.training_id}, expiry={existing_client.expiry}"
            )
           


            if admission_fee > 0:
                print(f"[add_client_data] Recording admission fee history with amount={final_admission_fees}")
                new_admission_fee = FeeHistory(
                        gym_id=gym_id,
                        client_id=existing_client.client_id,
                        type="admission",
                        fees_paid=final_admission_fees,
                        payment_date=date.today()
                    )
                db.add(new_admission_fee)

            new_fee_history = FeeHistory(
                gym_id=gym_id,
                client_id=existing_client.client_id,
                type="fees",
                fees_paid=total_amount,
                payment_date=date.today()
            )
            print(f"[add_client_data] Recording regular fee history with amount={total_amount}")
            db.add(new_fee_history)

            current_month = datetime.now().strftime("%Y-%m")
            existing_record = db.query(GymMonthlyData).filter(
                GymMonthlyData.gym_id == request.gym_id,
                GymMonthlyData.month_year.like(f"{current_month}%")  
            ).first()

            total_income = total_amount + (final_admission_fees if admission_fee > 0 else 0)
            if existing_record:
                existing_record.income += total_income
                existing_record.new_entrants += 1 
                print(f"[add_client_data] Updated existing monthly data for {current_month} with income delta {total_income}.")
            else:
                new_record = GymMonthlyData(
                    gym_id=request.gym_id,
                    month_year=datetime.now().strftime("%Y-%m-%d"),
                    income= total_income,
                    expenditure=0,
                    new_entrants=1
                )
                db.add(new_record)
                print("[add_client_data] Created new monthly summary entry.")

            # gym_analysis = db.query(GymAnalysis).filter(GymAnalysis.gym_id == gym_id).first()

            # if gym_analysis:
            #     data = gym_analysis.analysis
            #     if isinstance(data, str):
            #         try:
            #             data = json.loads(data)
            #         except json.JSONDecodeError:
            #             pass
            
            # else:
            #     gym_analysis=GymAnalysis(
            #         gym_id=gym_id
            #     )

            #     db.add(gym_analysis)
            #     db.commit()
            #     db.refresh(gym_analysis)
            #     data={}



            # if "goal_data" not in data:
            #     data["goal_data"] = {}

            # if "training_data" not in data:
            #     data["training_data"]={}

            # total_revenue = total_amount + (final_admission_fees if admission_fee > 0 else 0)
            # if existing_client.goals in data["goal_data"]:
            #     data["goal_data"][existing_client.goals] += total_revenue
            # else:
            #     data["goal_data"][existing_client.goals] = total_revenue

            # training_type= db.query(GymPlans).filter(GymPlans.id == existing_client.training_id).one().plans

            # if training_type in data["training_data"]:
            #     data["training_data"][training_type] += total_revenue
            # else:
            #     data["training_data"][training_type] = total_revenue


            # db.refresh(gym_analysis)
            # gym_analysis.analysis = data
            # db.commit()

            scheduler=db.query(ClientScheduler).filter(ClientScheduler.client_id==existing_client.client_id).first()
            if scheduler:
                db.delete(scheduler)
                # db.commit()  # REMOVED - commit happens at the end
                db.add(
                    ClientScheduler(
                        gym_id=gym_id,
                        client_id=existing_client.client_id
                    )
                )
                # db.commit()  # REMOVED - commit happens at the end

            

            collection_key = f"gym:{gym_id}:collection"
            if await redis.exists(collection_key):
                await redis.delete(collection_key)

            redis_key_monthly = f"gym:{gym_id}:monthly_data"
            if await redis.exists(redis_key_monthly):
                await redis.delete(redis_key_monthly)

            redis_key_analysis = f"gym:{gym_id}:analysis"
            if await redis.exists(redis_key_analysis):
                await redis.delete(redis_key_analysis)

            redis_key_hourly = f"gym:{gym_id}:hourlyagg"
            if await redis.exists(redis_key_hourly):
                await redis.delete(redis_key_hourly)


            members_key = f"gym:{gym_id}:members"
            if await redis.exists(members_key):
                await redis.delete(members_key)
    
            new_clients_key = f"gym:{gym_id}:new_clients"
            if await redis.exists(new_clients_key):
                await redis.delete(new_clients_key)

            redis_key = f"gym:{gym_id}:pendingClients"
            if await redis.exists(redis_key):
                await redis.delete(redis_key)

 
            fee_key=f"{existing_client.client_id}:fees"
            if await redis.exists(fee_key):
                await redis.delete(fee_key)
            
            client_key = f"gym:{gym_id}:clientdata"
            if await redis.exists(client_key):
                await redis.delete(client_key)
            
        

            gym = db.query(Gym).filter(Gym.gym_id == request.gym_id).first()
            gym_owner = db.query(GymOwner).filter(GymOwner.owner_id == gym.owner_id).first()
            plan = db.query(GymPlans).filter(GymPlans.id == existing_client.training_id).first()
            account= db.query(AccountDetails).filter(AccountDetails.gym_id == request.gym_id).first()

            # discount = (request.discounted_fee/plan.amount)*100
            discount = discount_amount if discount_type == "amount" else ((discount_percentage/100)* original_fee)
            today_date = date.today()

            def _parse_date(value):
                if isinstance(value, date):
                    return value
                if value:
                    try:
                        return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
                    except ValueError:
                        return None
                return None

            def _add_months_safe(base: date, months: int) -> date:
                months = int(months or 0)
                month_index = base.month - 1 + months
                year = base.year + month_index // 12
                month = (month_index % 12) + 1
                max_day = calendar.monthrange(year, month)[1]
                return date(year, month, min(base.day, max_day))

            # Determine cycle_start_date and 4095 based on entry_type
            cycle_start_date: date
            calculated_expiry_date: Optional[date] = None

            if entry_type in ("gym_membership", "personal_training"):
                # For gym_membership and pt: always use today as joined_date
                cycle_start_date = today_date

                # Calculate expiry from today
                plan_duration_months = int(plan.duration) if plan and plan.duration is not None else 0
                calculated_expiry_date = _add_months_safe(cycle_start_date, plan_duration_months)

                # Add bonus if exists
                if plan and plan.bonus and plan.bonus_type:
                    bonus_type = str(plan.bonus_type).lower().strip()
                    if bonus_type in ("day", "2"):
                        calculated_expiry_date = calculated_expiry_date + timedelta(days=int(plan.bonus))
                    elif bonus_type in ("month", "1"):
                        calculated_expiry_date = _add_months_safe(calculated_expiry_date, int(plan.bonus))

            elif entry_type == "normal":
                # Check if import data has valid expiry
                import_expiry_valid = (
                    import_data_expiry is not None
                    and import_data_expiry >= today_date
                )

                if import_expiry_valid:
                    # Use import expiry directly, don't recalculate
                    calculated_expiry_date = import_data_expiry
                    # Use import joined_at if available, otherwise use today's date
                    if import_data_joined_at is not None:
                        cycle_start_date = import_data_joined_at
                    else:
                        cycle_start_date = today_date
                else:
                    # NEW 4-CASE LOGIC for FittbotGymMembership joined_at and expires_at
                    print(f"[DEBUG-ADD_CLIENT] ===== DATE CALCULATION DEBUG START =====")
                    print(f"[DEBUG-ADD_CLIENT] today_date: {today_date}")
                    print(f"[DEBUG-ADD_CLIENT] request.fee_collection_start_date (raw): {request.fee_collection_start_date}")
                    print(f"[DEBUG-ADD_CLIENT] existing_client.expiry mode: {existing_client.expiry}")

                    # Parse provided fee_collection_start_date from request if available
                    provided_joined_date = None
                    if request.fee_collection_start_date:
                        provided_joined_date = request.fee_collection_start_date if isinstance(request.fee_collection_start_date, date) else _parse_date(request.fee_collection_start_date)
                        print(f"[DEBUG-ADD_CLIENT] provided_joined_date (parsed): {provided_joined_date}")
                    else:
                        print(f"[DEBUG-ADD_CLIENT] No fee_collection_start_date provided in request")

                    # Determine actual_joined_at and expiry_calculation_base based on 4 cases
                    expiry_calculation_base: date

                    if existing_client.expiry == "start_of_the_month":
                        print(f"[DEBUG-ADD_CLIENT] Mode: start_of_the_month")
                        # Case 1 & 2: start_of_month mode
                        if provided_joined_date:
                            # Case 2: start_of_month + joining_date GIVEN
                            print(f"[DEBUG-ADD_CLIENT] CASE 2: start_of_month + joining_date GIVEN")
                            actual_joined_at = provided_joined_date
                            expiry_calculation_base = date(provided_joined_date.year, provided_joined_date.month, 1)
                            print(f"[DEBUG-ADD_CLIENT] actual_joined_at: {actual_joined_at}")
                            print(f"[DEBUG-ADD_CLIENT] expiry_calculation_base (1st of month): {expiry_calculation_base}")
                        else:
                            # Case 1: start_of_month + NO joining_date given
                            print(f"[DEBUG-ADD_CLIENT] CASE 1: start_of_month + NO joining_date")
                            actual_joined_at = today_date
                            expiry_calculation_base = date(today_date.year, today_date.month, 1)
                            print(f"[DEBUG-ADD_CLIENT] actual_joined_at: {actual_joined_at}")
                            print(f"[DEBUG-ADD_CLIENT] expiry_calculation_base (1st of month): {expiry_calculation_base}")
                    else:
                        print(f"[DEBUG-ADD_CLIENT] Mode: joining_date (or default)")
                        # Case 3 & 4: joining_date mode (or default)
                        if provided_joined_date:
                            # Case 4: joined_date + joining_date GIVEN
                            print(f"[DEBUG-ADD_CLIENT] CASE 4: joining_date + joining_date GIVEN")
                            actual_joined_at = provided_joined_date
                            expiry_calculation_base = provided_joined_date
                            print(f"[DEBUG-ADD_CLIENT] actual_joined_at: {actual_joined_at}")
                            print(f"[DEBUG-ADD_CLIENT] expiry_calculation_base: {expiry_calculation_base}")
                        else:
                            # Case 3: joined_date + NO joining_date given
                            print(f"[DEBUG-ADD_CLIENT] CASE 3: joining_date + NO joining_date")
                            actual_joined_at = today_date
                            expiry_calculation_base = today_date
                            print(f"[DEBUG-ADD_CLIENT] actual_joined_at: {actual_joined_at}")
                            print(f"[DEBUG-ADD_CLIENT] expiry_calculation_base: {expiry_calculation_base}")

                    cycle_start_date = actual_joined_at  # For receipt and other uses
                    print(f"[DEBUG-ADD_CLIENT] cycle_start_date: {cycle_start_date}")

                    # Calculate expiry from expiry_calculation_base
                    plan_duration_months = int(plan.duration) if plan and plan.duration is not None else 0
                    print(f"[DEBUG-ADD_CLIENT] plan_duration_months: {plan_duration_months}")
                    calculated_expiry_date = _add_months_safe(expiry_calculation_base, plan_duration_months)
                    print(f"[DEBUG-ADD_CLIENT] calculated_expiry_date (after duration): {calculated_expiry_date}")

                    # Add bonus if exists
                    if plan and plan.bonus and plan.bonus_type:
                        bonus_type = str(plan.bonus_type).lower().strip()
                        print(f"[DEBUG-ADD_CLIENT] bonus: {plan.bonus}, bonus_type: {bonus_type}")
                        if bonus_type in ("day", "2"):
                            calculated_expiry_date = calculated_expiry_date + timedelta(days=int(plan.bonus))
                            print(f"[DEBUG-ADD_CLIENT] calculated_expiry_date (after day bonus): {calculated_expiry_date}")
                        elif bonus_type in ("month", "1"):
                            calculated_expiry_date = _add_months_safe(calculated_expiry_date, int(plan.bonus))
                            print(f"[DEBUG-ADD_CLIENT] calculated_expiry_date (after month bonus): {calculated_expiry_date}")
                    else:
                        print(f"[DEBUG-ADD_CLIENT] No bonus to apply")

                    print(f"[DEBUG-ADD_CLIENT] ===== FINAL DATES =====")
                    print(f"[DEBUG-ADD_CLIENT] FINAL actual_joined_at: {actual_joined_at}")
                    print(f"[DEBUG-ADD_CLIENT] FINAL calculated_expiry_date: {calculated_expiry_date}")
                    print(f"[DEBUG-ADD_CLIENT] ===== DATE CALCULATION DEBUG END =====")
            
            else:
                # Fallback for any other entry_type
                joined_date_value = _parse_date(existing_client.joined_date) or today_date
                cycle_start_date = joined_date_value
                plan_duration_months = int(plan.duration) if plan and plan.duration is not None else 0
                calculated_expiry_date = _add_months_safe(cycle_start_date, plan_duration_months)

            if not calculated_expiry_date:
                calculated_expiry_date = cycle_start_date



            due_date = calculated_expiry_date

            # Set payment_date to match joined_at used in FittbotGymMembership
            # gym_membership/personal_training uses date.today(), normal uses cycle_start_date
            if entry_type in ("gym_membership", "personal_training"):
                receipt_payment_date = date.today()
            else:
                receipt_payment_date = cycle_start_date

            new_receipt = FeesReceipt(
                client_id = existing_client.client_id,
                gym_id = existing_client.gym_id,
                client_name = existing_client.name,
                gym_name = gym.name,
                gym_logo = gym.logo,
                gym_contact = gym_owner.contact_number,
                gym_location = gym.location,
                plan_id = existing_client.training_id,
                plan_description = plan.plans,
                fees = plan.amount,
                discount = discount,
                discounted_fees = discounted_fee,
                due_date = due_date,
                invoice_number = None,  # Will be set after flush
                client_contact = existing_client.contact,
                bank_details = account.account_number if account else "",
                ifsc_code = account.account_ifsccode  if account else "",
                account_holder_name = account.account_holdername  if account else "",
                invoice_date = datetime.now().date(),
                payment_method = request.payment_method,
                gst_number = account.gst_number  if account else "",
                gst_type = gst_type if gst_type else "",
                gst_percentage = gst_percentage if gst_percentage else 0,
                total_amount = total_amount,
                client_email = existing_client.email,
                mail_status = False,
                payment_date = receipt_payment_date,
                payment_reference_number = request.payment_reference_number,
                fees_type = entry_type,
                created_at=datetime.now(),
                update_at=datetime.now()
            )

            if admission_fee > 0:
                admission_receipt = FeesReceipt(
                    client_id = existing_client.client_id,
                    gym_id = existing_client.gym_id,
                    client_name = existing_client.name,
                    gym_name = gym.name,
                    gym_logo = gym.logo,
                    gym_contact = gym_owner.contact_number,
                    gym_location = gym.location,
                    plan_id = existing_client.training_id,
                    plan_description = "Admission Fees",
                    fees = request.admission_fee,
                    discount = 0,
                    discounted_fees = admission_fee,
                    due_date = due_date,
                    invoice_number = None,  # Will be set after flush
                    client_contact = existing_client.contact,
                    bank_details = account.account_number if account else "",
                    ifsc_code = account.account_ifsccode  if account else "",
                    account_holder_name = account.account_holdername  if account else "",
                    invoice_date = datetime.now().date(),
                    payment_method = request.payment_method,
                    gst_number = account.gst_number  if account else "",
                    gst_type = gst_type if gst_type else "",
                    gst_percentage = gst_percentage if gst_percentage else 0,
                    total_amount = final_admission_fees,
                    client_email = existing_client.email,
                    mail_status = False,
                    payment_date = receipt_payment_date,
                    payment_reference_number = request.payment_reference_number,
                    fees_type = "admission",
                    created_at=datetime.now(),
                    update_at=datetime.now()
                )
                db.add(admission_receipt)
            db.add(new_receipt)
            db.flush()  # Get receipt_ids without committing

        
            gym_id = request.gym_id
            if admission_fee > 0:
                # Count includes both receipts we just added
                gym_receipt_count = db.query(FeesReceipt).filter(FeesReceipt.gym_id == gym_id).count()
                # Admission receipt was added first, so it gets count-1, new_receipt gets count
                admission_receipt.invoice_number = f"{gym.location[:3].upper()}-{gym.gym_id}-{gym_receipt_count - 1}"
                new_receipt.invoice_number = f"{gym.location[:3].upper()}-{gym.gym_id}-{gym_receipt_count}"
            else:
                # Only one receipt
                gym_receipt_count = db.query(FeesReceipt).filter(FeesReceipt.gym_id == gym_id).count()
                new_receipt.invoice_number = f"{gym.location[:3].upper()}-{gym.gym_id}-{gym_receipt_count}"

            import_data=db.query(GymImportData).filter(GymImportData.client_contact==existing_client.contact,GymImportData.gym_id==existing_client.gym_id).first()
            if import_data:
                db.delete(import_data)
                # db.commit()  # REMOVED - commit happens at the end

            pattern = f"gym{gym_id}:feesReceipt:*"
            async for key in redis.scan_iter(match=pattern):
                await redis.delete(key)

            # Invalidate analytics and monthly data caches (for real-time data)
            redis_key_analysis = f"gym:{gym_id}:analysis"
            if await redis.exists(redis_key_analysis):
                await redis.delete(redis_key_analysis)

            redis_key_monthly = f"gym:{gym_id}:monthly_data"
            if await redis.exists(redis_key_monthly):
                await redis.delete(redis_key_monthly)

            ladder: List[RewardGym] = (
                db.query(RewardGym)
                .filter(RewardGym.gym_id == request.gym_id)
                .order_by(RewardGym.xp.asc())
                .all()
            )

            reward=db.query(LeaderboardOverall).filter(LeaderboardOverall.client_id==existing_client.client_id).first()
            if reward:
                cur_xp=reward.xp
                tier        = _pick_next_reward(ladder, cur_xp)
                next_xp     = tier.xp   if tier else 0
                next_gift   = tier.gift if tier else None
    
                db.query(ClientNextXp).filter(
                    ClientNextXp.client_id == existing_client.client_id
                ).delete(synchronize_session=False)
    
                db.add(
                    ClientNextXp(
                        client_id=existing_client.client_id,
                        next_xp=next_xp,
                        gift=next_gift,
                    ))
                # db.commit()  # REMOVED - commit happens at the end

            fees=db.query(GymFees).filter(GymFees.client_id==existing_client.client_id).first()
            start_dt = request.fee_collection_start_date if request.fee_collection_start_date else cycle_start_date
            end_dt = calculated_expiry_date
            print(
                "[add_client_data] Fee schedule aligned: "
                f"start_dt={start_dt}, end_dt={end_dt}"
            )

            if fees:
                fees.start_date=start_dt
                fees.end_date= end_dt
                print("[add_client_data] Updated existing GymFees record.")
            else:
                print("[add_client_data] Creating new GymFees record.")
                db.add(
                    GymFees(
                        client_id=existing_client.client_id,
                        start_date=start_dt,
                        end_date=end_dt
                    )
                )



            if entry_type=="gym_membership" or entry_type=="personal_training":

                if request.membership_id:
                    # First get the membership to access its entitlement_id
                    membership_record = db.query(FittbotGymMembership).filter(
                        FittbotGymMembership.id == request.membership_id
                    ).first()

                    if membership_record and membership_record.entitlement_id:
                        # Get Entitlement to find order_item_id
                        entitlement = db.query(Entitlement).filter(
                            Entitlement.id == membership_record.entitlement_id
                        ).first()

                        if entitlement and entitlement.order_item_id:
                            # Get OrderItem to find order_id
                            order_item = db.query(OrderItem).filter(
                                OrderItem.id == entitlement.order_item_id
                            ).first()

                            if order_item and order_item.order_id:
                                # Lookup FittbotPayment by order_id (stored as entitlement_id)
                                payment_row = (
                                    db.query(FittbotPayment)
                                    .filter(FittbotPayment.entitlement_id == order_item.order_id)
                                    .first()
                                )
                                if payment_row:
                                    # Create Payout for this payment
                                    payout = Payout(
                                        payment_id=payment_row.id,
                                        gym_id=int(request.gym_id),
                                        gym_owner_id=None,
                                        amount_gross=payment_row.amount_net,
                                        amount_net=payment_row.amount_net,
                                        status="ready_for_transfer",
                                    )
                                    db.add(payout)
                                    print(f"[add_client_data] Created Payout: payment_id={payment_row.id}, gym_id={request.gym_id}, order_id={order_item.order_id}")
                                else:
                                    print(f"[add_client_data] FittbotPayment not found for order_id={order_item.order_id}")
                            else:
                                print(f"[add_client_data] OrderItem not found for order_item_id={entitlement.order_item_id if entitlement else 'N/A'}")
                        else:
                            print(f"[add_client_data] Entitlement not found for entitlement_id={membership_record.entitlement_id}")
                    else:
                        print(f"[add_client_data] Membership not found or no entitlement_id for membership_id={request.membership_id}")

                if request.membership_id is not None:
                    membership_status = db.query(FittbotGymMembership).filter(
                        FittbotGymMembership.id == request.membership_id
                    ).first()

                    membership_status.status = "active"
                    membership_status.amount=total_amount
                    membership_status.joined_at = date.today()
                    membership_status.expires_at = calculated_expiry_date
                    old_client=False

                db.commit()

                return {"status": 200, "message": "Client added to the Gym successfully."}
            
               

            elif entry_type=="normal":

                normal_entry=FittbotGymMembership(
                    gym_id=str(request.gym_id),
                    client_id=str(existing_client.client_id),
                    plan_id=request.training_type,
                    type="normal",
                    amount=total_amount,
                    status="active",
                    purchased_at=datetime.now(),
                    joined_at=cycle_start_date,
                    expires_at=calculated_expiry_date if request.new_expiry_date is  None else request.new_expiry_date,
                    old_client=request.old_client
                )

                db.add(normal_entry)
                # db.commit()  # REMOVED - commit happens at the end

                # Add admission fees entry if applicable
                if admission_fee > 0:
                    admission_entry = FittbotGymMembership(
                        gym_id=str(request.gym_id),
                        client_id=str(existing_client.client_id),
                        plan_id=request.training_type,
                        type="admission_fees",
                        amount=final_admission_fees,
                        status="active",
                        purchased_at=datetime.now(),
                        joined_at=cycle_start_date,
                        expires_at=calculated_expiry_date if request.new_expiry_date is  None else request.new_expiry_date,
                        old_client=request.old_client

                    )
                    db.add(admission_entry)

            

                join_request = db.query(GymJoinRequest).filter(
                    GymJoinRequest.gym_id == gym_id,
                    GymJoinRequest.client_id == existing_client.client_id,
                    GymJoinRequest.status=="pending"
                    ).order_by(GymJoinRequest.id.desc()).first()

                if join_request:
                        join_request.status = "onboarded"
                        join_request.updated_at = datetime.now()
                

                db.commit()

                return {"status": 200, "message": "Client added to the Gym successfully."}
            
            else:
                print("[add_client_data] No existing client found for provided contact.")
            

    except Exception as e:
        db.rollback()
        print(f"Error adding client data: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error adding client data")


class TemplateWorkoutRequest(BaseModel):
    name: str
    workoutPlan: Dict
    gym_id:int


@router.post("/gym/addworkouttemplate")
async def addWorkOutTemplate(
    template:TemplateWorkoutRequest,db: Session = Depends(get_db),redis: Redis = Depends(get_redis)
):
    try:
    
        new_template=TemplateWorkout(name=template.name,workoutPlan=template.workoutPlan,gym_id=template.gym_id)
        db.add(new_template)
        db.commit()
        workout_redis_key = f"gym:{template.gym_id}:all_workouts"

        if await redis.exists(workout_redis_key ):
            await redis.delete(workout_redis_key )
        client_redis_key = f"gym:{template.gym_id}:all_clients"
        if await redis.exists(client_redis_key):
            await redis.delete(client_redis_key)

        pattern = f"*:{template.gym_id}:assigned_plans"
        keys = await redis.keys(pattern)  

        if keys:  
            await redis.delete(*keys) 

        return{"status":200,"message":"Template added successfully"}
    except Exception as e:
        db.rollback()
        print(f"Error adding template: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error adding template: {str(e)}")


@router.get("/gym/addworkouttemplate")
async def getWorkOutTemplate(
    gym_id: int = Query(...),db: Session = Depends(get_db)
):
    try:
        template_names=db.query(TemplateWorkout).filter(TemplateWorkout.gym_id==gym_id).all()
        temp = [{"id": template.id, "name": template.name,"workoutPlan":template.workoutPlan} for template in template_names]
        
        return {"status": 200, "message": "Template added successfully","data":temp}
    except Exception as e:
        db.rollback()
        print(f"Error adding template: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error adding template: {str(e)}")


class UpdateWorkoutTemplateRequest(BaseModel):
    id: int
    gym_id:int
    workoutPlan: dict  

class editWorkoutTemplateRequest(BaseModel):
    id: int
    name: str 
    gym_id:int

@router.put("/gym/addworkouttemplate")
async def update_workout_template(
    request: UpdateWorkoutTemplateRequest, db: Session = Depends(get_db),redis: Redis = Depends(get_redis)
):
    try:
        
        template = db.query(TemplateWorkout).filter(TemplateWorkout.id == request.id).first()
        if not template:
            raise HTTPException(status_code=404, detail="Template not found with the given ID")

        if request.workoutPlan:
            template.workoutPlan = request.workoutPlan
        db.commit()
        db.refresh(template)
        workout_redis_key = f"gym:{request.gym_id}:all_workouts"
        if await redis.exists(workout_redis_key ):
            await redis.delete(workout_redis_key )

        client_redis_key = f"gym:{request.gym_id}:all_clients"
        if await redis.exists(client_redis_key):
            await redis.delete(client_redis_key)
        
        pattern = f"*:{request.gym_id}:assigned_plans"
        keys = await redis.keys(pattern)  

        if keys:  
            await redis.delete(*keys)
        

        return {
            "status": 200,
            "message": "Template updated successfully"
        }
    except Exception as e:
        db.rollback()
        print(f"Error updating template: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error updating template: {str(e)}")


@router.delete("/gym/addworkouttemplate")
async def delete_workout_template(
    id: int,
    gym_id:int,
    db: Session = Depends(get_db),redis: Redis = Depends(get_redis)
):
    try:
        db.query(TemplateWorkout).filter(
            TemplateWorkout.id == id
        ).delete()
        db.commit()
        
        workout_redis_key = f"gym:{gym_id}:all_workouts"
        
        if await redis.exists(workout_redis_key ):
            await redis.delete(workout_redis_key )

        client_redis_key = f"gym:{gym_id}:all_clients"
        if await redis.exists(client_redis_key):
            await redis.delete(client_redis_key)

        pattern = f"*:{gym_id}:assigned_plans"
        keys = await redis.keys(pattern)  

        if keys:  
            await redis.delete(*keys)
        return {
                "status": 200,
                "message": "Template deleted successfully"
            }
    except Exception as e:
        db.rollback()
        print(f"Error updating template: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error deleting workout template: {str(e)}")




@router.put("/gym/editworkouttemplate")
async def edit_workout_template(
    input :editWorkoutTemplateRequest,
    db: Session = Depends(get_db),redis: Redis = Depends(get_redis)
):
    try:
        result = db.query(TemplateWorkout).filter(
        TemplateWorkout.id == input.id
        ).update({"name": input.name})
        db.commit()
        workout_redis_key = f"gym:{input.gym_id}:all_workouts"
        if await redis.exists(workout_redis_key ):
            await redis.delete(workout_redis_key )

        client_redis_key = f"gym:{input.gym_id}:all_clients"
        if await redis.exists(client_redis_key):
            await redis.delete(client_redis_key)

        pattern = f"*:{input.gym_id}:assigned_plans"
        keys = await redis.keys(pattern)  

        if keys:  
            await redis.delete(*keys)

        if not result:
            raise HTTPException(status_code=404, detail="Template not found")

        return {
                "status": 200,
                "message": "Template updated successfully"
            }
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected error occured: {str(e)}")


def get_ist_time():
    ist = pytz.timezone("Asia/Kolkata")
    return datetime.now(ist)


@router.post("/create_post")
async def create_post(
    request: Request,
    gym_id: int = Form(...),
    client_id: Optional[int] = Form(None),
    content: str = Form(...),
    role: str = Form(...),
    media: str = Form("[]"),
    file: List[UploadFile] = File(None),
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis)
):
    try:

        # role = (role or "").lower()
        # actor_id, token_role = resolve_post_actor(
        #     request,
        #     db,
        #     gym_id=gym_id,
        #     declared_role=role,
        #     provided_client_id=client_id,
        # )
        # role = token_role
        # if role == "owner":
        #     client_id = None
        # else:
        #     client_id = actor_id

        user_name = None
        if role == "owner":
            owner = db.query(Gym).filter(Gym.gym_id == gym_id).first()
            if not owner:
                raise HTTPException(status_code=404, detail="Owner not found.")
            user_name = owner.name
        else:
            client = db.query(Client).filter(Client.client_id == client_id).first()
            if not client:
                raise HTTPException(status_code=404, detail="Client not found.")
            user_name = client.name

        now = get_ist_time()

        post = Post(
            gym_id=gym_id,
            client_id=client_id,
            content=content,
            created_at=now,
            updated_at=now,
        )
        db.add(post)
        db.flush()  # ensure post_id is populated without committing

        # Delete Redis cache immediately after post creation (before media)
        redis_key = f"gym:{gym_id}:posts"
        await redis.delete(redis_key)

        media_metadata = json.loads(media)
        media_entries = []


        if file:
            for index, upload_file in enumerate(file):
                if index >= len(media_metadata):
                    raise HTTPException(status_code=400, detail="Mismatch between media metadata and files.")

                file_meta = media_metadata[index]
                file_ext = upload_file.filename.split(".")[-1]
                unique_filename = f"{uuid.uuid4()}.{file_ext}"
                file_path = os.path.join(UPLOAD_DIR, unique_filename)

                
                with open(file_path, "wb") as buffer:
                    buffer.write(upload_file.file.read())

                media_entry = PostMedia(
                    post_id=post.post_id,
                    file_name=unique_filename,
                    file_type=file_meta["type"],
                    file_path=file_path,
                )
                db.add(media_entry)
                media_entries.append(media_entry)

        db.commit()
        db.refresh(post)

        # ✅ CACHE DELETION REMOVED - Lambda will invalidate cache after S3 upload
        # This prevents race condition where cache is built with local paths
        # instead of S3 URLs. Lambda deletes cache only after S3 URLs are in DB.

        return {"status": 200, "message": "Post created successfully."}

    except Exception as e:
        print(f"Error: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")



class CreateEditRequest(BaseModel):
    post_id: int
    gym_id: int
    content: str
    role:str
    client_id: Optional[int]=None
    

@router.put("/edit_post")
async def edit_post(
    payload: CreateEditRequest,
    request: Request,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis),
):
    try:
        actor_id, token_role = resolve_post_actor(
            request,
            db,
            gym_id=payload.gym_id,
            declared_role=payload.role,
            provided_client_id=payload.client_id,
        )

        post_id = payload.post_id
        gym_id = payload.gym_id
        content = payload.content

        post = db.query(Post).filter(Post.post_id == post_id, Post.gym_id == gym_id).first()
        if not post:
            return {"status": 404, "message": "Post not found."}

        enforce_post_mutation_permissions(
            token_role=token_role,
            actor_id=actor_id,
            post=post,
        )

        post.content = content
        post.updated_at = get_ist_time()
        db.commit()

        redis_key = f"gym:{gym_id}:posts"
        cached_posts = await redis.zrange(redis_key, 0, -1)
        for cached_post in cached_posts:
            post_data = json.loads(cached_post)
            if post_data["post_id"] == post_id:
                await redis.zrem(redis_key, cached_post)
                post_data["content"] = content
                post_data["updated_at"] = post.updated_at.isoformat()
                post_data["created_at"] = post.created_at.isoformat()
                await redis.zadd(redis_key, {json.dumps(post_data): post.created_at.timestamp()})
                break
        return {"status": 200, "message": "Post updated successfully."}
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected error occured: {str(e)}")

@router.delete("/delete_post")
async def delete_post(
    request: Request,
    post_id: int,
    gym_id: int,
    role: str,
    client_id: Optional[int] = None,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis),
):
    try:
        actor_id, token_role = resolve_post_actor(
            request,
            db,
            gym_id=gym_id,
            declared_role=role,
            provided_client_id=client_id,
        )

        post = db.query(Post).filter(Post.post_id == post_id, Post.gym_id == gym_id).first()
        if not post:
            return {"status": 404, "message": "Post not found."}

        enforce_post_mutation_permissions(
            token_role=token_role,
            actor_id=actor_id,
            post=post,
        )

        db.delete(post)
        db.commit()

        media_cache_key = f"post:{post_id}:media"
        await redis.delete(media_cache_key)

        redis_key = f"gym:{gym_id}:posts"
        cached_posts = await redis.zrange(redis_key, 0, -1)
        for cached_post in cached_posts:
            post_data = json.loads(cached_post)
            if post_data["post_id"] == post_id:
                await redis.zrem(redis_key, cached_post)
                break
        return {"status": 200, "message": "Post deleted successfully."}
    
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"An unexpected error occured: {str(e)}")

class CreatecommentRequest(BaseModel):
    role: str
    gym_id: int
    client_id: Optional[int]
    content: str
    post_id:int


@router.post("/create_comment")
async def add_comment(
    payload: CreatecommentRequest,
    request: Request,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis),
):
    try:
        if payload.gym_id is None:
            raise HTTPException(status_code=400, detail="gym_id is required")

        actor_id, token_role = resolve_post_actor(
            request,
            db,
            gym_id=payload.gym_id,
            declared_role=payload.role,
            provided_client_id=payload.client_id,
        )

        role = token_role
        post_id = payload.post_id
        content = payload.content
        if role=="client":
            client_id = actor_id
            gym_id = payload.gym_id

        else:
            client_id=None
            gym_id=payload.gym_id

        if not content.strip():
            return {"status": 400, "message": "Content cannot be empty."}

        now = get_ist_time()
        comment = Comment(
            post_id=payload.post_id,
            gym_id=payload.gym_id,
            client_id=actor_id if role == "client" else None,
            content=content,
            created_at=now,
            updated_at=now,
        )
        db.add(comment)
        db.commit()

        if role=="client":
                    client = db.query(Client).filter(Client.client_id == client_id).one()
                    user_name = client.name
        else:
            owner = db.query(Gym).filter(Gym.gym_id == gym_id).one()
            user_name = owner.name
        
        
        redis_key = f"post:{post_id}:comments"
        comment_data = {
            "gym_id":gym_id,
            "comment_id": comment.comment_id,
            "post_id": post_id,
            "client_id": client_id,
            "content": content,
            "user_name":user_name,
            "created_at": comment.created_at.isoformat(),
            "updated_at": comment.updated_at.isoformat()}
        await redis.rpush(redis_key, json.dumps(comment_data))

        comment_count_key = f"post:{post_id}:comment_count"
        await redis.incr(comment_count_key)

        return {"status": 200, "message": "Comment added successfully."}
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected error occured:{str(e)}")
    

@router.delete("/delete_comment")
async def delete_comment(
    comment_id: int,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis),
):
    try:
        comment = db.query(Comment).filter(Comment.comment_id == comment_id).first()
        if not comment:
            return {"status": 404, "message": "Comment not found."}

        db.delete(comment)
        db.commit()

        redis_key = f"post:{comment.post_id}:comments"
        cached_comments = await redis.lrange(redis_key, 0, -1)
        for cached_comment in cached_comments:
            comment_data = json.loads(cached_comment)
            if comment_data["comment_id"] == comment_id:
                await redis.lrem(redis_key, 0, cached_comment)
                break

        comment_count_key = f"post:{comment.post_id}:comment_count"
        await redis.decr(comment_count_key)

        return {"status": 200, "message": "Comment deleted successfully."}
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected error occured : {str(e)}")

class CreatelikeRequest(BaseModel):
    gym_id: Optional[int]
    client_id: Optional[int]
    post_id:int
    role:str


@router.post("/post_likes")
async def toggle_like(
    payload: CreatelikeRequest,
    request: Request,
    redis: Redis = Depends(get_redis),
    db: Session = Depends(get_db),
):
        
    try:
        if payload.gym_id is None:
            raise HTTPException(status_code=400, detail="gym_id is required")

        actor_id, token_role = resolve_post_actor(
            request,
            db,
            gym_id=payload.gym_id,
            declared_role=payload.role,
            provided_client_id=payload.client_id,
        )

        role = token_role
        post_id = payload.post_id
        like_count_key = f"post:{post_id}:like_count"
        liked_by_key = f"post:{post_id}:liked_by"
        if await redis.exists(liked_by_key):
            await redis.delete(liked_by_key)

        if not await redis.exists(like_count_key):
            await redis.set(like_count_key, 0)

        if role == "client":
            user_id = actor_id
            gym_id = payload.gym_id
            client_id = actor_id
            existing_like = db.query(Like).filter(
                Like.post_id == post_id,
                Like.client_id == client_id
            ).first()
        else:
            user_id = payload.gym_id
            gym_id = payload.gym_id
            client_id = None
            existing_like = db.query(Like).filter(
                Like.post_id == post_id,
                Like.gym_id == gym_id,
                Like.client_id == None  
            ).first()

        if existing_like:
            liked_by_key = f"post:{post_id}:liked_by"
            if role == "client":
                await redis.srem(liked_by_key, client_id)
            elif role == "owner":
                await redis.srem(liked_by_key, gym_id)

            db.delete(existing_like)
            db.commit()
            await redis.decr(like_count_key)
            like_count = await redis.get(like_count_key)

            return {
                "status": 200,
                "message": "Like removed.",
                "like_count": int(like_count),
                "liked": False,  
            }
        else:

            now = get_ist_time()
            new_like = Like(
                post_id=post_id,
                gym_id=gym_id,
                client_id=client_id,
                created_at=now,
            )
            db.add(new_like)
            db.commit()
            print("existing like count is",await redis.get(like_count_key))
            await redis.incr(like_count_key)
            like_count = await redis.get(like_count_key)
            print("like count after adding",like_count)

            return {
                "status": 200,
                "message": "Like added successfully.",
                "like_count": int(like_count),
                "liked": True,  
            }
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected error occured : {str(e)}")
    


@router.get("/fetch_comment")
async def fetch_post_comments(
    request: Request,
    gym_id: int,
    post_id: int,
    role: str,
    client_id: Optional[int] = None,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis),
):
    try:
        if role not in {"owner", "client"}:
            raise HTTPException(status_code=400, detail="Invalid role. Must be 'owner' or 'client'.")

        actor_id, token_role = resolve_post_actor(
            request,
            db,
            gym_id=gym_id,
            declared_role=role,
            provided_client_id=client_id,
        )
        role = token_role
        if role == "client":
            client_id = actor_id

        redis_key = f"post:{post_id}:comments"
        cached_comments = await redis.lrange(redis_key, 0, -1)
        comments = []

        usernames = await get_gym_usernames(gym_id, redis, db)

        gym_name_key = f"gym:{gym_id}:gym_name"
        gym_logo_key = f"gym:{gym_id}:logo"

        try:
            gym_name = await redis.get(gym_name_key)
            gym_logo = await redis.get(gym_logo_key)

            if not gym_name or not gym_logo:
                gym_details = db.query(Gym.name, Gym.logo).filter(Gym.gym_id == gym_id).one()
                gym_name = gym_details.name
                gym_logo = gym_details.logo if gym_details.logo else ""
                await redis.set(gym_name_key, gym_name, ex=86400)
                await redis.set(gym_logo_key, gym_logo, ex=86400)
        except Exception as e:
            print(str(e))
            raise HTTPException(status_code=500, detail="Error fetching gym details.")

        blocked_client_ids, blocked_gym_ids, blocked_list = [], [], []

        try:
            if role == "client":
                blocked_entry = db.query(BlockedUsers).filter(
                    BlockedUsers.user_id == client_id,
                    BlockedUsers.user_role == "client"
                ).first()

                if blocked_entry and blocked_entry.blocked_user_id:
                    blocked_user_id = json.loads(blocked_entry.blocked_user_id) if isinstance(
                        blocked_entry.blocked_user_id, str) else blocked_entry.blocked_user_id
                    blocked_client_ids = blocked_user_id.get("client", [])
                    blocked_gym_ids = blocked_user_id.get("owner", [])
            else:
                blocked_entry = db.query(BlockedUsers).filter(
                    BlockedUsers.user_id == gym_id,
                    BlockedUsers.user_role == "owner"
                ).first()

                if blocked_entry and blocked_entry.blocked_user_id:
                    blocked_user_id = json.loads(blocked_entry.blocked_user_id) if isinstance(
                        blocked_entry.blocked_user_id, str) else blocked_entry.blocked_user_id
                    blocked_list = blocked_user_id.get("client", [])
        except Exception:
            raise HTTPException(status_code=500, detail="Error fetching blocked users data.")

        if cached_comments:
            for cached_comment in reversed(cached_comments):
                try:
                    comment_data = json.loads(cached_comment)

                    if role == "client":
                        if comment_data["client_id"] is None and gym_id in blocked_gym_ids:
                            continue
                        if comment_data["client_id"] is not None and comment_data["client_id"] in blocked_client_ids:
                            continue

                    if role == "owner" and comment_data.get("client_id") is not None and comment_data["client_id"] in blocked_list:
                        continue

                    comment_data["is_editable"] = (role == "owner" and comment_data["client_id"] is None) or (role == "client" and client_id == comment_data["client_id"])
                    comment_data["user_name"] = "You" if comment_data["is_editable"] else (
                        gym_name if comment_data["client_id"] is None else usernames.get(comment_data["client_id"], "Unknown Client")
                    )
                    comment_data["profile_url"] = gym_logo if comment_data["client_id"] is None else await get_client_profile(comment_data["client_id"], redis, db)

                    comments.append(comment_data)
                except Exception as e:
                    print(str(e))
                    raise HTTPException(status_code=500, detail="Error processing cached comments.")
        else:
            try:
                comments_query = db.query(
                    Comment.comment_id,
                    Comment.post_id,
                    Comment.client_id,
                    Comment.gym_id,
                    Comment.content,
                    Comment.created_at,
                    Comment.updated_at,
                    Client.profile.label("client_profile_url")
                ).outerjoin(Client, Comment.client_id == Client.client_id).filter(
                    Comment.post_id == post_id,
                    Comment.gym_id == gym_id
                )

                if role == "client":
                    comments_query = comments_query.filter(
                        ~((Comment.client_id == None) & (Comment.gym_id.in_(blocked_gym_ids))),
                        ~(Comment.client_id.in_(blocked_client_ids))
                    )
                elif role == "owner":
                    comments_query = comments_query.filter(
                        ~(Comment.client_id.in_(blocked_list))
                    )

                comments_query = comments_query.order_by(Comment.created_at.desc()).all()
            except Exception as e:
                print(e)
                raise HTTPException(status_code=500, detail="Error fetching comments from the database.")

            for comment in comments_query:
                try:
                    user_name = "You" if (role == "owner" and comment.client_id is None) or (role == "client" and comment.client_id == client_id) else usernames.get(comment.client_id, gym_name)
                    is_editable = (role == "owner" and comment.client_id is None) or (role == "client" and comment.client_id == client_id)
                    profile_url = comment.client_profile_url if comment.client_id else gym_logo

                    comment_data = {
                        "comment_id": comment.comment_id,
                        "post_id": comment.post_id,
                        "client_id": comment.client_id,
                        "gym_id": comment.gym_id,
                        "content": comment.content,
                        "user_name": user_name,
                        "profile_url": profile_url,
                        "created_at": comment.created_at.isoformat(),
                        "updated_at": comment.updated_at.isoformat(),
                        "is_editable": is_editable,
                    }

                    redis_data = comment_data.copy()
                    redis_data.pop("user_name")
                    redis_data.pop("is_editable")
                    redis_data.pop("profile_url")
                    await redis.rpush(redis_key, json.dumps(redis_data))

                    comments.append(comment_data)
                except Exception:
                    raise HTTPException(status_code=500, detail="Error processing comment data.")

        return {"status": 200, "data": comments}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")


async def get_client_profile(client_id: int, redis: Redis, db: Session) -> str:
    try:
        redis_key = f"client:{client_id}:profile"
        profile_url = await redis.get(redis_key)

        if not profile_url:
            profile_url = db.query(Client.profile).filter(Client.client_id == client_id).scalar()
            if profile_url:
                await redis.set(redis_key, profile_url, ex=86400)

        return profile_url or ""
    except Exception as e:
        raise HTTPException(status_code=500, detail="Error fetching client profile URL.")


@router.get("/liked_by_names")
async def get_liked_by_names(
    gym_id: int,
    post_id: int,
    db: Session = Depends(get_db),
):
    try:
        
        likes = db.query(Like).filter(Like.gym_id == gym_id, Like.post_id == post_id).all()

        names = []

        for like in likes:
            if like.client_id is not None:
                client = db.query(Client).filter(Client.client_id == like.client_id).first()
                if client:
                    names.append({"name":client.name,"profile":client.profile})
            else:
                gym = db.query(Gym).filter(Gym.gym_id == like.gym_id).first()
                if gym:
                    names.append({"name":gym.name, "profile":gym.logo})

        return {"status":200,"names": names}

    except Exception as e:
        print(f"Error fetching liked by names: {e}")
        raise HTTPException(status_code=500, detail="An error occurred while fetching liked by names.")


async def get_gym_usernames(
    gym_id: int,
    redis: Redis = Depends(get_redis),
    db: Session = Depends(get_db),
):
    redis_key = f"gym:{gym_id}:usernames"
    cached_usernames = await redis.hgetall(redis_key)
    if cached_usernames:
       
        usernames = {int(k): v for k, v in cached_usernames.items()}
        return usernames

    try:
        clients = db.query(Client).filter(Client.gym_id == gym_id).all()
        if not clients:
            # Return empty dict if no clients found - this is a valid scenario
            return {}

        usernames = {client.client_id: client.name for client in clients}
        if usernames:
            for client_id, name in usernames.items():
                await redis.hset(redis_key, client_id, name)
            await redis.expire(redis_key, 86400)
        return usernames

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching gym usernames: {str(e)}")


@router.get("/get_post")
async def fetch_posts(
    request: Request,
    gym_id: int = Query(..., description="Gym ID"),
    role: str = Query(..., description="Role of the user (owner or client)"),
    client_id: Optional[int] = Query(None, description="Client ID (if applicable)"),
    page: int = Query(1, description="Page number (starting from 1)"),
    limit: int = Query(20, description="Number of posts per page (max 50)"),
    redis: Redis = Depends(get_redis),
    db: Session = Depends(get_db),
):
    try:

        
        if role not in {"owner", "client"}:
            return {
                "status": 400,
                "message": "Invalid role. Must be 'owner' or 'client'.",
                "data": None
            }

        if page < 1:
            return {
                "status": 400,
                "message": "Page number must be greater than 0",
                "data": None
            }

        if limit < 1 or limit > 50:
            return {
                "status": 400,
                "message": "Limit must be between 1 and 50",
                "data": None
            }

        # actor_id, token_role = resolve_post_actor(
        #     request,
        #     db,
        #     gym_id=gym_id,
        #     declared_role=role,
        #     provided_client_id=client_id,
        # )

        # role = token_role
        # if role == "client":
        #     client_id = actor_id

 
        post_qty = db.query(Post).filter(Post.gym_id == gym_id).all()


        if not post_qty:

            return {
                "status": 201,
                "message": "No Posts or feed available for this gym",
                "data": {
                    "isClient": (role == "client"),
                    "posts": [],
                    "pagination": {
                        "current_page": page,
                        "total_pages": 0,
                        "total_posts": 0,
                        "posts_per_page": limit,
                        "has_next_page": False,
                        "has_previous_page": False
                    }
                }
            }
        

      

 
        clients = db.query(Client).filter(Client.gym_id == gym_id).all()

 
        gym_name_key = f"gym:{gym_id}:gym_name"
        gym_logo_key = f"gym:{gym_id}:gym_logo"
        gym_name = await redis.get(gym_name_key)
        gym_logo = await redis.get(gym_logo_key)
        if not (gym_name or gym_logo):
            gym_object= db.query(Gym).filter(Gym.gym_id == gym_id).one()
            gym_name = gym_object.name if gym_object.name else ""
            gym_logo = gym_object.logo if gym_object.logo else ""
            await redis.set(gym_name_key, gym_name, ex=86400)
            await redis.set(gym_logo_key, gym_logo, ex=86400)
 
        if role == "client":
            if not client_id:
                return {
                    "status": 400,
                    "message": "client_id is required when role='client'",
                    "data": None
                }
 
            reported_post_ids = {
                r.post_id
                for r in db.query(Report)
                             .filter(Report.user_id == client_id, Report.user_role == "client")
                             .all()
            }
        else:  
            gym_obj = db.query(Gym).filter(Gym.gym_id == gym_id).first()
            if not gym_obj:
                return {"status": 400, "message": "Gym not found", "data": None}
 
            reported_post_ids = {
                r.post_id
                for r in db.query(Report).filter(Report.user_id == gym_obj.owner_id, Report.user_role == "owner").all()
            }


        
 
        owner_blocked_entry = db.query(BlockedUsers).filter(
            BlockedUsers.user_id == gym_id,
            BlockedUsers.user_role == "owner"
        ).first()

        owner_blocked_client_ids = []
        if owner_blocked_entry and owner_blocked_entry.blocked_user_id:
            owner_blocked_data = (
                owner_blocked_entry.blocked_user_id
                if isinstance(owner_blocked_entry.blocked_user_id, dict)
                else json.loads(owner_blocked_entry.blocked_user_id)
            )
            owner_blocked_client_ids = owner_blocked_data.get("client", [])
 
        if role == "client":
            blocked_entry = db.query(BlockedUsers).filter(
                BlockedUsers.user_id == client_id,
                BlockedUsers.user_role == "client"
            ).first()
 
            blocked_client_ids = []
            blocked_gym_ids = []
            if blocked_entry and blocked_entry.blocked_user_id:
                blocked_user_id = (
                    blocked_entry.blocked_user_id
                    if isinstance(blocked_entry.blocked_user_id, dict)
                    else json.loads(blocked_entry.blocked_user_id)
                )
                blocked_client_ids = blocked_user_id.get("client", [])
                blocked_gym_ids = blocked_user_id.get("owner", [])
 
        else:
            blocked_entry = db.query(BlockedUsers).filter(
                BlockedUsers.user_id == gym_id,
                BlockedUsers.user_role == "owner"
            ).first()
 
            blocked_list = []
            if blocked_entry and blocked_entry.blocked_user_id:
                blocked_user_id = (
                    blocked_entry.blocked_user_id
                    if isinstance(blocked_entry.blocked_user_id, dict)
                    else json.loads(blocked_entry.blocked_user_id)
                )
 
                blocked_list = blocked_user_id.get("client", [])
 


        offset = (page - 1) * limit
        
        total_count_query = db.query(Post).filter(Post.gym_id == gym_id)
        
        if reported_post_ids:
            total_count_query = total_count_query.filter(~Post.post_id.in_(reported_post_ids))
        
        if owner_blocked_client_ids:
            total_count_query = total_count_query.filter(
                or_(Post.client_id.is_(None), ~Post.client_id.in_(owner_blocked_client_ids))
            )

        
        
        total_posts_count = total_count_query.count()
        
        redis_key = f"gym:{gym_id}:posts"
        posts = []
        
        if page <= 3:
         

            cached_posts = await redis.zrevrange(redis_key, offset, offset + limit - 1)
            if not cached_posts:
                post_query = db.query(Post).filter(Post.gym_id == gym_id)
                
                if reported_post_ids:
                    post_query = post_query.filter(~Post.post_id.in_(reported_post_ids))
                
                if owner_blocked_client_ids:
                    post_query = post_query.filter(
                        or_(Post.client_id.is_(None), ~Post.client_id.in_(owner_blocked_client_ids))
                    )
                
                all_posts = post_query.order_by(Post.created_at.desc()).all()
                
                for p in all_posts:
                    redis_data = {
                        "post_id": p.post_id ,
                        "gym_id": p.gym_id,
                        "client_id": p.client_id,
                        "content": p.content if p.content else "",
                        "created_at": p.created_at.isoformat(),
                        "updated_at": p.updated_at.isoformat(),
                        "is_pinned": p.is_pinned if p.is_pinned else False
                    }
                    await redis.zadd(
                        redis_key,
                        {json.dumps(redis_data): p.created_at.timestamp()}
                    )
                
                cached_posts = await redis.zrevrange(redis_key, offset, offset + limit - 1)
        else:

            post_query = db.query(Post).filter(Post.gym_id == gym_id)
            
            if reported_post_ids:
                post_query = post_query.filter(~Post.post_id.in_(reported_post_ids))
            
            if owner_blocked_client_ids:
                post_query = post_query.filter(
                    or_(Post.client_id.is_(None), ~Post.client_id.in_(owner_blocked_client_ids))
                )
            
            paginated_posts = post_query.order_by(Post.created_at.desc()).offset(offset).limit(limit).all()
            
            cached_posts = []
            for p in paginated_posts:
                redis_data = {
                    "post_id": p.post_id,
                    "gym_id": p.gym_id,
                    "client_id": p.client_id,
                    "content": p.content if p.content else "",
                    "created_at": p.created_at.isoformat(),
                    "updated_at": p.updated_at.isoformat(),
                    "is_pinned": p.is_pinned if p.is_pinned else False
                }
                cached_posts.append(json.dumps(redis_data))

             
        client_profiles = {c.client_id: (c.name, c.profile) for c in db.query(Client).filter(Client.gym_id == gym_id).all()}
        for cached_post in cached_posts:
            post_data = json.loads(cached_post)
            if post_data["post_id"] in reported_post_ids:
                continue

            if post_data.get("client_id") is not None and post_data["client_id"] in owner_blocked_client_ids:
                continue
 
            if role == "client":
                if post_data.get("client_id") is None and gym_id in blocked_gym_ids:
                    continue
                if post_data.get("client_id") is not None and post_data["client_id"] in blocked_client_ids:
 
                    continue
 
            if role == "owner":
 
                if post_data.get("client_id") is not None and post_data["client_id"] in blocked_list:
                    continue

            media_cache_key = f"post:{post_data['post_id']}:media"
            cached_media = await redis.get(media_cache_key)

            media_list: List[Dict[str, Any]] = []
            cache_validated = False

            if cached_media:
                try:
                    decoded_media = json.loads(cached_media)
                    if isinstance(decoded_media, dict):
                        media_list = decoded_media.get("items", [])
                        cache_validated = decoded_media.get("validated", False)
                    elif isinstance(decoded_media, list):
                        media_list = decoded_media
                        cache_validated = False  # legacy cache without validation flag
                except json.JSONDecodeError:
                    media_list = []
                    cache_validated = False

            if not cache_validated:
                post_medias = db.query(PostMedia).filter(PostMedia.post_id == post_data["post_id"]).all()
                media_list = [
                    {
                        "file_name": media.file_name,
                        "file_type": media.file_type,
                        "file_url": media.file_path
                    }
                    for media in post_medias
                ]
                cache_payload = {
                    "items": media_list,
                    "validated": True,
                }
                ttl = MEDIA_CACHE_TTL_SECONDS if media_list else EMPTY_MEDIA_CACHE_TTL_SECONDS
                await redis.set(media_cache_key, json.dumps(cache_payload), ex=ttl)
            else:
                ttl = MEDIA_CACHE_TTL_SECONDS if media_list else EMPTY_MEDIA_CACHE_TTL_SECONDS
                await redis.expire(media_cache_key, ttl)

            post_data["media"] = media_list

            if role == "owner" and post_data["client_id"] is None:
                post_data["is_editable"] = True
                post_data["user_name"] = "You"
            elif role == "client" and client_id == post_data["client_id"]:
                post_data["is_editable"] = True
                post_data["user_name"] = "You"
            else:
                post_data["is_editable"] = False
                if post_data["client_id"] is None:
                    post_data["user_name"] = gym_name
                else:

                    redis_key_username = f"gym:{gym_id}:usernames"
                    usernames = await get_gym_usernames(gym_id, redis, db)
                    user_name = usernames.get(post_data["client_id"])

                    if user_name is None:
                        await redis.delete(redis_key_username)
                        usernames = await get_gym_usernames(gym_id, redis, db)
                        user_name = usernames.get(post_data["client_id"], "Unknown Client")

                    post_data["user_name"] = user_name
 
                    # post_data["user_name"] = (
                    #     (await get_gym_usernames(gym_id, redis, db))
                    #     .get(post_data["client_id"], "Unknown Client")
                    # )

                    
 
            like_count_key = f"post:{post_data['post_id']}:like_count"
            like_count = await redis.get(like_count_key)
            if like_count is None:
                like_count = db.query(Like).filter(Like.post_id == post_data["post_id"]).count()
                await redis.set(like_count_key, like_count, ex=86400)
            post_data["like_count"] = int(like_count)

            comments_query = db.query(Comment).filter(
                Comment.post_id == post_data['post_id'],
                Comment.gym_id == gym_id
            )

            if owner_blocked_client_ids:
                comments_query = comments_query.filter(
                    or_(Comment.client_id.is_(None), ~Comment.client_id.in_(owner_blocked_client_ids))
                )

            if role == "client":
                comments_query = comments_query.filter(
                    ~((Comment.client_id == None) & (Comment.gym_id.in_(blocked_gym_ids))),  
                    ~(Comment.client_id.in_(blocked_client_ids))  
                )
            elif role == "owner":
                comments_query = comments_query.filter(
                    ~(Comment.client_id.in_(blocked_list))  
                )

            comments_query = comments_query.order_by(Comment.created_at.desc()).all()
            comment_count = len(comments_query)
            post_data["comment_count"] = int(comment_count)
            liked_by_key = f"post:{post_data['post_id']}:liked_by"
            exists = await redis.exists(liked_by_key)

            if not exists:
                db_likes = db.query(Like).filter(Like.post_id == post_data["post_id"]).all()

                liked_users = []
                for like in db_likes:
                    if like.client_id is not None:
                        liked_users.append(str(like.client_id)) 
                    else:
                        liked_users.append(f"gym_{like.gym_id}")  

                if liked_users:
                    await redis.sadd(liked_by_key, *liked_users)  
                    await redis.expire(liked_by_key, 86400) 

            check_id = f"gym_{gym_id}" if role == "owner" else str(client_id)
            is_liked = await redis.sismember(liked_by_key, check_id)
            post_data["is_liked"] = bool(is_liked)
            post_data["profile_url"] = gym_logo if post_data["client_id"] is None else client_profiles.get(post_data["client_id"], ("", ""))[1]


            posts.append(post_data)
        
        total_pages = (total_posts_count + limit - 1) // limit  
        has_next_page = page < total_pages
        has_previous_page = page > 1

        data={
                "isClient": (role == "client"),
                "posts": posts,
                "pagination": {
                    "current_page": page,
                    "total_pages": total_pages,
                    "total_posts": total_posts_count,
                    "posts_per_page": limit,
                    "has_next_page": has_next_page,
                    "has_previous_page": has_previous_page
                }
            }


        return {
            "status": 200,
            "data": {
                "isClient": (role == "client"),
                "posts": posts,
                "pagination": {
                    "current_page": page,
                    "total_pages": total_pages,
                    "total_posts": total_posts_count,
                    "posts_per_page": limit,
                    "has_next_page": has_next_page,
                    "has_previous_page": has_previous_page
                }
            }
        }
 
    except Exception as e:
        print(e)
        raise HTTPException(status_code=500, detail=f"Unexpected error occurred: {str(e)}")
  

@router.get("/gym_assigned_data")
async def get_gym_details(
    gym_id: int = Query(...),
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis),
):
    client_redis_key = f"gym:{gym_id}:all_clients"
    trainer_redis_key = f"gym:{gym_id}:trainers"
    diet_redis_key = f"gym:{gym_id}:all_diets"
    workout_redis_key = f"gym:{gym_id}:all_workouts"
 
    response = {
        "status": 200,
        "message": "Successfully fetched.",
        "data": {
            "clients": [],
            "trainers": []
        }
    }

    try:
 
        cached_clients = await redis.get(client_redis_key)
        if cached_clients:
            clients_data = json.loads(cached_clients)
        else:
            clients = db.query(Client.client_id, Client.name, Client.training_id, Client.batch_id, Client.profile).filter(Client.gym_id == gym_id).all()
            if clients:                
                clients_data = []
                for client in clients:
                    gym_client_id = db.query(ClientGym).filter(ClientGym.client_id == client.client_id).first()
                    scheduler = db.query(ClientScheduler).filter(
                        ClientScheduler.gym_id == gym_id,
                        ClientScheduler.client_id == client.client_id
                    ).first()
                   
                    training = db.query(GymPlans).filter(GymPlans.id == client.training_id).first()
                    if training:
                        training_type = training.plans
                    else:
                        training_type = None
 
                    batch_t = db.query(GymBatches).filter(GymBatches.batch_id == client.batch_id).first()
                    if batch_t:
                        batch = batch_t.batch_name
                    else:
                        batch = None
 
                    client_data = {
                        "client_id": client.client_id,
                        "name": client.name,
                        "training": training_type,
                        "profile": client.profile,
                        "batch": batch,
                        "gym_client_id": gym_client_id.gym_client_id if gym_client_id else ''
                    }

                    assigned_trainer = None
                    assigned_trainer_info = None
                   
                    if scheduler:
                        if hasattr(scheduler, 'assigned_trainer_profile') and scheduler.assigned_trainer_profile:
                            trainer_profile = db.query(TrainerProfile).filter(
                                TrainerProfile.profile_id == scheduler.assigned_trainer_profile
                            ).first()
                            if trainer_profile:
                                assigned_trainer = trainer_profile.full_name
                                assigned_trainer_info = {
                                    "trainer_id": trainer_profile.trainer_id,
                                    "profile_id": trainer_profile.profile_id,
                                    "name": trainer_profile.full_name,
                                    "specialization": trainer_profile.specializations,
                                    "personal_trainer": trainer_profile.personal_trainer
                                }
 
                        elif scheduler.assigned_trainer:
                            trainer_profile = db.query(TrainerProfile).filter(
                                TrainerProfile.trainer_id == scheduler.assigned_trainer,
                                TrainerProfile.gym_id == gym_id
                            ).first()
                            if trainer_profile:
                                assigned_trainer = trainer_profile.full_name
                                assigned_trainer_info = {
                                    "trainer_id": trainer_profile.trainer_id,
                                    "profile_id": trainer_profile.profile_id,
                                    "name": trainer_profile.full_name,
                                    "specialization": trainer_profile.specializations,
                                    "personal_trainer": trainer_profile.personal_trainer
                                }
                            else:
                                trainer = db.query(Trainer).filter(Trainer.trainer_id == scheduler.assigned_trainer).first()
                                assigned_trainer = trainer.full_name if trainer else None
                                assigned_trainer_info = {
                                    "trainer_id": trainer.trainer_id if trainer else None,
                                    "profile_id": None,
                                    "name": trainer.full_name if trainer else None,
                                    "specialization": trainer.specializations if trainer else None,
                                    "personal_trainer": False
                                }
 
                    client_data["assigned_trainer"] = assigned_trainer
                    client_data["assigned_trainer_info"] = assigned_trainer_info
                       
                    if scheduler and scheduler.assigned_dietplan:
                        diet_plan = db.query(TemplateDiet).filter(TemplateDiet.template_id == scheduler.assigned_dietplan).first()
                        assigned_dietplan = diet_plan.template_name if diet_plan else None
                        client_data["assigned_dietplan"] = assigned_dietplan
                    else:
                        client_data["assigned_dietplan"] = None
 
                    if scheduler and scheduler.assigned_workoutplan:
                        workout_plan = db.query(TemplateWorkout).filter(TemplateWorkout.id == scheduler.assigned_workoutplan).first()
                        workout_plan_name = workout_plan.name if workout_plan else None
                        client_data["assigned_workoutplan"] = workout_plan_name
                    else:
                        client_data["assigned_workoutplan"] = None
 
                    clients_data.append(client_data)
 
                await redis.set(client_redis_key, json.dumps(clients_data), ex=3600)
            else:
                clients_data = []
 
        response["data"]["clients"] = clients_data
 
        cached_trainers = await redis.get(trainer_redis_key)
        if cached_trainers:
            trainers_data = json.loads(cached_trainers)
        else:
            trainers_query = db.query(Trainer, TrainerProfile).join(
                TrainerProfile, Trainer.trainer_id == TrainerProfile.trainer_id
            ).filter(TrainerProfile.gym_id == gym_id).all()
           
            if trainers_query:
                trainers_data = []
                for trainer, profile in trainers_query:
                    trainer_data = {
                        "trainer_id": trainer.trainer_id,
                        "profile_id": profile.profile_id,
                        "full_name": profile.full_name,  
                        "specialization": profile.specializations,  
                        "personal_trainer": profile.personal_trainer,
                        "can_view_client_data": profile.can_view_client_data,
                        "experience": profile.experience,
                        "email": profile.email,
                        "contact": trainer.contact
                    }
                    trainers_data.append(trainer_data)
               
                await redis.set(trainer_redis_key, json.dumps(trainers_data), ex=3600)
            else:
                trainers_data = []
 
        response["data"]["trainers"] = trainers_data
 
        cached_diets = await redis.get(diet_redis_key)
        if cached_diets:
            diets_data = json.loads(cached_diets)
        else:
            diets = db.query(TemplateDiet.template_id, TemplateDiet.template_name, TemplateDiet.template_details).filter(TemplateDiet.gym_id == gym_id).all()
            if diets:
                diets_data = [
                    {"template_id": diet.template_id, "template_name": diet.template_name}
                    for diet in diets
                ]
                await redis.set(diet_redis_key, json.dumps(diets_data), ex=3600)  
            else:
                diets_data = []
 
        response["data"]["diets"] = diets_data
 
        cached_workouts = await redis.get(workout_redis_key)
        if cached_workouts:
            workouts_data = json.loads(cached_workouts)
        else:
            workouts = db.query(TemplateWorkout.id, TemplateWorkout.name, TemplateWorkout.workoutPlan, TemplateWorkout.notes).filter(TemplateWorkout.gym_id == gym_id).all()
            if workouts:
                workouts_data = [
                    {"id": workout.id, "name": workout.name}
                    for workout in workouts
                ]
                await redis.set(workout_redis_key, json.dumps(workouts_data), ex=3600)  
            else:
                workouts_data = []
 
        response["data"]["workouts"] = workouts_data
 
        return response
 
    except Exception as e:
        print(e)
        raise HTTPException(status_code=500, detail=f"Unexpected error occurred: {str(e)}")
 




class AssignTrainerRequest(BaseModel):
    gym_id: int
    method: str
    diet_id: Optional[int] = None
    workout_id: Optional[int] = None
    trainer_id: Optional[int] = None  
    profile_id: Optional[int] = None  
    use_profile_id: Optional[bool] = False  
    batch_id: Optional[int] = None
    training_id: Optional[int] = None
    client_ids: Optional[List[int]] = None  
 
@router.post("/assign_trainer")
async def assign_trainer_to_clients(
    request: AssignTrainerRequest,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis),
):
   
    trainer_id = request.trainer_id
    profile_id = request.profile_id
    use_profile_id = request.use_profile_id
    gym_id = request.gym_id
    batch_id = request.batch_id
    training_id = request.training_id
    client_ids = request.client_ids
    method = request.method
    diet_id = request.diet_id
    workout_id = request.workout_id
 
    if method == "trainer":
        if use_profile_id and not profile_id:
            raise HTTPException(status_code=400, detail="profile_id is required when use_profile_id is True")
        elif not use_profile_id and not trainer_id:
            raise HTTPException(status_code=400, detail="trainer_id is required when use_profile_id is False")
       
        if use_profile_id:
            trainer_profile = db.query(TrainerProfile).filter(
                TrainerProfile.profile_id == profile_id,
                TrainerProfile.gym_id == gym_id
            ).first()
            if not trainer_profile:
                raise HTTPException(status_code=404, detail="Trainer profile not found for this gym")
        else:
            trainer_profile = db.query(TrainerProfile).filter(
                TrainerProfile.trainer_id == trainer_id,
                TrainerProfile.gym_id == gym_id
            ).first()
            if not trainer_profile:
                raise HTTPException(status_code=404, detail="Trainer not found for this gym")
 
    client_redis_key = f"gym:{gym_id}:clients"
 
    cached_clients = await redis.get(client_redis_key)
    if cached_clients:
        clients_data = json.loads(cached_clients)
    else:
        clients = db.query(Client.client_id, Client.name, Client.training_id, Client.batch_id).filter(
            Client.gym_id == gym_id
        ).all()
 
        if clients:
            clients_data = [
                {
                    "client_id": client.client_id,
                    "name": client.name,
                    "training_id": client.training_id,
                    "batch_id": client.batch_id,
                }
                for client in clients
            ]
            await redis.set(client_redis_key, json.dumps(clients_data), ex=3600)
        else:
            clients_data = []
 
    if client_ids:
        filtered_clients = [client for client in clients_data if client["client_id"] in client_ids]
    elif batch_id:
        filtered_clients = [client for client in clients_data if client["batch_id"] == batch_id]
    elif training_id:
        filtered_clients = [client for client in clients_data if client["training_id"] == training_id]
    else:
        filtered_clients = clients_data
 
    if not filtered_clients:
        raise HTTPException(status_code=404, detail="No clients found for the given criteria.")
 
    try:
        for client in filtered_clients:
            scheduler_entry = db.query(ClientScheduler).filter(
                ClientScheduler.client_id == client["client_id"],
                ClientScheduler.gym_id == gym_id
            ).first()
 
            if scheduler_entry:
                if method == "trainer":
                    if use_profile_id and hasattr(scheduler_entry, 'assigned_trainer_profile'):
                        scheduler_entry.assigned_trainer = profile_id
                    else:
                        scheduler_entry.assigned_trainer = trainer_id
                elif method == "diet":
                    scheduler_entry.assigned_dietplan = diet_id
                elif method == "workout":
                    scheduler_entry.assigned_workoutplan = workout_id
            else:
                if method == "trainer":
                    if use_profile_id and hasattr(ClientScheduler, 'assigned_trainer_profile'):
                       
                        new_entry = ClientScheduler(
                            gym_id=gym_id,
                            client_id=client["client_id"],
                            assigned_trainer_profile=profile_id,
                        )
                    else:
                        new_entry = ClientScheduler(
                            gym_id=gym_id,
                            client_id=client["client_id"],
                            assigned_trainer=trainer_id,
                        )
                    db.add(new_entry)
                elif method == "diet":
                    new_entry = ClientScheduler(
                        gym_id=gym_id,
                        client_id=client["client_id"],
                        assigned_dietplan=diet_id,
                    )
                    db.add(new_entry)
                elif method == "workout":
                    new_entry = ClientScheduler(
                        gym_id=gym_id,
                        client_id=client["client_id"],
                        assigned_workoutplan=workout_id,
                    )
                    db.add(new_entry)
       
            assigned_plans_key = f'{client["client_id"]}:{gym_id}:assigned_plans'
            if await redis.exists(assigned_plans_key):
                await redis.delete(assigned_plans_key)
       
        cache_keys = [
            f"gym:{gym_id}:all_clients",
            f"gym:{gym_id}:trainers",
            f"gym:{gym_id}:all_diets",
            f"gym:{gym_id}:all_workouts",
            f"gym:{gym_id}:clientdata",  
            f"trainer:{trainer_id}:gym:{gym_id}:assigned_clients" if trainer_id else None
        ]
       
        for key in cache_keys:
            if key and await redis.exists(key):
                await redis.delete(key)
 
        db.commit()
       
        assignment_type = "trainer profile" if use_profile_id else method
        return {
            "status": 200,
            "message": f"{assignment_type.title()} assignment successful",
            "details": {
                "assigned_clients": len(filtered_clients),
                "method": method,
                "use_profile_id": use_profile_id if method == "trainer" else None
            }
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error assigning {method}: {str(e)}")
 
 

@router.get("/gym_details")
async def get_gym_details(
    gym_id: int = Query(...),
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis),
):
    client_redis_key = f"gym:{gym_id}:all_clients"
    trainer_redis_key = f"gym:{gym_id}:trainers"
 
    response = {
        "status_code": 200,
        "message": "Successfully fetched.",
        "data": {
            "clients": [],
            "trainers": []
        }
    }
 
    try:
        cached_clients = await redis.get(client_redis_key)
        if cached_clients:
            clients_data = json.loads(cached_clients)
            client_message = "Client data fetched from Redis."
        else:
            clients = db.query(
                Client.client_id,
                Client.name,
                Client.training_id,
                Client.batch_id
            ).filter(Client.gym_id == gym_id).all()
           
            if clients:
                clients_data = []
                for client in clients:
                    scheduler = db.query(ClientScheduler).filter(
                        ClientScheduler.gym_id == gym_id,
                        ClientScheduler.client_id == client.client_id
                    ).first()
                   
                    client_data = {
                        "client_id": client.client_id,
                        "name": client.name,
                        "training_id": client.training_id,
                        "batch_id": client.batch_id,
                    }
                   
                    if scheduler:
                        if hasattr(scheduler, 'assigned_trainer_profile') and scheduler.assigned_trainer_profile:
                            
                            trainer_profile = db.query(TrainerProfile).filter(
                                TrainerProfile.profile_id == scheduler.assigned_trainer_profile
                            ).first()
                            if trainer_profile:
                                client_data["assigned_trainer"] = {
                                    "trainer_id": trainer_profile.trainer_id,
                                    "profile_id": trainer_profile.profile_id,
                                    "name": trainer_profile.full_name,
                                    "personal_trainer": trainer_profile.personal_trainer
                                }
                            else:
                                client_data["assigned_trainer"] = None
                        elif hasattr(scheduler, 'assigned_trainer') and scheduler.assigned_trainer:
                            
                            trainer_profile = db.query(TrainerProfile).filter(
                                TrainerProfile.trainer_id == scheduler.assigned_trainer,
                                TrainerProfile.gym_id == gym_id
                            ).first()
                            if trainer_profile:
                                client_data["assigned_trainer"] = {
                                    "trainer_id": trainer_profile.trainer_id,
                                    "profile_id": trainer_profile.profile_id,
                                    "name": trainer_profile.full_name,
                                    "personal_trainer": trainer_profile.personal_trainer
                                }
                            else:
                                client_data["assigned_trainer"] = None
                        else:
                            client_data["assigned_trainer"] = None
                    else:
                        client_data["assigned_trainer"] = None
 
                    clients_data.append(client_data)
 
                await redis.set(client_redis_key, json.dumps(clients_data), ex=3600)
                client_message = "Client data fetched from DB and cached in Redis."
            else:
                clients_data = []
                client_message = "No clients found for the given gym."
 
        response["data"]["clients"] = clients_data
 
        cached_trainers = await redis.get(trainer_redis_key)
        if cached_trainers:
            trainers_data = json.loads(cached_trainers)
            trainer_message = "Trainer data fetched from Redis."
        else:
            trainers_query = db.query(Trainer, TrainerProfile).join(
                TrainerProfile, Trainer.trainer_id == TrainerProfile.trainer_id
            ).filter(TrainerProfile.gym_id == gym_id).all()
           
            if trainers_query:
                trainers_data = []
                for trainer, profile in trainers_query:
                    trainer_data = {
                        "trainer_id": trainer.trainer_id,
                        "profile_id": profile.profile_id,
                        "full_name": profile.full_name,  
                        "email": profile.email,
                        "contact": trainer.contact,
                        "specialization": profile.specializations,
                        "experience": profile.experience,
                        "profile_image": profile.profile_image,
                        "can_view_client_data": profile.can_view_client_data,
                        "personal_trainer": profile.personal_trainer,
                        "availability": profile.availability,
                        "certifications": profile.certifications
                    }
                    trainers_data.append(trainer_data)
               
                await redis.set(trainer_redis_key, json.dumps(trainers_data), ex=3600)
                trainer_message = "Trainer data fetched from DB and cached in Redis."
            else:
                trainers_data = []
                trainer_message = "No trainers found for the given gym."
 
        response["data"]["trainers"] = trainers_data
 
        response["message"] = f"{client_message} {trainer_message}"
 
        return response
 
    except Exception as e:
        print(f"Error in get_gym_details: {str(e)}")
        return {
            "status": 500,
            "message": f"An error occurred: {str(e)}",
            "data": {
                "clients": [],
                "trainers": []
            }
        }
 
 

@router.get("/profile_data")
async def get_profile_data(
    role: str,
    client_id: int = None,
    owner_id: int = None,
    gym_id: int = None,
    trainer_id: int = None,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis),
):  
    """
    Get profile data for client, owner, or trainer with proper response structure
    """
    try:
        if role == "client":
            client = (
                db.query(Client)
                .filter(Client.client_id == client_id)
                .first()
                if client_id
                else None
            )
            return {
                "status":200,
                "message": "Client data retrieved successfully" if client else "No client data found",
                "client_data": client,
                "owner_data": None,
                "gym_data": None,
            }
 
        elif role == "owner":
            owner = (
                db.query(GymOwner)
                .filter(GymOwner.owner_id == owner_id)
                .first()
                if owner_id
                else None
            )
 
            if not owner:
                return {
                    "status":404,
                    "message": "Owner not found",
                    "client_data": None,
                    "owner_data": None,
                    "gym_data": None,
                }
 
            gyms = db.query(Gym).filter(Gym.owner_id == owner_id).all()

            if gym_id:
                gym = db.query(Gym).filter(Gym.gym_id == gym_id).first()
            else:
                gym = gyms[0] if gyms else None

            if not gym:
                return {
                    "status": 404,
                    "message": "Gym not found",
                    "client_data": None,
                    "owner_data": None,
                    "gym_data": None,
                }

            account = db.query(AccountDetails).filter(AccountDetails.gym_id == gym.gym_id).first()
            gym_details = db.query(GymDetails).filter(GymDetails.gym_id == gym.gym_id).first()
            gym_data = {}
 
            gyms_data=[]
 
            members_key = f"gym:{gym.gym_id}:members"
            cached_data = await redis.hgetall(members_key)
 
            if cached_data:
                total_members = int(cached_data.get("total_members", 0))
                access_count = int(cached_data.get("access_count", 0))
            else:
                total_members = db.query(Client).filter(Client.gym_id == gym.gym_id).count()
                access_count = db.query(Client).filter(
                    Client.gym_id == gym.gym_id, Client.access == True
                ).count()
 
            gym_data = {
                "gym_id": gym.gym_id,
                "owner_id":gym.owner_id,
                "name": gym.name,
                "location": gym.location,
                "cover_pic":gym.cover_pic,
                "contact_number":gym.contact_number,
                "services":gym.services,
                "operating_hours":gym.operating_hours,
                "logo":gym.logo,
                "address_street":gym.street,
                "address_area":gym.area,
                "address_city":gym.city,
                "address_state":gym.state,
                "address_pincode":gym.pincode,
                "address": {
                    "street": gym.street,
                    "area": gym.area,
                    "city": gym.city,
                    "state": gym.state,
                    "pincode": gym.pincode
                },
                "referral_id":gym.referal_id,
                "subscription_start_date": gym.subscription_start_date,
                "subscription_end_date": gym.subscription_end_date,
                "total_members": total_members,
                "access_count": access_count,
                "account_number":account.account_number if account else None,
                "account_holdername":account.account_holdername if account else None,
                "account_ifsccode":account.account_ifsccode if account else None,
                "account_branch":account.account_branch if account else None,
                "bank_name":account.bank_name if account else None,
                "account_id":account.account_id if account else None,
                "upi_id":account.upi_id if account else None,
                "gst_number":account.gst_number if account else None,
                "gst_type":account.gst_type if account else None,
                "gst_percentage":account.gst_percentage if account else None,
                "total_machineries": gym_details.total_machineries if gym_details else None,
                "floor_space": gym_details.floor_space if gym_details else None,
                "total_trainers": gym_details.total_trainers if gym_details else None,
                "yearly_membership_cost": gym_details.yearly_membership_cost if gym_details else None
            }
 
            for current_gym in gyms:
                current_account = db.query(AccountDetails).filter(AccountDetails.gym_id == current_gym.gym_id).first()
                current_gym_details = db.query(GymDetails).filter(GymDetails.gym_id == current_gym.gym_id).first()
                current_data = {
                    "gym_id": current_gym.gym_id,
                    "owner_id":current_gym.owner_id,
                    "referral_id":current_gym.referal_id,
                    "name": current_gym.name,
                    "location": current_gym.location,
                    "cover_pic":current_gym.cover_pic,
                    "logo":current_gym.logo,
                    "contact_number":current_gym.contact_number,
                    "services":current_gym.services,
                    "operating_hours":current_gym.operating_hours,
                    "logo":current_gym.logo,
                    "address_street":current_gym.street,
                    "address_area":current_gym.area,
                    "address_city":current_gym.city,
                    "address_state":current_gym.state,
                    "address_pincode":current_gym.pincode,
                    "address": {
                        "street": current_gym.street,
                        "area": current_gym.area,
                        "city": current_gym.city,
                        "state": current_gym.state,
                        "pincode": current_gym.pincode
                    },
                    "subscription_start_date": current_gym.subscription_start_date,
                    "subscription_end_date": current_gym.subscription_end_date,
                    "total_members": total_members,
                    "access_count": access_count,
                    "account_number":current_account.account_number if current_account else None,
                    "account_holdername":current_account.account_holdername if current_account else None,
                    "account_ifsccode":current_account.account_ifsccode if current_account else None,
                    "account_branch":current_account.account_branch if current_account else None,
                    "bank_name":current_account.bank_name if current_account else None,
                    "account_id":current_account.account_id if current_account else None,
                    "upi_id":current_account.upi_id if current_account else None,
                    "gst_number":current_account.gst_number if current_account else None,
                    "gst_type":current_account.gst_type if current_account else None,
                    "gst_percentage":current_account.gst_percentage if current_account else None,
                    "total_machineries": current_gym_details.total_machineries if current_gym_details else None,
                    "floor_space": current_gym_details.floor_space if current_gym_details else None,
                    "total_trainers": current_gym_details.total_trainers if current_gym_details else None,
                    "yearly_membership_cost": current_gym_details.yearly_membership_cost if current_gym_details else None
                }
 
                gyms_data.append(current_data)
           
            gyms_count = len(gyms_data)
 
           
            return {
                "status":200,
                "message": "gyms data retrieved successfully",
                "data":{
                "client_data": None,
                "owner_data": owner,
                "gym_data": gym_data,
                "gyms":gyms_data,
                "gyms_count":gyms_count
                }
            }
       
        elif role == "trainer":
            trainer = (
                db.query(Trainer)
                .filter(Trainer.trainer_id == trainer_id)
                .first()
                if trainer_id
                else None
            )
            if not trainer:
                return {
                    "status": 404,
                    "message": "Trainer not found",
                    "client_data": None,
                    "owner_data": None,
                    "gym_data": None,
                }
           
            gyms_query = db.query(Gym, TrainerProfile).join(
                TrainerProfile, Gym.gym_id == TrainerProfile.gym_id
            ).filter(TrainerProfile.trainer_id == trainer_id).all()
           
            gyms = [gym for gym, profile in gyms_query]
            profiles = [profile for gym, profile in gyms_query]
           
            gyms_data = []
            gym_data = {}
 
            gym = db.query(Gym).filter(Gym.gym_id == gym_id).first()
            account = db.query(AccountDetails).filter(AccountDetails.gym_id == gym_id).first()
 
            current_gym_profile = db.query(TrainerProfile).filter(
                TrainerProfile.trainer_id == trainer_id,
                TrainerProfile.gym_id == gym_id
            ).first()
 
            members_key = f"gym:{gym.gym_id}:members"
            cached_data = await redis.hgetall(members_key)
 
            if cached_data:
                total_members = int(cached_data.get("total_members", 0))
                access_count = int(cached_data.get("access_count", 0))
            else:
                total_members = db.query(Client).filter(Client.gym_id == gym.gym_id).count()
                access_count = db.query(Client).filter(
                    Client.gym_id == gym.gym_id, Client.access == True
                ).count()
 
            gym_data = {
                "gym_id": gym.gym_id,
                "owner_id": gym.owner_id,
                "name": gym.name,
                "location": gym.location,
                "cover_pic": gym.cover_pic,
                "logo": gym.logo,
                "contact_number":gym.contact_number,
                "services":gym.services,
                "operating_hours":gym.operating_hours,
                "logo":gym.logo,
                "address_street":gym.street,
                "address_area":gym.area,
                "address_city":gym.city,
                "address_state":gym.state,
                "address_pincode":gym.pincode,
                "address": {
                    "street": gym.street,
                    "area": gym.area,
                    "city": gym.city,
                    "state": gym.state,
                    "pincode": gym.pincode
                },
                "referal_id":gym.referal_id,
                "subscription_start_date": gym.subscription_start_date,
                "subscription_end_date": gym.subscription_end_date,
                "total_members": total_members,
                "access_count": access_count,
                "account_number":account.account_number if account else None,
                "account_holdername":account.account_holdername if account else None,
                "account_ifsccode":account.account_ifsccode if account else None,
                "account_branch":account.account_branch if account else None,
                "bank_name":account.bank_name if account else None,
                "account_id":account.account_id if account else None,
                "upi_id":account.upi_id if account else None,
                "gst_number":account.gst_number if account else None,
                "gst_type":account.gst_type if account else None,
                "gst_percentage":account.gst_percentage if account else None,
                "trainer_permissions": {
                    "can_view_client_data": current_gym_profile.can_view_client_data if current_gym_profile else False,
                    "personal_trainer": current_gym_profile.personal_trainer if current_gym_profile else False
                }
            }
           
            for gym, profile in gyms_query:
                members_key = f"gym:{gym.gym_id}:members"
                cached_data = await redis.hgetall(members_key)
               
                if cached_data:
                    total_members = int(cached_data.get("total_members", 0))
                    access_count = int(cached_data.get("access_count", 0))
                else:
                    total_members = db.query(Client).filter(Client.gym_id == gym.gym_id).count()
                    access_count = db.query(Client).filter(
                        Client.gym_id == gym.gym_id, Client.access == True
                    ).count()
               
                gym_info = {
                    "gym_id": gym.gym_id,
                    "owner_id": gym.owner_id,
                    "name": gym.name,
                    "location": gym.location,
                    "cover_pic": gym.cover_pic,
                    "logo": gym.logo,
                    "contact_number":gym.contact_number,
                    "services":gym.services,
                    "operating_hours":gym.operating_hours,
                    "logo":gym.logo,
                    "address_street":gym.street,
                    "address_area":gym.area,
                    "address_city":gym.city,
                    "address_state":gym.state,
                    "address_pincode":gym.pincode,
                    "address": {
                        "street": gym.street,
                        "area": gym.area,
                        "city": gym.city,
                        "state": gym.state,
                        "pincode": gym.pincode
                    },
                    "referral_id":gym.referal_id,
                    "subscription_start_date": gym.subscription_start_date,
                    "subscription_end_date": gym.subscription_end_date,
                    "total_members": total_members,
                    "access_count": access_count,
                    "trainer_permissions": {
                        "can_view_client_data": profile.can_view_client_data if profile else False,
                        "personal_trainer": profile.personal_trainer if profile else False,
                        "profile_id": profile.profile_id if profile else None
                    },
                    "gym_specific_trainer_info": {
                        "full_name": profile.full_name if profile else trainer.full_name,
                        "email": profile.email if profile else trainer.email,
                        "specialization": profile.specializations if profile else trainer.specializations,
                        "experience": profile.experience if profile else trainer.experience,
                        "certifications": profile.certifications if profile else trainer.certifications,
                        "work_timings": profile.work_timings if profile else trainer.work_timings,
                        "profile_image": profile.profile_image if profile else trainer.profile_image
                    }
                }
                gyms_data.append(gym_info)
           
            trainer_data = {
                "name": current_gym_profile.full_name if current_gym_profile else trainer.full_name,
                "email": current_gym_profile.email if current_gym_profile else trainer.email,
                "profileImage": current_gym_profile.profile_image if current_gym_profile else trainer.profile_image,
                "trainer_id": trainer.trainer_id,
                "gender": trainer.gender,
                "contact": trainer.contact,
                "specialization": current_gym_profile.specializations if current_gym_profile else trainer.specializations,
                "experience": current_gym_profile.experience if current_gym_profile else trainer.experience,
                "certifications": current_gym_profile.certifications if current_gym_profile else trainer.certifications,
                "work_timings": current_gym_profile.work_timings if current_gym_profile else trainer.work_timings,
                
                "permissions": {
                    "can_view_client_data": current_gym_profile.can_view_client_data if current_gym_profile else False,
                    "personal_trainer": current_gym_profile.personal_trainer if current_gym_profile else False
                },
                "profile_id": current_gym_profile.profile_id if current_gym_profile else None
            }
           
            return {
                "status": 200,
                "message": "Trainer data retrieved successfully",
                "data": {
                    "client_data": None,
                    "owner_data": trainer_data,
                    "gym_data": gym_data,
                    "gyms": gyms_data,  
                    "gyms_count": len(gyms_data)
                }
            }
 
        else:
            return {
                "status":400,
                "message": "Invalid role specified. Role must be either 'client', 'owner', or 'trainer'.",
                "client_data": None,
                "owner_data": None,
                "gym_data": None,
            }
       
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected error occured : {str(e)}")
 
 
 

class GymData(BaseModel):
    gym_id: Optional[int]=None
    name: Optional[str]=None
    location: Optional[str]=None
    referral_id:Optional[str]=None



class UpdateProfileRequest(BaseModel):
    owner_id:int
    method:str
    role:Optional[str]=None
    oldPassword:Optional[str]=None
    newPassword:Optional[str]=None
    client_data : Optional[str]= None
    gym_data : Optional[dict]=None
    owner_data: Optional[dict]=None
    
@router.put("/update_profile")
async def update_profile(request: UpdateProfileRequest, db: Session = Depends(get_db)):

    if request.method == "profile":
        if request.role != "owner":
            raise HTTPException(status_code=400, detail="Only owners can update profiles.")

        owner = db.query(GymOwner).filter(GymOwner.owner_id == request.owner_id).first()
        if not owner:
            raise HTTPException(status_code=404, detail="Owner not found.")

        is_changed=False
        if request.owner_data["contact_number"]:

   
            if not owner.contact_number == request.owner_data["contact_number"]:


                existing_owner = db.query(GymOwner).filter(
                    (GymOwner.contact_number == request.owner_data["contact_number"])
                ).first()
        
                if existing_owner:
                    raise HTTPException(status_code=400, detail="Mobile number already registered with different account")
                owner.verification= '{"mobile": false,"email": true}' 
                is_changed=True
            owner.contact_number = request.owner_data['contact_number']

        
        if request.owner_data["dob"]:
            if not owner.dob == request.owner_data["dob"]:
                today=date.today()  
                dob = datetime.strptime(str(request.owner_data["dob"]), "%Y-%m-%d").date()
                age=today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
                owner.age = age
            owner.dob = request.owner_data["dob"]
       
        if request.owner_data["email"]:
            if not owner.email == request.owner_data["email"]:
                existing_owner = db.query(GymOwner).filter(
                    (GymOwner.email == request.owner_data["email"])
                ).first()
        

                if existing_owner:
                        raise HTTPException(status_code=400, detail="Email already registered with different account")
                owner.verification='{"mobile": true,"email": false}'
                is_changed=True
            owner.email = request.owner_data["email"]

        if not (owner.contact_number == request.owner_data["contact_number"] and owner.email == request.owner_data["email"]):
            owner.verification='{"mobile": false,"email": false}'
            is_changed=True

        if request.owner_data["name"]:
            owner.name = request.owner_data["name"]        

        if request.gym_data:
            gym = db.query(Gym).filter(Gym.gym_id == request.gym_data["gym_id"], Gym.owner_id == request.owner_id).first()
            if not gym:
                raise HTTPException(status_code=404, detail=f'Gym with ID {request.gym_data["gym_id"]} not found or unauthorized.')
            if request.gym_data["gymName"]:
                gym.name = request.gym_data["gymName"]
            if request.gym_data["location"]:
                gym.location = request.gym_data["location"]

            if request.gym_data['referral_id'] and request.gym_data['referral_id'] is not None:
                database_gym = db.query(GymDatabase).filter(GymDatabase.referal_id == request.gym_data['referral_id']).first()
                if not database_gym:
                    raise HTTPException(status_code=404, detail=f"Invalid referal ID")
                gym.referal_id = request.gym_data['referral_id']
                gym.fittbot_verified=True

            if request.gym_data["contact_number"]:
                gym.contact_number = request.gym_data["contact_number"]
            if request.gym_data["services"]:
                gym.services = request.gym_data["services"]
            if request.gym_data["operating_hours"]:
                gym.operating_hours = request.gym_data["operating_hours"]
            if request.gym_data["address"]:
                address = request.gym_data["address"]
                gym.street = address.get("street", gym.street)
                gym.area = address.get("area", gym.area)
                gym.city = address.get("city", gym.city)
                gym.state = address.get("state", gym.state)
                gym.pincode = address.get("pincode", gym.pincode)


            if request.gym_data["account_id"]:
                account = db.query(AccountDetails).filter(AccountDetails.account_id == request.gym_data["account_id"]).first()
                if account:
                    account.account_number = request.gym_data["account_number"]
                    account.account_holdername = request.gym_data["account_holdername"]
                    account.account_ifsccode = request.gym_data["account_ifsccode"]
                    account.bank_name = request.gym_data["bank_name"]
                    account.account_branch = request.gym_data["account_branch"]
                    account.upi_id = request.gym_data["upi_id"]
                    account.gst_number = request.gym_data["gst_number"]
                    account.gst_type = request.gym_data["gst_type"]
                    account.gst_percentage = 18


            else:
                new_account = AccountDetails(
                    gym_id = request.gym_data['gym_id'],
                    account_number = request.gym_data["account_number"],
                    account_holdername = request.gym_data["account_holdername"],
                    account_ifsccode = request.gym_data["account_ifsccode"],
                    bank_name = request.gym_data["bank_name"],
                    account_branch = request.gym_data["account_branch"],
                    upi_id = request.gym_data["upi_id"],
                    gst_number = request.gym_data['gst_number'],
                    gst_type = request.gym_data['gst_type'],
                    gst_percentage = 18
                )

                db.add(new_account)

            # Handle gym details update or creation
            gym_details = db.query(GymDetails).filter(GymDetails.gym_id == request.gym_data['gym_id']).first()

            if gym_details:
                # Update existing gym details
                if "total_machineries" in request.gym_data:
                    gym_details.total_machineries = request.gym_data["total_machineries"]
                if "floor_space" in request.gym_data:
                    gym_details.floor_space = request.gym_data["floor_space"]
                if "total_trainers" in request.gym_data:
                    gym_details.total_trainers = request.gym_data["total_trainers"]
                if "yearly_membership_cost" in request.gym_data:
                    gym_details.yearly_membership_cost = request.gym_data["yearly_membership_cost"]
                gym_details.updated_at = datetime.now()
            else:
                # Create new gym details if they don't exist
                new_gym_details = GymDetails(
                    gym_id=request.gym_data['gym_id'],
                    total_machineries=request.gym_data.get("total_machineries"),
                    floor_space=request.gym_data.get("floor_space"),
                    total_trainers=request.gym_data.get("total_trainers"),
                    yearly_membership_cost=request.gym_data.get("yearly_membership_cost"),
                    created_at=datetime.now(),
                    updated_at=datetime.now()
                )
                db.add(new_gym_details)

        try:
            db.commit()
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Error updating profile: {str(e)}")

        # Prepare response with gym_data if it was updated
        response_data = {
            "status": 200,
            "message": "Profile updated successfully.",
            "is_changed": is_changed
        }

        if request.gym_data:
            gym = db.query(Gym).filter(Gym.gym_id == request.gym_data["gym_id"]).first()
            account = db.query(AccountDetails).filter(AccountDetails.gym_id == request.gym_data["gym_id"]).first()
            gym_details = db.query(GymDetails).filter(GymDetails.gym_id == request.gym_data["gym_id"]).first()

            gym_data_response = {
                "gym_id": gym.gym_id,
                "owner_id": gym.owner_id,
                "name": gym.name,
                "location": gym.location,
                "cover_pic": gym.cover_pic,
                "contact_number": gym.contact_number,
                "services": gym.services,
                "operating_hours": gym.operating_hours,
                "logo": gym.logo,
                "address_street": gym.street,
                "address_area": gym.area,
                "address_city": gym.city,
                "address_state": gym.state,
                "address_pincode": gym.pincode,
                "address": {
                    "street": gym.street,
                    "area": gym.area,
                    "city": gym.city,
                    "state": gym.state,
                    "pincode": gym.pincode
                },
                "referral_id": gym.referal_id,
                "account_number": account.account_number if account else None,
                "account_holdername": account.account_holdername if account else None,
                "account_ifsccode": account.account_ifsccode if account else None,
                "account_branch": account.account_branch if account else None,
                "bank_name": account.bank_name if account else None,
                "account_id": account.account_id if account else None,
                "upi_id": account.upi_id if account else None,
                "gst_number": account.gst_number if account else None,
                "gst_type": account.gst_type if account else None,
                "gst_percentage": account.gst_percentage if account else None,
                "total_machineries": gym_details.total_machineries if gym_details else None,
                "floor_space": gym_details.floor_space if gym_details else None,
                "total_trainers": gym_details.total_trainers if gym_details else None,
                "yearly_membership_cost": gym_details.yearly_membership_cost if gym_details else None
            }
            response_data["gym_data"] = gym_data_response

        print("is changed issss",is_changed)
        return response_data
        

    elif request.method == "password":
        if request.role != "owner":
            raise HTTPException(status_code=400, detail="Only owners can change passwords.")

        owner = db.query(GymOwner).filter(GymOwner.owner_id == request.owner_id).first()
        if not owner:
            raise HTTPException(status_code=404, detail="Owner not found.")

        if not verify_password(request.oldPassword, owner.password):

            raise HTTPException(status_code=400, detail="Incorrect old password.")
        else:
            hashed_password = get_password_hash(request.newPassword)

            owner.password = hashed_password

        try:
            db.commit()
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Error updating password: {str(e)}")

        return {"status":200,"message": "Profile updated successfully."}


class TemplateDietRequest(BaseModel):
    name: str
    dietPlan: dict  
    gym_id: int

@router.post("/gym/adddiettemplate")
async def add_diet_template(
    template: TemplateDietRequest, db: Session = Depends(get_db),redis : Redis =Depends(get_redis)
):
    try:
        
        new_template = TemplateDiet(
            template_name=template.name,
            template_details=template.dietPlan,
            gym_id=template.gym_id
        )

        
        db.add(new_template)
        db.commit()

        diet_redis_key = f"gym:{template.gym_id}:all_diets"
        if await redis.exists(diet_redis_key):
            await redis.delete(diet_redis_key)

        client_redis_key = f"gym:{template.gym_id}:all_clients"
        if await redis.exists(client_redis_key):
            await redis.delete(client_redis_key)

        pattern = f"*:{template.gym_id}:assigned_plans"
        keys = await redis.keys(pattern)  

        if keys:  
            await redis.delete(*keys)


        return {"status": 200, "message": "Diet template added successfully"}
    except Exception as e:
        db.rollback()
        print(f"Error adding diet template: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error adding diet template: {str(e)}")


@router.get("/gym/getdiettemplate")
async def getDietTemplate(
    gym_id: int = Query(...), db: Session = Depends(get_db)
):
    try:
        
        template_names = db.query(TemplateDiet).filter(TemplateDiet.gym_id == gym_id).all()

        temp = [
            {
                "id": template.template_id,
                "name": template.template_name,
                "dietPlan": template.template_details
            }
            for template in template_names
        ]
        
        return {"status": 200, "message": "Templates retrieved successfully", "data": temp}
    except Exception as e:
        db.rollback()
        print(f"Error retrieving templates: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error retrieving templates: {str(e)}")



class UpdateDietTemplateRequest(BaseModel):
    id: int
    dietPlan: dict 
    gym_id: int 

class EditDietTemplateRequest(BaseModel):
    id: int
    name: str  
    gym_id: int


@router.put("/gym/updatediettemplate")
async def update_diet_template(
    request: UpdateDietTemplateRequest, db: Session = Depends(get_db),redis : Redis =Depends(get_redis)
):
    try:
       
        template = db.query(TemplateDiet).filter(TemplateDiet.template_id == request.id).first()
        if not template:
            raise HTTPException(status_code=404, detail="Template not found with the given ID")

        if request.dietPlan:
            template.template_details = request.dietPlan
        db.commit()
        db.refresh(template)


        diet_redis_key = f"gym:{request.gym_id}:all_diets"
        if await redis.exists(diet_redis_key):
            await redis.delete(diet_redis_key)
        
        client_redis_key = f"gym:{request.gym_id}:all_clients"
        if await redis.exists(client_redis_key):
            await redis.delete(client_redis_key)

        pattern = f"*:{request.gym_id}:assigned_plans"
        keys = await redis.keys(pattern)  

        if keys:  
            await redis.delete(*keys)



        return {
            "status": 200,
            "message": "Diet template updated successfully"
        }
    except Exception as e:
        db.rollback()
        print(f"Error updating diet template: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error updating diet template: {str(e)}")


@router.delete("/gym/deletediettemplate")
async def delete_diet_template(
    id: int,
    gym_id:int,
    db: Session = Depends(get_db),
    redis : Redis =Depends(get_redis)
):
    try:
    
        result = db.query(TemplateDiet).filter(
            TemplateDiet.template_id == id
        ).delete()
        db.commit()
        diet_redis_key = f"gym:{gym_id}:all_diets"
        if await redis.exists(diet_redis_key):
            await redis.delete(diet_redis_key)
        client_redis_key = f"gym:{gym_id}:all_clients"
        if await redis.exists(client_redis_key):
            await redis.delete(client_redis_key)
        
        pattern = f"*:{gym_id}:assigned_plans"
        keys = await redis.keys(pattern)  

        if keys:  
            await redis.delete(*keys)  



        if result == 0:
            raise HTTPException(status_code=404, detail="Template not found with the given ID")

        return {
            "status": 200,
            "message": "Diet template deleted successfully"
        }
    except Exception as e:
        db.rollback()
        print(f"Error deleting diet template: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error deleting diet template: {str(e)}")


@router.put("/gym/editdiettemplate")
async def edit_diet_template(
    input: EditDietTemplateRequest,
    db: Session = Depends(get_db), redis : Redis =Depends(get_redis)
):
    try:
       
        result = db.query(TemplateDiet).filter(
            TemplateDiet.template_id == input.id
        ).update({"template_name": input.name})
        db.commit()

        if not result:
            raise HTTPException(status_code=404, detail="Template not found")
        
        diet_redis_key = f"gym:{input.gym_id}:all_diets"
        if await redis.exists(diet_redis_key):
            await redis.delete(diet_redis_key)

        client_redis_key = f"gym:{input.gym_id}:all_clients"
        if await redis.exists(client_redis_key):
            await redis.delete(client_redis_key)

        pattern = f"*:{input.gym_id}:assigned_plans"
        keys = await redis.keys(pattern)  

        if keys:  
            await redis.delete(*keys)  


        return {
            "status": 200,
            "message": "Diet template name updated successfully"
        }
    except Exception as e:
        db.rollback()
        print(f"Error updating diet template name: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error updating diet template name: {str(e)}")




@router.get("/feedback")
async def get_feedback_for_gym(gym_id: int, db: Session = Depends(get_db)):
    try:
        result = db.execute(
            select(Gym_Feedback, Client.name, Client.profile)
            .join(Client, Gym_Feedback.client_id == Client.client_id)
            .filter(Gym_Feedback.gym_id == gym_id)
            .order_by(desc(Gym_Feedback.timing))
        )
        feedbacks = result.all()  
        if not feedbacks:
            return {"status": 200, "data": []}  
 
        response_data = [
            {
                "feedback_id": fb.Gym_Feedback.id,
                "client_id": fb.Gym_Feedback.client_id,
                "client_name": fb.name,
                "tag": fb.Gym_Feedback.tag,
                "ratings": fb.Gym_Feedback.ratings,
                "feedback": fb.Gym_Feedback.feedback,
                "timing": fb.Gym_Feedback.timing,
                "client_image":fb.profile
            }
            for fb in feedbacks
        ]
 
        return {"status": 200, "data": response_data}
 
    except Exception as e:
        print(e)
        return {"status": 500, "message": f"Internal Server Error: {str(e)}"}
 

from sqlalchemy.orm import Session, aliased


@router.get("/conversations")
async def get_home_conversations(
    gym_id: int,
    user_id: int,
    db: Session = Depends(get_db),
):
    try:
        role = "owner"

        RecipientClient = aliased(Client)
        SenderClient = aliased(Client)

        all_messages = (
            db.query(
                Message.message_id,
                Message.sender_id,
                Message.recipient_id,
                Message.sender_role,
                Message.recipient_role,
                Message.message,
                Message.is_read,
                Message.sent_at,
                (Message.sender_role == role).label("is_self"),
                RecipientClient.name.label("recipient_name"),
                SenderClient.name.label("sender_name"),
            )
            .outerjoin(RecipientClient, Message.recipient_id == RecipientClient.client_id)
            .outerjoin(SenderClient, Message.sender_id == SenderClient.client_id)
            .filter(
                Message.gym_id == gym_id,
                or_(
                    and_(Message.sender_id == user_id, Message.sender_role == role),
                    and_(Message.recipient_id == user_id, Message.recipient_role == role),
                ),
            )
            .order_by(Message.sent_at.desc())
            .all()
        )

        grouped_conversations = {}
        for msg in all_messages:
            conversation_key = tuple(sorted([(msg.sender_id, msg.sender_role), (msg.recipient_id, msg.recipient_role)]))
            if conversation_key not in grouped_conversations:
                grouped_conversations[conversation_key] = msg

        conversations = [
            {
                "message_id": msg.message_id,
                "sender_id": msg.sender_id,
                "recipient_id": msg.recipient_id,
                "message": msg.message,
                "sent_at": msg.sent_at.isoformat(),
                "is_self": msg.is_self,
                "client_name": msg.recipient_name if msg.is_self else msg.sender_name,  
                "client_id": msg.recipient_id if msg.is_self else msg.sender_id
            }
            for msg in grouped_conversations.values()
        ]

        return {"status": 200, "data": conversations}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching conversations: {str(e)}")


@router.get("/owner_messages")
async def get_messages(
    gym_id: int,
    owner_id: int,
    user_id: int,
    db: Session = Depends(get_db),
):
 
    try:
 
        conversation_user_id=user_id
        user_role="owner"
        conv_user_role="client"
 
        query = db.query(Message).filter(
            Message.gym_id == gym_id,
            (
                (Message.sender_id == owner_id) &
                (Message.recipient_id == conversation_user_id) &
                (Message.sender_role == user_role) &
                (Message.recipient_role == conv_user_role)
            )
            |
            (
                (Message.sender_id == conversation_user_id) &
                (Message.recipient_id == owner_id) &
                (Message.sender_role == conv_user_role) &
                (Message.recipient_role == user_role)
            )
        )
 
        messages = query.order_by(Message.sent_at.asc()).all()
 
        return {
            "status": 200,
            "data": [
                {
                    "message_id": msg.message_id,
                    "sender_id": msg.sender_id,
                    "recipient_id": msg.recipient_id,
                    "message": msg.message,
                    "sent_at": msg.sent_at.isoformat(),
                    "is_self": msg.sender_role == "owner",
                }
                for msg in messages
            ],
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching messages: {str(e)}")
    
class SendMessageRequest(BaseModel):
    sender_id: int
    gym_id: int
    recipient_id:int
    message: str
 
@router.post("/send_message_owners")
async def send_message(request: SendMessageRequest, db: Session = Depends(get_db)):
    try:
 
        recipient_id=request.recipient_id
        sender_role="owner"
        recipient_role="client"
       
        new_message = Message(
            sender_id=request.sender_id,
            recipient_id=recipient_id,
            gym_id=request.gym_id,
            sender_role=sender_role,
            recipient_role=recipient_role,
            message=request.message,
            sent_at=datetime.now()
        )
        db.add(new_message)
        db.commit()
 
 
        return {"success": True, "message": "Message sent successfully","status": 200}
 
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred {e}")

@router.get("/gym/pending_clients")
async def get_pending_clients(gym_id:int, db: Session = Depends(get_db), redis : Redis =Depends(get_redis)):
 
    redis_key = f"gym:{gym_id}:pendingClients"
    try:
        cached_data = await redis.get(redis_key)
        if cached_data:
            print(f"Data fetched from Redis for gym_id: {gym_id}")
            return {
                "status": 200,
                "message": "Pending Clients retrived successfully from cache.",
                "data": json.loads(cached_data)
            }
 
        print("Data not found in Redis, fetching from database.")
        clients = db.query(Client).filter(Client.gym_id == gym_id).all()
 
        pending=[]
 
        if not clients:
            return {
            "status": 200,
            "message": "Data fetched successfully.",
            "data": []
        }
 
        for client in clients:
            verification= json.loads(client.verification)
            if not verification['mobile']:
                client_data={
                    "id":client.client_id,
                    "name":client.name,
                    "email":client.email,
                    "contact":client.contact
                }
                pending.append(client_data)
 
        await redis.set(redis_key, json.dumps(pending), ex=86400)
 
        return {
            "status": 200,
            "message": "Data fetched successfully.",
            "data": pending
        }
    except Exception as e:
        print(e)
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")

class pendingClient(BaseModel):
    id:int
    contact: str
    email: str
    name: str

@router.put("/gym/edit_pending_clients")
async def edit_pending_clients(request: pendingClient, db: Session = Depends(get_db), redis : Redis =Depends(get_redis)):
 
    try:
        client = db.query(Client).filter(Client.client_id == request.id).first()
 
        if not client:
            raise HTTPException(status_code=400, detail="Clients not found")
       
        verification = json.loads(client.verification)
        if verification['mobile']:
            raise HTTPException(status_code=400, detail="Client already Verified")
        
        is_contact_changed = request.contact != client.contact
        is_email_changed = request.email != client.email



        if is_contact_changed:
          
            contact_exists = db.query(Client).filter(Client.contact == request.contact).first() or db.query(GymOwner).filter(GymOwner.contact_number == request.contact).first()
            
            if contact_exists:
                raise HTTPException(status_code=400, detail="Mobile number already exists")
            client.contact = request.contact
            password=generate_password()
            hashed_password = get_password_hash(password)


           
            mailsent = send_welcome_email(
                user_name=request.name,
                company_name="Fittbot",
                mobile_number=request.contact,
                default_password=password,
                login_url="",
                recipient_email=request.email,
                support_email="support@fittbot.com"
            )
    
            if mailsent:
                print(f"Welcome mail send to {request.email}")

            smssent = send_welcome_sms(
                user_name=request.name,
                phone_number=request.contact,
                default_password=password,
                login_url=""
            )

            if smssent:
                print(f"welcome SMS send successfully to {request.contact}")

        if is_email_changed:
            email_exists = db.query(Client).filter(Client.email == request.email).first() or db.query(GymOwner).filter(GymOwner.email == request.email).first()

            if email_exists:
                raise HTTPException(status_code=400, detail="Email already exists")
            client.email = request.email
            password=generate_password()
          
            hashed_password = get_password_hash(password)
        
            client.password=hashed_password
           
            mailsent = send_welcome_email(
                user_name=request.name,
                company_name="Fittbot",
                mobile_number=request.contact,
                default_password=password,
                login_url="",
                recipient_email=request.email,
                support_email="support@fittbot.com"
            )
    
            if mailsent:
                print(f"Welcome mail send to {request.email}")

            smssent = send_welcome_sms(
                user_name=request.name,
                phone_number=request.contact,
                default_password=password,
                login_url=""
            )

            if smssent:
                print(f"welcome SMS send successfully to {request.contact}")


 
        client.name=request.name
        


        db.commit()

        gym_id = client.gym_id
        redis_key = f"gym:{gym_id}:pendingClients"
        if await redis.exists(redis_key):
            await redis.delete(redis_key)

        # Clear members count cache (in case contact/email changes affect counts)
        members_key = f"gym:{gym_id}:members"
        if await redis.exists(members_key):
            await redis.delete(members_key)

        # Clear client data cache
        client_data_key = f"gym:{gym_id}:clientdata"
        if await redis.exists(client_data_key):
            await redis.delete(client_data_key)

        return {
            "status": 200,
            "message": "Client Edited successfully."
        }
    except Exception as e:
        db.rollback
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")
    
@router.delete("/gym/delete_pending_client")
async def delete_pending_clients(client_id:int, db: Session = Depends(get_db), redis : Redis =Depends(get_redis)):
 
    try:
        client = db.query(Client).filter(Client.client_id == client_id).first()
 
        if not client:
            raise HTTPException(status_code=400, detail="Clients not found")
       
        verification = json.loads(client.verification)
        if verification['mobile']:
            raise HTTPException(status_code=400, detail="Client already Verified")
 
        gym_id = client.gym_id
        db.delete(client)
        db.commit()

        redis_key = f"gym:{gym_id}:pendingClients"
        if await redis.exists(redis_key):
            await redis.delete(redis_key)

        # Clear members count cache (total_members count is affected)
        members_key = f"gym:{gym_id}:members"
        if await redis.exists(members_key):
            await redis.delete(members_key)

        # Clear client data cache
        client_data_key = f"gym:{gym_id}:clientdata"
        if await redis.exists(client_data_key):
            await redis.delete(client_data_key)

        return {
            "status": 200,
            "message": "Client Deleted successfully."
        }
    except Exception as e:
        db.rollback
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")
    

@router.get("/gym/get_rewards")
async def get_rewards(gym_id:int, db: Session = Depends(get_db), redis : Redis =Depends(get_redis)):
 
    redis_key = f"gym:{gym_id}:gymRewards"
    try:
        cached_data = await redis.get(redis_key)
        if cached_data:
            print(f"Data fetched from Redis for gym_id: {gym_id}")
            return {
                "status": 200,
                "message": "Rewards data retrived successfully from cache.",
                "data": json.loads(cached_data)
            }
 
        print("Data not found in Redis, fetching from database.")
        rewards = db.query(RewardGym).filter(RewardGym.gym_id == gym_id).order_by(RewardGym.xp.asc())
 
        if not rewards:
            return {
                "status": 200,
                "message": "No rewards found for this gym",
                "data":[]
            } 
       
        reward_list=[]

        reward_list=[
            {
                "id": reward.id,
                "gym_id": reward.gym_id,
                "xp": reward.xp,
                "gift": reward.gift,
            }
            for reward in rewards
        ]
               
        await redis.set(redis_key, json.dumps(reward_list), ex=86400)
 
        return {
            "status": 200,
            "message": "rewards Data fetched successfully",
            "data":reward_list
        }
    except Exception as e:
        db.rollback
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")
    

class RewardData(BaseModel):
    gym_id:int
    reward:dict


def _pick_next_reward(
    ladder: List[RewardGym],
    current_xp: int,
):
    for tier in ladder:
        if tier.xp > current_xp:
            return tier
    return None

@router.post("/gym/create_rewards")
async def create_rewards(request:RewardData, db: Session = Depends(get_db), redis : Redis =Depends(get_redis)):
   
    redis_key = f"gym:{request.gym_id}:gymRewards"
    try:
        gym_id=request.gym_id
        reward=request.reward

        gym=db.query(Gym).filter(Gym.gym_id == gym_id).first()

        if not gym:
            raise HTTPException(status_code=400, details=f'There is no gym found for this gym Id')
        
        new_reward = RewardGym(
            gym_id=gym_id,
            xp = reward["xp"],
            gift = reward["gift"]
        )
        db.add(new_reward)
        db.commit()

        ladder: List[RewardGym] = (
                db.query(RewardGym)
                .filter(RewardGym.gym_id == request.gym_id)
                .order_by(RewardGym.xp.asc())
                .all()
            )

        active_client_ids = [
            cid for (cid,) in
            db.query(Client.client_id)
              .filter(
                  Client.gym_id == request.gym_id,
                  Client.status == "active",
              )
              .all()
        ]

        xp_map = {
            row.client_id: row.xp
            for row in db.query(
                    LeaderboardOverall.client_id,
                    LeaderboardOverall.xp,
                 )
                 .filter(LeaderboardOverall.client_id.in_(active_client_ids))
                 .all()
        }


        for cid in active_client_ids:
            cur_xp = xp_map.get(cid, 0)
            tier = _pick_next_reward(ladder, cur_xp)
            new_next_xp = tier.xp if tier else 0
            new_gift = tier.gift if tier else None

            db.query(ClientNextXp).filter(ClientNextXp.client_id == cid).delete(synchronize_session=False)

            db.add(
                ClientNextXp(
                    client_id=cid,
                    next_xp=new_next_xp,
                    gift=new_gift,
                )
            )



        if await redis.exists(redis_key):
            await redis.delete(redis_key)

        return {
            "status": 200,
            "message": "Reward Created successfully",
        }
    except Exception as e:
        db.rollback
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")



    

class updateRewardData(BaseModel):
    gym_id :int
    record_id : int
    updated_reward : dict

@router.put("/gym/update_rewards")
async def update_rewards(request:updateRewardData, db: Session = Depends(get_db), redis : Redis =Depends(get_redis)):
   
    redis_key = f"gym:{request.gym_id}:gymRewards"
    try:
        gym_id=request.gym_id
        reward_id=request.record_id
        reward=request.updated_reward

        gym=db.query(Gym).filter(Gym.gym_id == gym_id).first()

        if not gym:
            raise HTTPException(status_code=404, details=f'No gym found for this gym Id')
        
        rewards=db.query(RewardGym).filter(RewardGym.id == reward_id).first()

        if not rewards:
            raise HTTPException(status_code=404, details=f'No reward found for this Id')

        rewards.xp = reward["xp"]
        rewards.gift = reward["gift"]

        db.commit()

        ladder: List[RewardGym] = (
            db.query(RewardGym)
            .filter(RewardGym.gym_id == request.gym_id)
            .order_by(RewardGym.xp.asc())
            .all()
        )

        active_client_ids = [
            cid for (cid,) in
            db.query(Client.client_id)
              .filter(
                  Client.gym_id == request.gym_id,
                  Client.status == "active",
              )
              .all()
        ]

        xp_map = {
            row.client_id: row.xp
            for row in db.query(
                    LeaderboardOverall.client_id,
                    LeaderboardOverall.xp,
                 )
                 .filter(LeaderboardOverall.client_id.in_(active_client_ids))
                 .all()
        }


        for cid in active_client_ids:
            cur_xp = xp_map.get(cid, 0)
            tier = _pick_next_reward(ladder, cur_xp)
            new_next_xp = tier.xp if tier else 0
            new_gift = tier.gift if tier else None

            db.query(ClientNextXp).filter(ClientNextXp.client_id == cid).delete(synchronize_session=False)

            db.add(
                ClientNextXp(
                    client_id=cid,
                    next_xp=new_next_xp,
                    gift=new_gift,
                )
            )

       
        db.commit()


        if await redis.exists(redis_key):
            await redis.delete(redis_key)

        return {
            "status": 200,
            "message": "Reward Updated successfully",
        }
    except Exception as e:
        db.rollback
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")
    


@router.delete("/gym/delete_rewards")
async def delete_rewards(reward_id:int, gym_id:int, db: Session = Depends(get_db), redis : Redis =Depends(get_redis)):
 
    try:
        reward = db.query(RewardGym).filter(RewardGym.id == reward_id).first()
 
        if not reward:
            raise HTTPException(status_code=400, detail="Reward not found")
       
        db.delete(reward)
        db.commit()

        ladder: List[RewardGym] = (
            db.query(RewardGym)
            .filter(RewardGym.gym_id == gym_id)
            .order_by(RewardGym.xp.asc())
            .all()
        )

        active_client_ids = [
            cid for (cid,) in
            db.query(Client.client_id)
              .filter(
                  Client.gym_id == gym_id,
                  Client.status == "active",
              )
              .all()
        ]

        xp_map = {
            row.client_id: row.xp
            for row in db.query(
                    LeaderboardOverall.client_id,
                    LeaderboardOverall.xp,
                 )
                 .filter(LeaderboardOverall.client_id.in_(active_client_ids))
                 .all()
        }


        for cid in active_client_ids:
            cur_xp = xp_map.get(cid, 0)
            tier = _pick_next_reward(ladder, cur_xp)
            new_next_xp = tier.xp if tier else 0
            new_gift = tier.gift if tier else None

            db.query(ClientNextXp).filter(ClientNextXp.client_id == cid).delete(synchronize_session=False)

            db.add(
                ClientNextXp(
                    client_id=cid,
                    next_xp=new_next_xp,
                    gift=new_gift,
                )
            )

        db.commit()
       
        redis_key = f"gym:{reward.gym_id}:gymRewards"
        if await redis.exists(redis_key):
            await redis.delete(redis_key)
 
        return {
            "status": 200,
            "message": "Reward Deleted successfully."
        }
    except Exception as e:
        db.rollback
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")
    

@router.get("/gym/get_prize_list")
async def get_rewards(gym_id:int, db: Session = Depends(get_db), redis : Redis =Depends(get_redis)):
    try:

        redis_key_active = f"gym{gym_id}:activeprizes"
        redis_key_history = f"gym{gym_id}:prizehistory"

        active_prize_list = json.loads(await redis.get(redis_key_active))
        prize_history_list = json.loads( await redis.get(redis_key_history))

        if active_prize_list and prize_history_list:
            return{
                "status" :200,
                "message":"Prize list Fetched successfully",
                "data":{
                    "active_prizes":active_prize_list,
                    "prize_history":prize_history_list
                }
            }

        active_prize_list=[]
        prize_history_list=[]
        
        gym = db.query(Gym).filter(Gym.gym_id == gym_id).first()

        if not gym:
            raise HTTPException(status_code=400, details="No gym found for this gym Id")
        
        leaderboard = db.query(LeaderboardOverall).filter(LeaderboardOverall.gym_id == gym_id).all()

        if not leaderboard:
            raise HTTPException( status_code=400, detail="No leaderboard data found for this gym")

        gym_rewards = db.query(RewardGym).filter(RewardGym.gym_id == gym_id).order_by(RewardGym.xp.desc())

        if not gym_rewards:
            raise HTTPException(status_code= 400, detail= "No rewards found for this Gym")
        
        active_prizes=db.query(RewardPrizeHistory).filter(RewardPrizeHistory.gym_id == gym_id , RewardPrizeHistory.is_given == False).all()
        prize_history=db.query(RewardPrizeHistory).filter(RewardPrizeHistory.gym_id == gym_id , RewardPrizeHistory.is_given == True).all()

        if not active_prizes:
            active_prize_list =[]
        else:        
            for prizes in active_prizes:
                prize={
                    'id':prizes.id,
                    'client_id':prizes.client_id,
                    'gym_id':prizes.gym_id,
                    'client_name':prizes.client_name,
                    'achieved_date':prizes.achieved_date,
                    'gift':prizes.gift,
                    'given_date':prizes.given_date,
                    'is_given':prizes.is_given,
                    'xp':prizes.xp
                }
                active_prize_list.append(prize)

        if not prize_history:
            prize_history_list=[]

        else:
            for prizes in prize_history:
                prize={
                    'id':prizes.id,
                    'client_id':prizes.client_id,
                    'gym_id':prizes.gym_id,
                    'client_name':prizes.client_name,
                    'achieved_date':prizes.achieved_date,
                    'gift':prizes.gift,
                    'given_date':prizes.given_date,
                    'is_given':prizes.is_given,
                    'xp':prizes.xp
                }
                prize_history_list.append(prize)

        await redis.set(redis_key_active,json.dumps(active_prize_list), ex=86400)
        await redis.set(redis_key_history,json.dumps(prize_history_list), ex=86400)

        return{
            "status" :200,
            "message":"Prize list Fetched successfully",
            "data":{
                "active_prizes":active_prize_list,
                "prize_history":prize_history_list
            }
        }

    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")
    

@router.get("/gym/get_location")
async def get_rewards(gym_id:int, db: Session = Depends(get_db), redis : Redis =Depends(get_redis)):
    try:

        gym_location = db.query(GymLocation).filter(GymLocation.gym_id == gym_id).first()

        location=False
        if gym_location:
            latitude = float(gym_location.latitude)
            longitude = float(gym_location.longitude)
            location=True
        else:
            latitude = None
            longitude = None
            location=False
        
        gym_location={
            "latitude": latitude,
            "longitude": longitude
        }

  
        
        return {
            "status": 200,
            "data": location,
            "gym_location":gym_location
        }
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")
    
class gymLocationRequest(BaseModel):
    gym_id:int
    latitude:float
    longitude:float

@router.post("/gym/add_location")
async def add_gym_location(request:gymLocationRequest, db:Session = Depends(get_db), redis:Redis=Depends(get_redis)):
    try:
        redis_key = f"gym{request.gym_id}:gymLocation"
        if await redis.exists(redis_key):
            await redis.delete(redis_key)
        gym_location = db.query(GymLocation).filter(GymLocation.gym_id == request.gym_id).first()

        if gym_location:
            raise HTTPException(status_code=400, detail="Loaction already added for this gym")
        
        new_gym_location=GymLocation(
            gym_id = request.gym_id,
            latitude = request.latitude,
            longitude = request.longitude
        )

        db.add(new_gym_location)
        db.commit()
        return {
            "status":200,
            "message":"Gym location added successfully"
        }
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")



@router.put("/gym/edit_location")
async def edit_gym_location(request:gymLocationRequest, db:Session =Depends(get_db), redis:Redis=Depends(get_redis)):
    try:
        print(request)
        redis_key = f"gym{request.gym_id}:gymLocation"
        if await redis.exists(redis_key):
            await redis.delete(redis_key)
        gym_location = db.query(GymLocation).filter(GymLocation.gym_id == request.gym_id).first()

        if not gym_location:
            new_gym= GymLocation(
                gym_id=request.gym_id,
                latitude=request.latitude,
                longitude=request.longitude
            )
            db.add(new_gym)
            db.commit()

            #raise HTTPException(status_code=400, detail="There is no location data available for this gym")

        gym_location.latitude = request.latitude
        gym_location.longitude = request.longitude
        
        db.commit()
        return {
            "status":200,
            "message":"Gym location Updated successfully"
        }
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")
    


@router.get("/get_blocked_users")
async def get_blocked_users(client_id:int, role:str, db:Session=Depends(get_db)):
    try:

        records = db.query(BlockedUsers).filter(BlockedUsers.user_id == client_id ).first()

        blocked_list=[]
        if not records:
            return { "status":200, "message":"The user has not blocked any user", "data":blocked_list}
        
        blocked_users=json.loads(records.blocked_user_id) if isinstance(records.blocked_user_id, str) else records.blocked_user_id

        if role == "client":
            if blocked_users['client']:
                for client_id in blocked_users['client']:
                    client_record = db.query(Client).filter(Client.client_id == client_id).first()
                    client_data={
                        "name":client_record.name,
                        "client_id":client_record.client_id,
                        "gym_id":client_record.gym_id,
                        "role":"client",
                    }
                    blocked_list.append(client_data)
            if blocked_users['owner']:
                for gym in blocked_users['owner']:
                    owner_record = db.query(Gym).filter(Gym.gym_id == gym).first()
                    owner_data={
                        'name':owner_record.name,
                        'client_id':None,
                        "gym_id":owner_record.gym_id,
                        "role":"owner"
                    }
                    blocked_list.append(owner_data)
        
        if role =="owner":
            if blocked_users['client']:
                for client in blocked_users['client']:
                    client_record = db.query(Client).filter(Client.client_id == client).first()
                    client_data={
                        "name":client_record.name,
                        "client_id":client_record.client_id,
                        "gym_id":client_record.gym_id,
                        "role":"client",
                    }
                    blocked_list.append(client_data)

        
        return {
            "status":200,
            "message":"Data retrived Successfully",
            "data":blocked_list
        }


    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpectedd error occured, {str(e)}")
    
   

class UnblockUserReport(BaseModel):
    user_id: int
    gym_id: int
    blocked_user_id: int
    user_role: str
    blocked_user_role: str

@router.post("/unblock_users")
async def unblock_users(request:UnblockUserReport, db:Session = Depends(get_db), redis : Redis=Depends(get_redis)):
    try:
        redis_key = f'{request.user_role}{request.user_id}:blockedusers'
        if await redis.exists(redis_key):
            await redis.delete(redis_key)

        records = db.query(BlockedUsers).filter(BlockedUsers.user_id == request.user_id ).first()

        if not records:
            raise HTTPException(status_code=400, detail="The user has not blocked any user")
        
        blocked_users_record=json.loads(records.blocked_user_id) if isinstance(records.blocked_user_id, str) else records.blocked_user_id
        
        if request.blocked_user_role == "client":
           blocked_users_record['client'].remove(request.blocked_user_id)

        else :
            blocked_users_record['owner'].remove(request.blocked_user_id)

        records.blocked_user_id = json.dumps(blocked_users_record)

        db.commit()
        db.refresh(records)

        return{
            "status":200,
            "message":"User unblocked successfully"
        }

    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected error occured, {str(e)}")
    

@router.get('/check-plan-assignments')
async def check_plan_assignments(plan_id:int, db:Session = Depends(get_db), redis:Redis=Depends(get_redis)):
    try:

        assigned_clients=db.query(Client).filter(Client.training_id == plan_id).all()
        assigned_clients_list=[]

        if not assigned_clients:
            return {"status":200, "message":"No clients assigned to this plan", "data":assigned_clients_list}
        
        for client in assigned_clients:
            client_data={
                "name":client.name,
                "client_id":client.client_id
            }
            assigned_clients_list.append(client_data)

        return{
            "status":200,
            "message":"Data retrived successfully",
            "data":assigned_clients_list
        }
        
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected error occured, {str(e)}")
    

@router.get('/check-batch-assignments')
async def check_batch_assignments(batch_id:int, db:Session = Depends(get_db), redis: Redis=Depends(get_redis)):
    try:
        
        assigned_clients=db.query(Client).filter(Client.batch_id == batch_id).all()
        assigned_clients_list=[]

        if not assigned_clients:
            return {"status":200, "message":"No clients assigned to this plan", "data":assigned_clients_list}
        
        for client in assigned_clients:
            client_data={
                "name":client.name,
                "client_id":client.client_id
            }
            assigned_clients_list.append(client_data)


        return{
            "status":200,
            "message":"Data retrived successfully",
            "data":assigned_clients_list
        }
        
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected error occured, {str(e)}")
    
@router.get('/gym/clients-with-plans')
async def clients_with_Plans_batchs(gym_id:int, db: Session = Depends(get_db), redis:Redis = Depends(get_redis)):
    try:

        clients = db.query(Client).filter(Client.gym_id == gym_id).all()
        client_list=[]

        if not clients:
            return {"status":200, "message":"No clients available for this Gym", "data":client_list}
        
        for client in clients:
            client_data={
                "id":client.client_id,
                "plan_id":client.training_id,
                "batch_id":client.batch_id,
                "name":client.name,
                "email":client.email,
                "phone":client.contact
            }

            client_list.append(client_data)

        return{
            "status":200, "message":"Data Retrived successfully", "data":client_list
        }
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected error occured, f{str(e)}")
    
class updatePlanRequest(BaseModel):
    user_id: int
    plan_id:int

@router.put('/update-client-plan')
async def update_clients_plan(request:updatePlanRequest, db: Session = Depends(get_db), redis:Redis=Depends(get_redis)):
    try:
        client = db.query(Client).filter(Client.client_id == request.user_id).first()

        if not client:
            raise HTTPException(status_code=400, detail="User not found")
        
        client.training_id = request.plan_id
        db.commit()
        db.refresh(client)

        return{
            "status":200,
            "message":"Plan reassigned Successfully"
        }
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected error occured, {str(e)}")
    
class updateBatchRequest(BaseModel):
    user_id: int
    batch_id:int

@router.put('/update-client-batch')
async def update_clients_batch(request:updateBatchRequest, db: Session = Depends(get_db), redis:Redis=Depends(get_redis)):
    try:

        client = db.query(Client).filter(Client.client_id == request.user_id).first()

        if not client:
            raise HTTPException(status_code=400, detail="User not found")
        
        client.batch_id = request.batch_id
        db.commit()
        db.refresh(client)

        return{
            "status":200,
            "message":"Batch reassigned Successfully"
        }
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected error occured, {str(e)}")      
    
class updateGivenPrizeRequest(BaseModel):
    id:int
    given_date:datetime
    is_given:bool

@router.put('/update-given-prize')
async def update_given_prize(request:updateGivenPrizeRequest, db:Session=Depends(get_db), redis:Redis=Depends(get_redis)):
    try:
        prizeHistory= db.query(RewardPrizeHistory).filter(RewardPrizeHistory.id == request.id).first()
        redis_key_active = f"gym{prizeHistory.gym_id}:activeprizes"
        redis_key_history = f"gym{prizeHistory.gym_id}:prizehistory"

        if await redis.exists(redis_key_active) or await redis.exists(redis_key_history):
            await redis.delete(redis_key_history)
            await redis.delete(redis_key_active)

        if not prizeHistory:
            raise HTTPException(status_code=400, detail="Prize History not found")
        
        prizeHistory.given_date = datetime.now()
        prizeHistory.is_given = request.is_given

        db.commit()

        return{
            "status":200,
            "message":"Prize status updated successfully"
        }
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected error occoured: {str(e)}")
    
class AddExpenditureRequest(BaseModel):
    gym_id:int
    date:str
    type:str
    amount:float   

@router.post("/add-expenditure")
async def add_expenditure(request:AddExpenditureRequest, db:Session=Depends(get_db), redis : Redis =Depends(get_redis)):
    try:
        expenditure = Expenditure(
            gym_id=request.gym_id,
            expenditure_type=request.type,
            amount=request.amount,
            date = request.date
        )

        db.add(expenditure)
        db.commit()
        db.refresh(expenditure)

        current_month = datetime.now().strftime("%Y-%m")
        existing_record = db.query(GymMonthlyData).filter(
            GymMonthlyData.gym_id == request.gym_id,
            GymMonthlyData.month_year.like(f"{current_month}%")  
        ).first()

        if existing_record:
            existing_record.expenditure = existing_record.expenditure + request.amount
        else:
            new_record = GymMonthlyData(
                gym_id=request.gym_id,
                month_year=datetime.now().strftime("%Y-%m-%d"),
                income=0,
                expenditure=request.amount,
                new_entrants=0
            )
            db.add(new_record)
        
        db.commit()

        gym_analysis = db.query(GymAnalysis).filter(GymAnalysis.gym_id == request.gym_id).first()

        data=json.loads(gym_analysis.analysis)

        if "expenditure_data" not in data:
            data["expenditure_data"] = {}

        if request.type in data["expenditure_data"]:
            data["expenditure_data"][request.type] += request.amount
        else:
            data["expenditure_data"][request.type] = request.amount

        gym_analysis.analysis = json.dumps(data)

        db.commit()
        db.refresh(gym_analysis)

        collection_key = f"gym:{request.gym_id}:collection"
        if await redis.exists(collection_key):
            await redis.delete(collection_key)

        redis_key_monthly = f"gym:{request.gym_id}:monthly_data"
        if await redis.exists(redis_key_monthly):
            await redis.delete(redis_key_monthly)

        redis_key_analysis = f"gym:{request.gym_id}:analysis"
        if await redis.exists(redis_key_analysis):
            await redis.delete(redis_key_analysis)

        redis_key_hourly = f"gym:{request.gym_id}:hourlyagg"
        if await redis.exists(redis_key_hourly):
            await redis.delete(redis_key_hourly)

        return {
            "status":200,
            "message":"Expenditure added successfully"
        }
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected Error occured, {str(e)}")

class UpdateExpenditureRequest(BaseModel):
    expense_id :int
    gym_id:int
    date:str
    type:str
    amount:float   

@router.post("/update-expenditure")
async def update_expenditure(request:UpdateExpenditureRequest, db:Session=Depends(get_db), redis : Redis =Depends(get_redis)):
    try:
        expenditure = db.query(Expenditure).filter(Expenditure.expenditure_id == request.expense_id).first()
        older_amount = expenditure.amount
        expenditure.expenditure_type=request.type
        expenditure.amount=request.amount
        expenditure.date = request.date

        db.commit()
        db.refresh(expenditure)

        current_month = datetime.strptime(request.date, "%Y-%m-%d").strftime("%Y-%m")
        existing_record = db.query(GymMonthlyData).filter(
            GymMonthlyData.gym_id == request.gym_id,
            GymMonthlyData.month_year.like(f"{current_month}%")  
        ).first()

        if existing_record:
            existing_record.expenditure -= older_amount
            existing_record.expenditure += request.amount
        else:
            new_record = GymMonthlyData(
                gym_id=request.gym_id,
                month_year=datetime.strftime(request.date, "%Y-%m-%d"),
                income=0,
                expenditure=request.amount,
                new_entrants=0
            )
            db.add(new_record)
        
        db.commit()

        gym_analysis = db.query(GymAnalysis).filter(GymAnalysis.gym_id == request.gym_id).first()

        data=json.loads(gym_analysis.analysis)

        if "expenditure_data" not in data:
            data["expenditure_data"] = {}

        if request.type in data["expenditure_data"]:
            data["expenditure_data"][request.type] -= older_amount
            data["expenditure_data"][request.type] += request.amount
        else:
            data["expenditure_data"][request.type] = request.amount

        gym_analysis.analysis = json.dumps(data)

        db.commit()
        db.refresh(gym_analysis)

        collection_key = f"gym:{request.gym_id}:collection"
        if await redis.exists(collection_key):
            await redis.delete(collection_key)

        redis_key_monthly = f"gym:{request.gym_id}:monthly_data"
        if await redis.exists(redis_key_monthly):
            await redis.delete(redis_key_monthly)

        redis_key_analysis = f"gym:{request.gym_id}:analysis"
        if await redis.exists(redis_key_analysis):
            await redis.delete(redis_key_analysis)

        redis_key_hourly = f"gym:{request.gym_id}:hourlyagg"
        if await redis.exists(redis_key_hourly):
            await redis.delete(redis_key_hourly)

        return {
            "status":200,
            "message":"Expenditure updated successfully"
        }
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected Error occured, {str(e)}")

@router.delete("/delete-expenditure")
async def delete_expenditure(gym_id:int, expense_id:int, db: Session=Depends(get_db), redis : Redis=Depends(get_redis)):
    try:
        existing_expense = db.query(Expenditure).filter(Expenditure.expenditure_id == expense_id).one()

        if not existing_expense:
            raise HTTPException(status_code=400, detail="Unable to find the selected expense")
        
        existing_record = db.query(GymMonthlyData).filter(
            GymMonthlyData.gym_id == gym_id,
            GymMonthlyData.month_year.like(f"{existing_expense.date}%")  
        ).first()

        if existing_record:
            existing_record.expenditure -= existing_expense.amount

        gym_analysis = db.query(GymAnalysis).filter(GymAnalysis.gym_id == gym_id).first()

        data=json.loads(gym_analysis.analysis)

        if existing_expense.expenditure_type in data["expenditure_data"]:
            data["expenditure_data"][existing_expense.expenditure_type] -= existing_expense.amount

        gym_analysis.analysis = json.dumps(data)

        db.refresh(gym_analysis)
        
        db.delete(existing_expense)
        db.commit()
        collection_key = f"gym:{gym_id}:collection"
        if await redis.exists(collection_key):
            await redis.delete(collection_key)

        redis_key_monthly = f"gym:{gym_id}:monthly_data"
        if await redis.exists(redis_key_monthly):
            await redis.delete(redis_key_monthly)

        redis_key_analysis = f"gym:{gym_id}:analysis"
        if await redis.exists(redis_key_analysis):
            await redis.delete(redis_key_analysis)

        redis_key_hourly = f"gym:{gym_id}:hourlyagg"
        if await redis.exists(redis_key_hourly):
            await redis.delete(redis_key_hourly)

        return{
            "status":200,
            "message":"Expenditure deleted Successfully"
        }
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An unexpected error occured: {str(e)}")

@router.get("/get-client-data-qr")
async def get_client_data_qr(uuid:str, db:Session=Depends(get_db)):
    try:

        # decrypted_uuid = decrypt_uuid(uuid)
        decrypted_uuid = uuid

        client = db.query(Client).filter(Client.uuid_client == decrypted_uuid).one()

        if not client:
            raise HTTPException(status_code=404, detail="User not found")
        
        response={
            "goals":client.goals,
            "bmi":client.bmi,
            "height":client.height,
            "age":client.age,
            "lifestyle":client.lifestyle,
            "weight":client.weight,
            "profile":client.profile,
            "full_name":client.name,
            "client_id":client.client_id,
            "contact":client.contact,
            "email":client.email,
            "medical_issues":client.medical_issues,
            "gender":client.gender,
            "dob":client.dob,
            "uuid":client.uuid_client,
        }

        return{
            "status":200,
            "message":"Data retrived successfully",
            "data":response
        }
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occured:{str(e)}")

@router.get("/gym-enquiries")
async def get_enquiries(gym_id:int, db:Session=Depends(get_db), redis:Redis = Depends(get_redis)):

    try:

        redis_key=f"gym{gym_id}:enquiries"
        redis_incompleted =f"gym:{gym_id}:incompletedEnquiries"
        redis_completed =f"gym:{gym_id}:completedEnquiries"
        cached_data=await redis.get(redis_key)
        cached_data2 =await redis.get(redis_completed)
        cached_data3 = await redis.get(redis_incompleted)
        if cached_data and cached_data2 and cached_data3:
            return{
            "status":200,
            "message":"Data retrived successfully",
            "data":{
                "enquiry_data":json.loads(cached_data),
                "incomplete_enquiries":json.loads(cached_data3),
                "completed_enquiries":json.loads(cached_data2)
            }
        }
        enquiry_data={}
        incomplete_enquiry=[]
        completed_enquiry=[]
        incomplete_data={}
        completed_data={}
        pending_enquiries = db.query(GymEnquiry).filter(GymEnquiry.gym_id == gym_id, GymEnquiry.status == "Pending").order_by(GymEnquiry.updated_at.desc()).all()
        if pending_enquiries:
            enquiry_data["pending"] = [{
                "enquiry_id":enquiry.enquiry_id,
                "name":enquiry.name,
                "contact":enquiry.contact,
                "convenientTime":enquiry.convenientTime,
                "email":enquiry.email,
                "status":enquiry.status,
                "statusReason":enquiry.statusReason,
                "date":str(enquiry.created_at),
                "message":enquiry.message
            }for enquiry in pending_enquiries]
        else:
            enquiry_data["pending"]=[]

        followup_enquiries=db.query(GymEnquiry).filter(GymEnquiry.gym_id == gym_id, GymEnquiry.status == "Follow Up").order_by(GymEnquiry.updated_at.desc()).all()
        if followup_enquiries:
            enquiry_data["followUp"] = [{
                "enquiry_id":enquiry.enquiry_id,
                "name":enquiry.name,
                "contact":enquiry.contact,
                "convenientTime":enquiry.convenientTime,
                "email":enquiry.email,
                "status":enquiry.status,
                "statusReason":enquiry.statusReason,
                "date":str(enquiry.created_at),
                "message":enquiry.message
            }for enquiry in followup_enquiries]
        else:
            enquiry_data["followUp"]=[]

        rejected_enquiries = db.query(GymEnquiry).filter(GymEnquiry.gym_id == gym_id, GymEnquiry.status == "Rejected").order_by(GymEnquiry.updated_at.desc()).all()
        if rejected_enquiries:
            enquiry_data["rejected"] = [{
                "enquiry_id":enquiry.enquiry_id,
                "name":enquiry.name,
                "contact":enquiry.contact,
                "convenientTime":enquiry.convenientTime,
                "email":enquiry.email,
                "status":enquiry.status,
                "statusReason":enquiry.statusReason,
                "date":str(enquiry.created_at),
                "message":enquiry.message
            }for enquiry in rejected_enquiries]
        else:
            enquiry_data["rejected"]=[]

        joined_enquiries = db.query(GymEnquiry).filter(GymEnquiry.gym_id == gym_id, GymEnquiry.status == "Joined").order_by(GymEnquiry.updated_at.desc()).all()
        if joined_enquiries:
            enquiry_data["joined"] = [{
                "enquiry_id":enquiry.enquiry_id,
                "name":enquiry.name,
                "contact":enquiry.contact,
                "convenientTime":enquiry.convenientTime,
                "email":enquiry.email,
                "status":enquiry.status,
                "statusReason":enquiry.statusReason,
                "date":str(enquiry.created_at),
                "message":enquiry.message
            }for enquiry in joined_enquiries]
        else:
            enquiry_data["joined"]=[]

        incomplete_enquiries=db.query(GymEnquiry).filter(GymEnquiry.gym_id == gym_id).filter((GymEnquiry.status == "Pending") | (GymEnquiry.status == "Follow Up")).order_by(GymEnquiry.updated_at.desc()).all()
        if incomplete_enquiries:
           for enquiry in incomplete_enquiries:
                incomplete_data={
                    "enquiry_id":enquiry.enquiry_id,
                    "name":enquiry.name,
                    "contact":enquiry.contact,
                    "convenientTime":enquiry.convenientTime,
                    "email":enquiry.email,
                    "status":enquiry.status,
                    "statusReason":enquiry.statusReason,
                    "date":str(enquiry.created_at),
                    "message":enquiry.message
                }
                incomplete_enquiry.append(incomplete_data)
        
        completed_enquiries=db.query(GymEnquiry).filter(GymEnquiry.gym_id == gym_id).filter((GymEnquiry.status == "Rejected") | (GymEnquiry.status == "Joined")).order_by(GymEnquiry.updated_at.desc()).all()
        if completed_enquiries:
           for enquiry in completed_enquiries:
                completed_data={
                    "enquiry_id":enquiry.enquiry_id,
                    "name":enquiry.name,
                    "contact":enquiry.contact,
                    "convenientTime":enquiry.convenientTime,
                    "email":enquiry.email,
                    "status":enquiry.status,
                    "statusReason":enquiry.statusReason,
                    "date":str(enquiry.created_at),
                    "message":enquiry.message
                }
                completed_enquiry.append(completed_data)

        await redis.set(redis_key,json.dumps(enquiry_data), ex=86400)
        await redis.set(redis_completed,json.dumps(completed_enquiry), ex=86400)
        await redis.set(redis_incompleted,json.dumps(incomplete_enquiry), ex=86400)

        return{
            "status":200,
            "message":"Data retrived successfully",
            "data":{
                "enquiry_data":enquiry_data,
                "incomplete_enquiries":incomplete_enquiry,
                "completed_enquiries":completed_enquiry
            }
        }
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occured:{str(e)}")

class postEnquiryRequest(BaseModel):
    gym_id:int
    data:dict

@router.post("/gym-enquiries")
async def save_enquiries(request:postEnquiryRequest, db:Session=Depends(get_db), redis:Redis=Depends(get_redis)):
    try:
        existing_enquiry= db.query(GymEnquiry).filter(GymEnquiry.contact == request.data["contact"]).first()
        redis_key=f"gym{request.gym_id}:enquiries"
        redis_incompleted =f"gym:{request.gym_id}:incompletedEnquiries"
        redis_completed =f"gym:{request.gym_id}:completedEnquiries"

        
        new_enquiry=GymEnquiry(
            gym_id = request.gym_id,
            name = request.data["name"],
            contact = request.data["contact"],
            convenientTime= request.data["convenientTime"],
            email = request.data["email"],
            message = request.data["message"],
            status = request.data["status"],
            statusReason = request.data["statusReason"] if request.data["statusReason"] else None
        )

        db.add(new_enquiry)
        db.commit()
        db.refresh(new_enquiry)

        if "plan_id" in request.data and request.data.get("plan_id"):
            try:
                gym = db.query(Gym).filter(Gym.gym_id == request.gym_id).first()
                gym_owner = db.query(GymOwner).filter(GymOwner.owner_id == gym.owner_id).first()
                plan = db.query(GymPlans).filter(GymPlans.id == request.data["plan_id"]).first()
                account = db.query(AccountDetails).filter(AccountDetails.gym_id == request.gym_id).first()
                
                estimate_count = db.query(EnquiryEstimates).count()
                estimate_number = f"{gym.location[:3].upper()}-EST-{request.gym_id}-{estimate_count + 1}"
                
                fees = request.data.get("fees", plan.amount if plan else 0)
                admission_fees = request.data.get("admission_fees", 0)
                discount = request.data.get("discount", 0)
                discounted_fees = fees - discount
                gst_percentage = request.data.get("gst_percentage", 18)
                gst_type = request.data.get("gst_type", "inclusive")
                
                subtotal = discounted_fees + admission_fees
                
                if gst_type == "exclusive":
                    total_amount = subtotal + ((subtotal * gst_percentage) / 100)
                else:
                    total_amount = subtotal

                new_estimate = EnquiryEstimates(
                    enquiry_id=new_enquiry.enquiry_id,
                    gym_id=request.gym_id,
                    client_name=request.data["name"],
                    gym_name=gym.name if gym else "",
                    gym_logo=gym.logo if gym else "",
                    gym_contact=gym_owner.contact_number if gym_owner else "",
                    gym_location=gym.location if gym else "",
                    plan_id=request.data["plan_id"],
                    plan_description=plan.plans if plan else "",
                    fees=fees,
                    admission_fees=admission_fees,
                    fees_type=request.data.get("fees_type", "joining"),
                    discount=discount,
                    discounted_fees=discounted_fees,
                    estimate_number=estimate_number,
                    client_contact=request.data["contact"],
                    bank_details=account.account_number if account else "",
                    ifsc_code=account.account_ifsccode if account else "",
                    bank_name=account.bank_name if account else "",
                    upi_id=account.upi_id if account else "",
                    account_holder_name=account.account_holdername if account else "",
                    estimate_date=datetime.now().strftime("%Y-%m-%d"),
                    gst_number=account.gst_number if account else "",
                    client_email=request.data.get("email", ""),
                    mail_status=False,
                    created_at=datetime.now(),
                    update_at=datetime.now(),
                    gst_percentage=gst_percentage,
                    gst_type=gst_type,
                    branch=account.account_branch,
                    total_amount=total_amount
                )
                
                db.add(new_estimate)
                db.commit()
                db.refresh(new_estimate)
                
                estimate_data = estimate_to_dict(new_estimate)
                
                estimate_redis_keys = [
                    f"gym{request.gym_id}:enquiryEstimates:all:page*",
                    f"gym{request.gym_id}:enquiryEstimates:*"
                ]
                
                for key_pattern in estimate_redis_keys:
                    cursor = 0
                    while True:
                        cursor, keys = await redis.scan(cursor=cursor, match=key_pattern, count=100)
                        if keys:
                            await redis.delete(*keys)
                        if cursor == 0:
                            break
                            
            except Exception as estimate_error:
                print(f"Error creating estimate: {str(estimate_error)}")
                estimate_data = None
                
        members_key = f"gym:{request.gym_id}:members"

        if await redis.exists(redis_key):
            await redis.delete(redis_key)
        
        if await redis.exists(redis_incompleted):
            await redis.delete(redis_incompleted)

        if await redis.exists(redis_completed):
            await redis.delete(redis_completed)

        if await redis.exists(members_key):
            await redis.delete(members_key)

        message = "Enquiry Data created successfully"
        response_data = {
            "status": 200,
            "message": message,
            "enquiry_id": new_enquiry.enquiry_id
        }
        
        if "plan_id" in request.data and request.data.get("plan_id"):
            message = "Enquiry and Estimate created successfully"
            response_data["message"] = message
            if 'estimate_data' in locals() and estimate_data:
                response_data["estimate"] = estimate_data

        return{              
            "status":200,
            "message": message,
            "data":response_data
        }
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occured:{str(e)}")
    
class UpdateEnquiryRequest(BaseModel):
    enquiry_id:int
    gym_id:int
    status:str
    statusReason:str

@router.put("/update-enquiry-status")
async def update_enquiry_status(request:UpdateEnquiryRequest, db:Session=Depends(get_db), redis:Redis=Depends(get_redis)):
    try:
        redis_key=f"gym{request.gym_id}:enquiries"
        redis_incompleted =f"gym:{request.gym_id}:incompletedEnquiries"
        redis_completed =f"gym:{request.gym_id}:completedEnquiries"
        existing_enquiry = db.query(GymEnquiry).filter(GymEnquiry.enquiry_id == request.enquiry_id).first()

        if not existing_enquiry:
            raise HTTPException(status_code=404, detail="Enquiry data not found")
        
        existing_enquiry.status = request.status
        existing_enquiry.statusReason = request.statusReason

        db.commit()
        if await redis.exists(redis_key):
            await redis.delete(redis_key)

        members_key = f"gym:{request.gym_id}:members"

        if await redis.exists(members_key):
            await redis.delete(members_key)
        
        if await redis.exists(redis_incompleted):
            await redis.delete(redis_incompleted)

        if await redis.exists(redis_completed):
            await redis.delete(redis_completed)

        return{
            "status":200,
            "message":"Enquiry status updated successfully"
        }

    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occured : {str(e)}")

  
@router.post("/send-estimates_v1")
async def generate_and_send_estimates(expiry_ids:List[int], db: Session = Depends(get_db), redis: Redis = Depends(get_redis)):
    try:

        entries = (
            db.query(AboutToExpire)
            .filter(AboutToExpire.expiry_id.in_(expiry_ids))
            .all()
        )

        sent_ids = []  

        for entry in entries:
            event_payload = {
                "invoice_data": {
                    "receipt_id":        entry.expiry_id,
                    "client_id":         entry.client_id,
                    "gym_id":            entry.gym_id,
                    "client_name":       entry.client_name,
                    "gym_name":          entry.gym_name,
                    "gym_logo":          entry.gym_logo,
                    "gym_contact":       entry.gym_contact,
                    "gym_location":      entry.gym_location,
                    "plan_id":           entry.plan_id,
                    "plan_description":  entry.plan_description,
                    "fees":              int(entry.fees),
                    "discount":          int(entry.discount),
                    "discounted_fees":   int(entry.discounted_fees),
                    "due_date":          str(entry.due_date),
                    "invoice_number":    entry.invoice_number,
                    "client_contact":    entry.client_contact,
                    "bank_details":      entry.bank_details,
                    "ifsc_code":         entry.ifsc_code,
                    "account_holder":    entry.account_holder_name,
                    "client_email":      entry.email,
                    "upi_id":            entry.upi_id,
                    "bank_name":         entry.bank_name,
                    "gst_number":        entry.gst_number,
                    "branch":            entry.branch,
                    "gst_type":          "no_gst",
                    "discounted_price":  (entry.fees - entry.discounted_fees),
                    "invoice_type":      "invoice",
                }
            }

            try:
                # Use Lambda retry utility - automatically retries on throttling/errors
                invoke_lambda_with_retry(
                    lambda_client,
                    FunctionName=LAMBDA_FUNCTION_NAME,
                    InvocationType="Event",
                    Payload=json.dumps(event_payload).encode("utf-8"),
                )
                sent_ids.append(entry.expiry_id) 
                print("appended")    
            except Exception as e:
                raise HTTPException(
                    500, f"Could not enqueue receipt {entry.expiry_id} for mailing: {e}"
                )

        if sent_ids:
            print("sent_ids",sent_ids)
            (
                db.query(AboutToExpire)
                .filter(AboutToExpire.expiry_id.in_(sent_ids))
                .update({AboutToExpire.mail_status: True}, synchronize_session=False)
            )
            db.commit()

            # Update to match new date-based Redis key pattern
            about_key_pattern = f"gym:*:about_to_expire:*"
            expired_key_pattern = f"gym:*:expired:*"
            about_key = await redis.keys(about_key_pattern)
            if about_key:
                await redis.delete(*about_key)
                print("key dleted")

            expired_key = await redis.keys(expired_key_pattern)
            if expired_key:
                await redis.delete(*expired_key)
                print("key dleted")

            unpaid_key_pattern = f'*:unpaid_members:*:*'
            unpaid_key = await redis.keys(unpaid_key_pattern)

            if unpaid_key:
                await redis.delete(*unpaid_key)
                print("key dleted")


        return {"status":200, "message": "Process started in background"}

    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occured:{str(e)}")


@router.post("/send-estimates")
async def generate_and_send_estimates_v2(
    membership_ids: List[int],
    gym_id: Optional[int]=1,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis)
):

    try:
        
        today = date.today()
        five_days_later = today + timedelta(days=5)

        # Try to get data from Redis first
        invoice_key = f"gym:{gym_id}:invoice_data:{today.strftime('%Y-%m-%d')}"
        unpaid_key = f"gym:{gym_id}:unpaid_members:{today.strftime('%Y-%m-%d')}"

        cached_invoice = await redis.get(invoice_key)
        cached_unpaid = await redis.get(unpaid_key)

        # Helper function to decode cached data
        def _decode_cached(value: Any) -> str:
            if isinstance(value, (bytes, bytearray)):
                return value.decode()
            return value

        # Collect all memberships data from cache or DB
        all_memberships_data = []

        # Parse cached data if available
        if cached_invoice:
            try:
                invoice_data = json.loads(_decode_cached(cached_invoice))
                all_memberships_data.extend(invoice_data.get("send", []))
                all_memberships_data.extend(invoice_data.get("unsend", []))
            except json.JSONDecodeError:
                pass

        if cached_unpaid:
            try:
                unpaid_data = json.loads(_decode_cached(cached_unpaid))
                all_memberships_data.extend(unpaid_data.get("send", []))
                all_memberships_data.extend(unpaid_data.get("unsend", []))
            except json.JSONDecodeError:
                pass

        # If no cache, fetch from DB
        if not all_memberships_data:
            # Get gym and account details
            gym = db.query(Gym).filter(Gym.gym_id == gym_id).first()
            account = db.query(AccountDetails).filter(AccountDetails.gym_id == gym_id).first()

            # Query all memberships for the given IDs
            memberships = (
                db.query(FittbotGymMembership, Client, ClientGym, GymPlans)
                .select_from(FittbotGymMembership)
                .join(Client, Client.client_id == FittbotGymMembership.client_id)
                .join(ClientGym, ClientGym.client_id == Client.client_id, isouter=True)
                .join(GymPlans, GymPlans.id == FittbotGymMembership.plan_id, isouter=True)
                .filter(
                    FittbotGymMembership.id.in_(membership_ids),
                    FittbotGymMembership.gym_id == str(gym_id)
                )
                .all()
            )

            # Serialize memberships
            for membership, client, gym_client, plan in memberships:
                all_memberships_data.append({
                    "expiry_id": membership.id,
                    "client_id": int(membership.client_id) if membership.client_id else None,
                    "gym_id": int(membership.gym_id) if membership.gym_id else None,
                    "client_name": client.name if client else None,
                    "gym_name": gym.name if gym else None,
                    "gym_logo": gym.logo if gym else None,
                    "gym_contact": gym.contact_number if gym else None,
                    "gym_location": gym.location if gym else None,
                    "plan_id": membership.plan_id,
                    "plan_description": plan.plans if plan else None,
                    "fees": float(membership.amount) if membership.amount else 0.0,
                    "discount": 0,
                    "discounted_fees": float(membership.amount) if membership.amount else 0.0,
                    "due_date": str(membership.expires_at) if membership.expires_at else None,
                    "invoice_number": None,
                    "client_contact": client.contact if client else None,
                    "bank_details": account.account_number if account else None,
                    "ifsc_code": account.account_ifsccode if account else None,
                    "account_holder_name": account.account_holdername if account else None,
                    "paid": False,
                    "mail_send": False,
                    "expired": membership.expires_at < today if membership.expires_at else False,
                    "email": client.email if client else None,
                    "upi_id": account.upi_id if account else None,
                    "bank_name": account.bank_name if account else None,
                    "gst_number": account.gst_number if account else None,
                    "branch": account.account_branch if account else None,
                    "admission_fees": 0,  # Default to 0
                })

        # Filter the memberships we want to send based on membership_ids
        memberships_to_send = [m for m in all_memberships_data if m.get("expiry_id") in membership_ids]

        # Debug: Print all data in structured manner
        print("\n" + "="*80)
        print("DEBUG: SEND ESTIMATES V2 - DATA BEFORE SENDING")
        print("="*80)
        print(f"Gym ID: {gym_id}")
        print(f"Requested Membership IDs: {membership_ids}")
        print(f"Total memberships found in cache/DB: {len(all_memberships_data)}")
        print(f"Memberships to send (filtered): {len(memberships_to_send)}")
        print("-"*80)

        for idx, entry in enumerate(memberships_to_send, 1):
            print(f"\n--- Membership {idx}/{len(memberships_to_send)} ---")
            print(json.dumps({
                "expiry_id": entry.get("expiry_id"),
                "client_id": entry.get("client_id"),
                "client_name": entry.get("client_name"),
                "client_email": entry.get("email"),
                "client_contact": entry.get("client_contact"),
                "gym_name": entry.get("gym_name"),
                "gym_contact": entry.get("gym_contact"),
                "gym_location": entry.get("gym_location"),
                "plan_id": entry.get("plan_id"),
                "plan_description": entry.get("plan_description"),
                "fees": entry.get("fees"),
                "discount": entry.get("discount"),
                "discounted_fees": entry.get("discounted_fees"),
                "admission_fees": entry.get("admission_fees", 0),
                "due_date": entry.get("due_date"),
                "expired": entry.get("expired"),
                "bank_details": entry.get("bank_details"),
                "ifsc_code": entry.get("ifsc_code"),
                "account_holder_name": entry.get("account_holder_name"),
                "upi_id": entry.get("upi_id"),
                "bank_name": entry.get("bank_name"),
                "gst_number": entry.get("gst_number"),
                "branch": entry.get("branch"),
            }, indent=2))

        print("\n" + "="*80)
        print("END DEBUG - STARTING EMAIL SENDING PROCESS")
        print("="*80 + "\n")

        sent_ids = []

        # Send emails via Lambda
        for entry in memberships_to_send:
            event_payload = {
                "invoice_data": {
                    "receipt_id": entry.get("expiry_id"),
                    "client_id": entry.get("client_id"),
                    "gym_id": entry.get("gym_id"),
                    "client_name": entry.get("client_name"),
                    "gym_name": entry.get("gym_name"),
                    "gym_logo": entry.get("gym_logo"),
                    "gym_contact": entry.get("gym_contact"),
                    "gym_location": entry.get("gym_location"),
                    "plan_id": entry.get("plan_id"),
                    "plan_description": entry.get("plan_description"),
                    "fees": int(entry.get("fees", 0)),
                    "discount": int(entry.get("discount", 0)),
                    "discounted_fees": int(entry.get("discounted_fees", 0)),
                    "due_date": entry.get("due_date"),
                    "invoice_number": entry.get("invoice_number"),
                    "client_contact": entry.get("client_contact"),
                    "bank_details": entry.get("bank_details"),
                    "ifsc_code": entry.get("ifsc_code"),
                    "account_holder": entry.get("account_holder_name"),
                    "client_email": entry.get("email"),
                    "upi_id": entry.get("upi_id"),
                    "bank_name": entry.get("bank_name"),
                    "gst_number": entry.get("gst_number"),
                    "branch": entry.get("branch"),
                    "gst_type": "no_gst",
                    "discounted_price": (entry.get("fees", 0) - entry.get("discounted_fees", 0)),
                    "admission_fees": 0,  # Add admission_fees field for Lambda compatibility
                    "invoice_type": "enquiry_estimate",  # Changed from "enquiry_estimate" to "invoice"
                }
            }

            try:
                lambda_client.invoke(
                    FunctionName=LAMBDA_FUNCTION_NAME,
                    InvocationType="Event",
                    Payload=json.dumps(event_payload).encode("utf-8"),
                )
                sent_ids.append(entry.get("expiry_id"))
                print(f"Invoice sent for membership ID: {entry.get('expiry_id')}")
            except Exception as e:
                raise HTTPException(
                    500, f"Could not enqueue receipt {entry.get('expiry_id')} for mailing: {e}"
                )

        # Update mail_send status in FittbotGymMembership (if we want to track this)
        # For now, we'll just invalidate the Redis cache

        if sent_ids:
            print("sent_ids", sent_ids)

            # Invalidate Redis cache for the specific gym and date
            invoice_key_pattern = f"gym:{gym_id}:invoice_data:*"
            unpaid_key_pattern = f"gym:{gym_id}:unpaid_members:*"
            about_expire_pattern = f"gym:{gym_id}:about_to_expire:*"
            expired_pattern = f"gym:{gym_id}:expired:*"

            for pattern in [invoice_key_pattern, unpaid_key_pattern, about_expire_pattern, expired_pattern]:
                keys = await redis.keys(pattern)
                if keys:
                    await redis.delete(*keys)
                    print(f"Deleted Redis keys matching pattern: {pattern}")

        return {"status": 200, "message": "Process started in background", "sent_count": len(sent_ids)}

    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")


def send_invoice_email(to_email: str, pdf_path: str):
    sender_email = os.getenv("SMTP_EMAIL")
    sender_password = os.getenv("SMTP_PASSWORD")
    msg = EmailMessage()
    # subject = f"Welcome to {company_name}! - Your Account Details"
    # body = f"""
    # <html>
    # <body>
    #     <p>Hello {user_name},</p>
 
    #     <p>Welcome to <b>{company_name}</b>! Your account has been successfully created. Please find your login details below:</p>
 
    #     <p><b>Registered Mobile Number:</b> {mobile_number}<br>
    #     <b>Default Password:</b> {default_password}</p>
 
    #     <p>For your security, we recommend that you change your default password immediately after your first login.</p>
 
    #     <p>You can access your account here: <a href="{login_url}">{login_url}</a></p>
 
    #     <p>If you have any questions or need further assistance, please feel free to contact our support team at <a href="mailto:{support_email}">{support_email}</a>.</p>
 
    #     <p>Thank you for joining us!</p>
 
    #     <p>Best regards,<br>
    #     The {company_name} Team</p>
    # </body>
    # </html>
    # """

    # msg = MIMEMultipart()
    msg["Subject"] = "Your Invoice from FitZone Gym"
    msg.set_content("Dear Member,\n\nPlease find attached your invoice.\n\nRegards,\nFitZone Gym")

    with open(pdf_path, "rb") as f:
        pdf_data = f.read()
        msg.add_attachment(pdf_data, maintype="application", subtype="pdf", filename="invoice.pdf")
    msg["From"] = sender_email
    msg["To"] = to_email
    # msg["Subject"] = subject
    # msg.attach(MIMEText(body, "html"))
 
    try:
        server = smtplib.SMTP(os.getenv("SMTP_SERVER"), os.getenv("SMTP_PORT"))
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, to_email, msg.as_string())
        server.quit()
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False

async def process_invoice(entry, db: Session, redis:Redis):
    try:
        pdf_path = f"tmp/{entry.gym_location[:3].upper()}-{entry.gym_id}-{entry.expiry_id}.pdf"
        os.makedirs("tmp", exist_ok=True)
        #generate_pdf(entry, pdf_path)
        expiry = db.query(AboutToExpire).filter(AboutToExpire.expiry_id == entry.expiry_id).first()
        invoice_key = f'gym{expiry.gym_id}:invoice_data'
        unpaid_key = f'gym{expiry.gym_id}:unpaid_data'
        
        if send_invoice_email(expiry.email, pdf_path):
            print("email sent")
        else:
            print("unable to send email")

        if expiry:
            expiry.mail_status = True
            db.commit()

        if await redis.exists(invoice_key):
            await redis.delete(invoice_key)

        if await redis.exists(unpaid_key):
            await redis.delete(unpaid_key)

        os.remove(pdf_path)
    except Exception as e:
        print(f"Error processing invoice {entry.expiry_id}: {e}")

def receipt_to_dict(receipt: FeesReceipt):

    return {
        "receipt_id": receipt.receipt_id,
        "client_id": receipt.client_id,
        "gym_id": receipt.gym_id,
        "client_name": receipt.client_name,
        "gym_name": receipt.gym_name,
        "gym_logo": receipt.gym_logo,
        "gym_contact": receipt.gym_contact,
        "gym_location": receipt.gym_location,
        "plan_id": receipt.plan_id,
        "plan_description": receipt.plan_description,
        "fees": receipt.fees,
        "fees_type":receipt.fees_type,
        "discount": receipt.discount,
        "discounted_fees": receipt.discounted_fees,
        "due_date": str(receipt.due_date),
        "invoice_number": receipt.invoice_number,
        "client_contact": receipt.client_contact,
        "bank_details": receipt.bank_details,
        "ifsc_code": receipt.ifsc_code,
        "account_holder_name": receipt.account_holder_name,
        "invoice_date": str(receipt.invoice_date),
        "payment_method": receipt.payment_method,
        "gst_number": receipt.gst_number,
        "gst_type" : receipt.gst_type if receipt.gst_type else "",
        "gst_percentage" : receipt.gst_percentage if receipt.gst_percentage else 0,
        "total_amount" : receipt.total_amount,
        "client_email": receipt.client_email,
        "mail_status": receipt.mail_status,
        "created_at": str(receipt.created_at),
        "update_at": str(receipt.update_at),
        "payment_date": str(receipt.payment_date),
        "payment_reference_number": receipt.payment_reference_number,
        "branch":receipt.branch
    }
 
@router.get("/fees-receipts")
async def list_fees_receipts(
    gym_id: int,
    month: str = None,
    year: int = None,
    page: int = 1,
    limit: int = 25,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis)
):

    try:
        if not month or not year:
            offset = (page - 1) * limit
            redis_key = f"gym{gym_id}:feesReceipt:all:page{page}:limit{limit}"
            cached_data = await redis.get(redis_key)

            if cached_data:
                return json.loads(cached_data)

            total_sent = db.query(FeesReceipt).filter(
                FeesReceipt.gym_id == gym_id,
                FeesReceipt.mail_status == True
            ).count()

            total_unsent = db.query(FeesReceipt).filter(
                FeesReceipt.gym_id == gym_id,
                FeesReceipt.mail_status == False
            ).count()
 
            send_receipts = db.query(FeesReceipt).filter(
                FeesReceipt.gym_id == gym_id,
                FeesReceipt.mail_status == True
            ).order_by(FeesReceipt.payment_date.desc()).offset(offset).limit(limit).all()
 
            unsend_receipts = db.query(FeesReceipt).filter(
                FeesReceipt.gym_id == gym_id,
                FeesReceipt.mail_status == False
            ).order_by(FeesReceipt.created_at.desc()).offset(offset).limit(limit).all()
 
            send_receipt_list = [receipt_to_dict(r) for r in send_receipts]
            unsend_receipt_list = [receipt_to_dict(r) for r in unsend_receipts]
 
            receipt_list = {
                "send": send_receipt_list,
                "unsend": unsend_receipt_list,
                "pagination": {
                    "current_page": page,
                    "limit": limit,
                    "total_sent": total_sent,
                    "total_unsent": total_unsent,
                    "total_pages_sent": (total_sent + limit - 1) // limit,
                    "total_pages_unsent": (total_unsent + limit - 1) // limit,
                    "has_next_sent": page * limit < total_sent,
                    "has_prev_sent": page > 1,
                    "has_next_unsent": page * limit < total_unsent,
                    "has_prev_unsent": page > 1
                }
            }
 
            result = {
                "status": 200,
                "data": receipt_list,
                "message": "Paginated data retrieved successfully",
                "view_mode": "all"
            }
 
            await redis.set(redis_key, json.dumps(result), ex=3600)  
            return result
 
        else:
            month_lower = month.strip().lower()
            redis_key = f"gym{gym_id}:feesReceipt:{month_lower}:{year}"
            cached_data = await redis.get(redis_key)
            if cached_data:
                cached_result = json.loads(cached_data)
                cached_result["view_mode"] = "monthly"
                return cached_result
 
            try:
                month_number = datetime.strptime(month, "%B").month
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid month name")
 
            send_receipts = db.query(FeesReceipt).filter(
                FeesReceipt.gym_id == gym_id,
                FeesReceipt.mail_status == True,
                extract('month', FeesReceipt.payment_date) == month_number,
                extract('year', FeesReceipt.payment_date) == year
            ).order_by(FeesReceipt.payment_date.desc()).all()
 
            unsend_receipts = db.query(FeesReceipt).filter(
                FeesReceipt.gym_id == gym_id,
                FeesReceipt.mail_status == False,
                extract('month', FeesReceipt.payment_date) == month_number,
                extract('year', FeesReceipt.payment_date) == year
            ).order_by(FeesReceipt.created_at.desc()).all()
 
            send_receipt_list = [receipt_to_dict(r) for r in send_receipts]
            unsend_receipt_list = [receipt_to_dict(r) for r in unsend_receipts]
            receipt_list = {
                "send": send_receipt_list,
                "unsend": unsend_receipt_list
            }
 
            result = {
                "status": 200,
                "data": receipt_list,
                "message": "Monthly data retrieved successfully",
                "month": month,
                "year": year,
                "view_mode": "monthly"
            }

            await redis.set(redis_key, json.dumps(result), ex=86400) 

          
            return result
 
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")
 
@router.get("/feedback")
async def get_feedback_for_gym(gym_id: int, db: Session = Depends(get_db)):
    try:
        result = db.execute(
            select(Gym_Feedback, Client.name, Client.profile)
            .join(Client, Gym_Feedback.client_id == Client.client_id)
            .filter(Gym_Feedback.gym_id == gym_id)
            .order_by(desc(Gym_Feedback.timing))
        )
        feedbacks = result.all()  
        if not feedbacks:
            return {"status": 200, "data": []}  
 
        response_data = [
            {
                "feedback_id": fb.Gym_Feedback.id,
                "client_id": fb.Gym_Feedback.client_id,
                "client_name": fb.name,
                "tag": fb.Gym_Feedback.tag,
                "ratings": fb.Gym_Feedback.ratings,
                "feedback": fb.Gym_Feedback.feedback,
                "timing": fb.Gym_Feedback.timing,
                "client_image":fb.profile
            }
            for fb in feedbacks
        ]
 
        return {"status": 200, "data": response_data}
 
    except Exception as e:
        print(e)
        return {"status": 500, "message": f"Internal Server Error: {str(e)}"}
 
def generate_receipt(data, output_path: str):
    try:
        #template = env.get_template("receipt_template.html")
        # html_out = template.render(data=data)
        pass
        
    except Exception as e:
        print(str(e))

def send_receipt_email(to_email: str, pdf_path: str):
    sender_email = os.getenv("SMTP_EMAIL")
    sender_password = os.getenv("SMTP_PASSWORD")
    msg = EmailMessage()
    # subject = f"Welcome to {company_name}! - Your Account Details"
    # body = f"""
    # <html>
    # <body>
    #     <p>Hello {user_name},</p>
 
    #     <p>Welcome to <b>{company_name}</b>! Your account has been successfully created. Please find your login details below:</p>
 
    #     <p><b>Registered Mobile Number:</b> {mobile_number}<br>
    #     <b>Default Password:</b> {default_password}</p>
 
    #     <p>For your security, we recommend that you change your default password immediately after your first login.</p>
 
    #     <p>You can access your account here: <a href="{login_url}">{login_url}</a></p>
 
    #     <p>If you have any questions or need further assistance, please feel free to contact our support team at <a href="mailto:{support_email}">{support_email}</a>.</p>
 
    #     <p>Thank you for joining us!</p>
 
    #     <p>Best regards,<br>
    #     The {company_name} Team</p>
    # </body>
    # </html>
    # """

    # msg = MIMEMultipart()
    msg["Subject"] = "Your Invoice from FitZone Gym"
    msg.set_content("Dear Member,\n\nPlease find attached your invoice.\n\nRegards,\nFitZone Gym")

    with open(pdf_path, "rb") as f:
        pdf_data = f.read()
        msg.add_attachment(pdf_data, maintype="application", subtype="pdf", filename="invoice.pdf")
    msg["From"] = sender_email
    msg["To"] = to_email
    # msg["Subject"] = subject
    # msg.attach(MIMEText(body, "html"))
 
    try:
        server = smtplib.SMTP(os.getenv("SMTP_SERVER"), os.getenv("SMTP_PORT"))
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, to_email, msg.as_string())
        server.quit()
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False

async def process_receipt(entry, db: Session, redis:Redis):
    try:
        pdf_path = f"tmp/{entry.gym_name[:3].upper()}-{entry.gym_id}-{entry.receipt_id}.pdf"
        print(pdf_path)
        os.makedirs("tmp", exist_ok=True)
        generate_receipt(entry, pdf_path)
        receipt = db.query(FeesReceipt).filter(FeesReceipt.receipt_id == entry.receipt_id).first()
        invoice_key = f'gym{receipt.gym_id}:invoice_data'
        unpaid_key = f'gym{receipt.gym_id}:unpaid_data'
        
        if send_receipt_email(receipt.client_email, pdf_path):
            print("email sent")
        else:
            print("unable to send email")

        if receipt:
            receipt.mail_status = True
            db.commit()

        pattern = f"gym{receipt.gym_id}:feesReceipt:*"
        async for key in redis.scan_iter(match=pattern):
            await redis.delete(key)

        os.remove(pdf_path)
    except Exception as e:
        print(f"Error processing invoice {entry.receipt_id}: {e}")

def is_valid_mobile(number: str) -> bool:
    pattern = r"[0-9]\d{9}"
    result = re.fullmatch(pattern, number)
    print(f"  is_valid_mobile: input='{number}', pattern='{pattern}', match={result is not None}")
    return result is not None

def is_valid_gender(gender: str) -> bool:
    return gender.lower() in ["male", "female"]

@router.post("/import_gym_data")
async def import_client_data(
    gym_id: int = Form(...),
    file: UploadFile = Form(...),
    db: Session = Depends(get_db),
    redis:Redis = Depends(get_redis)
):  
    try:
        filename = file.filename.lower()
        corrupted_rows: List[Dict] = []
        imported_count = 0

        def process_row(row_data, row_num):
            Name, Email, Contact, Location, Gender, Status, AdmissionNumber, ExpiresAt, JoinedAt = row_data
            errors = []

            if not Name or not str(Name).strip():
                errors.append("Client name is missing")
                print("ERROR: Client name is missing")

            contact_str = str(Contact).strip()
            print(f"Contact after strip: '{contact_str}' (length: {len(contact_str)})")
            is_valid = is_valid_mobile(contact_str)
            print(f"is_valid_mobile result: {is_valid}")

            if not Contact or not is_valid:
                print(f"ERROR: Invalid mobile number - Contact value: '{Contact}'")
                errors.append("Invalid mobile number")

            email_str = str(Email).strip()
            print(f"Email after strip: '{email_str}' (has @: {'@' in email_str})")

            if not Email or "@" not in email_str:
                print(f"ERROR: Invalid email - Email value: '{Email}'")
                errors.append("Invalid email")

            if not Gender or not is_valid_gender(str(Gender)):
                print(f"ERROR: Invalid gender - Gender value: '{Gender}'")
                errors.append("Invalid gender")

            existing = db.query(GymImportData).filter(
                GymImportData.gym_id == gym_id,
                (GymImportData.client_email == str(Email)) | (GymImportData.client_contact == str(Contact))
            ).first()

            if existing:
                errors.append("Duplicate entry found (email or contact already exists)")

            # Parse expires_at date
            expires_at_date = None
            if ExpiresAt:
                try:
                    # Try to parse the date - handle both string and datetime objects
                    if isinstance(ExpiresAt, str):
                        expires_at_date = datetime.strptime(str(ExpiresAt).strip(), "%Y-%m-%d").date()
                    elif isinstance(ExpiresAt, datetime):
                        expires_at_date = ExpiresAt.date()
                    elif isinstance(ExpiresAt, date):
                        expires_at_date = ExpiresAt
                except Exception as e:
                    print(f"ERROR: Invalid expires_at date format - ExpiresAt value: '{ExpiresAt}'")
                    errors.append("Invalid expires_at date format (expected YYYY-MM-DD)")

            # Parse joined_at date
            joined_at_date = None
            if JoinedAt:
                try:
                    # Try to parse the date - handle both string and datetime objects
                    if isinstance(JoinedAt, str):
                        joined_at_date = datetime.strptime(str(JoinedAt).strip(), "%Y-%m-%d").date()
                    elif isinstance(JoinedAt, datetime):
                        joined_at_date = JoinedAt.date()
                    elif isinstance(JoinedAt, date):
                        joined_at_date = JoinedAt
                except Exception as e:
                    print(f"ERROR: Invalid joined_at date format - JoinedAt value: '{JoinedAt}'")
                    errors.append("Invalid joined_at date format (expected YYYY-MM-DD)")

            if errors:
                corrupted_rows.append({
                    "row": row_num,
                    "data": {
                        "Name": Name,
                        "Contact": Contact,
                        "Email": Email,
                        "Location": Location,
                        "Gender": Gender,
                        "Status": Status,
                        "AdmissionNumber": AdmissionNumber,
                        "ExpiresAt": ExpiresAt,
                        "JoinedAt": JoinedAt,
                    },
                    "errors": errors
                })
            else:
                import_record = GymImportData(
                    gym_id=gym_id,
                    client_name=str(Name),
                    client_contact=str(Contact),
                    client_email=str(Email),
                    client_location=str(Location) if Location else None,
                    status=str(Status) if Status else None,
                    gender=str(Gender),
                    admission_number=str(AdmissionNumber) if AdmissionNumber else None,
                    expires_at=expires_at_date,
                    joined_at=joined_at_date,
                    sms_status=False
                )
                db.add(import_record)
                db.flush()  # Get import_id for ClientGym

                # Add to ClientGym for unified admission number tracking
                if AdmissionNumber:
                    client_gym_record = ClientGym(
                        client_id=-1000000 - import_record.import_id,  # Large negative offset for imports
                        gym_id=gym_id,
                        gym_client_id=str(AdmissionNumber),
                        admission_number=str(AdmissionNumber)
                    )
                    db.add(client_gym_record)

                # Add to FittbotGymMembership only if expires_at is provided (for active members tracking)
                if expires_at_date:
                    membership_record = FittbotGymMembership(
                        gym_id=str(gym_id),
                        client_id=f"import_{import_record.import_id}",
                        plan_id=None,
                        type="imported",
                        amount=0,
                        purchased_at=datetime.now(),
                        status="active" if str(Status).lower() == "active" else "inactive",
                        joined_at=joined_at_date if joined_at_date else date.today(),
                        expires_at=expires_at_date,
                    )
                    db.add(membership_record)
                return True
            return False

        contents = await file.read()

        if filename.endswith((".xlsx", ".xls")):
            wb = load_workbook(filename=BytesIO(contents))
            sheet = wb.active
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if process_row(row, idx):
                    imported_count += 1

        elif filename.endswith(".csv"):
            decoded = contents.decode("utf-8")
            reader = csv.reader(StringIO(decoded))
            next(reader)
            for idx, row in enumerate(reader, start=2):
                if process_row(row, idx):
                    imported_count += 1

        else:
            raise HTTPException(status_code=400, detail="Unsupported file format")

        db.commit()


        all_key = f"gym{gym_id}:allImportData"
        active_key = f"gym{gym_id}:activeImportData"
        inactive_key = f"gym{gym_id}:inactiveImportData"
        client_redis_key = f"gym:{gym_id}:all_clients"
        trainer_redis_key = f"gym:{gym_id}:trainers"
        diet_redis_key = f"gym:{gym_id}:all_diets"
        workout_redis_key = f"gym:{gym_id}:all_workouts"
        today = date.today()
        attendance_key = f"gym:{gym_id}:attendance:{today.strftime('%Y-%m-%d')}"

        if await redis.exists(all_key):
            await redis.delete(all_key)

        if await redis.exists(active_key):
            await redis.delete(active_key)

        if await redis.exists(inactive_key):
            await redis.delete(inactive_key)


        members_key = f"gym:{gym_id}:members"
        if await redis.exists(members_key):
            await redis.delete(members_key)

        # Clear client data cache
        client_data_key = f"gym:{gym_id}:clientdata"
        if await redis.exists(client_data_key):
            await redis.delete(client_data_key)

        # Clear additional caches
        await redis.delete(client_redis_key, trainer_redis_key, diet_redis_key, workout_redis_key, attendance_key)

        return {
            "status": 200,
            "message":"Process started in background",
            "imported_count": imported_count,
            "corrupted_rows": corrupted_rows,
        }
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occured : {str(e)}")

    
def send_intimation_sms(entry):
    api_key = os.getenv("OTHER_API_KEY")
    sender_id = "NFCFIT"
    entity_id = "1701174022473316577"
    template_id = "1707174123946514804"
    app_url = "Fittbot"
    encoded_message = f"Hi {entry.client_name}, You can access your app here {app_url}. Regards- NFCFIT"
 
    url = (
        f"http://pwtpl.com/sms/V1/send-sms-api.php?"
        f"apikey={api_key}&senderid={sender_id}&templateid={template_id}"
        f"&entityid={entity_id}&number={entry.client_contact}&message={encoded_message}&format=json"
    )
 
    try:
        response = requests.get(url)
        if response.status_code == 200:
            json_response = response.json()
            print("SMS API Response:", json_response)
            if json_response.get('status') == 'OK':
                print("SMS sent successfully!")
                return True
            else:
                print("Failed to send SMS. Please check the API credentials and parameters.")
                return False
        else:
            print("HTTP error occurred:", response.status_code)
    except Exception as e:
        print("An error occurred:", e)
    
class intimationRequest(BaseModel):
    gym_id:int
    import_ids:List[int]

class ImportClientPayload(BaseModel):
    name: str
    contact: str
    email: Optional[str] = None
    location: Optional[str] = None
    status: Optional[str] = None
    gender: str
    admission_number: Optional[str] = None
    expires_at: Optional[Any] = None

class ImportClientPayloadRequest(BaseModel):
    gym_id: int
    clients: List[ImportClientPayload]

@router.post("/send-intimations")
async def generate_and_send_receipts(request:intimationRequest, db: Session = Depends(get_db), redis: Redis = Depends(get_redis)):


    try:
        entries = db.query(GymImportData).filter(GymImportData.import_id.in_(request.import_ids)).all()
        for entry in entries:
            if send_intimation_sms(entry):
                entry.sms_status = True
                db.commit()
            else:
                print("unable to send email")

        all_key = f"gym{request.gym_id}:allImportData"
        active_key = f"gym{request.gym_id}:activeImportData"
        inactive_key = f"gym{request.gym_id}:inactiveImportData"

        if await redis.exists(all_key):
            await redis.delete(all_key)
        if await redis.exists(active_key):
            await redis.delete(active_key)
        if await redis.exists(inactive_key):
            await redis.delete(inactive_key)

        return {"status":200, "message": "Process started in background"}

    except Exception as e: 
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occured:{str(e)}")

def get_data_as_dict(client):
    # Determine status based on expires_at if available
    today = date.today()
    expires_at = None
    if hasattr(client, 'expires_at') and client.expires_at:
        expires_at = str(client.expires_at) if client.expires_at else None
        if client.expires_at >= today:
            status = "active"
        else:
            status = "inactive"
    else:
        status = "inactive"

    return {
        "import_id":client.import_id,
        "gym_id":client.gym_id,
        "name":client.client_name,
        "contact":client.client_contact,
        "email":client.client_email,
        "location":client.client_location,
        "status":status,
        "gender":client.gender,
        "sms_status":client.sms_status,
        "admission_number":client.admission_number,
        "import_status":True,
        "expires_at":expires_at
    }


@router.get('/get-export-data')
async def get_export_data(gym_id : int, db:Session = Depends(get_db), redis:Redis=Depends(get_redis)):
    try:

        all_key = f"gym{gym_id}:allImportData"
        active_key = f"gym{gym_id}:activeImportData"
        inactive_key = f"gym{gym_id}:inactiveImportData"

        cached_all = await redis.get(all_key)
        cached_active = await redis.get(active_key)
        cached_inactive = await redis.get(inactive_key)

        all_client_data=[] 
        active_client_data=[]
        inactive_client_data=[]

        if cached_all:
            all_client_data = json.loads(cached_all)
        else:
            all_clients=db.query(GymImportData).filter(GymImportData.gym_id == gym_id).all()
            all_client_data=[get_data_as_dict(c) for c in all_clients]
            await redis.set(all_key, json.dumps(all_client_data), ex=86400)
        
        if cached_active:
            active_client_data = json.loads(cached_active)
        else:
            active_clients = db.query(GymImportData).filter(GymImportData.gym_id == gym_id, GymImportData.sms_status == True).all()
            active_client_data=[get_data_as_dict(c) for c in active_clients]
            await redis.set(active_key, json.dumps(active_client_data),ex=86400)

        if cached_inactive:
            inactive_client_data=json.loads(cached_inactive)
        else:
            inactive_clients = db.query(GymImportData).filter(GymImportData.gym_id == gym_id, GymImportData.sms_status == False).all()
            inactive_client_data=[get_data_as_dict(c) for c in inactive_clients]
            await redis.set(inactive_key, json.dumps(inactive_client_data), ex=86400)
        
        
        print("all_client_data",all_client_data)
        print("send_clients",active_client_data)
        print("unsend_clients",inactive_client_data)
        
        
        return{
            "status":200,
            "message":"Data retrived successfully",
            "import_clients":all_client_data,
  
        }
    
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occured : {str(e)}")


@router.post("/send-receipts")
async def generate_and_send_receipts(receipt_ids:List[int], db: Session = Depends(get_db), redis: Redis = Depends(get_redis)):
    try:
        pattern = f'gym*:feesReceipt:*'
        async for key in redis.scan_iter(match=pattern):
            await redis.delete(key)
        entries = db.query(FeesReceipt).filter(FeesReceipt.receipt_id.in_(receipt_ids)).all()
        for entry in entries:
            event_payload = {
 
                "invoice_data": {
                    "receipt_id":        entry.receipt_id,
                    "client_id":         entry.client_id,
                    "gym_id":            entry.gym_id,
                    "client_name":       entry.client_name,
                    "gym_name":          entry.gym_name,
                    "gym_logo":          entry.gym_logo,
                    "gym_contact":       entry.gym_contact,
                    "gym_location":      entry.gym_location,
                    "plan_id":           entry.plan_id,
                    "plan_description":  entry.plan_description,
                    "fees":              int(entry.fees),
                    "discount":          int(entry.discount),
                    "discounted_fees":   int(entry.discounted_fees),
                    "due_date":          str(entry.due_date),
                    "invoice_number":    entry.invoice_number,
                    "client_contact":    entry.client_contact,
                    "bank_details":      entry.bank_details,
                    "ifsc_code":         entry.ifsc_code,
                    "account_holder":    entry.account_holder_name,
                    "invoice_date":      str(entry.invoice_date),
                    "payment_method":    entry.payment_method,
                    "gst_number":        entry.gst_number,
                    "client_email":      entry.client_email,
                    "payment_date":      str(entry.payment_date),
                    "payment_reference": entry.payment_reference_number,
                    "gst_type":          entry.gst_type,
                    'gst_percentage':    entry.gst_percentage,
                    'branch': entry.branch,
                    "discounted_price":(entry.fees-entry.discounted_fees),
                    "invoice_type":'receipt'
                }
            }
 
            try:
                
                entry.mail_status=True
                db.commit()
                receipt_key_pattern = f"gym{entry.gym_id}:feesReceipt:all:page*"
                cursor = 0
                while True:
                    cursor, keys = await redis.scan(cursor=cursor, match=receipt_key_pattern, count=100)
                    if keys:
                        await redis.delete(*keys)
                    if cursor == 0:
                        break
                # Use Lambda retry utility - automatically retries on throttling/errors
                invoke_lambda_with_retry(
                    lambda_client,
                    FunctionName=LAMBDA_FUNCTION_NAME,
                    InvocationType="Event",
                    Payload=json.dumps(event_payload).encode("utf-8"),
                )
 
            except Exception as e:
                raise HTTPException(500, f"Could not enqueue receipt for mailing: {e}")
 
        return {"status":200, "message": "Process started in background"}
 
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occured:{str(e)}")
 
def estimate_to_dict(estimate: EnquiryEstimates):
    return {
        "id": estimate.id,
        "enquiry_id": estimate.enquiry_id,
        "gym_id": estimate.gym_id,
        "client_name": estimate.client_name,
        "gym_name": estimate.gym_name,
        "gym_logo": estimate.gym_logo,
        "gym_contact": estimate.gym_contact,
        "gym_location": estimate.gym_location,
        "plan_id": estimate.plan_id,
        "plan_description": estimate.plan_description,
        "fees": estimate.fees,
        "admission_fees": estimate.admission_fees,
        "fees_type": estimate.fees_type,
        "discount": estimate.discount,
        "discounted_fees": estimate.discounted_fees,
        "estimate_number": estimate.estimate_number,
        "client_contact": estimate.client_contact,
        "bank_details": estimate.bank_details,
        "ifsc_code": estimate.ifsc_code,
        "bank_name": estimate.bank_name,
        "upi_id": estimate.upi_id,
        "account_holder_name": estimate.account_holder_name,
        "estimate_date": estimate.estimate_date,
        "gst_number": estimate.gst_number,
        "client_email": estimate.client_email,
        "mail_status": estimate.mail_status,
        "created_at": str(estimate.created_at) if estimate.created_at else None,
        "update_at": str(estimate.update_at) if estimate.update_at else None,
        "gst_percentage": estimate.gst_percentage if estimate.gst_percentage else 18,
        "gst_type": estimate.gst_type if estimate.gst_type else "",
        "branch": estimate.branch,
        "total_amount": estimate.total_amount
    }

async def delete_invoice_keys(entry, redis):
    invoice_key_all = f'gym{entry.gym_id}:invoice_data:all'
    unpaid_key_pattern = f'gym{entry.gym_id}:unpaid_members:*'

    await redis.delete(invoice_key_all)

    cursor = 0
    while True:
        cursor, keys = await redis.scan(cursor=cursor, match=unpaid_key_pattern, count=100)
        if keys:
            await redis.delete(*keys)
        if cursor == 0:
            break

@router.get("/enquiry-estimates")
async def list_enquiry_estimates(
    gym_id: int,
    month: str = None,
    year: int = None,
    search: str = None,
    page: int = 1,
    limit: int = 25,
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis)
):
    try:
        offset = (page - 1) * limit
        
        base_query = db.query(EnquiryEstimates).filter(EnquiryEstimates.gym_id == gym_id)
        
        if month and year:
            try:
                month_number = datetime.strptime(month, "%B").month
                base_query = base_query.filter(
                    extract('month', EnquiryEstimates.created_at) == month_number,
                    extract('year', EnquiryEstimates.created_at) == year
                )
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid month name")
        
        if search:
            search_term = f"%{search}%"
            base_query = base_query.filter(
                or_(
                    EnquiryEstimates.client_name.ilike(search_term),
                    EnquiryEstimates.client_contact.ilike(search_term),
                    EnquiryEstimates.client_email.ilike(search_term),
                    EnquiryEstimates.estimate_number.ilike(search_term),
                    EnquiryEstimates.plan_description.ilike(search_term),
                    EnquiryEstimates.fees_type.ilike(search_term)
                )
            )
        
        redis_key = f"gym{gym_id}:enquiryEstimates"
        if month and year:
            redis_key += f":{month.lower()}:{year}"
        if search:
            redis_key += f":search:{search.lower()}"
        redis_key += f":page{page}:limit{limit}"
        
        cached_data = await redis.get(redis_key)
        if cached_data and not search:
            return json.loads(cached_data)
        
        total_sent = base_query.filter(EnquiryEstimates.mail_status == True).count()
        total_unsent = base_query.filter(EnquiryEstimates.mail_status == False).count()
        
        sent_estimates = base_query.filter(
            EnquiryEstimates.mail_status == True
        ).order_by(EnquiryEstimates.created_at.desc()).offset(offset).limit(limit).all()
        
        unsent_estimates = base_query.filter(
            EnquiryEstimates.mail_status == False
        ).order_by(EnquiryEstimates.created_at.desc()).offset(offset).limit(limit).all()
        
        sent_estimate_list = [estimate_to_dict(e) for e in sent_estimates]
        unsent_estimate_list = [estimate_to_dict(e) for e in unsent_estimates]
        
        estimate_list = {
            "sent": sent_estimate_list,
            "unsent": unsent_estimate_list,
            "pagination": {
                "current_page": page,
                "limit": limit,
                "total_sent": total_sent,
                "total_unsent": total_unsent,
                "total_pages_sent": (total_sent + limit - 1) // limit if total_sent > 0 else 1,
                "total_pages_unsent": (total_unsent + limit - 1) // limit if total_unsent > 0 else 1,
                "has_next_sent": page * limit < total_sent,
                "has_prev_sent": page > 1,
                "has_next_unsent": page * limit < total_unsent,
                "has_prev_unsent": page > 1
            }
        }
        
        view_mode = "all"
        if month and year:
            view_mode = "monthly"
        if search:
            view_mode = "search"
            
        result = {
            "status": 200,
            "data": estimate_list,
            "message": "Enquiry estimates retrieved successfully",
            "view_mode": view_mode,
            "search_query": search if search else None
        }
        
        if not search:
            await redis.set(redis_key, json.dumps(result), ex=3600)
        
        return result
        
    except Exception as e:
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")


@router.post("/send-enquiry-estimates")
async def generate_and_send_enquiry_estimates(estimate_ids:List[int], db: Session = Depends(get_db), redis: Redis = Depends(get_redis)):
    try:

        entries = (
            db.query(EnquiryEstimates)
            .filter(EnquiryEstimates.id.in_(estimate_ids))
            .all()
        )

        for entry in entries:
            event_payload = {
                "invoice_data": {
                    "receipt_id":        entry.id,
                    "enquiry_id":        entry.enquiry_id,
                    "gym_id":            entry.gym_id,
                    "client_name":       entry.client_name,
                    "gym_name":          entry.gym_name,
                    "gym_logo":          entry.gym_logo,
                    "gym_contact":       entry.gym_contact,
                    "gym_location":      entry.gym_location,
                    "plan_id":           entry.plan_id,
                    "plan_description":  entry.plan_description,
                    "fees":              int(entry.fees),
                    "admission_fees":    int(entry.admission_fees) if entry.admission_fees else 0,
                    "fees_type":         entry.fees_type,
                    "discount":          int(entry.discount),
                    "discounted_fees":   int(entry.discounted_fees),
                    "estimate_date":     str(entry.estimate_date),
                    "invoice_number":   entry.estimate_number,
                    "client_contact":    entry.client_contact,
                    "bank_details":      entry.bank_details,
                    "ifsc_code":         entry.ifsc_code,
                    "account_holder":    entry.account_holder_name,
                    "client_email":      entry.client_email,
                    "upi_id":            entry.upi_id,
                    "bank_name":         entry.bank_name,
                    "gst_number":        entry.gst_number,
                    "gst_type":          entry.gst_type,
                    "gst_percentage":    entry.gst_percentage,
                    "discounted_price":  ((entry.admission_fees + entry.fees) - entry.discount),
                    "invoice_type":      "enquiry_estimate",
                }
            }

            try:
                entry.mail_status = True
                db.commit()
                pattern = f"gym{entry.gym_id}:enquiryEstimates*"
                async for key in redis.scan_iter(match=pattern):
                    await redis.delete(key)
                # Use Lambda retry utility - automatically retries on throttling/errors
                invoke_lambda_with_retry(
                    lambda_client,
                    FunctionName=LAMBDA_FUNCTION_NAME,
                    InvocationType="Event",
                    Payload=json.dumps(event_payload).encode("utf-8"),
                )
                 
            except Exception as e:
                raise HTTPException(
                    500, f"Could not enqueue estimate {entry.id} for mailing: {e}"
                )

        return {"status":200, "message": "Process started in background"}

    except Exception as e: 
        print(str(e))
        raise HTTPException(status_code=500, detail=f"An error occured:{str(e)}")
