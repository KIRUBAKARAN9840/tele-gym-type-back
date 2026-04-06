# Purchases API for Admin Dashboard
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_, cast, Integer, func, union_all, literal, distinct, desc, asc
from typing import Optional
import json
import io
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from app.models.fittbot_models import Client, Gym, GymOwner, SessionPurchase, SessionBookingDay, FittbotGymMembership
from app.models.async_database import get_async_db
from app.models.dailypass_models import DailyPass, DailyPassDay
from app.fittbot_api.v1.payments.models.payments import Payment
from app.fittbot_api.v1.payments.models.orders import Order, OrderItem
from app.fittbot_api.v1.payments.models.subscriptions import Subscription

router = APIRouter(prefix="/api/admin/purchases", tags=["AdminPurchases"])


@router.get("/all-purchases")
async def get_all_purchases(
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, le=100, description="Items per page"),
    search: Optional[str] = Query(None, description="Search by client or gym name"),
    type: Optional[str] = Query(None, description="Filter by type: 'Session' or 'Daily Pass'"),
    start_date: Optional[str] = Query(None, description="Filter by start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="Filter by end date (YYYY-MM-DD)"),
    distinct_clients: Optional[bool] = Query(False, description="Show only distinct clients (1 booking across all types)"),
    distinct_gyms: Optional[bool] = Query(False, description="Show only distinct gyms (1 booking across all types)"),
    db: AsyncSession = Depends(get_async_db)
):
    try:
        search_pattern = f"%{search}%" if search else None

        # Parse dates if provided
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

        # Build DailyPass subquery with type label
        daily_pass_query = (
            select(
                DailyPass.id.label("id"),
                DailyPass.client_id.label("client_id"),
                DailyPass.gym_id.label("gym_id"),
                literal("Daily Pass").label("type"),
                DailyPass.days_total.label("days_total"),
                literal(None).label("total_sessions"),
                literal(None).label("scheduled_sessions"),
                (Payment.amount_minor / 100.0).label("amount"),
                DailyPass.created_at.label("purchased_at"),
                Client.name.label("client_name"),
                Gym.name.label("gym_name"),
                Gym.contact_number.label("gym_contact"),
                GymOwner.contact_number.label("owner_contact"),
                GymOwner.name.label("owner_name"),
                Gym.area.label("gym_area"),
                Client.contact.label("client_contact"),
                Client.platform.label("platform")
            )
            .select_from(DailyPass)
            .join(Gym, cast(DailyPass.gym_id, Integer) == Gym.gym_id)  # Inner join - exclude if gym not found
            .outerjoin(Client, cast(DailyPass.client_id, Integer) == Client.client_id)
            .outerjoin(GymOwner, Gym.owner_id == GymOwner.owner_id)
            .outerjoin(Payment, DailyPass.payment_id == Payment.provider_payment_id)
            .where(DailyPass.gym_id != "1")  # Exclude gym_id = 1
        )

        # Build SessionPurchase subquery with type label (only paid status)
        session_purchase_query = (
            select(
                SessionPurchase.id.label("id"),
                SessionPurchase.client_id.label("client_id"),
                SessionPurchase.gym_id.label("gym_id"),
                literal("Session").label("type"),
                literal(None).label("days_total"),
                SessionPurchase.sessions_count.label("total_sessions"),
                SessionPurchase.scheduled_sessions.label("scheduled_sessions"),
                SessionPurchase.payable_rupees.label("amount"),
                SessionPurchase.created_at.label("purchased_at"),
                Client.name.label("client_name"),
                Gym.name.label("gym_name"),
                Gym.contact_number.label("gym_contact"),
                GymOwner.contact_number.label("owner_contact"),
                GymOwner.name.label("owner_name"),
                Gym.area.label("gym_area"),
                Client.contact.label("client_contact"),
                Client.platform.label("platform")
            )
            .select_from(SessionPurchase)
            .join(Gym, SessionPurchase.gym_id == Gym.gym_id)  # Inner join - exclude if gym not found
            .outerjoin(Client, SessionPurchase.client_id == Client.client_id)
            .outerjoin(GymOwner, Gym.owner_id == GymOwner.owner_id)
            .where(SessionPurchase.status == "paid")
            .where(SessionPurchase.gym_id != 1)  # Exclude gym_id = 1
        )

        # Apply search filters to both subqueries if search is provided
        if search:
            daily_pass_query = daily_pass_query.where(
                or_(
                    Client.name.ilike(search_pattern),
                    Client.contact.ilike(search_pattern),
                    Gym.name.ilike(search_pattern)
                )
            )
            session_purchase_query = session_purchase_query.where(
                or_(
                    Client.name.ilike(search_pattern),
                    Client.contact.ilike(search_pattern),
                    Gym.name.ilike(search_pattern)
                )
            )

        # Apply date filters to both subqueries if provided
        if start_date_obj:
            daily_pass_query = daily_pass_query.where(func.date(DailyPass.created_at) >= start_date_obj)
            session_purchase_query = session_purchase_query.where(func.date(SessionPurchase.created_at) >= start_date_obj)
        if end_date_obj:
            daily_pass_query = daily_pass_query.where(func.date(DailyPass.created_at) <= end_date_obj)
            session_purchase_query = session_purchase_query.where(func.date(SessionPurchase.created_at) <= end_date_obj)

        # Combine both queries with UNION ALL
        combined_query = union_all(daily_pass_query, session_purchase_query).alias("combined_purchases")

        # Build distinct filtering subqueries (if distinct filter is requested)
        distinct_client_filter = None
        distinct_gym_filter = None

        if distinct_clients or distinct_gyms:
            # Get gym memberships base data for distinct calculation
            gym_membership_data_query = (
                select(
                    Order.customer_id.label("client_id"),
                    OrderItem.gym_id.label("gym_id"),
                    Order.order_metadata.label("order_metadata")
                )
                .select_from(Payment)
                .join(Order, Order.id == Payment.order_id)
                .join(OrderItem, OrderItem.order_id == Order.id)
                .where(Payment.status == "captured")
                .where(Order.status == "paid")
                .where(OrderItem.gym_id.isnot(None))
                .where(OrderItem.gym_id != "1")
            )

            # Apply date filters to gym memberships
            if start_date_obj:
                gym_membership_data_query = gym_membership_data_query.where(func.date(Payment.created_at) >= start_date_obj)
            if end_date_obj:
                gym_membership_data_query = gym_membership_data_query.where(func.date(Payment.created_at) <= end_date_obj)

            gm_result = await db.execute(gym_membership_data_query)
            gm_rows = gm_result.all()

            # Filter by metadata conditions
            valid_client_ids = []
            valid_gym_ids = []
            for row in gm_rows:
                metadata = row.order_metadata
                if not metadata or not isinstance(metadata, dict):
                    continue

                condition1 = (
                    metadata.get("audit") and isinstance(metadata.get("audit"), dict) and
                    metadata["audit"].get("source") == "dailypass_checkout_api"
                )
                condition2 = (
                    metadata.get("order_info") and isinstance(metadata.get("order_info"), dict) and
                    metadata["order_info"].get("flow") == "unified_gym_membership_with_sub"
                )
                condition3 = (
                    metadata.get("order_info") and isinstance(metadata.get("order_info"), dict) and
                    metadata["order_info"].get("flow") == "unified_gym_membership_with_free_fittbot"
                )

                if condition1 or condition2 or condition3:
                    if row.client_id:
                        try:
                            valid_client_ids.append(int(row.client_id))
                        except:
                            pass
                    if row.gym_id and row.gym_id.isdigit():
                        valid_gym_ids.append(int(row.gym_id))

            # Build distinct client filter (clients with exactly 1 booking total)
            if distinct_clients:
                # Count bookings per client from sessions/daily passes
                session_client_counts = (
                    select(
                        combined_query.c.client_name,
                        func.count().label("booking_count")
                    )
                    .select_from(combined_query)
                    .where(combined_query.c.client_name.isnot(None))
                    .group_by(combined_query.c.client_name)
                )
                if type:
                    session_client_counts = session_client_counts.where(combined_query.c.type == type)
                session_client_result = await db.execute(session_client_counts)
                session_client_counts_map = {row.client_name: row.booking_count for row in session_client_result.all()}

                # Count gym memberships per client
                gm_client_counts_map = {}
                if valid_client_ids:
                    gm_client_query = (
                        select(Client.name, func.count().label("gm_count"))
                        .where(Client.client_id.in_(valid_client_ids))
                        .where(Client.name.isnot(None))
                        .group_by(Client.name)
                    )
                    gm_client_result = await db.execute(gm_client_query)
                    gm_client_counts_map = {row.name: row.gm_count for row in gm_client_result.all()}

                # Find distinct clients (total count == 1)
                distinct_client_names = set()
                for client_name in set(session_client_counts_map.keys()) | set(gm_client_counts_map.keys()):
                    total = session_client_counts_map.get(client_name, 0) + gm_client_counts_map.get(client_name, 0)
                    if total == 1:
                        distinct_client_names.add(client_name)

                if distinct_client_names:
                    distinct_client_filter = distinct_client_names

            # Build distinct gym filter (gyms with exactly 1 booking total)
            if distinct_gyms:
                # Count bookings per gym from sessions/daily passes
                session_gym_counts = (
                    select(
                        combined_query.c.gym_name,
                        func.count().label("booking_count")
                    )
                    .select_from(combined_query)
                    .where(combined_query.c.gym_name.isnot(None))
                    .group_by(combined_query.c.gym_name)
                )
                if type:
                    session_gym_counts = session_gym_counts.where(combined_query.c.type == type)
                session_gym_result = await db.execute(session_gym_counts)
                session_gym_counts_map = {row.gym_name: row.booking_count for row in session_gym_result.all()}

                # Count gym memberships per gym
                gm_gym_counts_map = {}
                if valid_gym_ids:
                    gm_gym_query = (
                        select(Gym.name, func.count().label("gm_count"))
                        .where(Gym.gym_id.in_(valid_gym_ids))
                        .where(Gym.name.isnot(None))
                        .group_by(Gym.name)
                    )
                    gm_gym_result = await db.execute(gm_gym_query)
                    gm_gym_counts_map = {row.name: row.gm_count for row in gm_gym_result.all()}

                # Find distinct gyms (total count == 1)
                distinct_gym_names = set()
                for gym_name in set(session_gym_counts_map.keys()) | set(gm_gym_counts_map.keys()):
                    total = session_gym_counts_map.get(gym_name, 0) + gm_gym_counts_map.get(gym_name, 0)
                    if total == 1:
                        distinct_gym_names.add(gym_name)

                if distinct_gym_names:
                    distinct_gym_filter = distinct_gym_names

        # Apply type filter to combined query if type is provided
        if type:
            count_query = select(func.count()).select_from(combined_query).where(combined_query.c.type == type)
        else:
            count_query = select(func.count()).select_from(combined_query)

        # Apply distinct client filter
        if distinct_client_filter:
            count_query = count_query.where(combined_query.c.client_name.in_(distinct_client_filter))

        # Apply distinct gym filter
        if distinct_gym_filter:
            count_query = count_query.where(combined_query.c.gym_name.in_(distinct_gym_filter))

        count_result = await db.execute(count_query)
        total = count_result.scalar() or 0

        # Early return if no results
        if total == 0:
            return {
                "success": True,
                "data": {
                    "purchases": [],
                    "pagination": {
                        "total": 0,
                        "page": page,
                        "limit": limit,
                        "totalPages": 0,
                        "hasNext": False,
                        "hasPrev": False
                    }
                }
            }

        # Build the final paginated query from the union result
        final_query = (
            select(
                combined_query.c.id,
                combined_query.c.client_id,
                combined_query.c.gym_id,
                combined_query.c.type,
                combined_query.c.days_total,
                combined_query.c.total_sessions,
                combined_query.c.scheduled_sessions,
                combined_query.c.amount,
                combined_query.c.purchased_at,
                combined_query.c.client_name,
                combined_query.c.gym_name,
                combined_query.c.gym_contact,
                combined_query.c.owner_contact,
                combined_query.c.owner_name,
                combined_query.c.gym_area,
                combined_query.c.client_contact,
                combined_query.c.platform
            )
            .select_from(combined_query)
            .order_by(combined_query.c.purchased_at.desc())
            .offset((page - 1) * limit)
            .limit(limit)
        )

        # Apply type filter if provided
        if type:
            final_query = final_query.where(combined_query.c.type == type)

        # Apply distinct client filter
        if distinct_client_filter:
            final_query = final_query.where(combined_query.c.client_name.in_(distinct_client_filter))

        # Apply distinct gym filter
        if distinct_gym_filter:
            final_query = final_query.where(combined_query.c.gym_name.in_(distinct_gym_filter))

        # Execute query (async, non-blocking)
        result = await db.execute(final_query)
        rows = result.all()

        # Process scheduled_sessions to get unique dates count
        def get_session_display(scheduled_sessions_json):
            """Process scheduled_sessions JSON to get 'X / Y sessions' format."""
            if not scheduled_sessions_json:
                return None
            try:
                # Parse JSON string if needed
                if isinstance(scheduled_sessions_json, str):
                    sessions = json.loads(scheduled_sessions_json)
                elif isinstance(scheduled_sessions_json, list):
                    sessions = scheduled_sessions_json
                else:
                    return None

                if not sessions:
                    return None

                total_sessions = len(sessions)
                unique_dates = len(set(s.get("date") for s in sessions if s.get("date")))
                return f"{unique_dates} / {total_sessions} sessions"
            except Exception:
                return None

        # Extract session schedule (date and start_time only)
        def get_session_schedule(scheduled_sessions_json):
            """Extract date and start_time from scheduled_sessions."""
            if not scheduled_sessions_json:
                return []
            try:
                # Parse JSON string if needed
                if isinstance(scheduled_sessions_json, str):
                    sessions = json.loads(scheduled_sessions_json)
                elif isinstance(scheduled_sessions_json, list):
                    sessions = scheduled_sessions_json
                else:
                    return []

                if not sessions:
                    return []

                return [
                    {"date": s.get("date"), "start_time": s.get("start_time")}
                    for s in sessions
                    if s.get("date") and s.get("start_time")
                ]
            except Exception:
                return []

        # Fetch scheduled dates and status for daily passes and sessions
        daily_pass_ids = [row.id for row in rows if row.type == "Daily Pass"]
        session_ids = [row.id for row in rows if row.type == "Session"]

        daily_pass_dates = {}
        daily_pass_statuses = {}
        if daily_pass_ids:
            dp_dates_query = (
                select(DailyPassDay.pass_id, DailyPassDay.scheduled_date, DailyPassDay.status)
                .where(DailyPassDay.pass_id.in_(daily_pass_ids))
                .order_by(DailyPassDay.scheduled_date)
            )
            dp_dates_result = await db.execute(dp_dates_query)
            for dp_row in dp_dates_result.all():
                pass_id = dp_row.pass_id
                date_str = dp_row.scheduled_date.isoformat() if dp_row.scheduled_date else None
                daily_pass_dates.setdefault(pass_id, []).append(date_str)
                # Collect all statuses for this pass
                if pass_id not in daily_pass_statuses:
                    daily_pass_statuses[pass_id] = []
                if dp_row.status:
                    daily_pass_statuses[pass_id].append(dp_row.status)

        session_dates = {}
        session_statuses = {}
        if session_ids:
            sb_dates_query = (
                select(SessionBookingDay.purchase_id, SessionBookingDay.booking_date, SessionBookingDay.status)
                .where(SessionBookingDay.purchase_id.in_(session_ids))
                .order_by(SessionBookingDay.booking_date)
            )
            sb_dates_result = await db.execute(sb_dates_query)
            sb_rows = sb_dates_result.all()
            print(f"[ALL_PURCHASES] session_ids: {session_ids[:5]}... (showing first 5)")
            print(f"[ALL_PURCHASES] SessionBookingDay query returned {len(sb_rows)} rows")
            for sb_row in sb_rows:
                purchase_id = sb_row.purchase_id
                date_str = sb_row.booking_date.isoformat() if sb_row.booking_date else None
                session_dates.setdefault(purchase_id, []).append(date_str)
                # Collect all statuses for this purchase (status is always present)
                if purchase_id not in session_statuses:
                    session_statuses[purchase_id] = []
                session_statuses[purchase_id].append(sb_row.status)
            print(f"[ALL_PURCHASES] session_statuses keys: {list(session_statuses.keys())[:5]}... (showing first 5)")

        # Helper function to determine overall status
        # Priority: canceled > missed > rescheduled > scheduled > attended
        def get_overall_status(statuses):
            if not statuses:
                return None

            # Map database status values to frontend values
            status_mapping = {
                "booked": "scheduled",
                "cancelled": "canceled",
                "attended": "attended",
                "no_show": "missed",
                "refunded": "canceled"
            }

            # Map all statuses to frontend values
            mapped_statuses = [status_mapping.get(s, s) for s in statuses]

            status_priority = {
                "canceled": 5,
                "missed": 4,
                "rescheduled": 3,
                "scheduled": 2,
                "attended": 1
            }
            # Return the status with highest priority
            return max(mapped_statuses, key=lambda s: status_priority.get(s, 0))

        # Format response
        purchases = []
        for row in rows:
            purchase = {
                "id": row.id,
                "client_id": row.client_id,
                "client_name": row.client_name or "N/A",
                "gym_id": row.gym_id,
                "gym_name": row.gym_name or "N/A",
                "amount": float(row.amount) if row.amount else 0.0,
                "purchased_at": row.purchased_at,
                "type": row.type,
                "gym_contact": row.gym_contact or None,
                "owner_contact": row.owner_contact or None,
                "owner_name": row.owner_name or "N/A",
                "gym_area": row.gym_area or "N/A",
                "client_contact": row.client_contact or None,
                "platform": row.platform or None
            }

            if row.type == "Daily Pass":
                purchase["days_total"] = row.days_total
                purchase["session_display"] = None
                purchase["session_schedule"] = []
                purchase["scheduled_date"] = daily_pass_dates.get(row.id, [])
                # Determine overall status for daily pass
                statuses = daily_pass_statuses.get(row.id, [])
                purchase["status"] = get_overall_status(statuses)
            else:  # Session
                purchase["days_total"] = None
                purchase["session_display"] = get_session_display(row.scheduled_sessions)
                purchase["session_schedule"] = get_session_schedule(row.scheduled_sessions)
                purchase["scheduled_date"] = session_dates.get(row.id, [])
                # Determine overall status for session
                statuses = session_statuses.get(row.id, [])
                purchase["status"] = get_overall_status(statuses)

            purchases.append(purchase)

        # Calculate pagination info
        total_pages = (total + limit - 1) // limit if total > 0 else 0
        has_next = page < total_pages
        has_prev = page > 1

        # Get distinct clients from sessions/daily passes (SQL aggregation)
        session_client_counts = (
            select(
                combined_query.c.client_name,
                func.count().label("booking_count")
            )
            .select_from(combined_query)
            .where(combined_query.c.client_name.isnot(None))
            .group_by(combined_query.c.client_name)
        )
        if type:
            session_client_counts = session_client_counts.where(combined_query.c.type == type)

        session_client_result = await db.execute(session_client_counts)
        session_client_counts_map = {row.client_name: row.booking_count for row in session_client_result.all()}

        # Get distinct gyms from sessions/daily passes (SQL aggregation)
        session_gym_counts = (
            select(
                combined_query.c.gym_name,
                func.count().label("booking_count")
            )
            .select_from(combined_query)
            .where(combined_query.c.gym_name.isnot(None))
            .group_by(combined_query.c.gym_name)
        )
        if type:
            session_gym_counts = session_gym_counts.where(combined_query.c.type == type)

        session_gym_result = await db.execute(session_gym_counts)
        session_gym_counts_map = {row.gym_name: row.booking_count for row in session_gym_result.all()}

        # Get distinct clients and gyms from gym memberships
        # Fetch base data first, then aggregate with SQL

        # Step 1: Fetch gym membership base data with date filters
        gym_membership_data_query = (
            select(
                Order.customer_id.label("client_id"),
                OrderItem.gym_id.label("gym_id"),
                Order.order_metadata.label("order_metadata")
            )
            .select_from(Payment)
            .join(Order, Order.id == Payment.order_id)
            .join(OrderItem, OrderItem.order_id == Order.id)
            .where(Payment.status == "captured")
            .where(Order.status == "paid")
            .where(OrderItem.gym_id.isnot(None))
            .where(OrderItem.gym_id != "1")
        )

        # Apply date filters
        if start_date_obj:
            gym_membership_data_query = gym_membership_data_query.where(func.date(Payment.created_at) >= start_date_obj)
        if end_date_obj:
            gym_membership_data_query = gym_membership_data_query.where(func.date(Payment.created_at) <= end_date_obj)

        gm_result = await db.execute(gym_membership_data_query)
        gm_rows = gm_result.all()

        # Step 2: Filter by metadata conditions (in Python - simpler than complex SQL JSON queries)
        valid_client_ids = []
        valid_gym_ids = []
        for row in gm_rows:
            metadata = row.order_metadata
            if not metadata or not isinstance(metadata, dict):
                continue

            # Condition checks
            condition1 = (
                metadata.get("audit") and isinstance(metadata.get("audit"), dict) and
                metadata["audit"].get("source") == "dailypass_checkout_api"
            )
            condition2 = (
                metadata.get("order_info") and isinstance(metadata.get("order_info"), dict) and
                metadata["order_info"].get("flow") == "unified_gym_membership_with_sub"
            )
            condition3 = (
                metadata.get("order_info") and isinstance(metadata.get("order_info"), dict) and
                metadata["order_info"].get("flow") == "unified_gym_membership_with_free_fittbot"
            )

            if condition1 or condition2 or condition3:
                if row.client_id:
                    try:
                        valid_client_ids.append(int(row.client_id))
                    except:
                        pass
                if row.gym_id and row.gym_id.isdigit():
                    valid_gym_ids.append(int(row.gym_id))

        # Step 3: Count using SQL aggregation with the filtered IDs
        gm_client_counts_map = {}
        if valid_client_ids:
            gm_client_query = (
                select(
                    Client.name,
                    func.count().label("gm_count")
                )
                .where(Client.client_id.in_(valid_client_ids))
                .where(Client.name.isnot(None))
                .group_by(Client.name)
            )
            gm_client_result = await db.execute(gm_client_query)
            gm_client_counts_map = {row.name: row.gm_count for row in gm_client_result.all()}

        gm_gym_counts_map = {}
        if valid_gym_ids:
            gm_gym_query = (
                select(
                    Gym.name,
                    func.count().label("gm_count")
                )
                .where(Gym.gym_id.in_(valid_gym_ids))
                .where(Gym.name.isnot(None))
                .group_by(Gym.name)
            )
            gm_gym_result = await db.execute(gm_gym_query)
            gm_gym_counts_map = {row.name: row.gm_count for row in gm_gym_result.all()}

        # Merge counts from sessions/daily passes and gym memberships
        final_distinct_clients = []
        all_client_names = set(session_client_counts_map.keys()) | set(gm_client_counts_map.keys())
        for client_name in all_client_names:
            total_count = (
                session_client_counts_map.get(client_name, 0) +
                gm_client_counts_map.get(client_name, 0)
            )
            if total_count == 1:
                final_distinct_clients.append(client_name)

        final_distinct_gyms = []
        all_gym_names = set(session_gym_counts_map.keys()) | set(gm_gym_counts_map.keys())
        for gym_name in all_gym_names:
            total_count = (
                session_gym_counts_map.get(gym_name, 0) +
                gm_gym_counts_map.get(gym_name, 0)
            )
            if total_count == 1:
                final_distinct_gyms.append(gym_name)

        return {
            "success": True,
            "data": {
                "purchases": purchases,
                "pagination": {
                    "total": total,
                    "page": page,
                    "limit": limit,
                    "totalPages": total_pages,
                    "hasNext": has_next,
                    "hasPrev": has_prev
                },
                "distinctClients": final_distinct_clients,
                "distinctGyms": final_distinct_gyms
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        import logging
        logging.error(f"Error in get_all_purchases: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail="An error occurred while fetching purchases"
        )


@router.get("/gmv-summary")
async def get_gmv_summary(
    start_date: Optional[str] = Query(None, description="Filter by start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="Filter by end date (YYYY-MM-DD)"),
    db: AsyncSession = Depends(get_async_db)
):
    """
    Lean GMV summary: returns purchase count and total revenue for
    Daily Pass and Fitness Class (Session) only.
    All aggregation is done at the DB level — no rows are loaded into Python.
    """
    try:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

        # ── Daily Pass ─────────────────────────────────────────────────────────
        dp_conditions = [DailyPass.gym_id != "1"]
        if start_date_obj:
            dp_conditions.append(func.date(DailyPass.created_at) >= start_date_obj)
        if end_date_obj:
            dp_conditions.append(func.date(DailyPass.created_at) <= end_date_obj)

        dp_stmt = (
            select(
                func.count(DailyPass.id).label("count"),
                func.coalesce(func.sum(Payment.amount_minor / 100.0), 0).label("total_revenue")
            )
            .select_from(DailyPass)
            .join(Gym, cast(DailyPass.gym_id, Integer) == Gym.gym_id)  # INNER JOIN — exclude orphaned records (no gym found)
            .outerjoin(Payment, DailyPass.payment_id == Payment.provider_payment_id)
            .where(*dp_conditions)
        )
        dp_result = await db.execute(dp_stmt)
        dp_row = dp_result.one()

        # ── Fitness Class (Session) ─────────────────────────────────────────────
        sess_conditions = [
            SessionPurchase.status == "paid",
            SessionPurchase.gym_id != 1
        ]
        if start_date_obj:
            sess_conditions.append(func.date(SessionPurchase.created_at) >= start_date_obj)
        if end_date_obj:
            sess_conditions.append(func.date(SessionPurchase.created_at) <= end_date_obj)

        sess_stmt = (
            select(
                func.count(SessionPurchase.id).label("count"),
                func.coalesce(func.sum(SessionPurchase.payable_rupees), 0).label("total_revenue")
            )
            .select_from(SessionPurchase)
            .join(Gym, SessionPurchase.gym_id == Gym.gym_id)  # INNER JOIN — exclude orphaned records (no gym found)
            .where(*sess_conditions)
        )

        sess_result = await db.execute(sess_stmt)
        sess_row = sess_result.one()

        # ── Nutrition Plans ─────────────────────────────────────────────────────
        # Same filter as /nutritionist-plans endpoint:
        # Payment.status = "captured" + payment_metadata.flow = "nutrition_purchase_googleplay"
        # Excluded contacts: internal/test numbers that should not be counted
        EXCLUDED_CONTACTS = ["7373675762", "9486987082", "8667458723"]

        nutri_conditions = [
            Payment.status == "captured",
            func.json_extract(Payment.payment_metadata, "$.flow") == "nutrition_purchase_googleplay"
        ]
        if start_date_obj:
            nutri_conditions.append(func.date(Payment.captured_at) >= start_date_obj)
        if end_date_obj:
            nutri_conditions.append(func.date(Payment.captured_at) <= end_date_obj)

        nutri_stmt = (
            select(
                func.count(Payment.id).label("count"),
                func.coalesce(func.sum(Payment.amount_minor / 100.0), 0).label("total_revenue")
            )
            .select_from(Payment)
            .outerjoin(Client, Payment.customer_id == Client.client_id)
            .where(*nutri_conditions)
            .where(~Client.contact.in_(EXCLUDED_CONTACTS))
        )
        nutri_result = await db.execute(nutri_stmt)
        nutri_row = nutri_result.one()

        # ── Gym Membership ──────────────────────────────────────────────────────
        # Mirrors the listing page validation:
        # 1) Metadata filter (3 known flows via JSON extract)
        # 2) EXISTS on OrderItem + JOIN Gym  → confirms gym actually exists in Gym table
        # 3) INNER JOIN Client              → confirms client actually exists in Client table
        # Without (2) and (3) the count is 34 (metadata-only); with them it matches the 4 real rows.
        gym_meta_cond = or_(
            func.json_unquote(func.json_extract(Order.order_metadata, "$.audit.source")) == "dailypass_checkout_api",
            func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "unified_gym_membership_with_sub",
            func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "unified_gym_membership_with_free_fittbot"
        )

        # EXISTS on OrderItem — JOIN Gym confirms gym physically exists in Gym table
        gym_exists = (
            select(1)
            .select_from(OrderItem)
            .join(Gym, Gym.gym_id == cast(OrderItem.gym_id, Integer))
            .where(
                OrderItem.order_id == Order.id,
                OrderItem.gym_id.isnot(None),
                OrderItem.gym_id != "",
                OrderItem.gym_id != "1"
            )
            .exists()
        )

        gym_conditions = [
            Payment.status == "captured",
            Order.status == "paid",
            Order.customer_id.isnot(None),
            gym_meta_cond,
            gym_exists
        ]
        if start_date_obj:
            gym_conditions.append(func.date(Payment.captured_at) >= start_date_obj)
        if end_date_obj:
            gym_conditions.append(func.date(Payment.captured_at) <= end_date_obj)

        # Subquery deduplicates orders; INNER JOIN Client confirms client exists in Client table
        gym_subq = (
            select(
                Order.id.label("order_id"),
                Order.gross_amount_minor.label("gross_amount_minor")
            )
            .select_from(Payment)
            .join(Order, Order.id == Payment.order_id)
            .join(Client, Client.client_id == cast(Order.customer_id, Integer))
            .where(*gym_conditions)
            .distinct()
            .subquery()
        )

        gym_stmt = (
            select(
                func.count(gym_subq.c.order_id).label("count"),
                func.coalesce(func.sum(gym_subq.c.gross_amount_minor) / 100.0, 0).label("total_revenue")
            )
            .select_from(gym_subq)
        )


        gym_result = await db.execute(gym_stmt)
        gym_row = gym_result.one()

        return {
            "success": True,
            "data": {
                "daily_pass": {
                    "count": dp_row.count,
                    "total_revenue": float(dp_row.total_revenue)
                },
                "session": {
                    "count": sess_row.count,
                    "total_revenue": float(sess_row.total_revenue)
                },
                "nutrition_plan": {
                    "count": nutri_row.count,
                    "total_revenue": float(nutri_row.total_revenue)
                },
                "gym_membership": {
                    "count": gym_row.count,
                    "total_revenue": float(gym_row.total_revenue)
                }
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        import logging
        logging.error(f"Error in get_gmv_summary: {str(e)}")
        raise HTTPException(status_code=500, detail="An error occurred while fetching GMV summary")


@router.get("/today-schedule")
async def get_today_schedule(
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, le=100, description="Items per page"),
    db: AsyncSession = Depends(get_async_db)
):
   
    try:
        today = date.today()

        # Build DailyPassDay subquery with type label
        daily_pass_query = (
            select(
                DailyPassDay.id,
                DailyPassDay.scheduled_date.label("booking_date"),
                DailyPassDay.status.label("day_status"),
                literal(None).label("checkin_at"),
                DailyPass.days_total,
                (Payment.amount_minor / 100.0).label("amount"),
                DailyPass.created_at.label("purchased_at"),
                Client.name.label("client_name"),
                Gym.name.label("gym_name"),
                Client.platform.label("platform"),
                literal("Daily Pass").label("type"),
            )
            .select_from(DailyPassDay)
            .join(DailyPass, DailyPassDay.pass_id == DailyPass.id)
            .outerjoin(Client, cast(DailyPassDay.client_id, Integer) == Client.client_id)
            .outerjoin(Gym, cast(DailyPassDay.gym_id, Integer) == Gym.gym_id)
            .outerjoin(Payment, DailyPass.payment_id == Payment.provider_payment_id)
            .where(DailyPassDay.scheduled_date == today)
            .where(DailyPassDay.gym_id != 1)  # Exclude gym_id = 1
        )

        # Build SessionBookingDay subquery with type label
        session_booking_query = (
            select(
                SessionBookingDay.id,
                SessionBookingDay.booking_date,
                SessionBookingDay.status.label("day_status"),
                SessionBookingDay.scanned_at.label("checkin_at"),
                literal(None).label("days_total"),
                SessionPurchase.payable_rupees.label("amount"),
                SessionPurchase.created_at.label("purchased_at"),
                Client.name.label("client_name"),
                Gym.name.label("gym_name"),
                Client.platform.label("platform"),
                literal("Session").label("type"),
            )
            .select_from(SessionBookingDay)
            .join(SessionPurchase, SessionBookingDay.purchase_id == SessionPurchase.id)
            .outerjoin(Client, SessionBookingDay.client_id == Client.client_id)
            .outerjoin(Gym, SessionBookingDay.gym_id == Gym.gym_id)
            .where(SessionBookingDay.booking_date == today)
            .where(SessionBookingDay.gym_id != 1)  # Exclude gym_id = 1
        )

        # Combine both queries with UNION ALL
        combined_query = union_all(daily_pass_query, session_booking_query).alias("combined_schedule")

        # Count total records for pagination
        count_query = select(func.count()).select_from(combined_query)
        count_result = await db.execute(count_query)
        total = count_result.scalar() or 0

        # Early return if no results
        if total == 0:
            return {
                "success": True,
                "data": {
                    "schedule": [],
                    "date": today.isoformat(),
                    "pagination": {
                        "total": 0,
                        "page": page,
                        "limit": limit,
                        "totalPages": 0,
                        "hasNext": False,
                        "hasPrev": False
                    }
                }
            }

        # Build the final paginated query from the union result
        final_query = (
            select(
                combined_query.c.id,
                combined_query.c.booking_date,
                combined_query.c.day_status,
                combined_query.c.checkin_at,
                combined_query.c.days_total,
                combined_query.c.amount,
                combined_query.c.purchased_at,
                combined_query.c.client_name,
                combined_query.c.gym_name,
                combined_query.c.platform,
                combined_query.c.type
            )
            .select_from(combined_query)
            .order_by(combined_query.c.booking_date.desc(), combined_query.c.purchased_at.desc())
            .offset((page - 1) * limit)
            .limit(limit)
        )

        result = await db.execute(final_query)
        rows = result.all()

        # Format response
        schedule = [
            {
                "id": row.id,
                "client_name": row.client_name or "N/A",
                "gym_name": row.gym_name or "N/A",
                "scheduled_date": row.booking_date.isoformat() if row.booking_date else None,
                "status": row.day_status,
                "checkin_at": row.checkin_at.isoformat() if row.checkin_at else None,
                "days_total": row.days_total,
                "amount": float(row.amount) if row.amount else 0.0,
                "purchased_at": row.purchased_at,
                "type": row.type,
                "platform": row.platform
            }
            for row in rows
        ]

        # Calculate pagination info
        total_pages = (total + limit - 1) // limit if total > 0 else 0
        has_next = page < total_pages
        has_prev = page > 1

        return {
            "success": True,
            "data": {
                "schedule": schedule,
                "date": today.isoformat(),
                "pagination": {
                    "total": total,
                    "page": page,
                    "limit": limit,
                    "totalPages": total_pages,
                    "hasNext": has_next,
                    "hasPrev": has_prev
                }
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        import logging
        logging.error(f"Error in get_today_schedule: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail="An error occurred while fetching today's schedule"
        )


@router.get("/gym-memberships")
async def get_gym_memberships(
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, le=100, description="Items per page"),
    search: Optional[str] = Query(None, description="Search by client name, contact, or gym name"),
    start_date: Optional[str] = Query(None, description="Filter by start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="Filter by end date (YYYY-MM-DD)"),
    distinct_clients: Optional[bool] = Query(False, description="Show only distinct clients (1 booking across all types)"),
    distinct_gyms: Optional[bool] = Query(False, description="Show only distinct gyms (1 booking across all types)"),
    db: AsyncSession = Depends(get_async_db)
):
    """
    Get all gym memberships with pagination.
    Using same logic as Financials/Revenue Analytics APIs (Order-based approach).
    Excludes gym_id = 1, rows where client_id is not in clients table,
    and rows where gym_id is not in gyms table.
    """
    try:
        # Parse dates if provided
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
        # Fetch all gym memberships using Order-based approach (same as Financials API)
        gym_membership_stmt = (
            select(Payment, Order)
            .join(Order, Order.id == Payment.order_id)
            .where(Payment.status == "captured")
            .where(Order.status == "paid")
        )
        gym_membership_result = await db.execute(gym_membership_stmt)
        all_payments = gym_membership_result.all()

        # Collect order IDs to fetch gym info from order_items (exclude gym_id = 1)
        order_ids = [row.Order.id for row in all_payments]

        # Fetch order items to get gym_ids (exclude gym_id = 1)
        order_gym_mapping = {}
        order_gym_id_mapping = {}  # Store actual gym_id values
        if order_ids:
            order_items_stmt = (
                select(OrderItem)
                .where(OrderItem.order_id.in_(order_ids))
                .where(OrderItem.gym_id.isnot(None))
                .where(OrderItem.gym_id != "1")
            )
            order_items_result = await db.execute(order_items_stmt)
            order_items = order_items_result.scalars().all()

            # Create mapping from order_id to gym_id
            for item in order_items:
                if item.gym_id and item.gym_id.strip() and item.gym_id.isdigit():
                    order_gym_mapping[item.order_id] = int(item.gym_id)
                    order_gym_id_mapping[item.order_id] = item.gym_id

        # Filter by metadata conditions and collect valid orders
        valid_orders = []
        for row in all_payments:
            payment = row.Payment
            order = row.Order

            # Check order_metadata for specific conditions (same as Financials/Revenue Analytics)
            if not order.order_metadata or not isinstance(order.order_metadata, dict):
                continue

            metadata = order.order_metadata

            # Condition 1: audit.source = "dailypass_checkout_api"
            condition1 = False
            if metadata.get("audit") and isinstance(metadata.get("audit"), dict):
                if metadata["audit"].get("source") == "dailypass_checkout_api":
                    condition1 = True

            # Condition 2: order_info.flow = "unified_gym_membership_with_sub"
            condition2 = False
            if metadata.get("order_info") and isinstance(metadata.get("order_info"), dict):
                if metadata["order_info"].get("flow") == "unified_gym_membership_with_sub":
                    condition2 = True

            # Condition 3: order_info.flow = "unified_gym_membership_with_free_fittbot"
            condition3 = False
            if metadata.get("order_info") and isinstance(metadata.get("order_info"), dict):
                if metadata["order_info"].get("flow") == "unified_gym_membership_with_free_fittbot":
                    condition3 = True

            # Only include if any condition matches AND order has valid gym_id (not gym_id = 1)
            if not (condition1 or condition2 or condition3):
                continue

            if order.id not in order_gym_mapping:
                continue

            valid_orders.append({
                "order": order,
                "payment": payment,
                "gym_id": order_gym_id_mapping[order.id]
            })

        # Apply date filters if provided
        if start_date_obj or end_date_obj:
            filtered_orders = []
            for item in valid_orders:
                purchase_date = item["order"].created_at.date() if hasattr(item["order"].created_at, 'date') else item["order"].created_at
                if start_date_obj and purchase_date < start_date_obj:
                    continue
                if end_date_obj and purchase_date > end_date_obj:
                    continue
                filtered_orders.append(item)
            valid_orders = filtered_orders

        # ── Validate ALL orders (client + gym must exist) to get the TRUE count ──
        # Previously total was set before this check, causing count mismatch (e.g 34 vs 4)
        validated_orders = []
        for item in valid_orders:
            order = item["order"]
            gym_id_str = item["gym_id"]

            if not order.customer_id:
                continue

            # Client must exist
            client_stmt = select(Client).where(Client.client_id == int(order.customer_id))
            client_result = await db.execute(client_stmt)
            client = client_result.scalar_one_or_none()
            if not client:
                continue

            # Gym must exist
            if not gym_id_str or not gym_id_str.isdigit():
                continue
            gym_stmt = select(Gym).where(Gym.gym_id == int(gym_id_str))
            gym_result = await db.execute(gym_stmt)
            gym = gym_result.scalar_one_or_none()
            if not gym:
                continue

            # Cache resolved objects to avoid re-fetching during pagination
            item["_client"] = client
            item["_gym"] = gym
            validated_orders.append(item)

        # TRUE count — only orders that physically have a matching client + gym
        total = len(validated_orders)

        # Early return if no results
        if total == 0:
            return {
                "success": True,
                "data": {
                    "memberships": [],
                    "pagination": {
                        "total": 0,
                        "page": page,
                        "limit": limit,
                        "totalPages": 0,
                        "hasNext": False,
                        "hasPrev": False
                    }
                }
            }

        # Sort and paginate against the validated list
        validated_orders.sort(key=lambda x: x["order"].created_at, reverse=True)
        start_idx = (page - 1) * limit
        end_idx = start_idx + limit
        paginated_orders = validated_orders[start_idx:end_idx]

        # Build membership rows — client + gym already resolved above (no extra DB queries)
        memberships = []
        for item in paginated_orders:
            order = item["order"]
            client = item["_client"]
            gym = item["_gym"]

            client_name = client.name or "N/A"
            client_contact = client.contact
            gym_name = gym.name or "N/A"
            gym_contact = gym.contact_number
            gym_area = gym.area or "N/A"

            # Fetch gym owner details
            owner_name = "N/A"
            owner_contact = None
            if gym.owner_id:
                owner_stmt = select(GymOwner).where(GymOwner.owner_id == gym.owner_id)
                owner_result = await db.execute(owner_stmt)
                owner = owner_result.scalar_one_or_none()
                if owner:
                    owner_name = owner.name or "N/A"
                    owner_contact = owner.contact_number

            memberships.append({
                "id": order.id,
                "client_name": client_name,
                "gym_name": gym_name,
                "type": "gym_membership",
                "amount": float(order.gross_amount_minor / 100) if order.gross_amount_minor else 0.0,
                "purchased_at": order.created_at,
                "gym_contact": gym_contact,
                "owner_contact": owner_contact,
                "owner_name": owner_name,
                "gym_area": gym_area,
                "client_contact": client_contact,
                "platform": client.platform
            })

        # Apply search filter if search term is provided
        if search and search.strip():
            search_term_lower = search.lower().strip()
            memberships = [
                m for m in memberships
                if (m.get("client_name") and search_term_lower in m["client_name"].lower()) or
                   (m.get("client_contact") and search_term_lower in str(m["client_contact"])) or
                   (m.get("gym_name") and search_term_lower in m["gym_name"].lower())
            ]
            # Update total after filtering
            total = len(memberships)

        total_pages = (total + limit - 1) // limit if total > 0 else 0
        has_next = page < total_pages
        has_prev = page > 1

        # Calculate distinct clients and gyms across ALL booking types
        # Using same logic as all-purchases endpoint

        # Step 1: Get counts from Sessions/Daily Pass
        # Build DailyPass and SessionPurchase queries
        daily_pass_query = (
            select(
                DailyPass.id.label("id"),
                DailyPass.client_id.label("client_id"),
                DailyPass.gym_id.label("gym_id"),
                literal("Daily Pass").label("type"),
                Client.name.label("client_name"),
                Gym.name.label("gym_name")
            )
            .select_from(DailyPass)
            .join(Gym, cast(DailyPass.gym_id, Integer) == Gym.gym_id)
            .outerjoin(Client, cast(DailyPass.client_id, Integer) == Client.client_id)
            .where(DailyPass.gym_id != "1")
        )

        session_purchase_query = (
            select(
                SessionPurchase.id.label("id"),
                SessionPurchase.client_id.label("client_id"),
                SessionPurchase.gym_id.label("gym_id"),
                literal("Session").label("type"),
                Client.name.label("client_name"),
                Gym.name.label("gym_name")
            )
            .select_from(SessionPurchase)
            .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
            .outerjoin(Client, SessionPurchase.client_id == Client.client_id)
            .where(SessionPurchase.status == "paid")
            .where(SessionPurchase.gym_id != 1)
        )

        # Combine queries
        session_combined = union_all(daily_pass_query, session_purchase_query).alias("session_combined")

        # Count clients from sessions/daily passes
        session_client_counts = (
            select(
                session_combined.c.client_name,
                func.count().label("booking_count")
            )
            .select_from(session_combined)
            .where(session_combined.c.client_name.isnot(None))
            .group_by(session_combined.c.client_name)
        )
        session_client_result = await db.execute(session_client_counts)
        session_client_counts_map = {row.client_name: row.booking_count for row in session_client_result.all()}

        # Count gyms from sessions/daily passes
        session_gym_counts = (
            select(
                session_combined.c.gym_name,
                func.count().label("booking_count")
            )
            .select_from(session_combined)
            .where(session_combined.c.gym_name.isnot(None))
            .group_by(session_combined.c.gym_name)
        )
        session_gym_result = await db.execute(session_gym_counts)
        session_gym_counts_map = {row.gym_name: row.booking_count for row in session_gym_result.all()}

        # Step 2: Get counts from Gym Memberships (using valid_orders collected earlier)
        gm_valid_client_ids = []
        gm_valid_gym_ids = []
        for item in valid_orders:
            client_id = item["order"].customer_id
            gym_id = item["gym_id"]
            if client_id:
                try:
                    gm_valid_client_ids.append(int(client_id))
                except:
                    pass
            if gym_id and gym_id.isdigit():
                gm_valid_gym_ids.append(int(gym_id))

        # Count gym memberships per client
        gm_client_counts_map = {}
        if gm_valid_client_ids:
            gm_client_query = (
                select(
                    Client.name,
                    func.count().label("gm_count")
                )
                .where(Client.client_id.in_(gm_valid_client_ids))
                .where(Client.name.isnot(None))
                .group_by(Client.name)
            )
            gm_client_result = await db.execute(gm_client_query)
            gm_client_counts_map = {row.name: row.gm_count for row in gm_client_result.all()}

        # Count gym memberships per gym
        gm_gym_counts_map = {}
        if gm_valid_gym_ids:
            gm_gym_query = (
                select(
                    Gym.name,
                    func.count().label("gm_count")
                )
                .where(Gym.gym_id.in_(gm_valid_gym_ids))
                .where(Gym.name.isnot(None))
                .group_by(Gym.name)
            )
            gm_gym_result = await db.execute(gm_gym_query)
            gm_gym_counts_map = {row.name: row.gm_count for row in gm_gym_result.all()}

        # Step 3: Merge counts from all three types
        final_distinct_clients = []
        all_client_names = set(session_client_counts_map.keys()) | set(gm_client_counts_map.keys())
        for client_name in all_client_names:
            total_count = (
                session_client_counts_map.get(client_name, 0) +
                gm_client_counts_map.get(client_name, 0)
            )
            if total_count == 1:
                final_distinct_clients.append(client_name)

        final_distinct_gyms = []
        all_gym_names = set(session_gym_counts_map.keys()) | set(gm_gym_counts_map.keys())
        for gym_name in all_gym_names:
            total_count = (
                session_gym_counts_map.get(gym_name, 0) +
                gm_gym_counts_map.get(gym_name, 0)
            )
            if total_count == 1:
                final_distinct_gyms.append(gym_name)

        # Apply distinct filters (must be after distinct calculation)
        if distinct_clients or distinct_gyms:
            memberships = [
                m for m in memberships
                if (not distinct_clients or (m.get("client_name") and m["client_name"] in final_distinct_clients)) and
                   (not distinct_gyms or (m.get("gym_name") and m["gym_name"] in final_distinct_gyms))
            ]
            # Update total and pagination after filtering
            total = len(memberships)
            total_pages = (total + limit - 1) // limit if total > 0 else 0
            has_next = page < total_pages
            has_prev = page > 1

        return {
            "success": True,
            "data": {
                "memberships": memberships,
                "pagination": {
                    "total": total,
                    "page": page,
                    "limit": limit,
                    "totalPages": total_pages,
                    "hasNext": has_next,
                    "hasPrev": has_prev
                },
                "distinctClients": final_distinct_clients,
                "distinctGyms": final_distinct_gyms
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        import logging
        logging.error(f"Error in get_gym_memberships: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail="An error occurred while fetching gym memberships"
        )


@router.get("/export-purchases")
async def export_purchases(
    search: Optional[str] = Query(None, description="Search by client or gym name"),
    type: Optional[str] = Query(None, description="Filter by type: 'Session' or 'Daily Pass'"),
    start_date: Optional[str] = Query(None, description="Filter by start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="Filter by end date (YYYY-MM-DD)"),
    distinct_clients: Optional[bool] = Query(False, description="Show only distinct clients (1 booking across all types)"),
    distinct_gyms: Optional[bool] = Query(False, description="Show only distinct gyms (1 booking across all types)"),
    db: AsyncSession = Depends(get_async_db)
):
    """
    Export all purchases to CSV file.
    Returns all purchases (without pagination) for export purposes.
    """
    try:
        search_pattern = f"%{search}%" if search else None

        # Parse dates if provided
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None

        # Build DailyPass subquery with type label
        daily_pass_query = (
            select(
                DailyPass.id.label("id"),
                DailyPass.client_id.label("client_id"),
                DailyPass.gym_id.label("gym_id"),
                literal("Daily Pass").label("type"),
                DailyPass.days_total.label("days_total"),
                literal(None).label("total_sessions"),
                literal(None).label("scheduled_sessions"),
                (Payment.amount_minor / 100.0).label("amount"),
                DailyPass.created_at.label("purchased_at"),
                Client.name.label("client_name"),
                Gym.name.label("gym_name"),
                Gym.contact_number.label("gym_contact"),
                GymOwner.contact_number.label("owner_contact")
            )
            .select_from(DailyPass)
            .outerjoin(Client, cast(DailyPass.client_id, Integer) == Client.client_id)
            .outerjoin(Gym, cast(DailyPass.gym_id, Integer) == Gym.gym_id)
            .outerjoin(GymOwner, Gym.owner_id == GymOwner.owner_id)
            .outerjoin(Payment, DailyPass.payment_id == Payment.provider_payment_id)
        )

        # Build SessionPurchase subquery with type label (only paid status)
        session_purchase_query = (
            select(
                SessionPurchase.id.label("id"),
                SessionPurchase.client_id.label("client_id"),
                SessionPurchase.gym_id.label("gym_id"),
                literal("Session").label("type"),
                literal(None).label("days_total"),
                SessionPurchase.sessions_count.label("total_sessions"),
                SessionPurchase.scheduled_sessions.label("scheduled_sessions"),
                SessionPurchase.payable_rupees.label("amount"),
                SessionPurchase.created_at.label("purchased_at"),
                Client.name.label("client_name"),
                Gym.name.label("gym_name"),
                Gym.contact_number.label("gym_contact"),
                GymOwner.contact_number.label("owner_contact")
            )
            .select_from(SessionPurchase)
            .outerjoin(Client, SessionPurchase.client_id == Client.client_id)
            .outerjoin(Gym, SessionPurchase.gym_id == Gym.gym_id)
            .outerjoin(GymOwner, Gym.owner_id == GymOwner.owner_id)
            .where(SessionPurchase.status == "paid")
        )

        # Apply search filters to both subqueries if search is provided
        if search:
            daily_pass_query = daily_pass_query.where(
                or_(
                    Client.name.ilike(search_pattern),
                    Client.contact.ilike(search_pattern),
                    Gym.name.ilike(search_pattern)
                )
            )
            session_purchase_query = session_purchase_query.where(
                or_(
                    Client.name.ilike(search_pattern),
                    Client.contact.ilike(search_pattern),
                    Gym.name.ilike(search_pattern)
                )
            )

        # Apply date filters to both subqueries if provided
        if start_date_obj:
            daily_pass_query = daily_pass_query.where(func.date(DailyPass.created_at) >= start_date_obj)
            session_purchase_query = session_purchase_query.where(func.date(SessionPurchase.created_at) >= start_date_obj)
        if end_date_obj:
            daily_pass_query = daily_pass_query.where(func.date(DailyPass.created_at) <= end_date_obj)
            session_purchase_query = session_purchase_query.where(func.date(SessionPurchase.created_at) <= end_date_obj)

        # Combine both queries with UNION ALL
        combined_query = union_all(daily_pass_query, session_purchase_query).alias("combined_purchases")

        # Build distinct filtering subqueries (if distinct filter is requested)
        distinct_client_filter = None
        distinct_gym_filter = None

        if distinct_clients or distinct_gyms:
            # Get gym memberships base data for distinct calculation
            gym_membership_data_query = (
                select(
                    Order.customer_id.label("client_id"),
                    OrderItem.gym_id.label("gym_id"),
                    Order.order_metadata.label("order_metadata")
                )
                .select_from(Payment)
                .join(Order, Order.id == Payment.order_id)
                .join(OrderItem, OrderItem.order_id == Order.id)
                .where(Payment.status == "captured")
                .where(Order.status == "paid")
                .where(OrderItem.gym_id.isnot(None))
                .where(OrderItem.gym_id != "1")
            )

            # Apply date filters to gym memberships
            if start_date_obj:
                gym_membership_data_query = gym_membership_data_query.where(func.date(Payment.created_at) >= start_date_obj)
            if end_date_obj:
                gym_membership_data_query = gym_membership_data_query.where(func.date(Payment.created_at) <= end_date_obj)

            gm_result = await db.execute(gym_membership_data_query)
            gm_rows = gm_result.all()

            # Filter by metadata conditions
            valid_client_ids = []
            valid_gym_ids = []
            for row in gm_rows:
                metadata = row.order_metadata
                if not metadata or not isinstance(metadata, dict):
                    continue

                condition1 = (
                    metadata.get("audit") and isinstance(metadata.get("audit"), dict) and
                    metadata["audit"].get("source") == "dailypass_checkout_api"
                )
                condition2 = (
                    metadata.get("order_info") and isinstance(metadata.get("order_info"), dict) and
                    metadata["order_info"].get("flow") == "unified_gym_membership_with_sub"
                )
                condition3 = (
                    metadata.get("order_info") and isinstance(metadata.get("order_info"), dict) and
                    metadata["order_info"].get("flow") == "unified_gym_membership_with_free_fittbot"
                )

                if condition1 or condition2 or condition3:
                    if row.client_id:
                        try:
                            valid_client_ids.append(int(row.client_id))
                        except:
                            pass
                    if row.gym_id and row.gym_id.isdigit():
                        valid_gym_ids.append(int(row.gym_id))

            # Build distinct client filter (clients with exactly 1 booking total)
            if distinct_clients:
                # Count bookings per client from sessions/daily passes
                session_client_counts = (
                    select(
                        combined_query.c.client_name,
                        func.count().label("booking_count")
                    )
                    .select_from(combined_query)
                    .where(combined_query.c.client_name.isnot(None))
                    .group_by(combined_query.c.client_name)
                )
                if type:
                    session_client_counts = session_client_counts.where(combined_query.c.type == type)
                session_client_result = await db.execute(session_client_counts)
                session_client_counts_map = {row.client_name: row.booking_count for row in session_client_result.all()}

                # Count gym memberships per client
                gm_client_counts_map = {}
                if valid_client_ids:
                    gm_client_query = (
                        select(Client.name, func.count().label("gm_count"))
                        .where(Client.client_id.in_(valid_client_ids))
                        .where(Client.name.isnot(None))
                        .group_by(Client.name)
                    )
                    gm_client_result = await db.execute(gm_client_query)
                    gm_client_counts_map = {row.name: row.gm_count for row in gm_client_result.all()}

                # Find distinct clients (total count == 1)
                distinct_client_names = set()
                for client_name in set(session_client_counts_map.keys()) | set(gm_client_counts_map.keys()):
                    total = session_client_counts_map.get(client_name, 0) + gm_client_counts_map.get(client_name, 0)
                    if total == 1:
                        distinct_client_names.add(client_name)

                if distinct_client_names:
                    distinct_client_filter = distinct_client_names

            # Build distinct gym filter (gyms with exactly 1 booking total)
            if distinct_gyms:
                # Count bookings per gym from sessions/daily passes
                session_gym_counts = (
                    select(
                        combined_query.c.gym_name,
                        func.count().label("booking_count")
                    )
                    .select_from(combined_query)
                    .where(combined_query.c.gym_name.isnot(None))
                    .group_by(combined_query.c.gym_name)
                )
                if type:
                    session_gym_counts = session_gym_counts.where(combined_query.c.type == type)
                session_gym_result = await db.execute(session_gym_counts)
                session_gym_counts_map = {row.gym_name: row.booking_count for row in session_gym_result.all()}

                # Count gym memberships per gym
                gm_gym_counts_map = {}
                if valid_gym_ids:
                    gm_gym_query = (
                        select(Gym.name, func.count().label("gm_count"))
                        .where(Gym.gym_id.in_(valid_gym_ids))
                        .where(Gym.name.isnot(None))
                        .group_by(Gym.name)
                    )
                    gm_gym_result = await db.execute(gm_gym_query)
                    gm_gym_counts_map = {row.name: row.gm_count for row in gm_gym_result.all()}

                # Find distinct gyms (total count == 1)
                distinct_gym_names = set()
                for gym_name in set(session_gym_counts_map.keys()) | set(gm_gym_counts_map.keys()):
                    total = session_gym_counts_map.get(gym_name, 0) + gm_gym_counts_map.get(gym_name, 0)
                    if total == 1:
                        distinct_gym_names.add(gym_name)

                if distinct_gym_names:
                    distinct_gym_filter = distinct_gym_names

        # Build the final query from the union result (no pagination for export)
        final_query = (
            select(
                combined_query.c.id,
                combined_query.c.client_id,
                combined_query.c.gym_id,
                combined_query.c.type,
                combined_query.c.days_total,
                combined_query.c.total_sessions,
                combined_query.c.scheduled_sessions,
                combined_query.c.amount,
                combined_query.c.purchased_at,
                combined_query.c.client_name,
                combined_query.c.gym_name,
                combined_query.c.gym_contact,
                combined_query.c.owner_contact
            )
            .select_from(combined_query)
            .order_by(combined_query.c.purchased_at.desc())
        )

        # Apply type filter if provided
        if type:
            final_query = final_query.where(combined_query.c.type == type)

        # Apply distinct client filter
        if distinct_client_filter:
            final_query = final_query.where(combined_query.c.client_name.in_(distinct_client_filter))

        # Apply distinct gym filter
        if distinct_gym_filter:
            final_query = final_query.where(combined_query.c.gym_name.in_(distinct_gym_filter))

        # Execute query
        result = await db.execute(final_query)
        rows = result.all()

        # Process scheduled_sessions to get display format
        def get_session_display(scheduled_sessions_json):
            """Process scheduled_sessions JSON to get 'X / Y sessions' format."""
            if not scheduled_sessions_json:
                return "N/A"
            try:
                if isinstance(scheduled_sessions_json, str):
                    sessions = json.loads(scheduled_sessions_json)
                elif isinstance(scheduled_sessions_json, list):
                    sessions = scheduled_sessions_json
                else:
                    return "N/A"

                if not sessions:
                    return "N/A"

                total_sessions = len(sessions)
                unique_dates = len(set(s.get("date") for s in sessions if s.get("date")))
                return f"{unique_dates} / {total_sessions} sessions"
            except Exception:
                return "N/A"

        # Format purchase date
        def format_date(date_obj):
            if date_obj:
                return date_obj.strftime("%d-%b-%Y")
            return "N/A"

        # Format amount
        def format_amount(amount):
            return f"₹{float(amount):.2f}" if amount else "₹0.00"

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Purchases"

        # Write header
        headers = ["Client Name", "Gym Name", "Type", "Days / Sessions", "Amount", "Purchased At"]
        ws.append(headers)

        # Style the header row
        header_fill = PatternFill(start_color="FF5757", end_color="FF5757", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        # Write data rows
        for row in rows:
            if row.type == "Daily Pass":
                days_sessions = str(row.days_total) if row.days_total else "N/A"
            else:  # Session
                days_sessions = get_session_display(row.scheduled_sessions)

            ws.append([
                row.client_name or "N/A",
                row.gym_name or "N/A",
                row.type,
                days_sessions,
                format_amount(row.amount),
                format_date(row.purchased_at)
            ])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"purchases_export_{timestamp}.xlsx"

        # Return Excel file as streaming response
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        import logging
        logging.error(f"Error in export_purchases: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail="An error occurred while exporting purchases"
        )


@router.get("/export-today-schedule")
async def export_today_schedule(
    db: AsyncSession = Depends(get_async_db)
):
    """
    Export today's schedule to Excel file.
    Returns all daily pass days and session bookings scheduled for today (without pagination).
    """
    try:
        today = date.today()

        # Build DailyPassDay subquery with type label
        daily_pass_query = (
            select(
                DailyPassDay.id,
                DailyPassDay.scheduled_date.label("booking_date"),
                DailyPassDay.status.label("day_status"),
                literal(None).label("checkin_at"),
                DailyPass.days_total,
                (Payment.amount_minor / 100.0).label("amount"),
                DailyPass.created_at.label("purchased_at"),
                Client.name.label("client_name"),
                Gym.name.label("gym_name"),
                Client.platform.label("platform"),
                literal("Daily Pass").label("type"),
            )
            .select_from(DailyPassDay)
            .join(DailyPass, DailyPassDay.pass_id == DailyPass.id)
            .outerjoin(Client, cast(DailyPassDay.client_id, Integer) == Client.client_id)
            .outerjoin(Gym, cast(DailyPassDay.gym_id, Integer) == Gym.gym_id)
            .outerjoin(Payment, DailyPass.payment_id == Payment.provider_payment_id)
            .where(DailyPassDay.scheduled_date == today)
            .where(DailyPassDay.gym_id != 1)  # Exclude gym_id = 1
        )

        # Build SessionBookingDay subquery with type label
        session_booking_query = (
            select(
                SessionBookingDay.id,
                SessionBookingDay.booking_date,
                SessionBookingDay.status.label("day_status"),
                SessionBookingDay.scanned_at.label("checkin_at"),
                literal(None).label("days_total"),
                SessionPurchase.payable_rupees.label("amount"),
                SessionPurchase.created_at.label("purchased_at"),
                Client.name.label("client_name"),
                Gym.name.label("gym_name"),
                Client.platform.label("platform"),
                literal("Session").label("type"),
            )
            .select_from(SessionBookingDay)
            .join(SessionPurchase, SessionBookingDay.purchase_id == SessionPurchase.id)
            .outerjoin(Client, SessionBookingDay.client_id == Client.client_id)
            .outerjoin(Gym, SessionBookingDay.gym_id == Gym.gym_id)
            .where(SessionBookingDay.booking_date == today)
            .where(SessionBookingDay.gym_id != 1)  # Exclude gym_id = 1
        )

        # Combine both queries with UNION ALL
        combined_query = union_all(daily_pass_query, session_booking_query).alias("combined_schedule")

        # Build the final query from the union result (no pagination for export)
        final_query = (
            select(
                combined_query.c.id,
                combined_query.c.booking_date,
                combined_query.c.day_status,
                combined_query.c.checkin_at,
                combined_query.c.days_total,
                combined_query.c.amount,
                combined_query.c.purchased_at,
                combined_query.c.client_name,
                combined_query.c.gym_name,
                combined_query.c.type
            )
            .select_from(combined_query)
            .order_by(combined_query.c.booking_date.desc(), combined_query.c.purchased_at.desc())
        )

        # Execute query
        result = await db.execute(final_query)
        rows = result.all()

        # Format functions
        def format_date(date_obj):
            if date_obj:
                return date_obj.strftime("%d-%b-%Y")
            return "N/A"

        def format_datetime(date_obj):
            if date_obj:
                return date_obj.strftime("%d-%b-%Y %H:%M")
            return "N/A"

        def format_amount(amount):
            return f"₹{float(amount):.2f}" if amount else "₹0.00"

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Today's Schedule"

        # Write header
        headers = ["Client Name", "Gym Name", "Type", "Scheduled Date", "Status", "Check-in At", "Amount", "Purchased At", "Platform"]
        ws.append(headers)

        # Style the header row
        header_fill = PatternFill(start_color="FF5757", end_color="FF5757", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        # Write data rows
        for row in rows:
            ws.append([
                row.client_name or "N/A",
                row.gym_name or "N/A",
                row.type,
                format_date(row.booking_date),
                row.day_status or "N/A",
                format_datetime(row.checkin_at),
                format_amount(row.amount),
                format_datetime(row.purchased_at),
                (row.platform or "N/A").capitalize() if row.platform else "N/A"
            ])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"today_schedule_{timestamp}.xlsx"

        # Return Excel file as streaming response
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        import logging
        logging.error(f"Error in export_today_schedule: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail="An error occurred while exporting today's schedule"
        )


@router.get("/export-gym-memberships")
async def export_gym_memberships(
    search: Optional[str] = Query(None, description="Search by client or gym name"),
    start_date: Optional[str] = Query(None, description="Filter by start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="Filter by end date (YYYY-MM-DD)"),
    distinct_clients: Optional[bool] = Query(False, description="Show only distinct clients (1 booking across all types)"),
    distinct_gyms: Optional[bool] = Query(False, description="Show only distinct gyms (1 booking across all types)"),
    db: AsyncSession = Depends(get_async_db)
):
    """
    Export all gym memberships to Excel file.
    Using same logic as Financials/Revenue Analytics APIs (Order-based approach).
    Returns all memberships (without pagination) for export purposes.
    Excludes gym_id = 1, rows where client_id is not in clients table,
    and rows where gym_id is not in gyms table.
    Supports filtering by search, date range, and distinct clients/gyms.
    """
    try:
        # Fetch all gym memberships using Order-based approach (same as Financials API)
        gym_membership_stmt = (
            select(Payment, Order)
            .join(Order, Order.id == Payment.order_id)
            .where(Payment.status == "captured")
            .where(Order.status == "paid")
        )

        # Apply date filter
        if start_date:
            try:
                start_datetime = datetime.strptime(start_date, "%Y-%m-%d")
                gym_membership_stmt = gym_membership_stmt.where(Order.created_at >= start_datetime)
            except ValueError:
                pass

        if end_date:
            try:
                end_datetime = datetime.strptime(end_date, "%Y-%m-%d")
                # Include the entire end date
                end_datetime = end_datetime + timedelta(days=1)
                gym_membership_stmt = gym_membership_stmt.where(Order.created_at < end_datetime)
            except ValueError:
                pass

        gym_membership_result = await db.execute(gym_membership_stmt)
        all_payments = gym_membership_result.all()

        # Collect order IDs to fetch gym info from order_items (exclude gym_id = 1)
        order_ids = [row.Order.id for row in all_payments]

        # Fetch order items to get gym_ids (exclude gym_id = 1)
        order_gym_mapping = {}
        order_gym_id_mapping = {}  # Store actual gym_id values
        if order_ids:
            order_items_stmt = (
                select(OrderItem)
                .where(OrderItem.order_id.in_(order_ids))
                .where(OrderItem.gym_id.isnot(None))
                .where(OrderItem.gym_id != "1")
            )
            order_items_result = await db.execute(order_items_stmt)
            order_items = order_items_result.scalars().all()

            # Create mapping from order_id to gym_id
            for item in order_items:
                if item.gym_id and item.gym_id.strip() and item.gym_id.isdigit():
                    order_gym_mapping[item.order_id] = int(item.gym_id)
                    order_gym_id_mapping[item.order_id] = item.gym_id

        # Filter by metadata conditions and collect valid orders
        valid_orders = []
        for row in all_payments:
            payment = row.Payment
            order = row.Order

            # Check order_metadata for specific conditions (same as Financials/Revenue Analytics)
            if not order.order_metadata or not isinstance(order.order_metadata, dict):
                continue

            metadata = order.order_metadata

            # Condition 1: audit.source = "dailypass_checkout_api"
            condition1 = False
            if metadata.get("audit") and isinstance(metadata.get("audit"), dict):
                if metadata["audit"].get("source") == "dailypass_checkout_api":
                    condition1 = True

            # Condition 2: order_info.flow = "unified_gym_membership_with_sub"
            condition2 = False
            if metadata.get("order_info") and isinstance(metadata.get("order_info"), dict):
                if metadata["order_info"].get("flow") == "unified_gym_membership_with_sub":
                    condition2 = True

            # Condition 3: order_info.flow = "unified_gym_membership_with_free_fittbot"
            condition3 = False
            if metadata.get("order_info") and isinstance(metadata.get("order_info"), dict):
                if metadata["order_info"].get("flow") == "unified_gym_membership_with_free_fittbot":
                    condition3 = True

            # Only include if any condition matches AND order has valid gym_id (not gym_id = 1)
            if not (condition1 or condition2 or condition3):
                continue

            if order.id not in order_gym_mapping:
                continue

            valid_orders.append({
                "order": order,
                "gym_id": order_gym_id_mapping[order.id]
            })

        # Sort by purchased_at descending
        valid_orders.sort(key=lambda x: x["order"].created_at, reverse=True)

        # Format functions
        def format_date(date_obj):
            if date_obj:
                return date_obj.strftime("%d-%b-%Y")
            return "N/A"

        def format_amount(amount):
            return f"₹{float(amount):.2f}" if amount else "₹0.00"

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Gym Memberships"

        # Write header
        headers = ["Client Name", "Gym Name", "Type", "Amount", "Purchased At", "Platform"]
        ws.append(headers)

        # Style the header row
        header_fill = PatternFill(start_color="FF5757", end_color="FF5757", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        # Calculate distinct clients and gyms if filters are enabled
        distinct_clients_set = set()
        distinct_gyms_set = set()

        if distinct_clients or distinct_gyms:
            # Count sessions (fitness classes) by client/gym
            session_client_counts = {}
            session_gym_counts = {}

            session_orders_stmt = (
                select(Order)
                .join(Payment, Payment.order_id == Order.id)
                .where(Payment.status == "captured")
                .where(Order.status == "paid")
                .where(Order.type == "Session")
            )

            session_result = await db.execute(session_orders_stmt)
            session_orders = session_result.scalars().all()

            for order in session_orders:
                if order.customer_id:
                    session_client_counts[str(order.customer_id)] = session_client_counts.get(str(order.customer_id), 0) + 1
                if order.gym_id:
                    session_gym_counts[str(order.gym_id)] = session_gym_counts.get(str(order.gym_id), 0) + 1

            # Count daily passes by client/gym
            daily_pass_client_counts = {}
            daily_pass_gym_counts = {}

            daily_pass_orders_stmt = (
                select(Order)
                .join(Payment, Payment.order_id == Order.id)
                .where(Payment.status == "captured")
                .where(Order.status == "paid")
                .where(Order.type == "Daily Pass")
            )

            daily_pass_result = await db.execute(daily_pass_orders_stmt)
            daily_pass_orders = daily_pass_result.scalars().all()

            for order in daily_pass_orders:
                if order.customer_id:
                    daily_pass_client_counts[str(order.customer_id)] = daily_pass_client_counts.get(str(order.customer_id), 0) + 1
                if order.gym_id:
                    daily_pass_gym_counts[str(order.gym_id)] = daily_pass_gym_counts.get(str(order.gym_id), 0) + 1

            # Collect gym membership clients/gyms from valid_orders
            gym_membership_client_ids = set()
            gym_membership_gym_ids = set()

            for item in valid_orders:
                if item["order"].customer_id:
                    gym_membership_client_ids.add(str(item["order"].customer_id))
                if item["gym_id"]:
                    gym_membership_gym_ids.add(item["gym_id"])

            # Calculate total bookings and identify distinct clients
            all_client_ids = set(session_client_counts.keys()) | set(daily_pass_client_counts.keys()) | gym_membership_client_ids

            for client_id in all_client_ids:
                session_count = session_client_counts.get(client_id, 0)
                daily_pass_count = daily_pass_client_counts.get(client_id, 0)
                gym_membership_count = 1 if client_id in gym_membership_client_ids else 0
                total = session_count + daily_pass_count + gym_membership_count

                if total == 1:
                    distinct_clients_set.add(client_id)

            # Calculate total bookings and identify distinct gyms
            all_gym_ids = set(session_gym_counts.keys()) | set(daily_pass_gym_counts.keys()) | gym_membership_gym_ids

            for gym_id in all_gym_ids:
                session_count = session_gym_counts.get(gym_id, 0)
                daily_pass_count = daily_pass_gym_counts.get(gym_id, 0)
                gym_membership_count = 1 if gym_id in gym_membership_gym_ids else 0
                total = session_count + daily_pass_count + gym_membership_count

                if total == 1:
                    distinct_gyms_set.add(gym_id)

        # Write data rows
        for item in valid_orders:
            order = item["order"]
            gym_id_str = item["gym_id"]

            # Fetch client details
            # Skip row if client not found in clients table
            if not order.customer_id:
                continue

            try:
                client_stmt = select(Client).where(Client.client_id == int(order.customer_id))
                client_result = await db.execute(client_stmt)
                client = client_result.scalar_one_or_none()
                if not client:
                    # Skip this row if client not found in clients table
                    continue
                client_name = client.name or "N/A"
                client_platform = client.platform or "N/A"
            except:
                continue

            # Fetch gym details
            # Skip row if gym not found in gyms table
            if not gym_id_str or not gym_id_str.isdigit():
                continue

            try:
                gym_stmt = select(Gym).where(Gym.gym_id == int(gym_id_str))
                gym_result = await db.execute(gym_stmt)
                gym = gym_result.scalar_one_or_none()
                if not gym:
                    # Skip this row if gym not found in gyms table
                    continue
                gym_name = gym.name or "N/A"
            except:
                continue

            # Apply search filter
            if search:
                search_lower = search.lower()
                if search_lower not in (client_name or "").lower() and search_lower not in (gym_name or "").lower():
                    continue

            # Apply distinct clients filter
            if distinct_clients and order.customer_id:
                if str(order.customer_id) not in distinct_clients_set:
                    continue

            # Apply distinct gyms filter
            if distinct_gyms and gym_id_str:
                if gym_id_str not in distinct_gyms_set:
                    continue

            ws.append([
                client_name,
                gym_name,
                "Gym Membership",
                format_amount((order.gross_amount_minor / 100) if order.gross_amount_minor else 0),
                format_date(order.created_at),
                client_platform
            ])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"gym_memberships_{timestamp}.xlsx"

        # Return Excel file as streaming response
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        import logging
        logging.error(f"Error in export_gym_memberships: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail="An error occurred while exporting gym memberships"
        )


@router.get("/nutritionist-plans")
async def get_nutritionist_plans(
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, le=100, description="Items per page"),
    search: Optional[str] = Query(None, description="Search by client name, email, or mobile"),
    sort_order: str = Query("desc", description="Sort order for purchase date"),
    db: AsyncSession = Depends(get_async_db)
):
    """
    Get list of nutritionist plan purchases (one-time Google Play purchases).
    Fetches from payments.payments table where payment_metadata['flow'] = 'nutrition_purchase_googleplay'.

    Optimized with:
    - Fully async execution
    - Backend pagination (only fetches required page data)
    - Filter-aware querying (all filters applied in SQL)
    - No N+1 query patterns
    """
    try:
        import math

        # Contacts to always exclude (internal/test accounts)
        EXCLUDED_CONTACTS = ["7373675762", "9486987082", "8667458723"]

        # Build base query for Payment table — always join Client to apply exclusion
        base_payment_query = (
            select(Payment.customer_id)
            .select_from(Payment)
            .outerjoin(Client, Payment.customer_id == Client.client_id)
            .where(Payment.status == "captured")
            .where(func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_purchase_googleplay')
            .where(~Client.contact.in_(EXCLUDED_CONTACTS))  # Exclude internal/test contacts
        )

        # Apply search filter to base query
        if search:
            search_term = f"%{search.lower()}%"
            base_payment_query = base_payment_query.where(
                or_(
                    func.lower(Client.name).like(search_term),
                    Client.contact.like(search_term)
                )
            )

        # Get unique users count (matches home page logic)
        unique_count_subquery = base_payment_query.subquery()
        unique_count_stmt = select(func.count(distinct(unique_count_subquery.c.customer_id)))
        unique_count_result = await db.execute(unique_count_stmt)
        unique_users_count = unique_count_result.scalar() or 0

        # Build main query with all columns
        query = (
            select(
                Payment.id.label("purchase_id"),
                Payment.customer_id.label("customer_id"),
                Payment.captured_at.label("purchased_at"),
                Payment.amount_minor.label("amount_minor"),
                Payment.payment_metadata.label("payment_metadata"),
                Client.client_id.label("client_id"),
                Client.name.label("client_name"),
                Client.contact.label("client_contact"),
                Client.created_at.label("client_created_at")
            )
            .select_from(Payment)
            .outerjoin(Client, Payment.customer_id == Client.client_id)
            .where(Payment.status == "captured")
            .where(func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_purchase_googleplay')
            .where(~Client.contact.in_(EXCLUDED_CONTACTS))  # Exclude internal/test contacts
        )

        # Apply search filter
        if search:
            search_term = f"%{search.lower()}%"
            query = query.where(
                or_(
                    func.lower(Client.name).like(search_term),
                    Client.contact.like(search_term)
                )
            )

        # Get total count
        count_subquery = query.subquery()
        count_stmt = select(func.count()).select_from(count_subquery)
        count_result = await db.execute(count_stmt)
        total_count = count_result.scalar() or 0

        if total_count == 0:
            return {
                "success": True,
                "data": {
                    "users": [],
                    "total": 0,
                    "unique_users": 0,
                    "page": page,
                    "limit": limit,
                    "totalPages": 0,
                    "hasNext": False,
                    "hasPrev": False
                },
                "message": "Nutritionist plans fetched successfully"
            }

        # Apply sorting in SQL
        if sort_order == "asc":
            query = query.order_by(asc(Payment.captured_at))
        else:
            query = query.order_by(desc(Payment.captured_at))

        # Apply pagination in SQL (backend pagination)
        offset = (page - 1) * limit
        query = query.offset(offset).limit(limit)

        # Execute query (single query - no N+1)
        result = await db.execute(query)
        rows = result.all()

        # Format response with required columns
        nutritionist_plan_users = []
        for row in rows:
            # Extract booked_date from payment_metadata
            booked_date = None
            if row.payment_metadata and isinstance(row.payment_metadata, dict):
                booking_id = row.payment_metadata.get("booking_id")
                if booking_id:
                    # Format as date if it's a timestamp
                    try:
                        booked_date = datetime.fromisoformat(booking_id.replace('Z', '+00:00'))
                        booked_date = booked_date.strftime("%Y-%m-%d")
                    except:
                        booked_date = str(booking_id)

            # Format dates and amounts
            purchased_date = row.purchased_at.strftime("%Y-%m-%d") if row.purchased_at else "N/A"
            amount_rupees = float(row.amount_minor / 100) if row.amount_minor else 0.0

            user_data = {
                "id": row.purchase_id,
                "customer_id": row.customer_id,
                "client_id": row.client_id,
                "client_name": row.client_name or "N/A",
                "mobile": row.client_contact or "N/A",
                "gym_name": "N/A",
                "gym_location": "N/A",
                "purchased_date": purchased_date,
                "booked_date": booked_date or "N/A",
                "amount": amount_rupees
            }
            nutritionist_plan_users.append(user_data)

        # Calculate pagination info
        total_pages = math.ceil(total_count / limit)
        has_next = page < total_pages
        has_prev = page > 1

        return {
            "success": True,
            "data": {
                "users": nutritionist_plan_users,
                "total": total_count,
                "unique_users": unique_users_count,
                "page": page,
                "limit": limit,
                "totalPages": total_pages,
                "hasNext": has_next,
                "hasPrev": has_prev
            },
            "message": "Nutritionist plans fetched successfully"
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching nutritionist plans: {str(e)}"
        )


@router.get("/export-nutritionist-plans")
async def export_nutritionist_plans(
    db: AsyncSession = Depends(get_async_db)
):
    """
    Export nutritionist plans to Excel file.
    Returns all nutritionist plan purchases (one-time Google Play purchases) without pagination.
    """
    try:
        # Build optimized query for nutritionist plan purchases
        query = (
            select(
                Payment.id.label("purchase_id"),
                Payment.customer_id.label("customer_id"),
                Payment.captured_at.label("purchased_at"),
                Payment.amount_minor.label("amount_minor"),
                Payment.payment_metadata.label("payment_metadata"),
                Client.client_id.label("client_id"),
                Client.name.label("client_name"),
                Client.contact.label("client_contact"),
                Client.created_at.label("client_created_at")
            )
            .select_from(Payment)
            .outerjoin(Client, Payment.customer_id == Client.client_id)
            .where(Payment.status == "captured")
            .where(func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_purchase_googleplay')
            .order_by(desc(Payment.captured_at))
        )

        # Execute query
        result = await db.execute(query)
        rows = result.all()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Nutritionist Plans"

        # Write NEW header - matching the requirements
        headers = ["Client Name", "Contact", "Gym Name", "Purchased Date", "Booked Date", "Amount"]
        ws.append(headers)

        # Style the header row
        header_fill = PatternFill(start_color="FF5757", end_color="FF5757", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        # Helper function to format date
        def format_date(date_obj):
            if not date_obj:
                return "N/A"
            return date_obj.strftime("%Y-%m-%d")

        # Helper function to extract booked_date from metadata
        def get_booked_date(payment_metadata):
            if not payment_metadata or not isinstance(payment_metadata, dict):
                return "N/A"
            booking_id = payment_metadata.get("booking_id")
            if booking_id:
                # Try to parse as datetime
                try:
                    booked_date = datetime.fromisoformat(booking_id.replace('Z', '+00:00'))
                    return booked_date.strftime("%Y-%m-%d")
                except:
                    return str(booking_id)
            return "N/A"

        # Write data rows
        for row in rows:
            # Extract booked date from payment_metadata
            booked_date = get_booked_date(row.payment_metadata)

            # Format amount in rupees
            amount_rupees = float(row.amount_minor / 100) if row.amount_minor else 0.0

            ws.append([
                row.client_name or "N/A",
                row.client_contact or "N/A",
                "N/A",  # Gym Name - not available for nutritionist plans
                format_date(row.purchased_at),
                booked_date,
                f"{amount_rupees:.2f}"
            ])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"nutritionist_plans_{timestamp}.xlsx"

        # Return Excel file as streaming response
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        import logging
        logging.error(f"Error in export_nutritionist_plans: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail="An error occurred while exporting nutritionist plans"
        )
