# Backend Implementation Guide for Dashboard APIs
from fastapi import APIRouter, Depends, HTTPException, FastAPI, Query
from pydantic import BaseModel
from typing import Optional
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import datetime, timedelta
from sqlalchemy import func, and_, select, distinct, or_, text, String, desc, cast, Integer
from sqlalchemy.sql.expression import literal

from sqlalchemy.orm import Session
import io
import pandas as pd
from fastapi.responses import StreamingResponse
from openpyxl.styles import PatternFill, Font, Alignment

# Import centralized revenue service for consistency
from app.fittbot_admin_api.revenue_service import (
    get_revenue_breakdown,
    get_detailed_revenue_with_breakdowns,
    paise_to_rupees
)

from app.models.fittbot_models import (
    Client, Gym, ClientToken, OwnerToken, GymOwner, RewardInterest, RewardProgramOptIn,
    SessionSetting, GymPlans, GymStudiosPic, GymOnboardingPics,
    SessionBookingDay, SessionBooking, ClassSession, ClientFittbotAccess, ActiveUser, SessionPurchase
)
from app.models.adminmodels import(TicketAssignment, Employees, Admins)
from app.models.async_database import get_async_db
from app.fittbot_api.v1.payments.models.subscriptions import Subscription
from app.fittbot_api.v1.payments.models.catalog import CatalogProduct
from app.fittbot_api.v1.payments.models.payments import Payment
from app.fittbot_api.v1.payments.models.orders import Order, OrderItem
from app.models.dailypass_models import DailyPassPricing, get_dailypass_session, DailyPass

router = APIRouter(prefix="/api/admin/dashboard", tags=["AdminDashboard"])

async def get_monthly_active_users(db: AsyncSession, today: datetime.date) -> int:
    """
    Get count of distinct client_ids from active_users table
    where created_at is within the last 30 days
    Only include client_ids that have at least 2 rows with different dates
    Each qualifying client_id counts as 1
    """
    try:
        # Calculate date 30 days ago
        thirty_days_ago = today - timedelta(days=30)
        end_date_inclusive = today + timedelta(days=1)

        # Subquery: Find client_ids that have at least 2 distinct dates in the last 30 days
        subquery = select(ActiveUser.client_id).where(
            and_(
                ActiveUser.created_at >= thirty_days_ago,
                ActiveUser.created_at < end_date_inclusive
            )
        ).group_by(
            ActiveUser.client_id
        ).having(
            func.count(func.distinct(func.date(ActiveUser.created_at))) >= 2
        )

        # Main query: Count distinct client_ids (each qualifying client = 1)
        stmt = select(func.coalesce(func.count(distinct(ActiveUser.client_id)), 0)).where(
            ActiveUser.client_id.in_(subquery)
        )

        result = await db.execute(stmt)
        count = result.scalar()
        return int(count) if count is not None else 0
    except Exception as e:
        print(f"[MONTHLY_ACTIVE_USERS] Error fetching monthly active users: {e}")
        import traceback
        traceback.print_exc()
        return 0

async def get_total_paying_users(db: AsyncSession) -> int:
    """
    Get count of distinct customer_id from payments table
    Counts unique users who have made at least one payment
    """
    try:
        # Count distinct customer_id from payments table
        stmt = select(func.count(distinct(Payment.customer_id)))
        result = await db.execute(stmt)
        count = result.scalar()
        return count if count is not None else 0
    except Exception as e:
        print(f"[TOTAL_PAYING_USERS] Error fetching total paying users: {e}")
        return 0

async def get_fittbot_metrics(db: AsyncSession, filter_type='month'):

    today = datetime.now().date()

    # Total users based on filter
    stmt = select(func.count()).select_from(Client).filter(
        func.date(Client.created_at) == today
    )
    result = await db.execute(stmt)
    total_users_today = result.scalar() or 0

    # Yesterday users
    stmt = select(func.count()).select_from(Client).filter(
        func.date(Client.created_at) == today - timedelta(days=1)
    )
    result = await db.execute(stmt)
    total_users_yesterday = result.scalar() or 0

    stmt = select(func.count()).select_from(Client).filter(
        Client.created_at >= today - timedelta(days=7),
        Client.created_at < today + timedelta(days=1)
    )
    result = await db.execute(stmt)
    total_users_week = result.scalar() or 0

    stmt = select(func.count()).select_from(Client).filter(
        Client.created_at >= today - timedelta(days=30),
        Client.created_at < today + timedelta(days=1)
    )
    result = await db.execute(stmt)
    total_users_month = result.scalar() or 0

    stmt = select(func.count()).select_from(Client)
    result = await db.execute(stmt)
    total_users_overall = result.scalar() or 0

    # Revenue calculation
    revenue_data = await calculate_revenue(db, today)

    # Subscribed users calculation
    subscribed_users_data = await calculate_subscribed_users(db, today)

    # Monthly revenue trends for last 6 months
    monthly_revenue_trends = await calculate_monthly_revenue_trends(db, today)

    # Monthly active users (last 30 days)
    monthly_active_users = await get_monthly_active_users(db, today)

    # Total paying users (distinct customer_id from payments table)
    total_paying_users = await get_total_paying_users(db)

    return {
        "totalUsers": {
            "today": total_users_today,
            "yesterday": total_users_yesterday,
            "week": total_users_week,
            "month": total_users_month,
            "overall": total_users_overall
        },
        "revenue": revenue_data,
        "subscribedUsers": subscribed_users_data,
        "monthlyActiveUsers": monthly_active_users,
        "totalPayingUsers": total_paying_users,
        "monthlyRevenueTrends": monthly_revenue_trends
    }

async def get_fittbot_metrics_custom(db: AsyncSession, start_date: str, end_date: str):
    """
    Get Fittbot metrics for custom date range
    """
    try:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    if start_date_obj > end_date_obj:
        raise HTTPException(status_code=400, detail="Start date must be before end date")

    # Total users in custom range
    stmt = select(func.count()).select_from(Client).filter(
        func.date(Client.created_at) >= start_date_obj,
        func.date(Client.created_at) <= end_date_obj
    )
    result = await db.execute(stmt)
    total_users_custom = result.scalar() or 0

    # Revenue in custom range - using centralized revenue service
    revenue_data = await get_revenue_breakdown(
        db=db,
        start_date=start_date_obj,
        end_date=end_date_obj,
        exclude_gym_id_one=True
    )
    total_revenue = revenue_data.total_revenue

    # Subscribed users in custom range (subscriptions that started in custom range and are still active)
    now = datetime.now()
    stmt = select(func.count(distinct(Subscription.customer_id))).filter(
        Subscription.provider.in_(['razorpay_pg', 'google_play']),
        Subscription.status != 'pending',
        func.date(Subscription.active_from) >= start_date_obj,
        func.date(Subscription.active_from) <= end_date_obj,
        Subscription.active_until >= now
    )
    result = await db.execute(stmt)
    subscribed_custom = result.scalar() or 0

    # Monthly active users - use end_date to get active users up to the custom range end date
    monthly_active_users_custom = await get_monthly_active_users(db, end_date_obj)

    # Total paying users (distinct customer_id from payments table)
    total_paying_users_custom = await get_total_paying_users(db)

    return {
        "totalUsers": {
            "today": 0,
            "week": 0,
            "month": 0,
            "overall": 0,
            "custom": total_users_custom
        },
        "revenue": {
            "today": "₹0",
            "week": "₹0",
            "month": "₹0",
            "overall": "₹0",
            "custom": f"₹{paise_to_rupees(total_revenue):,.0f}"
        },
        "subscribedUsers": {
            "today": 0,
            "week": 0,
            "month": 0,
            "overall": 0,
            "custom": subscribed_custom
        },
        "monthlyActiveUsers": monthly_active_users_custom,
        "totalPayingUsers": total_paying_users_custom,
        "monthlyRevenueTrends": []
    }

async def get_business_metrics_custom(db: AsyncSession, start_date: str, end_date: str):
    """
    Get Business metrics for custom date range
    """
    try:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    if start_date_obj > end_date_obj:
        raise HTTPException(status_code=400, detail="Start date must be before end date")

    # Gym Owners in custom range
    stmt = select(func.count()).select_from(GymOwner).filter(
        func.date(GymOwner.created_at) >= start_date_obj,
        func.date(GymOwner.created_at) <= end_date_obj
    )
    result = await db.execute(stmt)
    gym_owners_custom = result.scalar() or 0

    # Gyms in custom range
    stmt = select(func.count()).select_from(Gym).filter(
        func.date(Gym.created_at) >= start_date_obj,
        func.date(Gym.created_at) <= end_date_obj
    )
    result = await db.execute(stmt)
    gyms_custom = result.scalar() or 0

    # Daily pass enabled gyms (total, not affected by date range)
    stmt = select(func.count()).select_from(Gym).filter(
        Gym.dailypass == True
    )
    result = await db.execute(stmt)
    daily_pass_gyms = result.scalar() or 0

    # Verified gyms (type contains "green") - total, not affected by date range
    # Count total gyms (excluding "dummy" type)
    stmt = select(func.count()).select_from(Gym).filter(
        or_(
            Gym.type.notlike("%dummy%"),
            Gym.type.is_(None)
        )
    )
    result = await db.execute(stmt)
    total_gyms_count = result.scalar() or 0

    # Count gyms where type contains "green"
    stmt = select(func.count()).select_from(Gym).filter(
        Gym.type.like("%green%")
    )
    result = await db.execute(stmt)
    verified_gyms_count = result.scalar() or 0

    # Unverified gyms (type contains "hold" or "red")
    stmt = select(func.count()).select_from(Gym).filter(
        or_(
            Gym.type.like("%hold%"),
            Gym.type.like("%red%")
        )
    )
    result = await db.execute(stmt)
    unverified_gyms_count = result.scalar() or 0

    # Count gyms where type = 'red'
    stmt = select(func.count()).select_from(Gym).filter(
        Gym.type == 'red'
    )
    result = await db.execute(stmt)
    red_gyms_count = result.scalar() or 0

    # Count gyms where type = 'hold'
    stmt = select(func.count()).select_from(Gym).filter(
        Gym.type == 'hold'
    )
    result = await db.execute(stmt)
    hold_gyms_count = result.scalar() or 0

    return {
        "gymOwners": {
            "today": 0,
            "week": 0,
            "month": 0,
            "overall": 0,
            "custom": gym_owners_custom
        },
        "gyms": {
            "today": 0,
            "week": 0,
            "month": 0,
            "overall": 0,
            "custom": gyms_custom
        },
        "dailyPassGyms": daily_pass_gyms,
        "verifiedGyms": {
            "verified": verified_gyms_count,
            "total": total_gyms_count
        },
        "unverifiedGyms": unverified_gyms_count,
        "unverifiedSplitup": {
            "red": red_gyms_count,
            "hold": hold_gyms_count
        }
    }

async def calculate_revenue(db: AsyncSession, today):
    """
    Calculate revenue for different time periods using centralized revenue service.

    Returns formatted revenue strings in rupees.
    """
    # Calculate revenue for different time periods using centralized service
    revenue_today = await get_revenue_breakdown(
        db=db,
        start_date=today,
        end_date=today,
        exclude_gym_id_one=True
    )

    revenue_week = await get_revenue_breakdown(
        db=db,
        start_date=today - timedelta(days=7),
        end_date=today,
        exclude_gym_id_one=True
    )

    revenue_month = await get_revenue_breakdown(
        db=db,
        start_date=today - timedelta(days=30),
        end_date=today,
        exclude_gym_id_one=True
    )

    # Overall revenue (all time)
    revenue_overall = await get_revenue_breakdown(
        db=db,
        start_date=datetime(2020, 1, 1).date(),
        end_date=today,
        exclude_gym_id_one=True
    )

    # Convert from paise to rupees and format
    return {
        "today": f"₹{paise_to_rupees(revenue_today.total_revenue):,.0f}",
        "week": f"₹{paise_to_rupees(revenue_week.total_revenue):,.0f}",
        "month": f"₹{paise_to_rupees(revenue_month.total_revenue):,.0f}",
        "overall": f"₹{paise_to_rupees(revenue_overall.total_revenue):,.0f}"
    }

async def calculate_subscribed_users(db: AsyncSession, today):
 
    now = datetime.now()

    # Today's subscribed users (subscriptions that started today and are still active)
    stmt = select(func.count(distinct(Subscription.customer_id))).filter(
        Subscription.provider.in_(['razorpay_pg', 'google_play']),
        Subscription.status != 'pending',
        func.date(Subscription.active_from) == today,
        Subscription.active_until >= now
    )
    result = await db.execute(stmt)
    subscribed_today = result.scalar() or 0

    # This week's subscribed users (subscriptions that started this week and are still active)
    stmt = select(func.count(distinct(Subscription.customer_id))).filter(
        Subscription.provider.in_(['razorpay_pg', 'google_play']),
        Subscription.status != 'pending',
        Subscription.active_from >= today - timedelta(days=7),
        Subscription.active_from < today + timedelta(days=1),
        Subscription.active_until >= now
    )
    result = await db.execute(stmt)
    subscribed_week = result.scalar() or 0

    # This month's subscribed users (subscriptions that started this month and are still active)
    stmt = select(func.count(distinct(Subscription.customer_id))).filter(
        Subscription.provider.in_(['razorpay_pg', 'google_play']),
        Subscription.status != 'pending',
        Subscription.active_from >= today - timedelta(days=30),
        Subscription.active_from < today + timedelta(days=1),
        Subscription.active_until >= now
    )
    result = await db.execute(stmt)
    subscribed_month = result.scalar() or 0

    # Overall subscribed users (all active subscriptions)
    stmt = select(func.count(distinct(Subscription.customer_id))).filter(
        Subscription.provider.in_(['razorpay_pg', 'google_play']),
        Subscription.status != 'pending',
        Subscription.active_until >= now
    )
    result = await db.execute(stmt)
    subscribed_overall = result.scalar() or 0

    return {
        "today": subscribed_today,
        "week": subscribed_week,
        "month": subscribed_month,
        "overall": subscribed_overall
    }

async def calculate_monthly_revenue_trends(db: AsyncSession, today):
   
    import calendar

    monthly_data = []
    current_date = datetime.now()

    # Get last 6 months
    for i in range(5, -1, -1):
        # Calculate the month
        target_month = current_date.month - i
        target_year = current_date.year

        # Handle year rollover
        while target_month <= 0:
            target_month += 12
            target_year -= 1

        # Get first and last day of the month
        first_day = datetime(target_year, target_month, 1).date()
        last_day = datetime(target_year, target_month, calendar.monthrange(target_year, target_month)[1]).date()

        # Calculate revenue for this month
        stmt = select(
            func.coalesce(func.sum(CatalogProduct.base_amount_minor), 0)
        ).select_from(Subscription).join(
            CatalogProduct, Subscription.product_id == CatalogProduct.sku
        ).filter(
            Subscription.provider.in_(['razorpay_pg', 'google_play']),
            Subscription.status != 'pending',
            Subscription.active_from >= first_day,
            Subscription.active_from <= last_day
        )
        result = await db.execute(stmt)
        monthly_revenue = result.scalar() or 0

        # Convert from paise to thousands (1000 rupees = 100000 paise)
        revenue_in_thousands = monthly_revenue / 100000

        # Get month abbreviation
        month_abbr = calendar.month_abbr[target_month]

        monthly_data.append({
            "month": month_abbr,
            "revenue": round(revenue_in_thousands, 2)
        })

    return monthly_data

async def get_plans_metrics(db: AsyncSession):
    """
    Get Plans metrics (nutritionist plan purchases data)

    NEW LOGIC: Counts payments from payments.payments table
    where payment_metadata['flow'] = 'nutrition_purchase_googleplay'
    """
    # EXCLUDED_CONTACTS for test/internal accounts
    EXCLUDED_CONTACTS = ["7373675762", "9486987082", "8667458723", "9840633149"]

    # Nutritionist Plans - count payments where flow = 'nutrition_purchase_googleplay'
    # Exclude internal/test contacts to match the inner page logic
    stmt = (
        select(func.count(distinct(Payment.customer_id)))
        .select_from(Payment)
        .outerjoin(Client, Payment.customer_id == Client.client_id)
        .where(
            Payment.status == "captured",
            func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_purchase_googleplay',
            ~Client.contact.in_(EXCLUDED_CONTACTS)
        )
    )
    result = await db.execute(stmt)
    nutritionist_total = result.scalar() or 0

    # For one-time purchases, all are counted as "total"
    # Gold/Platinum/Diamond breakdown is not applicable for one-time purchases
    return {
        "fittbotSubscriptions": {
            "total": nutritionist_total,
            "gold": nutritionist_total,  # All plans count toward each category
            "platinum": nutritionist_total,
            "diamond": nutritionist_total
        }
    }

async def get_business_metrics(db: AsyncSession, filter_type='month'):
    """
    Get Business metrics (gym-related data)
    """
    today = datetime.now().date()

    # Gym Owners count based on filter
    stmt = select(func.count()).select_from(GymOwner).filter(
        func.date(GymOwner.created_at) == today
    )
    result = await db.execute(stmt)
    gym_owners_today = result.scalar() or 0

    stmt = select(func.count()).select_from(GymOwner).filter(
        GymOwner.created_at >= today - timedelta(days=7),
        GymOwner.created_at < today + timedelta(days=1)
    )
    result = await db.execute(stmt)
    gym_owners_week = result.scalar() or 0

    stmt = select(func.count()).select_from(GymOwner).filter(
        GymOwner.created_at >= today - timedelta(days=30),
        GymOwner.created_at < today + timedelta(days=1)
    )
    result = await db.execute(stmt)
    gym_owners_month = result.scalar() or 0

    stmt = select(func.count()).select_from(GymOwner)
    result = await db.execute(stmt)
    gym_owners_overall = result.scalar() or 0

    # Gyms count based on filter
    stmt = select(func.count()).select_from(Gym).filter(
        func.date(Gym.created_at) == today
    )
    result = await db.execute(stmt)
    gyms_today = result.scalar() or 0

    stmt = select(func.count()).select_from(Gym).filter(
        Gym.created_at >= today - timedelta(days=7),
        Gym.created_at < today + timedelta(days=1)
    )
    result = await db.execute(stmt)
    gyms_week = result.scalar() or 0

    stmt = select(func.count()).select_from(Gym).filter(
        Gym.created_at >= today - timedelta(days=30),
        Gym.created_at < today + timedelta(days=1)
    )
    result = await db.execute(stmt)
    gyms_month = result.scalar() or 0

    stmt = select(func.count()).select_from(Gym)
    result = await db.execute(stmt)
    gyms_overall = result.scalar() or 0

    # Daily pass enabled gyms (no time filter, just total count where dailypass = 1/True)
    stmt = select(func.count()).select_from(Gym).filter(
        Gym.dailypass == True
    )
    result = await db.execute(stmt)
    daily_pass_gyms = result.scalar() or 0

    # Verified gyms (type contains "green")
    # Count total gyms (excluding "dummy" type)
    stmt = select(func.count()).select_from(Gym).filter(
        or_(
            Gym.type.notlike("%dummy%"),
            Gym.type.is_(None)
        )
    )
    result = await db.execute(stmt)
    total_gyms_count = result.scalar() or 0

    # Count gyms where type contains "green"
    stmt = select(func.count()).select_from(Gym).filter(
        Gym.type.like("%green%")
    )
    result = await db.execute(stmt)
    verified_gyms_count = result.scalar() or 0

    # Unverified gyms (type contains "hold" or "red")
    stmt = select(func.count()).select_from(Gym).filter(
        or_(
            Gym.type.like("%hold%"),
            Gym.type.like("%red%")
        )
    )
    result = await db.execute(stmt)
    unverified_gyms_count = result.scalar() or 0

    # Count gyms where type = 'red'
    stmt = select(func.count()).select_from(Gym).filter(
        Gym.type == 'red'
    )
    result = await db.execute(stmt)
    red_gyms_count = result.scalar() or 0

    # Count gyms where type = 'hold'
    stmt = select(func.count()).select_from(Gym).filter(
        Gym.type == 'hold'
    )
    result = await db.execute(stmt)
    hold_gyms_count = result.scalar() or 0

    return {
        "gymOwners": {
            "today": gym_owners_today,
            "week": gym_owners_week,
            "month": gym_owners_month,
            "overall": gym_owners_overall
        },
        "gyms": {
            "today": gyms_today,
            "week": gyms_week,
            "month": gyms_month,
            "overall": gyms_overall
        },
        "dailyPassGyms": daily_pass_gyms,
        "verifiedGyms": {
            "verified": verified_gyms_count,
            "total": total_gyms_count
        },
        "unverifiedGyms": unverified_gyms_count,
        "unverifiedSplitup": {
            "red": red_gyms_count,
            "hold": hold_gyms_count
        }
    }

async def get_support_tickets(db: AsyncSession):
    """
    Get Support tickets data
    """
    today = datetime.now().date()

    stmt = select(func.count()).select_from(ClientToken)
    result = await db.execute(stmt)
    total_client_tickets = result.scalar() or 0

    stmt = select(func.count()).select_from(OwnerToken)
    result = await db.execute(stmt)
    total_gym_tickets = result.scalar() or 0

    stmt = select(func.count()).select_from(ClientToken).filter(
        ClientToken.resolved == False
    )
    result = await db.execute(stmt)
    unresolved_client_tickets = result.scalar() or 0

    stmt = select(func.count()).select_from(OwnerToken).filter(
        OwnerToken.resolved == False
    )
    result = await db.execute(stmt)
    unresolved_gym_tickets = result.scalar() or 0

    # Client tokens resolved today
    stmt_client = select(func.count()).select_from(ClientToken).filter(
        func.date(ClientToken.updated_at) == today,
        ClientToken.resolved == True
    )
    result_client = await db.execute(stmt_client)
    resolved_client_today = result_client.scalar() or 0

    # Owner tokens resolved today
    stmt_owner = select(func.count()).select_from(OwnerToken).filter(
        func.date(OwnerToken.updated_at) == today,
        OwnerToken.resolved == True
    )
    result_owner = await db.execute(stmt_owner)
    resolved_owner_today = result_owner.scalar() or 0

    resolved_today = resolved_client_today + resolved_owner_today

    return {
        "totalTickets": {
            "gym": total_gym_tickets,
            "client": total_client_tickets
        },
        "unresolvedTickets": {
            "gym": unresolved_gym_tickets,
            "client": unresolved_client_tickets
        },
        "resolvedToday": resolved_today
    }

async def get_rewards_metrics(db: AsyncSession):
    """
    Get Rewards metrics from reward_interest table
    """
    # Total count of reward_interest records
    stmt = select(func.count()).select_from(RewardInterest)
    result = await db.execute(stmt)
    total_count = result.scalar() or 0

    # Count of interested (where interested = True)
    stmt = select(func.count()).select_from(RewardInterest).filter(
        RewardInterest.interested == True
    )
    result = await db.execute(stmt)
    interested_count = result.scalar() or 0

    return {
        "total": total_count,
        "interested": interested_count
    }

async def get_reward_program_participants(db: AsyncSession):
    """
    Get Reward Program Participants count from reward_program_opt_ins table
    Counts the total number of records based on client_id
    """
    # Count distinct client_ids from reward_program_opt_ins
    stmt = select(func.count(distinct(RewardProgramOptIn.client_id)))
    result = await db.execute(stmt)
    total_participants = result.scalar() or 0

    return {
        "totalParticipants": total_participants
    }

async def get_gym_plans_metrics(db: AsyncSession):
    """
    Get Gym Plans metrics - count of gyms with session plans, membership plans, and daily pass pricing
    """
    # Get total gym count first
    stmt = select(func.count()).select_from(Gym)
    result = await db.execute(stmt)
    total_gyms = result.scalar() or 0

    # Count gyms with session plans
    stmt = select(func.count(distinct(SessionSetting.gym_id)))
    result = await db.execute(stmt)
    gyms_with_session_plans = result.scalar() or 0

    # Count gyms with membership plans
    stmt = select(func.count(distinct(GymPlans.gym_id)))
    result = await db.execute(stmt)
    gyms_with_membership_plans = result.scalar() or 0

    stmt = select(Gym.gym_id)
    result = await db.execute(stmt)
    all_gym_ids = [row[0] for row in result.all()]
    all_gym_id_strings = [str(gym_id) for gym_id in all_gym_ids]

    # Get all gym_id strings from DailyPassPricing that match our gyms
    stmt = select(DailyPassPricing.gym_id).filter(
        DailyPassPricing.gym_id.in_(all_gym_id_strings)
    )
    result = await db.execute(stmt)
    daily_pass_gym_ids = result.all()

    # Count unique gym IDs (convert to int to avoid duplicates from string representation)
    gyms_with_daily_pass = len(set([int(gym_id[0]) for gym_id in daily_pass_gym_ids]))

    return {
        "sessionPlans": gyms_with_session_plans,
        "membershipPlans": gyms_with_membership_plans,
        "dailyPass": gyms_with_daily_pass,
        "totalGyms": total_gyms
    }

async def get_gym_photos_metrics(db: AsyncSession):
    """
    Calculate gym photos metrics using mutually exclusive logic.
    Logic matches the gymdetails page:
    1. Verified Studio: Gyms with photos in gym_studios_pic (priority)
    2. Pending Photo verification: Gyms with photos ONLY in gym_onboarding_pics (NO studio photos)
    3. Photos Not Uploaded: Gyms with NEITHER studio NOR onboarding photos
    """

    # Get all gym IDs
    stmt = select(Gym.gym_id)
    result = await db.execute(stmt)
    all_gym_ids = [row[0] for row in result.all()]

    # Get gym IDs with studio photos
    stmt = select(GymStudiosPic.gym_id)
    result = await db.execute(stmt)
    gym_ids_with_studio = set([row[0] for row in result.all()])

    # Get gym IDs with onboarding photos
    stmt = select(GymOnboardingPics.gym_id)
    result = await db.execute(stmt)
    gym_ids_with_onboarding = set([row[0] for row in result.all()])

    # Verified Studio: Gyms with studio photos (priority given to studio)
    verified_studio_count = len(gym_ids_with_studio)

    # Pending Photo verification: Gyms with ONLY onboarding photos (excluding those with studio photos)
    gym_ids_only_onboarding = gym_ids_with_onboarding - gym_ids_with_studio
    pending_verification_count = len(gym_ids_only_onboarding)

    # Photos Not Uploaded: Gyms with neither studio nor onboarding photos
    gym_ids_with_any_photos = gym_ids_with_studio.union(gym_ids_with_onboarding)
    no_uploads_count = len(all_gym_ids) - len(gym_ids_with_any_photos)

    return {
        "studio": verified_studio_count,
        "onboard": pending_verification_count,
        "noUploads": no_uploads_count
    }

@router.get("/overview")
async def get_dashboard_overview(
    fittbot_filter: str = "month",
    business_filter: str = "month",
    custom_start_date: str = None,
    custom_end_date: str = None,
    db: AsyncSession = Depends(get_async_db)
):
    try:
        # Use custom date range if provided for fittbot
        if custom_start_date and custom_end_date and fittbot_filter == "custom":
            fittbot_data = await get_fittbot_metrics_custom(db, custom_start_date, custom_end_date)
        else:
            fittbot_data = await get_fittbot_metrics(db, fittbot_filter)

        # Use custom date range if provided for business
        if custom_start_date and custom_end_date and business_filter == "custom":
            business_data = await get_business_metrics_custom(db, custom_start_date, custom_end_date)
        else:
            business_data = await get_business_metrics(db, business_filter)

        plans_data = await get_plans_metrics(db)
        support_data = await get_support_tickets(db)
        rewards_data = await get_rewards_metrics(db)
        gym_plans_data = await get_gym_plans_metrics(db)
        gym_photos_data = await get_gym_photos_metrics(db)
        reward_program_data = await get_reward_program_participants(db)

        return {
            "success": True,
            "data": {
                "fittbot": fittbot_data,
                "business": business_data,
                "plans": plans_data,
                "support": support_data,
                "rewards": rewards_data,
                "gymPlans": gym_plans_data,
                "gymPhotos": gym_photos_data,
                "rewardProgram": reward_program_data
            },
            "message": "Dashboard data fetched successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/fittbot-metrics")
async def get_fittbot_metrics_endpoint(filter: str = "month", db: AsyncSession = Depends(get_async_db)):
    try:
        data = await get_fittbot_metrics(db, filter)
        return {
            "success": True,
            "data": data,
            "message": "Fittbot metrics fetched successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/business-metrics")
async def get_business_metrics_endpoint(filter: str = "month", db: AsyncSession = Depends(get_async_db)):
    try:
        data = await get_business_metrics(db, filter)
        return {
            "success": True,
            "data": data,
            "message": "Business metrics fetched successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/support-tickets")
async def get_support_tickets_endpoint(db: AsyncSession = Depends(get_async_db)):
    try:
        data = await get_support_tickets(db)
        return {
            "success": True,
            "data": data,
            "message": "Support tickets data fetched successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/revenue-analytics")
async def get_revenue_analytics(
    start_date: str = None,
    end_date: str = None,
    source: str = None,
    gym_id: int = None,
    db: AsyncSession = Depends(get_async_db)
):
    """
    Get revenue analytics data using centralized revenue service.

    Provides:
    - Total revenue
    - Source-wise breakdown
    - Daily revenue over time
    - Gym-wise revenue breakdown
    """
    try:
        # Parse dates if provided
        if start_date:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        else:
            start_date_obj = datetime(2020, 1, 1).date()

        if end_date:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        else:
            end_date_obj = datetime.now().date()

        # Use centralized revenue service for detailed breakdowns
        revenue_data = await get_detailed_revenue_with_breakdowns(
            db=db,
            start_date=start_date_obj,
            end_date=end_date_obj,
            source=source,
            specific_gym_id=gym_id,
            exclude_gym_id_one=True
        )

        # Build response data
        analytics_data = {
            "totalRevenue": revenue_data.total_revenue,
            "sourceSplit": {
                "daily_pass": revenue_data.source_split["daily_pass"],
                "sessions": revenue_data.source_split["sessions"],
                "fittbot_subscription": revenue_data.source_split["fittbot_subscription"],
                "gym_membership": revenue_data.source_split["gym_membership"],
                "ai_credits": revenue_data.source_split.get("ai_credits", 0)
            },
            "sourceSplitRupees": {
                "daily_pass": revenue_data.source_split_rupees.get("daily_pass", 0),
                "sessions": revenue_data.source_split_rupees.get("sessions", 0),
                "fittbot_subscription": revenue_data.source_split_rupees.get("fittbot_subscription", 0),
                "gym_membership": revenue_data.source_split_rupees.get("gym_membership", 0),
                "ai_credits": revenue_data.source_split_rupees.get("ai_credits", 0)
            },
            "revenueOverTime": [
                {
                    "date": point.date,
                    "revenue": point.revenue
                }
                for point in revenue_data.daily_revenue
            ],
            "gymBreakdown": [
                {
                    "gym_id": point.gym_id,
                    "gym_name": point.gym_name,
                    "revenue": point.revenue
                }
                for point in revenue_data.gym_breakdown
            ],
            "filters": {
                "startDate": start_date_obj.isoformat(),
                "endDate": end_date_obj.isoformat(),
                "source": source or "all",
                "gymId": gym_id or "all"
            }
        }

        return {
            "success": True,
            "data": analytics_data,
            "message": "Revenue analytics fetched successfully"
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    except Exception as e:
        print(f"[DASHBOARD] Error in revenue-analytics: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/recurring-subscribers")
async def get_recurring_subscribers(db: AsyncSession = Depends(get_async_db)):
    """
    Get count of recurring subscribers.
    NEW LOGIC: Returns clients who have purchased Nutritionist Plan more than once.
    Queries payments.payments where payment_metadata['flow'] = 'nutrition_purchase_googleplay'
    """
    try:
        from collections import defaultdict

        # Dictionary to store nutritionist plan purchase count per customer_id
        customer_purchase_count = defaultdict(int)

        # NEW LOGIC: Query payments table for nutritionist plan purchases
        # Filters:
        # - payment_metadata['flow'] = 'nutrition_purchase_googleplay'
        # - status = 'captured'
        payment_stmt = (
            select(Payment.customer_id)
            .where(Payment.status == "captured")
            .where(func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_purchase_googleplay')
        )

        payment_result = await db.execute(payment_stmt)
        payments = payment_result.all()

        # Count each payment as a nutritionist plan purchase
        for payment in payments:
            customer_purchase_count[payment.customer_id] += 1

        # Filter only customers with more than 1 nutritionist plan purchase
        recurring_customer_ids = [
            customer_id for customer_id, count in customer_purchase_count.items()
            if count > 1
        ]

        # Convert string customer_ids to integers and check which ones exist in Client table
        customer_ids_int = []
        for customer_id_str in recurring_customer_ids:
            try:
                customer_ids_int.append(int(customer_id_str))
            except (ValueError, TypeError):
                pass

        # Fetch only valid client IDs that exist in Client table
        valid_count = 0
        if customer_ids_int:
            stmt = select(func.count()).select_from(Client).where(Client.client_id.in_(customer_ids_int))
            result = await db.execute(stmt)
            valid_count = result.scalar() or 0

        return {
            "success": True,
            "data": {
                "total": valid_count
            },
            "message": "Recurring subscribers fetched successfully"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching recurring subscribers: {str(e)}")


@router.get("/recurring-subscribers/details")
async def get_recurring_subscribers_details(
    page: int = 1,
    limit: int = 10,
    search: str = None,
    db: AsyncSession = Depends(get_async_db)
):
    """
    Get detailed list of recurring subscribers with pagination.
    Returns clients who have purchased Nutritionist Plan (Fittbot subscription) more than once.
    NEW LOGIC: Query payments.payments where payment_metadata['flow'] = 'nutrition_purchase_googleplay'
    """
    try:
        from collections import defaultdict
        from sqlalchemy import or_

        # Dictionary to store subscription info per customer_id
        customer_subscriptions = defaultdict(lambda: {"count": 0, "payments": []})

        # NEW LOGIC: Query payments table for nutritionist plan purchases
        # Filters:
        # - payment_metadata['flow'] = 'nutrition_purchase_googleplay'
        # - status = 'captured'
        # Extract: customer_id, id, amount_minor, captured_at
        payment_stmt = (
            select(Payment.customer_id, Payment.id, Payment.amount_minor, Payment.captured_at)
            .where(Payment.status == "captured")
            .where(func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_purchase_googleplay')
        )

        payment_result = await db.execute(payment_stmt)
        payments = payment_result.all()

        # Count each payment as a nutritionist plan purchase
        for payment in payments:
            customer_subscriptions[payment.customer_id]["count"] += 1
            customer_subscriptions[payment.customer_id]["payments"].append({
                "id": payment.id,
                "date": payment.captured_at.isoformat() if payment.captured_at else None,
                "amount": payment.amount_minor
            })

        # Filter only customers with more than 1 nutritionist plan purchase
        recurring_customers = {
            customer_id: data
            for customer_id, data in customer_subscriptions.items()
            if data["count"] > 1
        }

        # Get customer IDs and convert to integers for matching with Client table
        # The customer_id from payments is string, but Client.client_id is integer
        customer_ids_int = []
        customer_id_mapping = {}  # Maps int -> string (original customer_id)

        for customer_id_str in recurring_customers.keys():
            try:
                customer_id_int = int(customer_id_str)
                customer_ids_int.append(customer_id_int)
                customer_id_mapping[customer_id_int] = customer_id_str
            except (ValueError, TypeError):
                # Skip if customer_id cannot be converted to int
                pass

        # First, fetch all valid clients to get accurate count
        valid_client_ids = []
        if customer_ids_int:
            stmt = select(Client.client_id).where(Client.client_id.in_(customer_ids_int))
            result = await db.execute(stmt)
            valid_client_ids = [row[0] for row in result.all()]

        # Filter to only include customers that exist in Client table
        customer_ids_int = [cid for cid in customer_ids_int if cid in valid_client_ids]

        # Apply search filter if provided (search on ID, name, or contact)
        if search:
            # Fetch clients for filtering
            if customer_ids_int:
                stmt = select(Client).where(Client.client_id.in_(customer_ids_int))
                # Add search conditions
                search_pattern = f"%{search}%"
                stmt = stmt.where(
                    (Client.client_id.like(search_pattern)) |
                    (Client.name.like(search_pattern)) |
                    (Client.contact.like(search_pattern))
                )
                result = await db.execute(stmt)
                matching_clients = result.scalars().all()
                customer_ids_int = [client.client_id for client in matching_clients]

        # Get total count for pagination
        total_count = len(customer_ids_int)

        # Apply pagination
        start_idx = (page - 1) * limit
        end_idx = start_idx + limit
        paginated_customer_ids_int = customer_ids_int[start_idx:end_idx]

        # Fetch client details for the paginated customers
        subscribers_data = []
        if paginated_customer_ids_int:
            stmt = select(Client).where(Client.client_id.in_(paginated_customer_ids_int))
            result = await db.execute(stmt)
            clients = result.scalars().all()

            # Create a mapping of client_id to client data
            clients_map = {client.client_id: client for client in clients}

            for customer_id_int in paginated_customer_ids_int:
                client = clients_map.get(customer_id_int)
                if client:
                    # Get the original string customer_id for looking up subscriptions
                    customer_id_str = customer_id_mapping[customer_id_int]
                    subscriptions = recurring_customers[customer_id_str]

                    # Calculate total amount spent (using amount_minor from payments table)
                    total_amount = sum(
                        p["amount"] for p in subscriptions["payments"]
                    )

                    # Get first and last subscription date using captured_at
                    # Select the earliest entry for first_subscription
                    all_dates = [p["date"] for p in subscriptions["payments"] if p["date"]]
                    all_dates.sort()

                    subscribers_data.append({
                        "customer_id": customer_id_int,  # Use integer for display
                        "name": client.name or "N/A",
                        "contact": client.contact or "N/A",
                        "subscription_count": subscriptions["count"],
                        "total_amount": total_amount,
                        "first_subscription": all_dates[0] if all_dates else None,  # Earliest captured_at
                        "last_subscription": all_dates[-1] if all_dates else None,   # Latest captured_at
                    })

        return {
            "success": True,
            "data": {
                "subscribers": subscribers_data,
                "pagination": {
                    "page": page,
                    "limit": limit,
                    "total": total_count,
                    "total_pages": (total_count + limit - 1) // limit
                }
            },
            "message": "Recurring subscribers details fetched successfully"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching recurring subscribers details: {str(e)}")


@router.get("/support-tickets-list")
async def get_support_tickets_list(
    source: str = Query(None, description="Filter by source: 'Fittbot' for client, 'Fittbot Business' for gym"),
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, le=100, description="Items per page"),
    status: str = Query(None, description="Filter by status: 'all', 'resolved', 'unresolved', or 'follow_up'"),
    search: str = Query(None, description="Search by token or email"),
    db: AsyncSession = Depends(get_async_db)
):

    import math
    try:

        if source == "Fittbot Business":
            # Query gym owner support tokens
            base_model = OwnerToken
            query = select(
                OwnerToken.id,
                OwnerToken.token.label('ticket_id'),
                literal("Fittbot Business").label('source'),
                Gym.name.label('name'),
                OwnerToken.email,
                OwnerToken.subject,
                OwnerToken.issue,
                OwnerToken.followed_up,
                OwnerToken.resolved,
                OwnerToken.comments,
                OwnerToken.created_at,
                OwnerToken.resolved_at,
                Employees.name.label('assigned_to'),
            ).outerjoin(
                Gym, OwnerToken.gym_id == Gym.gym_id
            ).outerjoin(
                TicketAssignment,
                (TicketAssignment.ticket_id == OwnerToken.id) & (TicketAssignment.ticket_source == "Fittbot Business") & (TicketAssignment.status == "active")
            ).outerjoin(
                Employees, Employees.id == TicketAssignment.employee_id
            )
        elif source == "Fittbot":
            # Query client support tokens
            base_model = ClientToken
            query = select(
                ClientToken.id,
                ClientToken.token.label('ticket_id'),
                literal("Fittbot").label('source'),
                Client.name.label('name'),
                ClientToken.email,
                ClientToken.subject,
                ClientToken.issue,
                ClientToken.followed_up,
                ClientToken.resolved,
                ClientToken.comments,
                ClientToken.created_at,
                ClientToken.resolved_at,
                Employees.name.label('assigned_to'),
            ).outerjoin(
                Client, ClientToken.client_id == Client.client_id
            ).outerjoin(
                TicketAssignment,
                (TicketAssignment.ticket_id == ClientToken.id) & (TicketAssignment.ticket_source == "Fittbot") & (TicketAssignment.status == "active")
            ).outerjoin(
                Employees, Employees.id == TicketAssignment.employee_id
            )
        else:
            raise HTTPException(status_code=400, detail="Invalid source. Use 'Fittbot' or 'Fittbot Business'")

        if status == "Resolved":
            query = query.filter(base_model.resolved == True)
        elif status == "Pending":
            query = query.filter(base_model.resolved == False)
        elif status == "Follow Up":
            query = query.filter(base_model.followed_up == True, base_model.resolved == False)
        elif status == "all" or status is None:
            pass

        # Apply search filter
        if search:
            search_term = f"%{search.lower()}%"
            query = query.filter(
                or_(
                    func.lower(base_model.token).like(search_term),
                    func.lower(base_model.email).like(search_term)
                )
            )

        # Order by created_at descending
        query = query.order_by(base_model.created_at.desc())

        # Get total count
        count_query = select(func.count()).select_from(base_model)
        if status == "resolved":
            count_query = count_query.filter(base_model.resolved == True)
        elif status == "unresolved":
            count_query = count_query.filter(base_model.resolved == False)
        elif status == "follow_up":
            count_query = count_query.filter(base_model.followed_up == True, base_model.resolved == False)
        
        if search:
            search_term = f"%{search.lower()}%"
            count_query = count_query.filter(
                or_(
                    func.lower(base_model.token).like(search_term),
                    func.lower(base_model.email).like(search_term)
                )
            )

        count_result = await db.execute(count_query)
        total_count = count_result.scalar() or 0

        # Apply pagination
        offset = (page - 1) * limit
        query = query.offset(offset).limit(limit)

        # Execute query
        result = await db.execute(query)
        tickets_data = result.all()

        # Format response
        def map_status(followed_up, resolved):
           
            if resolved:
                return "Resolved"
            elif followed_up:
                return "Follow Up"
            else:
                return "Pending"

        tickets = []
        for row in tickets_data:
            tickets.append({
                "id": row.id,
                "ticket_id": row.ticket_id or f"ticket-{row.id}",
                "source": row.source,
                "name": row.name or "N/A",
                "email": row.email or "N/A",
                "subject": row.subject,
                "issue_type": row.subject,
                "issue": row.issue,
                "status": map_status(row.followed_up, row.resolved),
                "comments": row.comments,
                "created_at": row.created_at.isoformat() if row.created_at else None,
                "resolved_at": row.resolved_at.isoformat() if row.resolved_at else None,
                "assigned_to": row.assigned_to or None,
            })

        total_pages = (total_count + limit - 1) // limit if total_count > 0 else 1
        has_next = page < total_pages
        has_prev = page > 1

        return {
            "success": True,
            "message": "Tickets fetched successfully",
            "data": {
                "tickets": tickets,
                "total": total_count,
                "page": page,
                "limit": limit,
                "totalPages": total_pages,
                "hasNext": has_next,
                "hasPrev": has_prev
            }
        }
    except Exception as e:
        print(f"Error fetching support tickets: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error fetching support tickets: {str(e)}")

@router.get("/support-tickets-export")
async def export_support_tickets(
    source: str = Query(..., description="Fittbot or Fittbot Business"),
    start_date: str = Query(None),
    end_date: str = Query(None),
    db: AsyncSession = Depends(get_async_db)
):
    """
    Export support tickets to Excel with 4 sheets based on status.
    Applies date range filter.
    """
    try:
        import io
        import pandas as pd
        from fastapi.responses import StreamingResponse

        # Determine models based on source
        if source == "Fittbot Business":
            base_model = OwnerToken
            user_model = Gym
            join_col = OwnerToken.gym_id == Gym.gym_id
            ticket_source_val = "Fittbot Business"
        elif source == "Fittbot":
            base_model = ClientToken
            user_model = Client
            join_col = ClientToken.client_id == Client.client_id
            ticket_source_val = "Fittbot"
        else:
            raise HTTPException(status_code=400, detail="Invalid source. Use 'Fittbot' or 'Fittbot Business'")

        # Build base query
        stmt = select(
            base_model.token.label('ticket_id'),
            base_model.subject,
            base_model.issue.label('issue_description'),
            base_model.comments,
            base_model.followed_up,
            base_model.resolved,
            base_model.created_at,
            base_model.resolved_at,
            Employees.name.label('assigned_to')
        ).outerjoin(
            user_model, join_col
        ).outerjoin(
            TicketAssignment,
            (TicketAssignment.ticket_id == base_model.id) & (TicketAssignment.ticket_source == ticket_source_val) & (TicketAssignment.status == "active")
        ).outerjoin(
            Employees, Employees.id == TicketAssignment.employee_id
        )

        # Apply date filters
        if start_date:
            stmt = stmt.filter(func.date(base_model.created_at) >= start_date)
        if end_date:
            stmt = stmt.filter(func.date(base_model.created_at) <= end_date)

        # Order by created_at descending
        stmt = stmt.order_by(base_model.created_at.desc())

        # Execute query
        result = await db.execute(stmt)
        all_tickets = result.all()

        # Group data for sheets
        def get_status(followed_up, resolved):
            if resolved: return "Resolved"
            elif followed_up: return "Follow Up"
            else: return "Pending"

        data_all = []
        data_resolved = []
        data_unresolved = []
        data_followup = []

        for row in all_tickets:
            status = get_status(row.followed_up, row.resolved)
            item = {
                "Ticket ID": row.ticket_id,
                "Subject": row.subject or "N/A",
                "Status": status,
                "Assigned To": row.assigned_to or "N/A",
                "Issue Description": row.issue_description or "N/A",
                "Comments": row.comments or "N/A",
                "Created At": row.created_at.strftime("%Y-%m-%d %H:%M:%S") if row.created_at else "N/A",
                "Resolved At": row.resolved_at.strftime("%Y-%m-%d %H:%M:%S") if row.resolved_at else "N/A"
            }
            
            data_all.append(item)
            if status == "Resolved":
                data_resolved.append(item)
            elif status == "Pending":
                data_unresolved.append(item)
            elif status == "Follow Up":
                data_followup.append(item)

        # Create Excel in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheets configuration
            sheets = [
                ("All Tickets", data_all),
                ("Resolved", data_resolved),
                ("Unresolved", data_unresolved),
                ("Follow Up", data_followup)
            ]
            
            for sheet_name, data in sheets:
                df = pd.DataFrame(data)
                if df.empty:
                    # Create even if empty to satisfy the requirement of 4 sheets
                    df = pd.DataFrame(columns=["Ticket ID", "Subject", "Status", "Assigned To", "Issue Description", "Comments", "Created At", "Resolved At"])
                
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                
                # Auto-adjust column widths and style header
                worksheet = writer.sheets[sheet_name]
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Steel Blue
                header_font = Font(color="FFFFFF", bold=True)
                
                for idx, col in enumerate(df.columns, 1):
                    # Style header cell
                    cell = worksheet.cell(row=1, column=idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                    
                    series = df[col]
                    max_len = max(
                        series.astype(str).map(len).max(),
                        len(str(col))
                    ) + 2
                    col_letter = worksheet.cell(row=1, column=idx).column_letter
                    worksheet.column_dimensions[col_letter].width = min(max_len, 60)

        output.seek(0)
        
        # Filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_source = source.replace(" ", "_")
        filename = f"Support_Tickets_{safe_source}_{timestamp}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"Error exporting support tickets: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error exporting support tickets: {str(e)}")
    

@router.get("/gym-ticket-detail")
async def get_gym_ticket_detail(
    ticket_id: str = Query(..., description="Ticket token string"),
    db: AsyncSession = Depends(get_async_db)
):
    """Get gym (owner) ticket details by token string"""
    try:
        query = select(
            OwnerToken.id,
            OwnerToken.token.label('ticket_id'),
            literal("Fittbot Business").label('source'),
            Gym.name.label('name'),
            OwnerToken.email,
            OwnerToken.subject,
            OwnerToken.issue,
            OwnerToken.followed_up,
            OwnerToken.resolved,
            OwnerToken.comments,
            OwnerToken.created_at
        ).outerjoin(
            Gym, OwnerToken.gym_id == Gym.gym_id
        ).filter(
            OwnerToken.token == ticket_id
        )

        result = await db.execute(query)
        ticket_data = result.first()

        if not ticket_data:
            raise HTTPException(status_code=404, detail="Ticket not found")

        def map_status(followed_up, resolved):
            if resolved:
                return "resolved"
            else:
                return "pending"

        ticket = {
            "id": ticket_data.id,
            "ticket_id": ticket_data.ticket_id or f"ticket-{ticket_data.id}",
            "source": ticket_data.source,
            "name": ticket_data.name or "N/A",
            "email": ticket_data.email or "N/A",
            "subject": ticket_data.subject,
            "issue": ticket_data.issue,
            "status": map_status(ticket_data.followed_up, ticket_data.resolved),
            "comments": ticket_data.comments,
            "created_at": ticket_data.created_at.isoformat() if ticket_data.created_at else None,
        }

        return {
            "success": True,
            "data": ticket,
            "message": "Ticket details fetched successfully"
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching ticket details: {str(e)}")



@router.get("/client-ticket-detail")
async def get_client_ticket_detail(
    ticket_id: str = Query(..., description="Ticket token string"),
    db: AsyncSession = Depends(get_async_db)
):
   
    import sys
    try:

        query = select(
            ClientToken.id,
            ClientToken.token.label('ticket_id'),
            literal("Fittbot").label('source'),
            Client.name.label('name'),
            ClientToken.email,
            ClientToken.subject,
            ClientToken.issue,
            ClientToken.followed_up,
            ClientToken.resolved,
            ClientToken.comments,
            ClientToken.created_at
        ).outerjoin(
            Client, ClientToken.client_id == Client.client_id
        ).filter(
            ClientToken.token == ticket_id
        )

        result = await db.execute(query)
        ticket_data = result.first()

        if not ticket_data:
            raise HTTPException(status_code=404, detail="Ticket not found")

        def map_status(followed_up, resolved):
           
            if resolved:
                return "Resolved"
            elif followed_up:
                return "Follow Up"
            else:
                return "Pending"

        ticket = {
            "id": ticket_data.id,
            "ticket_id": ticket_data.ticket_id or f"ticket-{ticket_data.id}",
            "source": ticket_data.source,
            "name": ticket_data.name or "N/A",
            "email": ticket_data.email or "N/A",
            "subject": ticket_data.subject,
            "issue": ticket_data.issue,
            "status": map_status(ticket_data.followed_up, ticket_data.resolved),
            "comments": ticket_data.comments,
            "created_at": ticket_data.created_at.isoformat() if ticket_data.created_at else None,
        }

        return {
            "success": True,
            "data": ticket,
            "message": "Ticket details fetched successfully"
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc(file=sys.stderr)
        print(f"[CLIENT-TICKET-DETAIL-ERROR] {str(e)}", file=sys.stderr)
        raise HTTPException(status_code=500, detail=f"Error fetching ticket details: {str(e)}")

class MarkResolvedRequest(BaseModel):
    ticket_id: str


@router.post("/gym-ticket-resolve")
async def mark_gym_ticket_resolved(
    request: MarkResolvedRequest,
    db: AsyncSession = Depends(get_async_db)
):
    """Mark gym ticket as resolved"""
    try:
        query = select(OwnerToken).filter(OwnerToken.token == request.ticket_id)
        result = await db.execute(query)
        ticket = result.scalar_one_or_none()

        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket not found")

        ticket.resolved = True
        ticket.followed_up = True
        ticket.updated_at = datetime.now()

        await db.commit()

        return {
            "success": True,
            "message": "Ticket marked as resolved"
        }

    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error marking ticket as resolved: {str(e)}")


@router.post("/client-ticket-resolve")
async def mark_client_ticket_resolved(
    request: MarkResolvedRequest,
    db: AsyncSession = Depends(get_async_db)
):
    """Mark client ticket as resolved"""
    try:
        query = select(ClientToken).filter(ClientToken.token == request.ticket_id)
        result = await db.execute(query)
        ticket = result.scalar_one_or_none()

        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket not found")

        ticket.resolved = True
        ticket.followed_up = True
        ticket.updated_at = datetime.now()

        await db.commit()

        return {
            "success": True,
            "message": "Ticket marked as resolved"
        }

    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error marking ticket as resolved: {str(e)}")


class TicketFollowUpRequest(BaseModel):
    ticket_id: str
    source: str  # "client" or "owner"
    comment: Optional[str] = None
    status: Optional[str] = None  # "followup"


@router.post("/ticket_followup")
async def ticket_followup(
    request: TicketFollowUpRequest,
    db: AsyncSession = Depends(get_async_db)
):
    """
    Mark a ticket for follow-up or add comments to a ticket
    """
    import sys
    try:
        print(f"[TICKET-FOLLOWUP] ticket_id={request.ticket_id}, source={request.source}, comment={request.comment}, status={request.status}", file=sys.stderr)

        if not request.comment and not request.status:
            raise HTTPException(status_code=400, detail="At least one of comment or status must be provided")

        # Select the correct table based on source
        if request.source == "client":
            query = select(ClientToken).filter(ClientToken.token == request.ticket_id)
        elif request.source == "owner":
            query = select(OwnerToken).filter(OwnerToken.token == request.ticket_id)
        else:
            raise HTTPException(status_code=400, detail="Invalid source. Use 'client' or 'owner'")

        result = await db.execute(query)
        ticket = result.scalar_one_or_none()

        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket not found")

        # Update status to followup if requested
        if request.status and request.status == "follow_up":
            ticket.followed_up = True
            ticket.resolved = False

        # Add comment if provided
        if request.comment:
            ticket.comments = request.comment

        ticket.updated_at = datetime.now()
        await db.commit()

        return {
            "success": True,
            "message": "Ticket updated successfully"
        }

    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        import traceback
        traceback.print_exc(file=sys.stderr)
        print(f"[TICKET-FOLLOWUP-ERROR] {str(e)}", file=sys.stderr)
        raise HTTPException(status_code=500, detail=f"Error updating ticket: {str(e)}")


@router.get("/purchase-analytics")
async def get_purchase_analytics(
    start_date: str = None,
    end_date: str = None,
    source: str = None,
    gym_id: int = None,
    location: str = None,
    db: AsyncSession = Depends(get_async_db)
):

    try:
        # Debug: Log the gym_id parameter
        import logging
        logging.info(f"Purchase analytics called with gym_id: {gym_id} (type: {type(gym_id)}), start_date: {start_date}, end_date: {end_date}, source: {source}")

        # Parse dates if provided
        if start_date:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        else:
            # Default to early date for overall data
            start_date_obj = datetime(2020, 1, 1).date()

        if end_date:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        else:
            # Default to today
            end_date_obj = datetime.now().date()

        # Get client IDs filtered by location if location is provided
        location_client_ids = set()
        if location and location != "all":
            try:
                location_stmt = select(Client.client_id).where(Client.location == location)
                location_result = await db.execute(location_stmt)
                location_client_ids = set(row[0] for row in location_result.all())
            except Exception as e:
                location_client_ids = set()
        else:
            # Debug: Check how many purchases have NULL or invalid client_id
            try:
                dailypass_session = get_dailypass_session()

                # Count total DailyPass purchases
                total_dp = dailypass_session.query(func.count()).scalar()

                # Count DailyPass with NULL client_id
                null_client_dp = dailypass_session.query(func.count()).filter(DailyPass.client_id.is_(None)).scalar()

                # Count DailyPass with client_id NOT in Client table
                valid_client_ids = select(Client.client_id)
                not_in_client = dailypass_session.query(func.count()).filter(~DailyPass.client_id.in_(valid_client_ids)).scalar()

                dailypass_session.close()
            except Exception as e:
                pass

        # Initialize result structure
        category_breakdown = {
            "daily_pass": {"purchases": 0, "unique_users": 0, "purchases_over_time": []},
            "sessions": {"purchases": 0, "unique_users": 0, "purchases_over_time": []},
            "fittbot_subscription": {"purchases": 0, "unique_users": 0, "purchases_over_time": []},
            "ai_credits": {"purchases": 0, "unique_users": 0, "purchases_over_time": []},
            "gym_membership": {"purchases": 0, "unique_users": 0, "purchases_over_time": []}
        }

        all_purchases_over_time = {}  # date -> total count
        gym_purchases = {}  # gym_id -> purchase count
        location_purchases = {}  # location -> purchase count

        # 1. DAILY PASS PURCHASES - Single aggregated query
        # GMV fix: INNER JOIN with Gym table to exclude orphaned DailyPass records (gym_id not in Gym table)
        if not source or source == "daily_pass":
            try:
                import logging
                logging.info(f"Daily pass purchases: gym_id={gym_id}, source={source}")
                dailypass_session = get_dailypass_session()

                # Build base query — JOIN Gym table to exclude records with no matching gym
                # GMV fix: SUM(days_total) so multi-day passes count each day (not just 1 per transaction)
                all_query = dailypass_session.query(
                    func.coalesce(func.sum(DailyPass.days_total), 0).label('total_purchases'),
                    func.count(distinct(DailyPass.client_id)).label('total_unique_users')
                ).join(
                    Gym, Gym.gym_id == func.cast(DailyPass.gym_id, Integer), isouter=False
                ).filter(
                    func.date(DailyPass.created_at) >= start_date_obj,
                    func.date(DailyPass.created_at) <= end_date_obj,
                    DailyPass.gym_id != "1"
                )

                # Apply gym filter if provided
                if gym_id:
                    gym_id_str = str(gym_id)
                    all_query = all_query.filter(DailyPass.gym_id == gym_id_str)
                    logging.info(f"Daily pass query with gym filter: gym_id_str={gym_id_str}")

                # Apply location filter if provided
                if location_client_ids:
                    all_query = all_query.filter(DailyPass.client_id.in_(location_client_ids))

                # Get totals
                total_result = all_query.first()

                if total_result:
                    category_breakdown["daily_pass"]["purchases"] = total_result.total_purchases or 0
                    category_breakdown["daily_pass"]["unique_users"] = total_result.total_unique_users or 0

                # Build separate query for purchases over time (grouped by date) — same Gym JOIN
                # GMV fix: SUM(days_total) so multi-day passes count each day in time-series too
                base_query = dailypass_session.query(
                    func.date(DailyPass.created_at).label('purchase_date'),
                    func.coalesce(func.sum(DailyPass.days_total), 0).label('purchase_count')
                ).join(
                    Gym, Gym.gym_id == func.cast(DailyPass.gym_id, Integer), isouter=False
                ).filter(
                    func.date(DailyPass.created_at) >= start_date_obj,
                    func.date(DailyPass.created_at) <= end_date_obj,
                    DailyPass.gym_id != "1"
                )

                # Apply gym filter if provided
                if gym_id:
                    base_query = base_query.filter(DailyPass.gym_id == str(gym_id))

                # Apply location filter if provided
                if location_client_ids:
                    base_query = base_query.filter(DailyPass.client_id.in_(location_client_ids))

                # Group by date and get results
                base_query = base_query.group_by(func.date(DailyPass.created_at))
                time_result = base_query.all()

                if time_result:
                    # Build purchases over time
                    for row in time_result:
                        date_key = row.purchase_date.isoformat() if row.purchase_date else None
                        if date_key:
                            if date_key not in all_purchases_over_time:
                                all_purchases_over_time[date_key] = 0
                            all_purchases_over_time[date_key] += row.purchase_count
                            category_breakdown["daily_pass"]["purchases_over_time"].append({
                                "date": date_key,
                                "purchases": row.purchase_count
                            })

                # Sort purchases over time by date
                category_breakdown["daily_pass"]["purchases_over_time"].sort(key=lambda x: x["date"])

                # Get gym-wise purchases
                # When gym filter is applied, track just that gym; otherwise get all gyms
                if gym_id:
                    # Track purchases for the filtered gym
                    gym_count_query = dailypass_session.query(
                        func.count().label('purchase_count')
                    ).filter(
                        func.date(DailyPass.created_at) >= start_date_obj,
                        func.date(DailyPass.created_at) <= end_date_obj,
                        DailyPass.gym_id == str(gym_id)
                    )

                    # Apply location filter if provided
                    if location_client_ids:
                        gym_count_query = gym_count_query.filter(DailyPass.client_id.in_(location_client_ids))

                    gym_count_result = gym_count_query.first()
                    if gym_count_result and gym_count_result.purchase_count:
                        gym_key = int(gym_id)
                        if gym_key not in gym_purchases:
                            gym_purchases[gym_key] = 0
                        gym_purchases[gym_key] += gym_count_result.purchase_count
                        logging.info(f"Daily pass for gym {gym_id}: {gym_count_result.purchase_count} purchases")
                else:
                    # Get all gyms when no filter
                    gym_query = dailypass_session.query(
                        DailyPass.gym_id,
                        func.count().label('purchase_count')
                    ).filter(
                        func.date(DailyPass.created_at) >= start_date_obj,
                        func.date(DailyPass.created_at) <= end_date_obj,
                        DailyPass.gym_id.isnot(None),
                        DailyPass.gym_id != "1"  # Exclude gym_id = 1
                    )

                    # Apply location filter if provided
                    if location_client_ids:
                        gym_query = gym_query.filter(DailyPass.client_id.in_(location_client_ids))

                    gym_query = gym_query.group_by(DailyPass.gym_id)
                    gym_result = gym_query.all()

                    for row in gym_result:
                        try:
                            gym_key = int(row.gym_id)
                            if gym_key not in gym_purchases:
                                gym_purchases[gym_key] = 0
                            gym_purchases[gym_key] += row.purchase_count
                        except (ValueError, TypeError):
                            pass

                logging.info(f"Daily pass gym_purchases: {gym_purchases}")

                # Get location-wise purchases (only when no location filter is applied)
                # GMV fix: add Gym INNER JOIN so orphaned DailyPass records are excluded
                if not location or location == "all":
                    try:
                        # Get all DailyPass client_ids — with Gym JOIN to exclude orphans
                        daily_pass_records = dailypass_session.query(
                            DailyPass.client_id
                        ).join(
                            Gym, Gym.gym_id == func.cast(DailyPass.gym_id, Integer), isouter=False
                        ).filter(
                            func.date(DailyPass.created_at) >= start_date_obj,
                            func.date(DailyPass.created_at) <= end_date_obj,
                            DailyPass.client_id.isnot(None),
                            DailyPass.gym_id != "1"
                        )

                        if gym_id:
                            daily_pass_records = daily_pass_records.filter(DailyPass.gym_id == str(gym_id))

                        daily_pass_records = daily_pass_records.all()

                        client_ids = list(set([str(r.client_id) for r in daily_pass_records if r.client_id]))

                        if client_ids:
                            client_location_stmt = select(Client.client_id, Client.location).where(
                                Client.client_id.in_(client_ids),
                                Client.location.isnot(None),
                                Client.location != ''
                            )
                            client_location_result = await db.execute(client_location_stmt)
                            client_locations = {str(row[0]): row[1] for row in client_location_result.all()}

                            for record in daily_pass_records:
                                if record.client_id and str(record.client_id) in client_locations:
                                    raw_loc = client_locations[str(record.client_id)]
                                    normalized_loc = raw_loc.strip().replace(' ', '_')
                                    if normalized_loc not in location_purchases:
                                        location_purchases[normalized_loc] = 0
                                    location_purchases[normalized_loc] += 1

                    except Exception as e:
                        import logging
                        logging.error(f"Error getting location-wise daily pass purchases: {str(e)}")


                dailypass_session.close()
            except Exception as e:
                # Log error for debugging daily pass purchases
                import logging
                logging.error(f"Purchase analytics - Daily pass error: {str(e)}")
                pass

        # 2. SESSION PURCHASES - Use SessionPurchase (status='paid') + Gym INNER JOIN (matches GMV logic)
        if not source or source == "sessions":
            try:
                # GMV fix: SessionPurchase with status='paid' + INNER JOIN Gym (no orphans)
                # SUM(sessions_count) so multi-class purchases count all classes (not just 1 per transaction)
                total_stmt = (
                    select(
                        func.coalesce(func.sum(SessionPurchase.sessions_count), 0).label('total_purchases'),
                        func.count(distinct(SessionPurchase.client_id)).label('total_unique_users')
                    )
                    .select_from(SessionPurchase)
                    .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
                    .where(
                        SessionPurchase.status == "paid",
                        SessionPurchase.gym_id != 1,
                        func.date(SessionPurchase.created_at) >= start_date_obj,
                        func.date(SessionPurchase.created_at) <= end_date_obj,
                    )
                )

                # Apply gym filter if provided
                if gym_id:
                    total_stmt = total_stmt.where(SessionPurchase.gym_id == gym_id)

                # Apply location filter if provided
                if location_client_ids:
                    total_stmt = total_stmt.where(SessionPurchase.client_id.in_(location_client_ids))

                total_result = await db.execute(total_stmt)
                total_row = total_result.first()

                if total_row:
                    category_breakdown["sessions"]["purchases"] = total_row.total_purchases or 0
                    category_breakdown["sessions"]["unique_users"] = total_row.total_unique_users or 0

                # Then get purchases over time (grouped by date) — same model/filters
                # GMV fix: SUM(sessions_count) per day in time-series
                time_stmt = (
                    select(
                        func.date(SessionPurchase.created_at).label('purchase_date'),
                        func.coalesce(func.sum(SessionPurchase.sessions_count), 0).label('purchase_count')
                    )
                    .select_from(SessionPurchase)
                    .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
                    .where(
                        SessionPurchase.status == "paid",
                        SessionPurchase.gym_id != 1,
                        func.date(SessionPurchase.created_at) >= start_date_obj,
                        func.date(SessionPurchase.created_at) <= end_date_obj,
                    )
                )

                # Apply gym filter if provided
                if gym_id:
                    time_stmt = time_stmt.where(SessionPurchase.gym_id == gym_id)

                # Apply location filter if provided
                if location_client_ids:
                    time_stmt = time_stmt.where(SessionPurchase.client_id.in_(location_client_ids))

                # Group by date and execute
                time_stmt = time_stmt.group_by(func.date(SessionPurchase.created_at))
                time_result = await db.execute(time_stmt)
                session_results = time_result.all()

                if session_results:
                    # Build purchases over time
                    for row in session_results:
                        date_key = row.purchase_date.isoformat() if row.purchase_date else None
                        if date_key:
                            if date_key not in all_purchases_over_time:
                                all_purchases_over_time[date_key] = 0
                            all_purchases_over_time[date_key] += row.purchase_count
                            category_breakdown["sessions"]["purchases_over_time"].append({
                                "date": date_key,
                                "purchases": row.purchase_count
                            })

                # Sort purchases over time by date
                category_breakdown["sessions"]["purchases_over_time"].sort(key=lambda x: x["date"])

                # Get gym-wise purchases — same SessionPurchase model
                if gym_id:
                    gym_count_stmt = (
                        select(func.count().label('purchase_count'))
                        .select_from(SessionPurchase)
                        .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
                        .where(
                            SessionPurchase.status == "paid",
                            func.date(SessionPurchase.created_at) >= start_date_obj,
                            func.date(SessionPurchase.created_at) <= end_date_obj,
                            SessionPurchase.gym_id == gym_id
                        )
                    )

                    if location_client_ids:
                        gym_count_stmt = gym_count_stmt.where(SessionPurchase.client_id.in_(location_client_ids))

                    gym_count_result = await db.execute(gym_count_stmt)
                    gym_count_row = gym_count_result.first()

                    if gym_count_row and gym_count_row.purchase_count:
                        gym_key = int(gym_id)
                        if gym_key not in gym_purchases:
                            gym_purchases[gym_key] = 0
                        gym_purchases[gym_key] += gym_count_row.purchase_count
                        logging.info(f"Sessions for gym {gym_id}: {gym_count_row.purchase_count} purchases")
                else:
                    gym_wise_stmt = (
                        select(
                            SessionPurchase.gym_id,
                            func.count().label('purchase_count')
                        )
                        .select_from(SessionPurchase)
                        .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
                        .where(
                            SessionPurchase.status == "paid",
                            func.date(SessionPurchase.created_at) >= start_date_obj,
                            func.date(SessionPurchase.created_at) <= end_date_obj,
                            SessionPurchase.gym_id.isnot(None),
                            SessionPurchase.gym_id != 1
                        )
                    )

                    if location_client_ids:
                        gym_wise_stmt = gym_wise_stmt.where(SessionPurchase.client_id.in_(location_client_ids))

                    gym_wise_stmt = gym_wise_stmt.group_by(SessionPurchase.gym_id)
                    gym_wise_result = await db.execute(gym_wise_stmt)
                    gym_wise_rows = gym_wise_result.all()

                    for row in gym_wise_rows:
                        if row.gym_id:
                            if row.gym_id not in gym_purchases:
                                gym_purchases[row.gym_id] = 0
                            gym_purchases[row.gym_id] += row.purchase_count

                # Get location-wise purchases (only when no location filter is applied)
                if not location or location == "all":
                    try:
                        session_records_stmt = (
                            select(SessionPurchase.client_id)
                            .select_from(SessionPurchase)
                            .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
                            .where(
                                SessionPurchase.status == "paid",
                                func.date(SessionPurchase.created_at) >= start_date_obj,
                                func.date(SessionPurchase.created_at) <= end_date_obj,
                                SessionPurchase.client_id.isnot(None),
                                SessionPurchase.gym_id != 1
                            )
                        )

                        if gym_id:
                            session_records_stmt = session_records_stmt.where(SessionPurchase.gym_id == gym_id)

                        if location_client_ids:
                            session_records_stmt = session_records_stmt.where(SessionPurchase.client_id.in_(location_client_ids))

                        session_records_result = await db.execute(session_records_stmt)
                        session_records = session_records_result.all()

                        # Get unique client_ids
                        client_ids = list(set([str(row[0]) for row in session_records]))

                        if client_ids:
                            # Query Client table to get locations
                            client_location_stmt = select(Client.client_id, Client.location).where(
                                Client.client_id.in_(client_ids),
                                Client.location.isnot(None),
                                Client.location != ''
                            )
                            client_location_result = await db.execute(client_location_stmt)
                            client_locations = {str(row[0]): row[1] for row in client_location_result.all()}

                            # Count purchases per location (normalize location names)
                            for row in session_records:
                                if row[0] and str(row[0]) in client_locations:
                                    raw_loc = client_locations[str(row[0])]
                                    # Normalize location: trim whitespace and replace spaces with underscores
                                    normalized_loc = raw_loc.strip().replace(' ', '_')
                                    if normalized_loc not in location_purchases:
                                        location_purchases[normalized_loc] = 0
                                    location_purchases[normalized_loc] += 1

                    except Exception as e:
                        import logging
                        logging.error(f"Error getting location-wise session purchases: {str(e)}")

            except Exception:
                pass

        # 3. NUTRITIONIST PLAN (FITTBOT SUBSCRIPTION) PURCHASES
        # GMV fix: added excluded contacts filter to match gmv-summary logic
        # NOTE: Skip when gym filter is applied (not gym-specific purchases)
        EXCLUDED_CONTACTS = ["7373675762", "9486987082", "8667458723", "9840633149"]
        if (not source or source == "fittbot_subscription") and not gym_id:
            try:
                nutritionist_stmt = (
                    select(
                        func.date(Payment.captured_at).label('purchase_date'),
                        func.count().label('purchase_count'),
                        func.count(distinct(Payment.customer_id)).label('unique_users')
                    )
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .where(Payment.status == "captured")
                    .where(func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_purchase_googleplay')
                    .where(func.date(Payment.captured_at) >= start_date_obj)
                    .where(func.date(Payment.captured_at) <= end_date_obj)
                    .where(~Client.contact.in_(EXCLUDED_CONTACTS))
                )

                if location_client_ids:
                    location_customer_ids = {str(cid) for cid in location_client_ids}
                    nutritionist_stmt = nutritionist_stmt.where(Payment.customer_id.in_(location_customer_ids))

                nutritionist_stmt = nutritionist_stmt.group_by(func.date(Payment.captured_at))

                result = await db.execute(nutritionist_stmt)
                nutritionist_results = result.all()

                # Get total unique users across all dates
                unique_users_stmt = (
                    select(func.count(distinct(Payment.customer_id)))
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .where(Payment.status == "captured")
                    .where(func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_purchase_googleplay')
                    .where(func.date(Payment.captured_at) >= start_date_obj)
                    .where(func.date(Payment.captured_at) <= end_date_obj)
                    .where(~Client.contact.in_(EXCLUDED_CONTACTS))
                )

                if location_client_ids:
                    location_customer_ids = {str(cid) for cid in location_client_ids}
                    unique_users_stmt = unique_users_stmt.where(Payment.customer_id.in_(location_customer_ids))

                unique_result = await db.execute(unique_users_stmt)
                unique_users_count = unique_result.scalar() or 0

                # Calculate totals and build purchases over time
                total_purchases = 0
                for row in nutritionist_results:
                    date_key = row.purchase_date.isoformat() if row.purchase_date else None
                    if date_key:
                        total_purchases += row.purchase_count
                        if date_key not in all_purchases_over_time:
                            all_purchases_over_time[date_key] = 0
                        all_purchases_over_time[date_key] += row.purchase_count
                        category_breakdown["fittbot_subscription"]["purchases_over_time"].append({
                            "date": date_key,
                            "purchases": row.purchase_count
                        })

                category_breakdown["fittbot_subscription"]["purchases"] = total_purchases
                category_breakdown["fittbot_subscription"]["unique_users"] = unique_users_count

            except Exception as e:
                import logging
                logging.error(f"Purchase analytics - Nutritionist Plan error: {str(e)}")
                pass

        # 3.5. AI CREDITS PURCHASES
        # GMV fix: added excluded contacts filter to match gmv-summary logic
        # NOTE: Skip when gym filter is applied (not gym-specific purchases)
        if (not source or source == "ai_credits") and not gym_id:
            try:
                ai_credits_stmt = (
                    select(
                        func.date(Payment.captured_at).label('purchase_date'),
                        func.count().label('purchase_count'),
                        func.count(distinct(Payment.customer_id)).label('unique_users')
                    )
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .where(Payment.status == "captured")
                    .where(func.json_extract(Payment.payment_metadata, '$.flow') == 'food_scanner_credits')
                    .where(func.date(Payment.captured_at) >= start_date_obj)
                    .where(func.date(Payment.captured_at) <= end_date_obj)
                    .where(~Client.contact.in_(EXCLUDED_CONTACTS))
                )

                if location_client_ids:
                    location_customer_ids = {str(cid) for cid in location_client_ids}
                    ai_credits_stmt = ai_credits_stmt.where(Payment.customer_id.in_(location_customer_ids))

                ai_credits_stmt = ai_credits_stmt.group_by(func.date(Payment.captured_at))

                result = await db.execute(ai_credits_stmt)
                ai_credits_results = result.all()

                # Get total unique users across all dates
                ai_credits_unique_users_stmt = (
                    select(func.count(distinct(Payment.customer_id)))
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .where(Payment.status == "captured")
                    .where(func.json_extract(Payment.payment_metadata, '$.flow') == 'food_scanner_credits')
                    .where(func.date(Payment.captured_at) >= start_date_obj)
                    .where(func.date(Payment.captured_at) <= end_date_obj)
                    .where(~Client.contact.in_(EXCLUDED_CONTACTS))
                )

                if location_client_ids:
                    location_customer_ids = {str(cid) for cid in location_client_ids}
                    ai_credits_unique_users_stmt = ai_credits_unique_users_stmt.where(Payment.customer_id.in_(location_customer_ids))

                ai_credits_unique_result = await db.execute(ai_credits_unique_users_stmt)
                ai_credits_unique_users_count = ai_credits_unique_result.scalar() or 0

                # Calculate totals and build purchases over time
                ai_credits_total_purchases = 0
                for row in ai_credits_results:
                    date_key = row.purchase_date.isoformat() if row.purchase_date else None
                    if date_key:
                        ai_credits_total_purchases += row.purchase_count
                        if date_key not in all_purchases_over_time:
                            all_purchases_over_time[date_key] = 0
                        all_purchases_over_time[date_key] += row.purchase_count
                        category_breakdown["ai_credits"]["purchases_over_time"].append({
                            "date": date_key,
                            "purchases": row.purchase_count
                        })

                category_breakdown["ai_credits"]["purchases"] = ai_credits_total_purchases
                category_breakdown["ai_credits"]["unique_users"] = ai_credits_unique_users_count

            except Exception as e:
                import logging
                logging.error(f"Purchase analytics - AI Credits error: {str(e)}")
                pass

        # 4. GYM MEMBERSHIP PURCHASES
        # GMV fix: SQL EXISTS subquery with Gym JOIN + Client JOIN (replaces Python-loop in-memory filtering)
        if not source or source == "gym_membership":
            try:
                gym_meta_cond = or_(
                    func.json_unquote(func.json_extract(Order.order_metadata, "$.audit.source")) == "dailypass_checkout_api",
                    func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "unified_gym_membership_with_sub",
                    func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "unified_gym_membership_with_free_fittbot"
                )

                # EXISTS on OrderItem + Gym JOIN — confirms gym physically exists
                gym_exists_cond = (
                    select(1)
                    .select_from(OrderItem)
                    .join(Gym, Gym.gym_id == cast(OrderItem.gym_id, Integer))
                    .where(
                        OrderItem.order_id == Order.id,
                        OrderItem.gym_id.isnot(None),
                        OrderItem.gym_id != "",
                        OrderItem.gym_id != "1"
                    )
                    .exists()
                )

                gm_base_conditions = [
                    Payment.status == "captured",
                    Order.status == "paid",
                    Order.customer_id.isnot(None),
                    gym_meta_cond,
                    gym_exists_cond,
                    func.date(Payment.captured_at) >= start_date_obj,
                    func.date(Payment.captured_at) <= end_date_obj,
                ]

                if gym_id:
                    gm_base_conditions.append(
                        select(1).select_from(OrderItem)
                        .where(OrderItem.order_id == Order.id, OrderItem.gym_id == str(gym_id))
                        .exists()
                    )

                if location_client_ids:
                    location_customer_ids = {str(cid) for cid in location_client_ids}
                    gm_base_conditions.append(Order.customer_id.in_(location_customer_ids))

                # Deduped subquery for counts
                gm_subq = (
                    select(
                        Order.id.label("order_id"),
                        Payment.captured_at.label("captured_at"),
                        Order.customer_id.label("customer_id")
                    )
                    .select_from(Payment)
                    .join(Order, Order.id == Payment.order_id)
                    .join(Client, Client.client_id == cast(Order.customer_id, Integer))
                    .where(*gm_base_conditions)
                    .distinct()
                    .subquery()
                )

                gm_total_stmt = select(
                    func.count(gm_subq.c.order_id).label("purchases"),
                    func.count(distinct(gm_subq.c.customer_id)).label("unique_users")
                ).select_from(gm_subq)

                gm_total_result = await db.execute(gm_total_stmt)
                gm_total_row = gm_total_result.one()
                category_breakdown["gym_membership"]["purchases"] = gm_total_row.purchases or 0
                category_breakdown["gym_membership"]["unique_users"] = gm_total_row.unique_users or 0

                # Time-series
                gm_time_stmt = (
                    select(
                        func.date(gm_subq.c.captured_at).label("purchase_date"),
                        func.count().label("purchase_count")
                    )
                    .select_from(gm_subq)
                    .group_by(func.date(gm_subq.c.captured_at))
                )
                gm_time_result = await db.execute(gm_time_stmt)
                for row in gm_time_result.all():
                    date_key = row.purchase_date.isoformat() if row.purchase_date else None
                    if date_key:
                        if date_key not in all_purchases_over_time:
                            all_purchases_over_time[date_key] = 0
                        all_purchases_over_time[date_key] += row.purchase_count
                        category_breakdown["gym_membership"]["purchases_over_time"].append({
                            "date": date_key,
                            "purchases": row.purchase_count
                        })
                category_breakdown["gym_membership"]["purchases_over_time"].sort(key=lambda x: x["date"])

                # Location tracking — re-added: get Client.location for each gym_membership order
                if not location or location == "all":
                    try:
                        gm_customer_ids_stmt = select(gm_subq.c.customer_id).select_from(gm_subq)
                        gm_customer_ids_result = await db.execute(gm_customer_ids_stmt)
                        gm_customer_ids = list(set([str(row[0]) for row in gm_customer_ids_result.all() if row[0]]))

                        if gm_customer_ids:
                            gm_loc_stmt = select(Client.client_id, Client.location).where(
                                Client.client_id.in_(gm_customer_ids),
                                Client.location.isnot(None),
                                Client.location != ''
                            )
                            gm_loc_result = await db.execute(gm_loc_stmt)
                            gm_client_locations = {str(row[0]): row[1] for row in gm_loc_result.all()}

                            for cid in gm_customer_ids:
                                if cid in gm_client_locations:
                                    normalized_loc = gm_client_locations[cid].strip().replace(' ', '_')
                                    if normalized_loc not in location_purchases:
                                        location_purchases[normalized_loc] = 0
                                    location_purchases[normalized_loc] += 1
                    except Exception as e:
                        import logging
                        logging.error(f"Error getting location-wise gym membership purchases: {str(e)}")

            except Exception as e:
                import logging
                logging.error(f"Purchase analytics - Gym Membership error: {str(e)}")
                pass

        # Convert all purchases over time to sorted array
        purchases_over_time = [
            {
                "date": date,
                "purchases": count
            }
            for date, count in sorted(all_purchases_over_time.items())
        ]

        # Calculate total purchases across all categories
        total_purchases = sum(cat_data["purchases"] for cat_data in category_breakdown.values())

        # Debug logging
        import logging
        logging.info(f"Purchase analytics total_purchases: {total_purchases}, gym_purchases: {gym_purchases}")

        # Build gym breakdown
        gym_breakdown = []
        if not gym_id and gym_purchases:
            # Build full gym breakdown when no gym filter is applied
            gym_names = {}
            gym_ids = list(gym_purchases.keys())
            gym_stmt = select(Gym.gym_id, Gym.name).where(Gym.gym_id.in_(gym_ids))
            gym_result = await db.execute(gym_stmt)
            for gym_id_val, gym_name in gym_result.all():
                gym_names[gym_id_val] = gym_name

            gym_breakdown = [
                {
                    "gym_id": gym_id,
                    "gym_name": gym_names.get(gym_id, f"Gym {gym_id}"),
                    "revenue": gym_purchases[gym_id]
                }
                for gym_id in sorted(gym_purchases.keys(), key=lambda x: gym_purchases[x], reverse=True)
            ]
        elif gym_id:
            # When gym filter is applied, include the filtered gym in breakdown
            gym_names = {}
            gym_stmt = select(Gym.gym_id, Gym.name).where(Gym.gym_id == gym_id)
            gym_result = await db.execute(gym_stmt)
            for gym_id_val, gym_name in gym_result.all():
                gym_names[gym_id_val] = gym_name

            # Use the gym_purchases count if available, otherwise 0
            gym_purchases_count = gym_purchases.get(gym_id, 0)

            gym_breakdown = [{
                "gym_id": gym_id,
                "gym_name": gym_names.get(gym_id, f"Gym {gym_id}"),
                "revenue": gym_purchases_count
            }]
            logging.info(f"Gym breakdown for filtered gym {gym_id}: {gym_breakdown}")

        # Build location breakdown
        location_breakdown = []
        if location_purchases:
            # Sort locations by purchase count (descending)
            location_breakdown = [
                {
                    "location": loc,
                    "purchases": count
                }
                for loc, count in sorted(location_purchases.items(), key=lambda x: x[1], reverse=True)
            ]
            logging.info(f"Location breakdown: {location_breakdown}")

        # Build revenue by city breakdown — uses same GMV source logic as compute_gmv_totals()
        # Sums revenue from all 5 sources grouped by Gym.city
        revenue_by_city = []
        try:
            EXCLUDED_CONTACTS_SET = ["7373675762", "9486987082", "8667458723", "9840633149"]
            city_revenue_map = {}  # city -> total_revenue (rupees)

            # ── 1. Daily Pass ────────────────────────────────────────────────────
            try:
                _dp_session = get_dailypass_session()
                dp_city_q = (
                    _dp_session.query(
                        Gym.city.label("city"),
                        func.coalesce(func.sum(Payment.amount_minor / 100.0), 0).label("revenue")
                    )
                    .join(Gym, Gym.gym_id == func.cast(DailyPass.gym_id, Integer), isouter=False)
                    .outerjoin(Payment, DailyPass.payment_id == Payment.provider_payment_id)
                    .filter(
                        DailyPass.gym_id != "1",
                        func.date(DailyPass.created_at) >= start_date_obj,
                        func.date(DailyPass.created_at) <= end_date_obj,
                    )
                    .group_by(Gym.city)
                )
                if gym_id:
                    dp_city_q = dp_city_q.filter(DailyPass.gym_id == str(gym_id))
                for row in dp_city_q.all():
                    city = (row.city or "Unknown").strip()
                    city_revenue_map[city] = city_revenue_map.get(city, 0) + float(row.revenue or 0)
                _dp_session.close()
            except Exception as e:
                logging.error(f"[RevCity] daily_pass error: {e}")

            # ── 2. Sessions (SessionPurchase) ────────────────────────────────────
            try:
                sess_city_stmt = (
                    select(
                        func.coalesce(Gym.city, "Unknown").label("city"),
                        func.coalesce(func.sum(SessionPurchase.payable_rupees), 0).label("revenue")
                    )
                    .select_from(SessionPurchase)
                    .join(Gym, SessionPurchase.gym_id == Gym.gym_id)
                    .where(
                        SessionPurchase.status == "paid",
                        SessionPurchase.gym_id != 1,
                        func.date(SessionPurchase.created_at) >= start_date_obj,
                        func.date(SessionPurchase.created_at) <= end_date_obj,
                    )
                    .group_by(Gym.city)
                )
                if gym_id:
                    sess_city_stmt = sess_city_stmt.where(SessionPurchase.gym_id == gym_id)
                sess_city_result = await db.execute(sess_city_stmt)
                for row in sess_city_result.all():
                    city = (row.city or "Unknown").strip()
                    city_revenue_map[city] = city_revenue_map.get(city, 0) + float(row.revenue or 0)
            except Exception as e:
                logging.error(f"[RevCity] sessions error: {e}")

            # ── 3. Nutrition Plans ───────────────────────────────────────────────
            # No city/gym dimension — assign to "App" bucket
            try:
                nutri_city_stmt = (
                    select(func.coalesce(func.sum(Payment.amount_minor / 100.0), 0).label("revenue"))
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .where(
                        Payment.status == "captured",
                        func.json_extract(Payment.payment_metadata, "$.flow") == "nutrition_purchase_googleplay",
                        func.date(Payment.captured_at) >= start_date_obj,
                        func.date(Payment.captured_at) <= end_date_obj,
                        ~Client.contact.in_(EXCLUDED_CONTACTS_SET),
                    )
                )
                nutri_result = await db.execute(nutri_city_stmt)
                nutri_rev = float(nutri_result.scalar() or 0)
                if nutri_rev > 0:
                    city_revenue_map["App"] = city_revenue_map.get("App", 0) + nutri_rev
            except Exception as e:
                logging.error(f"[RevCity] nutrition error: {e}")

            # ── 4. AI Credits ────────────────────────────────────────────────────
            # No city/gym dimension — assign to "App" bucket
            try:
                ai_city_stmt = (
                    select(func.coalesce(func.sum(Payment.amount_minor / 100.0), 0).label("revenue"))
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .where(
                        Payment.status == "captured",
                        func.json_extract(Payment.payment_metadata, "$.flow") == "food_scanner_credits",
                        func.date(Payment.captured_at) >= start_date_obj,
                        func.date(Payment.captured_at) <= end_date_obj,
                        ~Client.contact.in_(EXCLUDED_CONTACTS_SET),
                    )
                )
                ai_result = await db.execute(ai_city_stmt)
                ai_rev = float(ai_result.scalar() or 0)
                if ai_rev > 0:
                    city_revenue_map["App"] = city_revenue_map.get("App", 0) + ai_rev
            except Exception as e:
                logging.error(f"[RevCity] ai_credits error: {e}")

            # ── 5. Gym Membership ────────────────────────────────────────────────
            try:
                gm_meta_cond = or_(
                    func.json_unquote(func.json_extract(Order.order_metadata, "$.audit.source")) == "dailypass_checkout_api",
                    func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "unified_gym_membership_with_sub",
                    func.json_unquote(func.json_extract(Order.order_metadata, "$.order_info.flow")) == "unified_gym_membership_with_free_fittbot",
                )
                gm_exists = (
                    select(1)
                    .select_from(OrderItem)
                    .join(Gym, Gym.gym_id == cast(OrderItem.gym_id, Integer))
                    .where(
                        OrderItem.order_id == Order.id,
                        OrderItem.gym_id.isnot(None),
                        OrderItem.gym_id != "",
                        OrderItem.gym_id != "1",
                    )
                    .exists()
                )
                gm_city_conditions = [
                    Payment.status == "captured",
                    Order.status == "paid",
                    Order.customer_id.isnot(None),
                    gm_meta_cond,
                    gm_exists,
                    func.date(Payment.captured_at) >= start_date_obj,
                    func.date(Payment.captured_at) <= end_date_obj,
                ]
                if gym_id:
                    gm_city_conditions.append(
                        select(1).select_from(OrderItem)
                        .where(OrderItem.order_id == Order.id, OrderItem.gym_id == str(gym_id))
                        .exists()
                    )

                # Join through OrderItem to get the gym, then Gym.city
                gm_city_subq = (
                    select(
                        Order.id.label("order_id"),
                        Order.gross_amount_minor.label("gross_amount_minor"),
                        OrderItem.gym_id.label("item_gym_id"),
                    )
                    .select_from(Payment)
                    .join(Order, Order.id == Payment.order_id)
                    .join(Client, Client.client_id == cast(Order.customer_id, Integer))
                    .join(OrderItem, and_(
                        OrderItem.order_id == Order.id,
                        OrderItem.gym_id.isnot(None),
                        OrderItem.gym_id != "",
                        OrderItem.gym_id != "1",
                    ))
                    .where(*gm_city_conditions)
                    .distinct()
                    .subquery()
                )

                gm_city_stmt = (
                    select(
                        func.coalesce(Gym.city, "Unknown").label("city"),
                        func.coalesce(func.sum(gm_city_subq.c.gross_amount_minor / 100.0), 0).label("revenue"),
                    )
                    .select_from(gm_city_subq)
                    .join(Gym, Gym.gym_id == cast(gm_city_subq.c.item_gym_id, Integer))
                    .group_by(Gym.city)
                )
                gm_city_result = await db.execute(gm_city_stmt)
                for row in gm_city_result.all():
                    city = (row.city or "Unknown").strip()
                    city_revenue_map[city] = city_revenue_map.get(city, 0) + float(row.revenue or 0)
            except Exception as e:
                logging.error(f"[RevCity] gym_membership error: {e}")

            # Build sorted output
            revenue_by_city = sorted(
                [{"city": city, "amount": round(amt, 2)} for city, amt in city_revenue_map.items() if city],
                key=lambda x: x["amount"],
                reverse=True
            )[:20]

            logging.info(f"Final revenue_by_city: {revenue_by_city}")

        except Exception as e:
            logging.error(f"Error building revenue_by_city: {str(e)}")
            import traceback
            traceback.print_exc()


        analytics_data = {
            "totalPurchases": total_purchases,
            "categoryBreakdown": category_breakdown,
            "purchasesOverTime": purchases_over_time,
            "gymBreakdown": gym_breakdown,
            "locationBreakdown": location_breakdown,
            "revenueByCity": revenue_by_city,
            "filters": {
                "startDate": start_date_obj.isoformat(),
                "endDate": end_date_obj.isoformat(),
                "source": source or "all",
                "gymId": gym_id or "all",
                "location": location or "all"
            }
        }

        return {
            "success": True,
            "data": analytics_data,
            "message": "Purchase analytics fetched successfully"
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/booking-averages")
async def get_booking_averages(
    db: AsyncSession = Depends(get_async_db)
):
    """
    Get average bookings for different time periods with source breakdown.
    Returns monthly average (last 3 months), weekly average (last 3 weeks),
    and daily average (last 3 days), each broken down by source.
    """
    try:
        today = datetime.now().date()

        # Helper function to get purchases by source for a date range
        async def get_purchases_by_source(start_date, end_date):
            """Get purchases from all sources for a date range, broken down by source."""
            result = {
                "daily_pass": 0,
                "sessions": 0,
                "gym_membership": 0,
                "fittbot_subscription": 0,
                "ai_credits": 0
            }

            # 1. Daily Pass purchases
            try:
                dailypass_session = get_dailypass_session()
                dp_count = dailypass_session.query(func.count()).filter(
                    and_(
                        func.date(DailyPass.created_at) >= start_date,
                        func.date(DailyPass.created_at) <= end_date,
                        DailyPass.gym_id != "1"
                    )
                ).scalar()
                result["daily_pass"] = dp_count or 0
                dailypass_session.close()
            except Exception:
                pass

            # 2. Session purchases (via SessionBookingDay)
            try:
                session_count_stmt = select(func.count()).select_from(SessionBookingDay).where(
                    and_(
                        func.date(SessionBookingDay.booking_date) >= start_date,
                        func.date(SessionBookingDay.booking_date) <= end_date,
                        SessionBookingDay.gym_id != 1
                    )
                )
                session_result = await db.execute(session_count_stmt)
                result["sessions"] = session_result.scalar() or 0
            except Exception:
                pass

            # 3. Gym membership purchases (from Order/OrderItem with metadata)
            try:
                gym_membership_start = datetime.combine(start_date, datetime.min.time())
                gym_membership_end = datetime.combine(end_date, datetime.min.time()).replace(hour=23, minute=59, second=59)

                order_ids_subquery = (
                    select(Order.id)
                    .join(Payment, Payment.order_id == Order.id)
                    .where(
                        and_(
                            Payment.status == "captured",
                            Order.status == "paid",
                            Payment.captured_at >= gym_membership_start,
                            Payment.captured_at <= gym_membership_end
                        )
                    )
                )

                orders_result = await db.execute(order_ids_subquery)
                order_ids = [row[0] for row in orders_result.all()]

                if order_ids:
                    order_items_stmt = select(OrderItem).where(
                        and_(
                            OrderItem.order_id.in_(order_ids),
                            OrderItem.gym_id.isnot(None),
                            OrderItem.gym_id != "1"
                        )
                    )
                    order_items_result = await db.execute(order_items_stmt)
                    order_items = order_items_result.scalars().all()

                    for item in order_items:
                        if item.order_id:
                            order_stmt = select(Order).where(Order.id == item.order_id)
                            order_result = await db.execute(order_stmt)
                            order = order_result.scalar_one_or_none()

                            if order and order.order_metadata and isinstance(order.order_metadata, dict):
                                metadata = order.order_metadata
                                condition1 = False
                                if metadata.get("audit") and isinstance(metadata.get("audit"), dict):
                                    if metadata["audit"].get("source") == "dailypass_checkout_api":
                                        condition1 = True

                                condition2 = False
                                if metadata.get("order_info") and isinstance(metadata.get("order_info"), dict):
                                    if metadata["order_info"].get("flow") == "unified_gym_membership_with_sub":
                                        condition2 = True

                                condition3 = False
                                if metadata.get("order_info") and isinstance(metadata.get("order_info"), dict):
                                    if metadata["order_info"].get("flow") == "unified_gym_membership_with_free_fittbot":
                                        condition3 = True

                                if condition1 or condition2 or condition3:
                                    result["gym_membership"] += 1
            except Exception:
                pass

            # 4. Nutritionist Plan (Fittbot subscription) purchases
            # Filter: payment_metadata['flow'] == 'nutrition_purchase_googleplay'
            # Exclude internal/test contacts: 7373675762, 9486987082, 8667458723
            try:
                EXCLUDED_CONTACTS_NUTRI = ["7373675762", "9486987082", "8667458723", "9840633149"]
                subscription_start = datetime.combine(start_date, datetime.min.time())
                subscription_end = datetime.combine(end_date, datetime.min.time()).replace(hour=23, minute=59, second=59)

                nutritionist_stmt = (
                    select(func.count(Payment.id))
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .where(
                        and_(
                            Payment.status == "captured",
                            func.json_extract(Payment.payment_metadata, '$.flow') == 'nutrition_purchase_googleplay',
                            Payment.captured_at >= subscription_start,
                            Payment.captured_at <= subscription_end,
                            ~Client.contact.in_(EXCLUDED_CONTACTS_NUTRI)
                        )
                    )
                )
                query_result = await db.execute(nutritionist_stmt)
                result["fittbot_subscription"] = query_result.scalar() or 0
            except Exception:
                pass

            # 5. AI Credits purchases
            # Filter: payment_metadata['flow'] == 'food_scanner_credits' (exact match)
            # Exclude internal/test contacts: 7373675762, 9486987082, 8667458723
            try:
                EXCLUDED_CONTACTS = ["7373675762", "9486987082", "8667458723", "9840633149"]
                ai_start = datetime.combine(start_date, datetime.min.time())
                ai_end = datetime.combine(end_date, datetime.min.time()).replace(hour=23, minute=59, second=59)

                ai_stmt = (
                    select(func.count(Payment.id))
                    .select_from(Payment)
                    .outerjoin(Client, Payment.customer_id == Client.client_id)
                    .where(
                        and_(
                            Payment.status == "captured",
                            func.json_extract(Payment.payment_metadata, '$.flow') == 'food_scanner_credits',
                            Payment.captured_at >= ai_start,
                            Payment.captured_at <= ai_end,
                            ~Client.contact.in_(EXCLUDED_CONTACTS)
                        )
                    )
                )
                ai_result = await db.execute(ai_stmt)
                result["ai_credits"] = ai_result.scalar() or 0
            except Exception:
                pass

            return result

        # Helper to calculate average from list of source-wise data
        def calculate_source_averages(data_list):
            """Calculate average for each source across data points."""
            if not data_list:
                return {"daily_pass": 0, "sessions": 0, "gym_membership": 0, "fittbot_subscription": 0, "ai_credits": 0}

            num_points = len(data_list)
            return {
                "daily_pass": round(sum(d["daily_pass"] for d in data_list) / num_points, 2),
                "sessions": round(sum(d["sessions"] for d in data_list) / num_points, 2),
                "gym_membership": round(sum(d["gym_membership"] for d in data_list) / num_points, 2),
                "fittbot_subscription": round(sum(d["fittbot_subscription"] for d in data_list) / num_points, 2),
                "ai_credits": round(sum(d["ai_credits"] for d in data_list) / num_points, 2)
            }

        # Calculate Daily Average (last 3 days) with source breakdown
        daily_source_data = []
        for i in range(3):
            day_date = today - timedelta(days=i)
            day_data = await get_purchases_by_source(day_date, day_date)
            daily_source_data.append(day_data)
        daily_average = calculate_source_averages(daily_source_data)

        # Calculate Weekly Average (last 3 FULL weeks) with source breakdown
        weekly_source_data = []
        for i in range(3):
            # Last 3 FULL weeks - exclude current incomplete week
            # Week 1: 7-13 days ago (full week)
            # Week 2: 14-20 days ago (full week)
            # Week 3: 21-27 days ago (full week)
            week_end = today - timedelta(days=(i * 7) + 1)  # Start from yesterday to ensure full week
            week_start = week_end - timedelta(days=6)
            week_data = await get_purchases_by_source(week_start, week_end)
            weekly_source_data.append(week_data)
        weekly_average = calculate_source_averages(weekly_source_data)

        # Calculate Monthly Average (last 3 FULL months) with source breakdown
        monthly_source_data = []
        for i in range(3):
            # Last 3 FULL months - exclude current incomplete month
            # Month 1: Previous full month
            # Month 2: 2 months ago (full month)
            # Month 3: 3 months ago (full month)
            month_date = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
            for _ in range(i):
                month_date = (month_date - timedelta(days=1)).replace(day=1)
            month_start = month_date
            # Get last day of the month
            next_month = month_date.replace(day=28) + timedelta(days=4)
            month_end = next_month - timedelta(days=next_month.day)

            month_data = await get_purchases_by_source(month_start, month_end)
            monthly_source_data.append(month_data)
        monthly_average = calculate_source_averages(monthly_source_data)

        # Calculate totals for display
        daily_total = round(sum(daily_average.values()), 2)
        weekly_total = round(sum(weekly_average.values()), 2)
        monthly_total = round(sum(monthly_average.values()), 2)

        return {
            "success": True,
            "data": {
                "dailyAverage": daily_total,
                "weeklyAverage": weekly_total,
                "monthlyAverage": monthly_total,
                "dailyBreakdown": daily_average,
                "weeklyBreakdown": weekly_average,
                "monthlyBreakdown": monthly_average
            },
            "message": "Booking averages fetched successfully"
        }

    except Exception as e:
        import logging
        logging.error(f"Error in booking-averages: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


